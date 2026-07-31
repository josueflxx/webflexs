"""Generate a standalone Excel preview using the production logo logic.

This is a manual utility, not an automated test. Keeping all work behind
``main()`` prevents Django/unittest discovery from writing an XLSX file merely
by importing this module.
"""


def main():
    import os
    from io import BytesIO

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "flexs_project.settings.local")

    import django
    from django.conf import settings
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from PIL import Image as PILImage

    django.setup()

    logo_path = os.path.join(
        settings.BASE_DIR,
        "core",
        "static",
        "core",
        "img",
        "flexs-logo.png",
    )
    print(f"Logo: {logo_path} (exists={os.path.exists(logo_path)})")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "INDICE"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.sheet_properties.tabColor = "FF6B3A"

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    worksheet["A1"] = "Catalogo FLEXS"
    worksheet["A1"].font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=16)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FF6B3A")
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    worksheet.row_dimensions[1].height = 50

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    worksheet["A2"] = (
        "Lista digital de productos activos, visibles para clientes y con precio publicable."
    )
    worksheet["A2"].font = Font(name="Segoe UI", color="E5E7EB", italic=True)
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="111827")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    worksheet.row_dimensions[2].height = 22

    card_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    labels = ["Productos", "Hojas", "Version", "Generado", "Vigente desde"]
    values = ["1250", "8", "catalogo-20260528", "28/05/2026 14:30", "28/05/2026 14:30"]
    for column_index, (label, value) in enumerate(zip(labels, values), start=1):
        label_cell = worksheet.cell(row=4, column=column_index, value=label)
        value_cell = worksheet.cell(row=5, column=column_index, value=value)
        for cell in (label_cell, value_cell):
            cell.fill = card_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label_cell.font = Font(name="Segoe UI", color="374151", bold=True)
        value_cell.font = Font(name="Segoe UI", color="111827", bold=True, size=14)

    worksheet.column_dimensions["A"].width = 38
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 30
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 35

    print("Inserting logo with RGBA->RGB conversion via BytesIO...")
    try:
        pil_image = PILImage.open(logo_path)
        print(f"  Original mode: {pil_image.mode}, size: {pil_image.size}")

        if pil_image.mode == "RGBA":
            background = PILImage.new("RGB", pil_image.size, (255, 255, 255))
            background.paste(pil_image, mask=pil_image.split()[3])
            pil_image.close()
            pil_image = background
            print("  Converted to RGB")

        logo_buffer = BytesIO()
        pil_image.save(logo_buffer, "PNG")
        pil_image.close()
        logo_buffer.seek(0)
        print(f"  BytesIO size: {logo_buffer.getbuffer().nbytes} bytes")

        logo = Image(logo_buffer)
        logo.width = 240
        logo.height = 50
        worksheet.add_image(logo, "F1")
        print("  Logo added at F1 successfully!")
    except Exception as exc:
        print(f"  ERROR: {exc}")
        import traceback

        traceback.print_exc()

    output_path = os.path.join(settings.BASE_DIR, "test_logo_final.xlsx")
    workbook.save(output_path)
    print(f"\nSaved: {output_path}")
    print("Open this file in Excel to verify the logo appears at F1!")


if __name__ == "__main__":
    main()
