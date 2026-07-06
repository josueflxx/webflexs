"""
Import script for Movigom products and embedded photos from Excel.
Run with: python import_movigom_excel.py [options]
"""
import os
import sys
import argparse
from decimal import Decimal

# Add current project root to python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Initialize Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'flexs_project.settings.local')
import django
django.setup()

from django.core.files.base import ContentFile
from django.utils.text import slugify
from django.db import transaction
import openpyxl

from catalog.models import Category, Product, Supplier


def parse_price(val):
    if val is None:
        return Decimal("0.00")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    
    # Strip spaces and currency symbols
    s = str(val).strip().replace("$", "").replace(" ", "")
    if not s:
        return Decimal("0.00")
        
    # Handle thousands and decimal separators
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            # 1.234,56 -> 1234.56
            s = s.replace(".", "").replace(",", ".")
        else:
            # 1,234.56 -> 1234.56
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
            
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0.00")


def get_or_create_category(name, parent=None):
    name = name.strip()
    if not name:
        return None
        
    cat = Category.objects.filter(name=name, parent=parent).first()
    if cat:
        return cat
        
    # Generate unique slug
    base_slug = slugify(name)
    if not base_slug:
        base_slug = "cat"
        
    slug = base_slug
    counter = 1
    while Category.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1
        
    return Category.objects.create(name=name, parent=parent, slug=slug)


def clean_sku(val):
    if val is None:
        return ""
    return str(val).strip()


def main():
    parser = argparse.ArgumentParser(description="Import Movigom products and embedded images from Excel.")
    parser.add_argument(
        "--file", 
        default=r"C:\Users\Brian\Desktop\Movigom trabajo\fotos\LISTAS_con_PDF_y_FOTOS_2027.xlsx",
        help="Path to the Excel file."
    )
    parser.add_argument(
        "--dry-run", 
        action="store_true", 
        help="Simulate the import without writing to the database or saving files."
    )
    parser.add_argument(
        "--limit", 
        type=int, 
        default=None, 
        help="Limit number of rows processed per sheet (for testing)."
    )
    parser.add_argument(
        "--sheet", 
        default=None, 
        help="Specify a single sheet to import (e.g. 'BUJES')."
    )
    args = parser.parse_args()

    excel_path = args.file
    print(f"Loading workbook from: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    
    # Check sheet names
    sheet_names = wb.sheetnames
    if args.sheet:
        if args.sheet not in sheet_names:
            print(f"Sheet '{args.sheet}' not found in workbook. Available: {sheet_names}")
            sys.exit(1)
        sheets_to_process = [args.sheet]
    else:
        sheets_to_process = sheet_names

    # Ensure Supplier "MOVIGOM S.R.L." exists
    supplier_name = "MOVIGOM S.R.L."
    if not args.dry_run:
        supplier_obj, _ = Supplier.objects.get_or_create(name=supplier_name)
    else:
        supplier_obj = None
        print(f"[Dry-run] Would ensure Supplier '{supplier_name}' exists.")

    total_created = 0
    total_updated = 0
    total_skipped = 0
    total_images_saved = 0

    for sheet_name in sheets_to_process:
        sheet = wb[sheet_name]
        print(f"\nProcessing Sheet: '{sheet_name}' (rows: {sheet.max_row})")
        
        # Build map of row -> image object (only for Column H, which is index 7)
        row_images = {}
        for img in getattr(sheet, '_images', []):
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                r = img.anchor._from.row + 1
                c = img.anchor._from.col
                if c == 7: # Column H
                    row_images[r] = img

        print(f"Found {len(row_images)} embedded images in Column H of sheet '{sheet_name}'.")

        # Keep track of active subheader (e.g. "▸ AGRALE" -> "AGRALE")
        active_subheader = ""
        
        # Read from row 2
        processed_in_sheet = 0
        for r in range(2, sheet.max_row + 1):
            if args.limit and processed_in_sheet >= args.limit:
                print(f"Reached limit of {args.limit} rows for sheet '{sheet_name}'.")
                break

            val_a = sheet.cell(row=r, column=1).value
            
            # Check if this is a subheader row
            is_subheader = False
            if isinstance(val_a, str):
                val_a_str = val_a.strip()
                if val_a_str.startswith(('▸', '▪', '•', '▫', '◦', '▪️', '-', '*')):
                    is_subheader = True
                elif len(val_a_str) > 40 and not sheet.cell(row=r, column=3).value:
                    is_subheader = True
            
            if is_subheader:
                clean_header = val_a.strip()
                for b in ('▸', '▪', '•', '▫', '◦', '▪️', '-', '*'):
                    clean_header = clean_header.replace(b, '')
                active_subheader = clean_header.strip()
                continue
                
            # If SKU is completely empty, skip the row
            sku = clean_sku(val_a)
            if not sku:
                continue
                
            # Skip if SKU is too long (> 50 characters) to avoid database errors
            if len(sku) > 50:
                print(f"  Warning: Skipping row {r} because SKU '{sku[:20]}...' exceeds 50 chars.")
                continue

            # This is a product row!
            processed_in_sheet += 1
            
            marca_seccion = sheet.cell(row=r, column=2).value or ""
            descripcion = sheet.cell(row=r, column=3).value or ""
            ref = sheet.cell(row=r, column=4).value or ""
            precio_raw = sheet.cell(row=r, column=5).value
            nombre_pdf = sheet.cell(row=r, column=6).value or ""
            pag_pdf = sheet.cell(row=r, column=7).value or ""

            # Determine product name
            clean_nombre_pdf = str(nombre_pdf).strip()
            if clean_nombre_pdf and clean_nombre_pdf != "— no figura en el PDF —":
                name = clean_nombre_pdf
            else:
                name = str(descripcion).strip()
                
            if not name:
                name = f"Producto {sku}"

            price = parse_price(precio_raw)
            
            # Determine categories
            parent_cat_name = sheet_name
            child_cat_name = str(marca_seccion).strip()
            if not child_cat_name:
                child_cat_name = active_subheader
            
            # Prepare attributes
            attrs = {}
            if ref:
                attrs["referencia_proveedor"] = str(ref).strip()
            if pag_pdf:
                attrs["pagina_pdf"] = str(pag_pdf).strip()

            # Check if there is an image for this row
            img = row_images.get(r)
            img_format = getattr(img, 'format', 'jpeg') if img else None
            
            # Print row detail
            action_desc = "Updating" if Product.objects.filter(sku=sku).exists() else "Creating"
            has_img_str = f"with image ({img_format})" if img else "no image"
            print(f"  [{action_desc}] SKU: {sku} | Name: {name[:40]} | Category: {parent_cat_name} -> {child_cat_name} | Price: {price} | {has_img_str}")

            if not args.dry_run:
                with transaction.atomic():
                    # 1. Categories
                    parent_cat = get_or_create_category(parent_cat_name)
                    child_cat = get_or_create_category(child_cat_name, parent=parent_cat) if child_cat_name else parent_cat

                    # 2. Product
                    product, created = Product.objects.get_or_create(sku=sku, defaults={
                        "name": name,
                        "price": price,
                        "cost": Decimal("0.00"),
                        "supplier": supplier_name,
                        "supplier_ref": supplier_obj,
                        "description": str(descripcion).strip(),
                        "category": child_cat,
                        "attributes": attrs,
                    })

                    if not created:
                        product.name = name
                        product.price = price
                        product.supplier = supplier_name
                        product.supplier_ref = supplier_obj
                        product.description = str(descripcion).strip()
                        product.category = child_cat
                        
                        # Merge attributes
                        merged_attrs = {**(product.attributes or {}), **attrs}
                        product.attributes = merged_attrs
                        total_updated += 1
                    else:
                        total_created += 1

                    # Ensure child_cat is in categories M2M
                    product.categories.add(child_cat)
                    if child_cat != parent_cat:
                        product.categories.add(parent_cat)

                    # 3. Handle image save
                    if img:
                        try:
                            # Read raw bytes using img._data()
                            raw_bytes = img._data()
                            filename = f"{sku}.{img_format}"
                            product.image.save(filename, ContentFile(raw_bytes), save=False)
                            total_images_saved += 1
                        except Exception as e:
                            print(f"    Warning: Error saving image for SKU {sku}: {e}")
                            
                    product.save()
            else:
                # Dry run tallies
                if Product.objects.filter(sku=sku).exists():
                    total_updated += 1
                else:
                    total_created += 1
                if img:
                    total_images_saved += 1

    print("\n=========================================")
    print("Import Summary:")
    print("=========================================")
    print(f"Products Created: {total_created}")
    print(f"Products Updated: {total_updated}")
    print(f"Products Skipped: {total_skipped}")
    print(f"Images Extracted & Saved: {total_images_saved}")
    if args.dry_run:
        print("\n*** NOTE: This was a DRY-RUN. No changes were saved to the database. ***")
    print("=========================================")


if __name__ == "__main__":
    main()
