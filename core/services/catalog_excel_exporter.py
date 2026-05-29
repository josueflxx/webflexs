"""Catalog Excel export builder using configurable templates."""

import json
from types import SimpleNamespace
from decimal import Decimal

from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from catalog.models import Category, CategoryProductOrder, Product, PriceListItem
from catalog.services.clamp_measure_parser import (
    clamp_measure_group_label,
    parse_product_clamp_measure,
    sort_clamp_measure_results,
)
from core.models import (
    CATALOG_EXPORT_COLUMN_CHOICES,
    CATALOG_EXPORT_SPECIAL_GROUPING_CLAMP_MEASURE,
)


SORT_MAP = {
    "name_asc": ("name", "id"),
    "name_desc": ("-name", "-id"),
    "sku_asc": ("sku", "id"),
    "sku_desc": ("-sku", "-id"),
    "updated_desc": ("-updated_at", "-id"),
    "price_desc": ("-price", "-id"),
    "price_asc": ("price", "id"),
}

MONEY_KEYS = {"price", "cost"}
INTEGER_KEYS = {"stock"}
STATUS_KEYS = {"is_active", "is_visible_in_catalog"}
DATE_KEYS = {"created_at", "updated_at"}
TEXT_WRAP_KEYS = {"description", "attributes_json", "categories"}
RIGHT_ALIGNED_KEYS = MONEY_KEYS | INTEGER_KEYS
CLAMP_MEASURE_COLUMN_KEYS = [
    "sku",
    "name",
    "price",
]
CLAMP_MEASURE_EXPORT_HEADERS = ["Codigo", "Nombre", "Precio"]
COLUMN_WIDTH_RULES = {
    "sku": (12, 18),
    "name": (34, 58),
    "description": (32, 72),
    "supplier": (18, 32),
    "supplier_normalized": (20, 34),
    "price": (14, 16),
    "cost": (14, 16),
    "stock": (10, 12),
    "is_active": (12, 16),
    "is_visible_in_catalog": (16, 22),
    "primary_category": (22, 34),
    "categories": (28, 58),
    "filter_1": (16, 32),
    "filter_2": (16, 32),
    "filter_3": (16, 32),
    "filter_4": (16, 32),
    "filter_5": (16, 32),
    "created_at": (18, 22),
    "updated_at": (18, 22),
    "attributes_json": (32, 72),
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(name="Segoe UI", color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center")
GROUP_TITLE_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")
GROUP_TITLE_FONT = Font(name="Segoe UI", color="92400E", bold=True, size=12)
GROUP_PRIMARY_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
GROUP_PRIMARY_FONT = Font(name="Segoe UI", color="1D4ED8", bold=True, size=12)
GROUP_SUBCATEGORY_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")
GROUP_SUBCATEGORY_FONT = Font(name="Segoe UI", color="92400E", bold=True, size=12)
GROUP_NESTED_FILL = PatternFill(fill_type="solid", fgColor="DCFCE7")
GROUP_NESTED_FONT = Font(name="Segoe UI", color="166534", bold=True, size=12)
GROUP_DEEP_FILL = PatternFill(fill_type="solid", fgColor="EDE9FE")
GROUP_DEEP_FONT = Font(name="Segoe UI", color="5B21B6", bold=True, size=12)
GROUP_LEVEL_STYLES = (
    (GROUP_PRIMARY_FILL, GROUP_PRIMARY_FONT, 0),
    (GROUP_SUBCATEGORY_FILL, GROUP_SUBCATEGORY_FONT, 1),
    (GROUP_NESTED_FILL, GROUP_NESTED_FONT, 2),
    (GROUP_DEEP_FILL, GROUP_DEEP_FONT, 3),
)
GROUP_REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FEE2E2")
GROUP_REVIEW_FONT = Font(name="Segoe UI", color="991B1B", bold=True, size=12)
GROUP_TITLE_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="medium", color="94A3B8"),
    bottom=Side(style="thin", color="CBD5E1"),
)
GROUP_BACKLINK_FONT = Font(name="Segoe UI", color="2563EB", bold=True, underline="single")
GROUP_BACKLINK_ALIGNMENT = Alignment(horizontal="right", vertical="center")
GROUP_TITLE_HEIGHT = 24
INDEX_SHEET_TITLE = "INDICE"
INDEX_TITLE_FILL = PatternFill(fill_type="solid", fgColor="FF6B3A")
INDEX_TITLE_FONT = Font(name="Segoe UI", color="FFFFFF", bold=True, size=16)
INDEX_SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="111827")
INDEX_SUBTITLE_FONT = Font(name="Segoe UI", color="E5E7EB", italic=True)
INDEX_CARD_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
INDEX_CARD_VALUE_FONT = Font(name="Segoe UI", color="111827", bold=True, size=14)
INDEX_CARD_ACCENT_FONT = Font(name="Segoe UI", color="C2410C", bold=True, size=14)
INDEX_MUTED_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")
INDEX_MUTED_FONT = Font(name="Segoe UI", color="374151", bold=True)
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
STATUS_OK_FILL = PatternFill(fill_type="solid", fgColor="D1FAE5")
STATUS_BAD_FILL = PatternFill(fill_type="solid", fgColor="FEE2E2")
CLAMP_REVIEW_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFFBEB")
SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
SEPARATOR_BORDER = Border(
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
SEPARATOR_ROW_HEIGHT = 28
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
STANDARD_FONT = Font(name="Segoe UI", size=10)


def _prepare_worksheet(worksheet, tab_color=None):
    worksheet.sheet_view.showGridLines = True
    worksheet.sheet_view.zoomScale = 90
    if tab_color:
        worksheet.sheet_properties.tabColor = tab_color


def _decimal_to_excel(value):
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return float(value)
    return value


def _yes_no(value):
    return "Si" if value else "No"


def _resolve_product_export_price(product, price_map=None, discount_percentage=None):
    base_price = product.price
    if price_map is not None:
        base_price = price_map.get(product.id, product.price)
    if base_price is None:
        return Decimal("0")
    if not isinstance(base_price, Decimal):
        base_price = Decimal(str(base_price or 0))
    if discount_percentage is not None and discount_percentage != 0:
        discount_decimal = Decimal(str(discount_percentage)) / Decimal("100")
        base_price = base_price * (Decimal("1") - discount_decimal)
    return base_price


def _product_has_public_export_price(product, price_map=None, discount_percentage=None):
    return _resolve_product_export_price(
        product,
        price_map=price_map,
        discount_percentage=discount_percentage,
    ) > 0


def _serialize_product_value(product, key, price_map=None, discount_percentage=None):
    if key == "sku":
        return product.sku
    if key == "name":
        return product.name
    if key == "description":
        return product.description or ""
    if key == "supplier":
        return product.supplier or ""
    if key == "supplier_normalized":
        return product.supplier_ref.name if product.supplier_ref_id else ""
    if key == "price":
        return _decimal_to_excel(
            _resolve_product_export_price(
                product,
                price_map=price_map,
                discount_percentage=discount_percentage,
            )
        )
    if key == "cost":
        return _decimal_to_excel(product.cost)
    if key == "stock":
        return product.stock
    if key == "is_active":
        return _yes_no(product.is_active)
    if key == "is_visible_in_catalog":
        return _yes_no(product.is_visible_in_catalog())
    if key == "primary_category":
        primary = product.get_primary_category()
        return primary.name if primary else ""
    if key == "categories":
        linked = product.get_linked_categories()
        if not linked:
            return ""
        return " | ".join(sorted({cat.name for cat in linked}))
    if key == "filter_1":
        return product.filter_1 or ""
    if key == "filter_2":
        return product.filter_2 or ""
    if key == "filter_3":
        return product.filter_3 or ""
    if key == "filter_4":
        return product.filter_4 or ""
    if key == "filter_5":
        return product.filter_5 or ""
    if key == "created_at":
        if not product.created_at:
            return ""
        return timezone.localtime(product.created_at).replace(tzinfo=None)
    if key == "updated_at":
        if not product.updated_at:
            return ""
        return timezone.localtime(product.updated_at).replace(tzinfo=None)
    if key == "attributes_json":
        return json.dumps(product.attributes or {}, ensure_ascii=False)
    return ""


def _is_public_category(category, category_lookup=None):
    current = category
    seen = set()
    while current and current.pk not in seen:
        if not current.is_active or not current.visible_in_catalog:
            return False
        seen.add(current.pk)
        if category_lookup is None:
            current = current.parent
        else:
            current = category_lookup.get(current.parent_id)
    return bool(category)


def _selected_category_ids(sheet):
    return set(sheet.categories.values_list("id", flat=True))


def _selected_supplier_ids(sheet):
    return set(sheet.suppliers.values_list("id", flat=True))


def _sheet_search_query(sheet):
    return (getattr(sheet, "search_query", "") or "").strip()


def _sheet_requires_public_catalog(sheet):
    template = getattr(sheet, "template", None)
    return bool(sheet.only_catalog_visible or (template and template.is_client_download_enabled))


def _sheet_requires_public_price(sheet):
    template = getattr(sheet, "template", None)
    return bool(template and template.is_client_download_enabled)


def _sheet_has_explicit_scope(sheet):
    return bool(
        _selected_category_ids(sheet)
        or _selected_supplier_ids(sheet)
        or _sheet_search_query(sheet)
    )


def _sheet_should_export(sheet):
    selected_category_ids = _selected_category_ids(sheet)
    if _sheet_requires_public_catalog(sheet):
        template = getattr(sheet, "template", None)
        if template and template.is_client_download_enabled and not _sheet_has_explicit_scope(sheet):
            return False
        if selected_category_ids:
            return bool(_resolve_category_ids(sheet, selected_ids=selected_category_ids))
    return True


def _public_category_ids_with_products():
    all_categories = {
        category.id: category
        for category in Category.objects.select_related("parent")
    }
    visible_categories = {
        category.id: category
        for category in all_categories.values()
        if category.is_active and category.visible_in_catalog
    }
    if not visible_categories:
        return set()

    products = Product.catalog_visible(Product.objects.all(), include_uncategorized=False)
    linked_ids = set(
        products.exclude(category_id__isnull=True).values_list("category_id", flat=True)
    )
    linked_ids.update(
        products.exclude(categories__id__isnull=True).values_list("categories__id", flat=True)
    )

    public_ids = set()
    for category_id in linked_ids:
        node = visible_categories.get(category_id)
        if not node:
            continue
        chain = []
        cursor = node
        is_public_path = True
        while cursor:
            if not cursor.is_active or not cursor.visible_in_catalog:
                is_public_path = False
                break
            chain.append(cursor)
            cursor = all_categories.get(cursor.parent_id)
        if is_public_path:
            public_ids.update(category.id for category in chain)
    return public_ids


def _category_public_sort_key(category):
    return (
        category.public_order,
        category.order,
        (category.display_name or category.name or "").lower(),
        category.id,
    )


def _public_root_for_category(category, category_lookup):
    if category is None:
        return None

    current = category
    root = category
    seen = set()
    while current and current.pk not in seen:
        seen.add(current.pk)
        root = current
        parent = category_lookup.get(current.parent_id)
        if parent is None:
            break
        current = parent
    return root


class _ListRelationAdapter:
    def __init__(self, items):
        self._items = list(items)

    def values_list(self, field_name, flat=False):
        values = [getattr(item, field_name) for item in self._items]
        if flat:
            return values
        return [(value,) for value in values]

    def all(self):
        return list(self._items)


class _ColumnRelationAdapter:
    def __init__(self, items):
        self._items = list(items)

    def filter(self, **kwargs):
        items = self._items
        for field_name, expected in kwargs.items():
            items = [
                item
                for item in items
                if getattr(item, field_name, None) == expected
            ]
        return _ColumnRelationAdapter(items)

    def order_by(self, *fields):
        items = list(self._items)
        for field in reversed(fields):
            reverse = field.startswith("-")
            field_name = field[1:] if reverse else field
            items.sort(
                key=lambda item: getattr(item, field_name, None),
                reverse=reverse,
            )
        return _ColumnRelationAdapter(items)

    def __iter__(self):
        return iter(self._items)


def _build_synthetic_category_sheet(template, category, columns, order, name=None):
    return SimpleNamespace(
        template=template,
        name=name or category.name[:31] or "Categoria",
        order=order,
        include_header=True,
        only_active_products=True,
        only_catalog_visible=True,
        include_descendant_categories=True,
        group_by_subcategories=True,
        special_grouping="",
        search_query="",
        max_rows=None,
        sort_by="name_asc",
        categories=_ListRelationAdapter([category]),
        suppliers=_ListRelationAdapter([]),
        columns=_ColumnRelationAdapter(columns),
        _synthetic=True,
    )


def _complete_public_catalog_sheets(template, sheets):
    if not template.is_client_download_enabled:
        return sheets

    public_ids = _public_category_ids_with_products()
    if not public_ids:
        return sheets

    public_categories = {
        category.id: category
        for category in Category.objects.filter(
            id__in=public_ids,
            is_active=True,
            visible_in_catalog=True,
        ).select_related("parent")
    }
    public_roots = [
        category
        for category in public_categories.values()
        if category.parent_id not in public_categories
    ]
    public_roots.sort(key=_category_public_sort_key)

    sheet_root_map = {}

    def sheet_primary_public_root(sheet):
        sheet_key = id(sheet)
        if sheet_key in sheet_root_map:
            return sheet_root_map[sheet_key]

        roots = []
        for category_id in _selected_category_ids(sheet):
            root = _public_root_for_category(public_categories.get(category_id), public_categories)
            if root:
                roots.append(root)

        roots.sort(key=_category_public_sort_key)
        sheet_root_map[sheet_key] = roots[0] if roots else None
        return sheet_root_map[sheet_key]

    selected_root_ids = {
        root.id
        for sheet in sheets
        for root in [sheet_primary_public_root(sheet)]
        if root is not None
    }

    base_columns = []
    for sheet in sheets:
        base_columns = list(sheet.columns.filter(is_active=True).order_by("order", "id"))
        if base_columns:
            break

    completed_sheets = list(sheets)
    next_order = len(completed_sheets) + 1

    for category in public_roots:
        if category.id in selected_root_ids:
            continue
        synthetic_sheet = _build_synthetic_category_sheet(
            template,
            category,
            base_columns,
            next_order,
            name=category.display_name or category.name,
        )
        sheet_root_map[id(synthetic_sheet)] = category
        completed_sheets.append(synthetic_sheet)
        next_order += 1

    completed_sheets.sort(
        key=lambda sheet: (
            sheet_primary_public_root(sheet) is None,
            _category_public_sort_key(sheet_primary_public_root(sheet))
            if sheet_primary_public_root(sheet) is not None
            else (999999, 999999, (sheet.name or "").lower(), getattr(sheet, "id", 0) or 0),
            getattr(sheet, "order", 0) or 0,
            getattr(sheet, "id", 0) or 0,
        )
    )

    deduped_sheets = []
    seen_root_ids = set()
    for sheet in completed_sheets:
        root = sheet_primary_public_root(sheet)
        if root is not None:
            if root.id in seen_root_ids:
                continue
            seen_root_ids.add(root.id)
        deduped_sheets.append(sheet)

    used_names = set()

    def unique_sheet_name(base_name):
        base_name = (str(base_name or "Categoria").strip()[:31] or "Categoria")
        candidate = base_name
        counter = 2
        while candidate in used_names:
            suffix = f" {counter}"
            candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
            counter += 1
        used_names.add(candidate)
        return candidate

    for sheet in deduped_sheets:
        root = sheet_primary_public_root(sheet)
        base_name = (root.display_name if root is not None else sheet.name) or "Categoria"
        sheet._export_name = unique_sheet_name(base_name)

    return deduped_sheets


def _filter_public_category_ids(category_ids):
    if not category_ids:
        return set()
    category_lookup = Category.objects.all().in_bulk()
    return {
        category_id
        for category_id in category_ids
        if _is_public_category(category_lookup.get(category_id), category_lookup=category_lookup)
    }


def _resolve_category_ids(sheet, selected_ids=None):
    """Resolve sheet categories, honoring public visibility for client catalog exports."""
    category_ids = selected_ids if selected_ids is not None else _selected_category_ids(sheet)
    if not category_ids:
        return set()

    if not sheet.include_descendant_categories:
        resolved_ids = category_ids
    else:
        categories = Category.objects.filter(id__in=category_ids)
        resolved_ids = set()
        for category in categories:
            resolved_ids.update(category.get_descendant_ids(include_self=True))

    if _sheet_requires_public_catalog(sheet):
        return _filter_public_category_ids(resolved_ids)
    return resolved_ids


def _apply_sheet_filters(sheet):
    queryset = Product.objects.select_related("category", "supplier_ref").prefetch_related("categories").all()

    if sheet.only_active_products:
        queryset = queryset.filter(is_active=True)

    if _sheet_requires_public_catalog(sheet):
        queryset = Product.catalog_visible(queryset=queryset, include_uncategorized=False)

    selected_category_ids = _selected_category_ids(sheet)
    category_ids = _resolve_category_ids(sheet, selected_ids=selected_category_ids)
    if selected_category_ids and not category_ids:
        return queryset.none()
    if category_ids:
        queryset = queryset.filter(
            Q(category_id__in=category_ids) | Q(categories__id__in=category_ids)
        ).distinct()

    supplier_ids = list(sheet.suppliers.values_list("id", flat=True))
    if supplier_ids:
        queryset = queryset.filter(supplier_ref_id__in=supplier_ids)

    search_query = (sheet.search_query or "").strip()
    if search_query:
        queryset = queryset.filter(
            Q(sku__icontains=search_query)
            | Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(supplier__icontains=search_query)
            | Q(filter_1__icontains=search_query)
            | Q(filter_2__icontains=search_query)
            | Q(filter_3__icontains=search_query)
            | Q(filter_4__icontains=search_query)
            | Q(filter_5__icontains=search_query)
        )

    order_by = SORT_MAP.get(sheet.sort_by) or SORT_MAP["name_asc"]
    queryset = queryset.order_by(*order_by)

    if sheet.max_rows:
        queryset = queryset[: sheet.max_rows]

    return queryset


def _string_len_for_width(value):
    if value is None:
        return 0
    if isinstance(value, float):
        return len(f"{value:,.2f}")
    if isinstance(value, (int, Decimal)):
        return len(f"{value:,}")
    if hasattr(value, "strftime"):
        return 16
    return len(str(value))


def _apply_header_styles(worksheet, total_columns, row=1, column_keys=None):
    worksheet.row_dimensions[row].height = 22
    for col_idx in range(1, total_columns + 1):
        cell = worksheet.cell(row=row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        key = column_keys[col_idx - 1] if column_keys and col_idx <= len(column_keys) else ""
        if key in RIGHT_ALIGNED_KEYS:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_row_styles(worksheet, excel_row, row_values, column_keys):
    worksheet.row_dimensions[excel_row].height = 20
    for col_idx, value in enumerate(row_values, start=1):
        key = column_keys[col_idx - 1]
        cell = worksheet.cell(row=excel_row, column=col_idx)
        cell.font = STANDARD_FONT

        if excel_row % 2 == 0:
            cell.fill = ALT_ROW_FILL

        cell.border = THIN_BORDER

        if key in MONEY_KEYS:
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif key in INTEGER_KEYS:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif key in DATE_KEYS:
            cell.number_format = "yyyy-mm-dd hh:mm"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif key in STATUS_KEYS:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if str(value).strip().lower() == "si":
                cell.fill = STATUS_OK_FILL
            elif str(value).strip().lower() == "no":
                cell.fill = STATUS_BAD_FILL
        else:
            wrap = key in TEXT_WRAP_KEYS
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _set_auto_column_widths(worksheet, column_widths, column_keys=None):
    for idx, width in enumerate(column_widths, start=1):
        key = column_keys[idx - 1] if column_keys and idx <= len(column_keys) else None
        min_width, max_width = COLUMN_WIDTH_RULES.get(key, (10, 72))
        adjusted = max(min_width, min(width + 2, max_width))
        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted


def _safe_sheet_link(sheet_name, cell_reference="A1"):
    escaped = str(sheet_name).replace("'", "''")
    return f"#'{escaped}'!{cell_reference}"


def _append_index_sheet(workbook, template, stats, generated_at):
    from django.conf import settings
    from openpyxl.drawing.image import Image
    import os

    title = INDEX_SHEET_TITLE
    if title in workbook.sheetnames:
        counter = 2
        while f"{title} {counter}" in workbook.sheetnames:
            counter += 1
        title = f"{title} {counter}"

    worksheet = workbook.create_sheet(title, 0)
    _prepare_worksheet(worksheet, "FF6B3A")
    worksheet.sheet_view.showGridLines = False

    local_generated_at = timezone.localtime(generated_at).replace(tzinfo=None)
    version = local_generated_at.strftime("catalogo-%Y%m%d-%H%M%S")
    valid_from_label = local_generated_at.strftime("%d/%m/%Y %H:%M")
    rows_by_sheet = stats.get("rows_by_sheet", {})

    template_label = (template.name or "").strip() or "General"
    title_text = template_label if template_label.lower().startswith("catalogo") else f"Catalogo {template_label}"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    worksheet["A1"] = title_text
    worksheet["A1"].font = INDEX_TITLE_FONT
    worksheet["A1"].fill = INDEX_TITLE_FILL
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    worksheet.row_dimensions[1].height = 50

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    worksheet["A2"] = (
        f"Lista digital vigente desde {valid_from_label}: productos activos, visibles para clientes y con precio publicable."
        if template.is_client_download_enabled
        else f"Exportacion vigente desde {valid_from_label}, generada segun la configuracion de la plantilla."
    )
    worksheet["A2"].font = INDEX_SUBTITLE_FONT
    worksheet["A2"].fill = INDEX_SUBTITLE_FILL
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    worksheet.row_dimensions[2].height = 22

    stat_cards = [
        ("Productos", stats.get("total_rows", 0), INDEX_CARD_ACCENT_FONT),
        ("Hojas", len(rows_by_sheet), INDEX_CARD_VALUE_FONT),
        ("Version", version, INDEX_CARD_VALUE_FONT),
        ("Generado", local_generated_at.strftime("%d/%m/%Y %H:%M"), INDEX_CARD_VALUE_FONT),
        ("Vigente desde", valid_from_label, INDEX_CARD_ACCENT_FONT),
    ]
    for col_idx, (label, value, value_font) in enumerate(stat_cards, start=1):
        label_cell = worksheet.cell(row=4, column=col_idx, value=label)
        value_cell = worksheet.cell(row=5, column=col_idx, value=value)
        for cell in (label_cell, value_cell):
            cell.fill = INDEX_CARD_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label_cell.font = INDEX_MUTED_FONT
        value_cell.font = value_font
    worksheet.row_dimensions[4].height = 20
    worksheet.row_dimensions[5].height = 26

    header_row = 7
    headers = ["Hoja", "Productos", "Vigencia", "Abrir"]
    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=col_idx, value=header)
    _apply_header_styles(worksheet, len(headers), row=header_row)

    row_idx = header_row + 1
    for sheet_name, row_count in rows_by_sheet.items():
        worksheet.cell(row=row_idx, column=1, value=sheet_name)
        worksheet.cell(row=row_idx, column=2, value=row_count)
        worksheet.cell(row=row_idx, column=3, value=valid_from_label)
        link_cell = worksheet.cell(row=row_idx, column=4, value=f"Ir a {sheet_name}")
        link_cell.hyperlink = _safe_sheet_link(sheet_name)
        link_cell.style = "Hyperlink"
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row_idx].height = 22
        row_idx += 1

    worksheet.freeze_panes = "A8"
    worksheet.column_dimensions["A"].width = 38
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 30
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 35

    # Insert brand logo AFTER all cells/merges are set up to avoid
    # conflicts with merged cell ranges. Convert RGBA PNG to RGB
    # (white background) so Excel renders it reliably.
    logo_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'core', 'img', 'flexs-logo.png')
    if os.path.exists(logo_path):
        try:
            from io import BytesIO
            from PIL import Image as PILImage

            pil_img = PILImage.open(logo_path)
            if pil_img.mode == 'RGBA':
                background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                background.paste(pil_img, mask=pil_img.split()[3])
                pil_img.close()
                pil_img = background

            logo_buffer = BytesIO()
            pil_img.save(logo_buffer, 'PNG')
            pil_img.close()
            logo_buffer.seek(0)

            img = Image(logo_buffer)
            img.width = 240
            img.height = 50
            worksheet.add_image(img, 'F1')
        except Exception:
            pass


def _worksheet_next_row(worksheet):
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        return 1
    return worksheet.max_row + 1


def _append_separator_row(worksheet, total_columns, height=SEPARATOR_ROW_HEIGHT):
    worksheet.append([" "] + [""] * max(total_columns - 1, 0))
    row_index = worksheet.max_row
    worksheet.row_dimensions[row_index].height = height
    for col_idx in range(1, total_columns + 1):
        cell = worksheet.cell(row=row_index, column=col_idx)
        cell.fill = SEPARATOR_FILL
        cell.border = SEPARATOR_BORDER
    return row_index


def _set_group_outline(worksheet, start_row, end_row, level=1):
    if end_row < start_row:
        return
    worksheet.sheet_properties.outlinePr.summaryBelow = False
    worksheet.sheet_properties.outlinePr.applyStyles = True
    outline_level = min(max(int(level), 1), 7)
    for row_index in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_index].outlineLevel = outline_level


def _category_depth(category, category_lookup):
    depth = 0
    parent_id = getattr(category, "parent_id", None)
    seen = set()
    while parent_id and parent_id not in seen:
        seen.add(parent_id)
        parent = category_lookup.get(parent_id)
        if parent is None:
            break
        depth += 1
        parent_id = parent.parent_id
    return depth


def _category_path(category, category_lookup, public=False):
    if category is None:
        return []
    path = []
    current = category
    seen = set()
    while current and current.pk not in seen:
        seen.add(current.pk)
        label = current.display_name if public else current.name
        path.insert(0, label)
        current = category_lookup.get(current.parent_id)
    return path


def _is_descendant_of(category, root_id, category_lookup):
    current = category
    seen = set()
    while current and current.pk not in seen:
        if current.pk == root_id:
            return True
        seen.add(current.pk)
        current = category_lookup.get(current.parent_id)
    return False


def _category_sort_key(category, category_lookup):
    path = []
    current = category
    seen = set()
    while current and current.pk not in seen:
        seen.add(current.pk)
        path.insert(
            0,
            (
                current.public_order,
                current.order,
                (current.display_name or current.name or "").lower(),
                current.pk,
            ),
        )
        current = category_lookup.get(current.parent_id)
    return tuple(path)


def _client_export_uses_canonical_categories(sheet):
    template = getattr(sheet, "template", None)
    return bool(template and template.is_client_download_enabled and _selected_category_ids(sheet))


def _select_canonical_public_category(product, category_lookup):
    linked_categories = [
        category
        for category in product.get_linked_categories()
        if category and _is_public_category(category, category_lookup=category_lookup)
    ]
    if not linked_categories:
        return None

    primary_category = category_lookup.get(product.category_id) if product.category_id else None
    if primary_category and _is_public_category(primary_category, category_lookup=category_lookup):
        same_branch = [
            category
            for category in linked_categories
            if _is_descendant_of(category, primary_category.pk, category_lookup)
        ]
        if same_branch:
            same_branch.sort(
                key=lambda category: (
                    -_category_depth(category, category_lookup),
                    _category_sort_key(category, category_lookup),
                )
            )
            return same_branch[0]
        return primary_category

    linked_categories.sort(
        key=lambda category: (
            -_category_depth(category, category_lookup),
            _category_sort_key(category, category_lookup),
        )
    )
    return linked_categories[0]


def _product_has_public_category(product, category_lookup):
    return any(
        category and _is_public_category(category, category_lookup=category_lookup)
        for category in product.get_linked_categories()
    )


def _iter_sheet_products(sheet, price_map=None, discount_percentage=None):
    queryset = _apply_sheet_filters(sheet)
    requires_public_catalog = _sheet_requires_public_catalog(sheet)
    requires_public_price = _sheet_requires_public_price(sheet)
    if not _client_export_uses_canonical_categories(sheet):
        category_lookup = Category.objects.select_related("parent").in_bulk() if requires_public_catalog else None
        for product in queryset.iterator(chunk_size=500):
            if requires_public_catalog and not _product_has_public_category(product, category_lookup):
                continue
            if requires_public_price and not _product_has_public_export_price(
                product,
                price_map=price_map,
                discount_percentage=discount_percentage,
            ):
                continue
            yield product
        return

    resolved_ids = _resolve_category_ids(sheet, selected_ids=_selected_category_ids(sheet))
    category_lookup = Category.objects.select_related("parent").in_bulk()
    matched_products = []
    matched_product_ids = []
    matched_category_ids = []
    for product in queryset.iterator(chunk_size=500):
        if requires_public_price and not _product_has_public_export_price(
            product,
            price_map=price_map,
            discount_percentage=discount_percentage,
        ):
            continue
        canonical_category = _select_canonical_public_category(product, category_lookup)
        if canonical_category and canonical_category.pk in resolved_ids:
            product._catalog_export_category_id = canonical_category.pk
            matched_products.append(product)
            matched_product_ids.append(product.pk)
            matched_category_ids.append(canonical_category.pk)

    order_map = {
        (category_id, product_id): (block_order, sort_order)
        for category_id, product_id, block_order, sort_order in CategoryProductOrder.objects.filter(
            category_id__in=set(matched_category_ids),
            product_id__in=matched_product_ids,
        ).values_list("category_id", "product_id", "block_order", "sort_order")
    }
    for product in matched_products:
        category_id = getattr(product, "_catalog_export_category_id", None)
        block_order, sort_order = order_map.get((category_id, product.pk), (999999999, 999999999))
        product._catalog_export_block_order = block_order
        product._catalog_export_sort_order = sort_order

    matched_products.sort(
        key=lambda product: (
            getattr(product, "_catalog_export_block_order", 999999999),
            getattr(product, "_catalog_export_sort_order", 999999999),
            (product.name or "").lower(),
            product.sku or "",
            product.pk,
        )
    )
    yield from matched_products


def _build_grouping_context(sheet):
    selected_ids = _selected_category_ids(sheet)
    resolved_ids = _resolve_category_ids(sheet, selected_ids=selected_ids)
    category_lookup = Category.objects.all().in_bulk()

    selected_categories = [
        category_lookup[category_id]
        for category_id in selected_ids
        if category_id in category_lookup
    ]
    selected_categories.sort(key=lambda category: _category_sort_key(category, category_lookup))

    return {
        "selected_ids": selected_ids,
        "resolved_ids": resolved_ids,
        "only_catalog_visible": _sheet_requires_public_catalog(sheet),
        "category_lookup": category_lookup,
        "selected_categories": selected_categories,
    }


def _select_product_group_category(product, grouping_context):
    resolved_ids = grouping_context["resolved_ids"]
    selected_ids = grouping_context["selected_ids"]
    only_catalog_visible = grouping_context["only_catalog_visible"]
    if selected_ids and not resolved_ids:
        return None

    export_category_id = getattr(product, "_catalog_export_category_id", None)
    if export_category_id and (not resolved_ids or export_category_id in resolved_ids):
        export_category = grouping_context["category_lookup"].get(export_category_id)
        if export_category:
            return export_category

    linked_categories = [
        category
        for category in product.get_linked_categories()
        if category
        and (not resolved_ids or category.pk in resolved_ids)
        and (not only_catalog_visible or _is_public_category(category, grouping_context["category_lookup"]))
    ]
    if not linked_categories and not resolved_ids:
        linked_categories = [
            category
            for category in product.get_linked_categories()
            if category and (not only_catalog_visible or _is_public_category(category, grouping_context["category_lookup"]))
        ]
    if not linked_categories:
        return None

    category_lookup = grouping_context["category_lookup"]
    linked_categories.sort(
        key=lambda category: (
            -_category_depth(category, category_lookup),
            _category_sort_key(category, category_lookup),
        )
    )
    return linked_categories[0]


def _category_group_label(category, grouping_context):
    if category is None:
        return "Sin categoria"

    category_lookup = grouping_context["category_lookup"]
    selected_categories = grouping_context["selected_categories"]
    matching_roots = [
        root
        for root in selected_categories
        if _is_descendant_of(category, root.pk, category_lookup)
    ]

    if len(selected_categories) == 1 and matching_roots:
        root = matching_roots[0]
        if category.pk == root.pk:
            return "Categoria principal"
        root_path = _category_path(root, category_lookup, public=True)
        category_path = _category_path(category, category_lookup, public=True)
        relative_path = category_path[len(root_path):]
        return " > ".join(relative_path) if relative_path else category.display_name

    return " > ".join(_category_path(category, category_lookup, public=True)) or category.display_name


def _category_group_level(category, grouping_context):
    if category is None:
        return None

    category_lookup = grouping_context["category_lookup"]
    selected_categories = grouping_context["selected_categories"]
    category_depth = _category_depth(category, category_lookup)
    matching_roots = [
        root
        for root in selected_categories
        if _is_descendant_of(category, root.pk, category_lookup)
    ]
    if matching_roots:
        matching_roots.sort(
            key=lambda root: _category_depth(root, category_lookup),
            reverse=True,
        )
        return max(0, category_depth - _category_depth(matching_roots[0], category_lookup))
    return category_depth


def _group_title_style(label, category, grouping_context):
    if label == "Sin categoria":
        return INDEX_MUTED_FILL, INDEX_MUTED_FONT, 0

    level = _category_group_level(category, grouping_context)
    if level is None:
        level = 1
    if label == "Categoria principal":
        level = 0
    style_index = min(max(int(level), 0), len(GROUP_LEVEL_STYLES) - 1)
    fill, font, indent = GROUP_LEVEL_STYLES[style_index]
    return fill, font, min(max(int(level), indent), 4)


def _format_group_title(label, product_count, level):
    if label == "Categoria principal":
        return f"Categoria principal ({product_count} productos)"
    if label == "Sin categoria":
        return f"Sin categoria ({product_count} productos)"
    if level == 0:
        return f"Categoria: {label} ({product_count} productos)"
    if level and level >= 2:
        level_label = "4+" if level >= 4 else str(level)
        return f"Subcategoria nivel {level_label}: {label} ({product_count} productos)"
    return f"Subcategoria: {label} ({product_count} productos)"


def _append_section_title_row(
    worksheet,
    row_index,
    title,
    total_columns,
    fill,
    font,
    indent_level=0,
):
    title_cell = worksheet.cell(row=row_index, column=1, value=title)
    title_cell.font = font
    title_cell.fill = fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent_level)
    title_cell.border = GROUP_TITLE_BORDER

    title_end_column = total_columns
    link_column = None
    if total_columns >= 3:
        title_end_column = total_columns - 1
        link_column = total_columns

    for col_idx in range(2, title_end_column + 1):
        cell = worksheet.cell(row=row_index, column=col_idx)
        cell.fill = fill
        cell.border = GROUP_TITLE_BORDER

    if title_end_column > 1:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=title_end_column,
        )

    if link_column:
        link_cell = worksheet.cell(row=row_index, column=link_column, value="Indice")
        link_cell.hyperlink = _safe_sheet_link(INDEX_SHEET_TITLE)
        link_cell.font = GROUP_BACKLINK_FONT
        link_cell.fill = fill
        link_cell.alignment = GROUP_BACKLINK_ALIGNMENT
        link_cell.border = GROUP_TITLE_BORDER
    else:
        title_cell.hyperlink = _safe_sheet_link(INDEX_SHEET_TITLE)

    worksheet.row_dimensions[row_index].height = GROUP_TITLE_HEIGHT


def _append_group_title(worksheet, row_index, label, product_count, total_columns, category, grouping_context):
    fill, font, indent_level = _group_title_style(label, category, grouping_context)
    level = _category_group_level(category, grouping_context)
    _append_section_title_row(
        worksheet,
        row_index,
        _format_group_title(label, product_count, level),
        total_columns,
        fill,
        font,
        indent_level=indent_level,
    )


def _sheet_uses_clamp_measure_grouping(sheet):
    return getattr(sheet, "special_grouping", "") == CATALOG_EXPORT_SPECIAL_GROUPING_CLAMP_MEASURE


def _append_clamp_measure_group_title(worksheet, row_index, label, product_count, total_columns):
    if label == "PARA REVISAR":
        title = f"Para revisar ({product_count} productos)"
        fill = GROUP_REVIEW_FILL
        font = GROUP_REVIEW_FONT
    else:
        title = f"Diametro {label} ({product_count} productos)"
        fill = GROUP_PRIMARY_FILL
        font = GROUP_PRIMARY_FONT

    _append_section_title_row(worksheet, row_index, title, total_columns, fill, font)


def _clamp_measure_needs_review(result):
    if result is None:
        return False
    observations = str(getattr(result, "observaciones", "") or "").lower()
    return bool(
        not getattr(result, "diametro", "")
        or "revisar" in observations
        or "falta" in observations
        or "no coincide" in observations
    )


def _apply_clamp_measure_row_styles(worksheet, row_index, result=None):
    worksheet.row_dimensions[row_index].height = 20
    needs_review = _clamp_measure_needs_review(result)
    for col_idx in range(1, len(CLAMP_MEASURE_EXPORT_HEADERS) + 1):
        cell = worksheet.cell(row=row_index, column=col_idx)
        cell.font = STANDARD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=col_idx == 2)
        if row_index % 2 == 0:
            cell.fill = ALT_ROW_FILL
        if needs_review:
            cell.fill = CLAMP_REVIEW_ROW_FILL
        if col_idx == len(CLAMP_MEASURE_EXPORT_HEADERS):
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="top")


def _append_clamp_measure_products(
    worksheet,
    sheet_config,
    price_map,
    discount_percentage,
):
    products = list(
        _iter_sheet_products(
            sheet_config,
            price_map=price_map,
            discount_percentage=discount_percentage,
        )
    )
    
    parsed_results = []
    for product in products:
        price = _resolve_product_export_price(
            product, price_map=price_map, discount_percentage=discount_percentage
        )
        parsed_results.append(
            parse_product_clamp_measure(product, price=_decimal_to_excel(price))
        )
        
    results = sort_clamp_measure_results(parsed_results)
    headers = CLAMP_MEASURE_EXPORT_HEADERS
    total_columns = len(headers)
    column_widths = [16, 52, 15]

    if not results:
        if sheet_config.include_header:
            worksheet.append(headers)
            _apply_header_styles(worksheet, total_columns, column_keys=CLAMP_MEASURE_COLUMN_KEYS)
        _set_auto_column_widths(worksheet, column_widths, CLAMP_MEASURE_COLUMN_KEYS)
        return 0

    counts_by_group = {}
    for result in results:
        label = clamp_measure_group_label(result)
        counts_by_group[label] = counts_by_group.get(label, 0) + 1

    row_count = 0
    current_group = None
    current_group_detail_start = None
    for result in results:
        group_label = clamp_measure_group_label(result)
        if group_label != current_group:
            if current_group is not None:
                _set_group_outline(worksheet, current_group_detail_start, worksheet.max_row)
                _append_separator_row(worksheet, total_columns)
            next_row = _worksheet_next_row(worksheet)
            _append_clamp_measure_group_title(
                worksheet,
                next_row,
                group_label,
                counts_by_group[group_label],
                total_columns,
            )
            current_group = group_label
            current_group_detail_start = next_row + 1

            if sheet_config.include_header:
                worksheet.append(headers)
                _apply_header_styles(
                    worksheet,
                    total_columns,
                    row=worksheet.max_row,
                    column_keys=CLAMP_MEASURE_COLUMN_KEYS,
                )

        row_values = [
            result.codigo_original,
            result.nombre_original,
            result.precio if result.precio is not None else "",
        ]
        worksheet.append(row_values)
        row_index = worksheet.max_row
        _apply_clamp_measure_row_styles(worksheet, row_index, result=result)
        for idx, value in enumerate(row_values):
            column_widths[idx] = max(column_widths[idx], _string_len_for_width(value))
        row_count += 1

    _set_group_outline(worksheet, current_group_detail_start, worksheet.max_row)
    worksheet.freeze_panes = None
    _set_auto_column_widths(worksheet, column_widths, CLAMP_MEASURE_COLUMN_KEYS)
    return row_count


def _append_product_row(
    worksheet,
    excel_row,
    product,
    columns,
    column_keys,
    column_widths,
    price_map,
    discount_percentage,
):
    row_values = [
        _serialize_product_value(
            product,
            col.key,
            price_map=price_map,
            discount_percentage=discount_percentage,
        )
        for col in columns
    ]
    worksheet.append(row_values)
    for idx, value in enumerate(row_values):
        column_widths[idx] = max(column_widths[idx], _string_len_for_width(value))
    _apply_row_styles(worksheet, excel_row, row_values, column_keys)


def _append_grouped_products(
    worksheet,
    sheet_config,
    columns,
    headers,
    column_keys,
    column_widths,
    price_map,
    discount_percentage,
):
    grouping_context = _build_grouping_context(sheet_config)
    grouped_products = {}

    products = list(
        _iter_sheet_products(
            sheet_config,
            price_map=price_map,
            discount_percentage=discount_percentage,
        )
    )
    for product in products:
        group_category = _select_product_group_category(product, grouping_context)
        group_key = group_category.pk if group_category else 0
        if group_key not in grouped_products:
            grouped_products[group_key] = {
                "category": group_category,
                "products": [],
            }
        grouped_products[group_key]["products"].append(product)

    sorted_groups = sorted(
        grouped_products.values(),
        key=lambda group: (
            group["category"] is None,
            _category_sort_key(group["category"], grouping_context["category_lookup"])
            if group["category"] is not None
            else ((999999, 999999, "sin categoria", 0),),
        ),
    )

    if not sorted_groups:
        if sheet_config.include_header:
            worksheet.append(headers)
            _apply_header_styles(worksheet, len(headers), column_keys=column_keys)
        return 0

    row_count = 0
    has_appended_group = False
    for group in sorted_groups:
        products_in_group = group["products"]
        if not products_in_group:
            continue

        if has_appended_group:
            _append_separator_row(worksheet, len(headers))

        title = _category_group_label(group["category"], grouping_context)
        next_row = _worksheet_next_row(worksheet)
        _append_group_title(
            worksheet,
            next_row,
            title,
            len(products_in_group),
            len(headers),
            group["category"],
            grouping_context,
        )
        column_widths[0] = max(column_widths[0], _string_len_for_width(title) + 12)

        if sheet_config.include_header:
            worksheet.append(headers)
            header_row = worksheet.max_row
            _apply_header_styles(worksheet, len(headers), row=header_row, column_keys=column_keys)

        detail_start_row = next_row + 1
        for product in products_in_group:
            excel_row = worksheet.max_row + 1
            _append_product_row(
                worksheet,
                excel_row,
                product,
                columns,
                column_keys,
                column_widths,
                price_map,
                discount_percentage,
            )
            row_count += 1

        group_level = _category_group_level(group["category"], grouping_context)
        _set_group_outline(
            worksheet,
            detail_start_row,
            worksheet.max_row,
            level=(group_level or 0) + 1,
        )
        has_appended_group = True

    return row_count


def build_catalog_workbook(template, price_list=None, discount_percentage=None):
    """
    Build an XLSX workbook from one CatalogExcelTemplate instance.
    Returns (workbook, stats_dict).
    """
    generated_at = timezone.now()
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = template.name
    workbook.properties.creator = "WebFlexs"
    workbook.properties.created = timezone.localtime(generated_at).replace(tzinfo=None)

    rows_by_sheet = {}
    export_columns_map = dict(CATALOG_EXPORT_COLUMN_CHOICES)

    sheets = list(
        template.sheets.prefetch_related("columns", "categories", "suppliers").order_by("order", "id")
    )
    sheets = _complete_public_catalog_sheets(template, sheets)
    if not sheets:
        default_sheet = workbook.create_sheet("Catalogo")
        _prepare_worksheet(default_sheet, "94A3B8")
        default_sheet.append(["Mensaje"])
        default_sheet.append(["La plantilla no tiene hojas configuradas."])
        rows_by_sheet["Catalogo"] = 1
        stats = {"rows_by_sheet": rows_by_sheet, "total_rows": 1, "skipped_sheets": []}
        _append_index_sheet(workbook, template, stats, generated_at)
        return workbook, stats

    price_map = None
    if price_list:
        price_map = {
            product_id: price
            for product_id, price in PriceListItem.objects.filter(price_list=price_list).values_list(
                "product_id", "price"
            )
        }

    exported_any_sheet = False
    skipped_sheets = []

    for sheet_config in sheets:
        if not _sheet_should_export(sheet_config):
            skipped_sheets.append(getattr(sheet_config, "_export_name", None) or sheet_config.name)
            continue

        sheet_name = getattr(sheet_config, "_export_name", None) or sheet_config.name
        worksheet = workbook.create_sheet(sheet_name[:31] or "Hoja")
        _prepare_worksheet(worksheet, "1F2937")
        columns = list(
            sheet_config.columns.filter(is_active=True).order_by("order", "id")
        )
        if not columns:
            fallback_columns = [
                ("sku", "SKU"),
                ("name", "Articulo"),
                ("price", "Precio"),
            ]
            columns = [
                type("TmpColumn", (), {"key": key, "header": header})
                for key, header in fallback_columns
            ]

        headers = [
            (col.header or export_columns_map.get(col.key, col.key))
            for col in columns
        ]
        column_keys = [col.key for col in columns]
        column_widths = [_string_len_for_width(header) for header in headers]

        if _sheet_uses_clamp_measure_grouping(sheet_config):
            row_count = _append_clamp_measure_products(
                worksheet,
                sheet_config,
                price_map,
                discount_percentage,
            )
            if _sheet_requires_public_catalog(sheet_config) and row_count == 0:
                workbook.remove(worksheet)
                skipped_sheets.append(sheet_name)
                continue
            exported_any_sheet = True
            rows_by_sheet[sheet_name] = row_count
            continue

        if sheet_config.group_by_subcategories:
            row_count = _append_grouped_products(
                worksheet,
                sheet_config,
                columns,
                headers,
                column_keys,
                column_widths,
                price_map,
                discount_percentage,
            )
            _set_auto_column_widths(worksheet, column_widths, column_keys)
            if _sheet_requires_public_catalog(sheet_config) and row_count == 0:
                workbook.remove(worksheet)
                skipped_sheets.append(sheet_name)
                continue
            exported_any_sheet = True
            rows_by_sheet[sheet_name] = row_count
            continue

        if sheet_config.include_header:
            worksheet.append(headers)
            worksheet.freeze_panes = "A2"
            _apply_header_styles(worksheet, len(headers), column_keys=column_keys)

        row_count = 0
        for product in _iter_sheet_products(
            sheet_config,
            price_map=price_map,
            discount_percentage=discount_percentage,
        ):
            excel_row = row_count + (1 if sheet_config.include_header else 0) + 1
            _append_product_row(
                worksheet,
                excel_row,
                product,
                columns,
                column_keys,
                column_widths,
                price_map,
                discount_percentage,
            )
            row_count += 1

        if sheet_config.include_header and row_count > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        _set_auto_column_widths(worksheet, column_widths, column_keys)

        if _sheet_requires_public_catalog(sheet_config) and row_count == 0:
            workbook.remove(worksheet)
            skipped_sheets.append(sheet_name)
            continue

        exported_any_sheet = True
        rows_by_sheet[sheet_name] = row_count

    if not exported_any_sheet:
        default_sheet = workbook.create_sheet("Catalogo")
        _prepare_worksheet(default_sheet, "94A3B8")
        default_sheet.append(["Mensaje"])
        default_sheet.append(["No hay categorias visibles para exportar."])
        rows_by_sheet["Catalogo"] = 0

    total_rows = sum(rows_by_sheet.values())
    stats = {
        "rows_by_sheet": rows_by_sheet,
        "total_rows": total_rows,
        "skipped_sheets": skipped_sheets,
    }
    _append_index_sheet(workbook, template, stats, generated_at)
    return workbook, stats


def build_export_filename(template):
    stamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    return f"catalogo_{template.slug}_{stamp}.xlsx"
