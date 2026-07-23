"""Company-scoped API v1 endpoints."""

import csv
import io
import unicodedata
import uuid
from pathlib import Path

from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import connection, transaction
from django.db.models import Count, Q
from django.db.models.functions import Lower
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from rest_framework import generics, permissions, status
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.response import Response
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from accounts.models import ClientProfile
from catalog.models import (
    Category,
    Product,
    Supplier,
    SupplierImportProfile,
    SupplierPriceListBatch,
    SupplierPriceListRow,
)
from catalog.services.supplier_price_lists import (
    MAPPING_FIELDS,
    apply_supplier_price_list,
    generate_supplier_price_list_preview,
    hash_uploaded_file,
    inspect_source_file,
    mapping_uses_coordinates,
    report_rows,
    rollback_supplier_price_list,
    update_row_decisions,
)
from core.api_v1.permissions import (
    HasAdminCapability,
    HasRequiredCapabilityWhenStaff,
    IsStaffUser,
)
from core.api_v1.authentication import BearerTokenAuthentication
from core.api_v1.serializers import (
    CategorySerializer,
    ClientProfileSerializer,
    OrderListSerializer,
    ProductClientSerializer,
    ProductStaffSerializer,
)
from orders.models import Order
from core.models import (
    AdminAuditLog,
    ExternalEditorDraft,
    ExternalEditorJob,
    ExternalEditorSavedView,
    WebhookDelivery,
    WebhookEndpoint,
)
from core.services.authorization import (
    CAP_CHANGE_PRICES,
    CAP_GLOBAL_SEARCH,
    CAP_MANAGE_INTEGRATIONS,
    CAP_MANAGE_ORDERS,
    CAP_MANAGE_PRODUCTS,
    CAP_RUN_IMPORTS,
    capability_required,
    has_capability,
)
from orders.services.workflow import (
    get_allowed_next_statuses_for_user,
    get_order_queue_queryset_for_user,
    get_role_queue_statuses,
    get_user_order_roles,
    resolve_user_order_role,
)
from core.services.company_context import (
    get_active_company,
    get_user_companies,
    user_has_company_access,
)
from core.services.external_editor import (
    MAX_SELECTION_IDS,
    PRICE_FIELDS,
    ExternalEditorConflict,
    apply_editor_product_patch,
    build_editor_product_queryset,
    serialize_editor_product,
)
from core.services.external_editor_jobs import (
    ExternalEditorJobPayloadConflict,
    create_external_editor_job,
    execute_external_editor_job,
    preview_external_editor_job,
    rollback_external_editor_job,
    serialize_external_editor_job,
)


def _parse_bool_param(raw_value):
    if raw_value is None:
        return None
    value = str(raw_value).strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return None


def _resolve_request_company(request):
    user = getattr(request, "user", None)
    requested_company_id = str(
        request.query_params.get("company_id")
        or request.query_params.get("company")
        or ""
    ).strip()
    if requested_company_id:
        if not requested_company_id.isdigit():
            return None
        from core.models import Company

        requested_company = Company.objects.filter(pk=int(requested_company_id), is_active=True).first()
        if (
            user
            and getattr(user, "is_authenticated", False)
            and requested_company
            and user_has_company_access(user, requested_company)
        ):
            return requested_company
        return None

    return get_active_company(request)


class ApiV1BaseListView(generics.ListAPIView):
    """Common list behavior for API v1 endpoints."""

    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_default"


class ApiHealthView(APIView):
    """Lightweight health endpoint for smoke checks and deploy validation."""

    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"

    def get(self, request):
        db_ok = True
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                cursor.fetchone()
        except Exception:
            db_ok = False

        return Response(
            {
                "ok": db_ok,
                "api_version": "v1",
                "feature_api_v1_enabled": bool(getattr(settings, "FEATURE_API_V1_ENABLED", False)),
                "external_editor_enabled": bool(
                    getattr(settings, "FEATURE_EXTERNAL_EDITOR_ENABLED", False)
                ),
                "external_editor_writes": bool(
                    getattr(settings, "FEATURE_EXTERNAL_EDITOR_WRITES", False)
                ),
                "db_ok": db_ok,
            },
            status=200 if db_ok else 503,
        )


def _external_editor_disabled_response():
    return Response(
        {"detail": "El editor externo no esta habilitado en este entorno."},
        status=status.HTTP_404_NOT_FOUND,
    )


def _external_editor_writes_disabled_response():
    return Response(
        {"detail": "La escritura del editor externo esta deshabilitada."},
        status=status.HTTP_503_SERVICE_UNAVAILABLE,
    )


def _validation_error_response(exc):
    details = getattr(exc, "message_dict", None) or getattr(exc, "messages", None) or str(exc)
    return Response({"detail": details}, status=status.HTTP_400_BAD_REQUEST)


def _dispatch_external_editor_job(job, created):
    if created:
        sync_limit = int(getattr(settings, "EXTERNAL_EDITOR_SYNC_LIMIT", 250) or 250)
        if getattr(settings, "FEATURE_BACKGROUND_JOBS_ENABLED", False) or job.total > sync_limit:
            from core.services.background_jobs import dispatch_external_editor_job

            dispatch_external_editor_job(job.pk)
        else:
            job = execute_external_editor_job(job.pk)
    job.refresh_from_db()
    return job


def _serialize_editor_saved_view(saved_view):
    return {
        "id": saved_view.pk,
        "name": saved_view.name,
        "filters": saved_view.filters,
        "createdAt": saved_view.created_at.isoformat(),
        "updatedAt": saved_view.updated_at.isoformat(),
    }


def _serialize_editor_draft(draft):
    return {
        "id": draft.pk,
        "name": draft.name,
        "status": draft.status,
        "changes": draft.changes,
        "changeCount": len(draft.changes or []),
        "publishedJobId": draft.published_job_id,
        "createdAt": draft.created_at.isoformat(),
        "updatedAt": draft.updated_at.isoformat(),
        "publishedAt": draft.published_at.isoformat() if draft.published_at else None,
    }


def _serialize_supplier_price_list_batch(batch):
    return {
        "id": batch.pk,
        "supplierId": batch.supplier_id,
        "supplierName": batch.supplier.name,
        "companyId": batch.company_id,
        "companyName": batch.company.name,
        "profileId": batch.profile_id,
        "fileName": batch.original_filename,
        "fileSize": batch.file_size,
        "sheetName": batch.sheet_name,
        "headerRow": batch.header_row,
        "columnMapping": batch.column_mapping,
        "defaultCurrency": batch.default_currency,
        "pricingMode": batch.pricing_mode,
        "status": batch.status,
        "statusLabel": batch.get_status_display(),
        "summary": batch.summary,
        "error": batch.error_message,
        "createdBy": batch.created_by.username if batch.created_by else "",
        "appliedBy": batch.applied_by.username if batch.applied_by else "",
        "createdAt": batch.created_at.isoformat(),
        "previewedAt": batch.previewed_at.isoformat() if batch.previewed_at else None,
        "appliedAt": batch.applied_at.isoformat() if batch.applied_at else None,
        "rolledBackAt": batch.rolled_back_at.isoformat() if batch.rolled_back_at else None,
        "canApply": batch.status == SupplierPriceListBatch.STATUS_PREVIEWED,
        "canRollback": batch.status == SupplierPriceListBatch.STATUS_APPLIED,
    }


def _serialize_supplier_price_list_row(row):
    return {
        "id": row.pk,
        "rowNumber": row.row_number,
        "rowType": row.row_type,
        "supplierCode": row.supplier_code,
        "description": row.supplier_description,
        "productId": row.matched_product_id,
        "sku": row.matched_product.sku if row.matched_product else "",
        "productName": row.matched_product.name if row.matched_product else "",
        "matchMethod": row.match_method,
        "confidence": row.match_confidence,
        "changeType": row.change_type,
        "changeLabel": row.get_change_type_display(),
        "previousCost": str(row.previous_cost) if row.previous_cost is not None else None,
        "proposedCost": str(row.proposed_cost) if row.proposed_cost is not None else None,
        "proposedFinalCost": (
            str(row.proposed_final_cost) if row.proposed_final_cost is not None else None
        ),
        "difference": str(row.difference_amount) if row.difference_amount is not None else None,
        "differencePercentage": (
            str(row.difference_percentage) if row.difference_percentage is not None else None
        ),
        "currency": row.currency,
        "decision": row.decision,
        "warnings": row.warnings,
        "applied": row.applied,
    }


def _editor_supplier_batch(request, batch_id):
    batch = (
        SupplierPriceListBatch.objects.select_related(
            "supplier", "company", "profile", "created_by", "applied_by", "rolled_back_by"
        )
        .filter(pk=batch_id)
        .first()
    )
    if not batch or not user_has_company_access(request.user, batch.company):
        return None
    return batch


def _supplier_import_permission_response(request, *, changes_prices=False):
    if not has_capability(request.user, CAP_RUN_IMPORTS):
        return Response(
            {"detail": "No tienes permiso para ejecutar importaciones."},
            status=status.HTTP_403_FORBIDDEN,
        )
    if changes_prices and not has_capability(request.user, CAP_CHANGE_PRICES):
        return Response(
            {"detail": "No tienes permiso para modificar costos o precios."},
            status=status.HTTP_403_FORBIDDEN,
        )
    return None


class ExternalEditorBaseView(APIView):
    authentication_classes = [SessionAuthentication, BearerTokenAuthentication, TokenAuthentication]
    permission_classes = [IsStaffUser, HasAdminCapability]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"
    required_capability = CAP_MANAGE_PRODUCTS

    def editor_is_enabled(self):
        return bool(getattr(settings, "FEATURE_EXTERNAL_EDITOR_ENABLED", False))

    def editor_writes_are_enabled(self):
        return bool(getattr(settings, "FEATURE_EXTERNAL_EDITOR_WRITES", False))


class ApiEditorProductListView(ExternalEditorBaseView):
    """Official server-side grid source for the external editor."""

    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()

        queryset = build_editor_product_queryset(request.query_params)
        try:
            page_number = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(1, int(request.query_params.get("pageSize", 50))))
        except (TypeError, ValueError):
            return Response({"detail": "Paginacion invalida."}, status=status.HTTP_400_BAD_REQUEST)

        paginator = Paginator(queryset, page_size)
        page = paginator.get_page(page_number)
        return Response(
            {
                "items": [serialize_editor_product(product) for product in page.object_list],
                "total": paginator.count,
                "totalCount": paginator.count,
                "page": page.number,
                "pageSize": page_size,
                "totalPages": paginator.num_pages,
            }
        )


class ApiEditorProductSelectionIdsView(ExternalEditorBaseView):
    """Resolve a filtered grid selection on the server, not only the visible page."""

    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()

        queryset = build_editor_product_queryset(request.query_params)
        ids = list(queryset.values_list("id", flat=True)[: MAX_SELECTION_IDS + 1])
        if len(ids) > MAX_SELECTION_IDS:
            return Response(
                {"detail": f"La seleccion supera el limite de {MAX_SELECTION_IDS} productos."},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )
        return Response({"ids": ids, "count": len(ids)})


class ApiEditorProductDetailView(ExternalEditorBaseView):
    def get(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        product = build_editor_product_queryset({}).filter(pk=product_id).first()
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        return Response(serialize_editor_product(product))

    def patch(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        if PRICE_FIELDS.intersection(request.data) and not has_capability(request.user, CAP_CHANGE_PRICES):
            return Response(
                {"detail": "No tienes permiso para modificar costos, margenes o precios."},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            with transaction.atomic():
                product = (
                    Product.objects.select_for_update(of=("self",))
                    .select_related("category", "category__parent", "supplier_ref")
                    .prefetch_related("categories")
                    .filter(pk=product_id)
                    .first()
                )
                if not product:
                    return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
                product = apply_editor_product_patch(
                    product=product,
                    payload=request.data,
                    user=request.user,
                )
        except ExternalEditorConflict as exc:
            current = build_editor_product_queryset({}).filter(pk=product_id).first()
            return Response(
                {
                    "detail": str(exc),
                    "code": "product_version_conflict",
                    "current": serialize_editor_product(current) if current else None,
                },
                status=status.HTTP_409_CONFLICT,
            )
        except ValidationError as exc:
            return _validation_error_response(exc)

        return Response(serialize_editor_product(product))

    put = patch


class ApiEditorCategoryListView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        categories = Category.objects.filter(is_active=True).select_related("parent").order_by("order", "name")
        roots = []
        children = {}
        for category in categories:
            item = {
                "id": category.pk,
                "name": category.name,
                "parentId": category.parent_id,
                "parentName": category.parent.name if category.parent_id else "",
                "status": "active",
            }
            if category.parent_id:
                children.setdefault(category.parent_id, []).append(item)
            else:
                roots.append(item)
        for root in roots:
            root["subcategories"] = children.get(root["id"], [])
        return Response(roots)


class ApiEditorSupplierListView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        suppliers = Supplier.objects.filter(is_active=True).order_by("name")
        return Response(
            [
                {"id": supplier.pk, "name": supplier.name, "status": "active"}
                for supplier in suppliers
            ]
        )


class ApiEditorLoginView(ObtainAuthToken):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"

    def post(self, request, *args, **kwargs):
        if not getattr(settings, "FEATURE_EXTERNAL_EDITOR_ENABLED", False):
            return _external_editor_disabled_response()
        serializer = self.serializer_class(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        if not user.is_staff or not has_capability(user, CAP_MANAGE_PRODUCTS):
            return Response(
                {"message": "No tienes permiso para usar el editor externo."},
                status=status.HTTP_403_FORBIDDEN,
            )
        token, _created = Token.objects.get_or_create(user=user)
        return Response({"token": token.key})


class ApiEditorProfileView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        return Response(
            {
                "id": request.user.pk,
                "username": request.user.username,
                "email": request.user.email,
                "role": "Admin" if request.user.is_superuser else "Operador",
                "permissions": {
                    "manageProducts": has_capability(request.user, CAP_MANAGE_PRODUCTS),
                    "changePrices": has_capability(request.user, CAP_CHANGE_PRICES),
                    "publishDrafts": has_capability(request.user, CAP_MANAGE_PRODUCTS),
                    "manageImages": has_capability(request.user, CAP_MANAGE_PRODUCTS),
                },
            }
        )


def _external_editor_payload_changes_prices(payload):
    changes = payload.get("changes") or {}
    if PRICE_FIELDS.intersection(changes):
        return True
    for item in payload.get("items") or []:
        if isinstance(item, dict) and PRICE_FIELDS.intersection(item.get("changes") or {}):
            return True
    price_rule_fields = {"cost", "margin", "saleprice", "price"}
    return any(
        str(rule.get("field") or "").strip().lower() in price_rule_fields
        for rule in (payload.get("rules") or [])
        if isinstance(rule, dict)
    )


class ApiEditorBulkPreviewView(ExternalEditorBaseView):
    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if _external_editor_payload_changes_prices(request.data) and not has_capability(
            request.user,
            CAP_CHANGE_PRICES,
        ):
            return Response(
                {"detail": "No tienes permiso para modificar costos, margenes o precios."},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            return Response(preview_external_editor_job(request.data))
        except ValidationError as exc:
            return _validation_error_response(exc)


class ApiEditorBulkView(ExternalEditorBaseView):
    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        if _external_editor_payload_changes_prices(request.data) and not has_capability(
            request.user,
            CAP_CHANGE_PRICES,
        ):
            return Response(
                {"detail": "No tienes permiso para modificar costos, margenes o precios."},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            job, created = create_external_editor_job(
                payload=request.data,
                user=request.user,
                idempotency_key=request.headers.get("Idempotency-Key"),
            )
            job = _dispatch_external_editor_job(job, created)
        except ExternalEditorJobPayloadConflict as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_409_CONFLICT)
        except ValidationError as exc:
            return _validation_error_response(exc)

        job.refresh_from_db()
        response_status = (
            status.HTTP_202_ACCEPTED
            if job.status in {ExternalEditorJob.STATUS_PENDING, ExternalEditorJob.STATUS_RUNNING}
            else status.HTTP_200_OK
        )
        return Response(serialize_external_editor_job(job), status=response_status)


class ApiEditorJobDetailView(ExternalEditorBaseView):
    def get_job(self, request, job_id):
        queryset = ExternalEditorJob.objects.select_related("created_by", "rolled_back_by")
        if not request.user.is_superuser:
            queryset = queryset.filter(created_by=request.user)
        return queryset.filter(pk=job_id).first()

    def get(self, request, job_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        job = self.get_job(request, job_id)
        if not job:
            return Response({"detail": "Trabajo no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        return Response(serialize_external_editor_job(job))


class ApiEditorJobRollbackView(ApiEditorJobDetailView):
    def post(self, request, job_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        job = self.get_job(request, job_id)
        if not job:
            return Response({"detail": "Trabajo no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        if _external_editor_payload_changes_prices(job.request_payload) and not has_capability(
            request.user,
            CAP_CHANGE_PRICES,
        ):
            return Response(
                {"detail": "No tienes permiso para revertir cambios de precios."},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            job = rollback_external_editor_job(job=job, user=request.user)
        except ValidationError as exc:
            return _validation_error_response(exc)
        return Response(serialize_external_editor_job(job))


class ApiEditorJobListView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        queryset = ExternalEditorJob.objects.select_related("created_by", "rolled_back_by")
        if not request.user.is_superuser:
            queryset = queryset.filter(created_by=request.user)
        try:
            limit = min(100, max(1, int(request.query_params.get("limit", 30))))
        except (TypeError, ValueError):
            limit = 30
        return Response({"items": [serialize_external_editor_job(job, include_items=False) for job in queryset[:limit]]})


class ApiEditorJobRedoView(ApiEditorJobDetailView):
    def post(self, request, job_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        source = self.get_job(request, job_id)
        if not source:
            return Response({"detail": "Trabajo no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        payload = {
            key: value
            for key, value in source.request_payload.items()
            if key != "resolvedProductIds"
        }
        if _external_editor_payload_changes_prices(payload) and not has_capability(request.user, CAP_CHANGE_PRICES):
            return Response({"detail": "No tienes permiso para repetir cambios de precios."}, status=status.HTTP_403_FORBIDDEN)
        try:
            job, created = create_external_editor_job(
                payload=payload,
                user=request.user,
                idempotency_key=request.headers.get("Idempotency-Key") or f"redo-{job_id}-{uuid.uuid4()}",
            )
            job = _dispatch_external_editor_job(job, created)
        except (ValidationError, ExternalEditorJobPayloadConflict) as exc:
            return _validation_error_response(exc)
        return Response(serialize_external_editor_job(job))


def _editor_product_issues(product):
    issues = []
    if not product.sku.strip():
        issues.append("SKU vacio")
    if not product.name.strip():
        issues.append("Nombre vacio")
    if not product.category_id and not product.categories.exists():
        issues.append("Sin categoria")
    if not product.supplier_ref_id and not product.supplier.strip():
        issues.append("Sin proveedor")
    if product.cost <= 0:
        issues.append("Costo en cero")
    if product.price <= 0:
        issues.append("Precio en cero")
    if product.price < product.cost:
        issues.append("Precio menor al costo")
    if product.stock < 0:
        issues.append("Stock negativo")
    return issues


class ApiEditorWorkspaceView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        products = build_editor_product_queryset({})
        duplicate_groups = (
            products.annotate(normalized_name=Lower("name"))
            .values("normalized_name")
            .annotate(count=Count("id"))
            .filter(count__gt=1)
            .count()
        )
        tag_counts = {}
        for attributes in products.values_list("attributes", flat=True):
            if not isinstance(attributes, dict):
                continue
            for tag in attributes.get("editor_tags") or []:
                clean_tag = str(tag).strip()
                if clean_tag:
                    tag_counts[clean_tag] = tag_counts.get(clean_tag, 0) + 1
        return Response(
            {
                "totalProducts": products.count(),
                "activeProducts": products.filter(is_active=True).count(),
                "inactiveProducts": products.filter(is_active=False).count(),
                "uncategorizedProducts": products.filter(category__isnull=True, categories__isnull=True).distinct().count(),
                "withoutSupplierProducts": products.filter(supplier_ref__isnull=True, supplier="").count(),
                "zeroCostProducts": products.filter(cost__lte=0).count(),
                "zeroPriceProducts": products.filter(price__lte=0).count(),
                "outOfStockProducts": products.filter(stock__lte=0).count(),
                "duplicateGroups": duplicate_groups,
                "trashProducts": build_editor_product_queryset({"trash": "true"}).count(),
                "drafts": ExternalEditorDraft.objects.filter(created_by=request.user, status=ExternalEditorDraft.STATUS_DRAFT).count(),
                "pendingJobs": ExternalEditorJob.objects.filter(
                    created_by=request.user,
                    status__in=[ExternalEditorJob.STATUS_PENDING, ExternalEditorJob.STATUS_RUNNING],
                ).count(),
                "tags": [
                    {"name": name, "count": count}
                    for name, count in sorted(tag_counts.items(), key=lambda item: (-item[1], item[0].casefold()))[:30]
                ],
            }
        )


class ApiEditorValidationView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        queryset = build_editor_product_queryset(request.query_params)
        items = []
        invalid_count = 0
        for product in queryset.iterator(chunk_size=500):
            issues = _editor_product_issues(product)
            if issues:
                invalid_count += 1
                if len(items) < 250:
                    items.append({"product": serialize_editor_product(product), "issues": issues})
        return Response({"total": queryset.count(), "invalid": invalid_count, "valid": queryset.count() - invalid_count, "items": items})


class ApiEditorDuplicateView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        products = build_editor_product_queryset({})
        groups = list(
            products.annotate(normalized_name=Lower("name"))
            .values("normalized_name")
            .annotate(count=Count("id"))
            .filter(count__gt=1)
            .order_by("-count", "normalized_name")[:100]
        )
        group_keys = [group["normalized_name"] for group in groups]
        grouped_products = {key: [] for key in group_keys}
        if group_keys:
            matches = products.annotate(normalized_name=Lower("name")).filter(normalized_name__in=group_keys)
            for product in matches:
                grouped_products.setdefault(product.normalized_name, []).append(product)
        payload = []
        for group in groups:
            payload.append(
                {
                    "key": group["normalized_name"],
                    "count": group["count"],
                    "products": [
                        serialize_editor_product(product)
                        for product in grouped_products.get(group["normalized_name"], [])[:25]
                    ],
                }
            )
        return Response({"groups": payload, "totalGroups": len(groups)})


class ApiEditorProductCreateView(ExternalEditorBaseView):
    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        sku = str(request.data.get("internalCode") or request.data.get("sku") or "").strip()
        name = str(request.data.get("name") or "").strip()
        if not sku or not name:
            return Response({"detail": "El SKU y el nombre son obligatorios."}, status=status.HTTP_400_BAD_REQUEST)
        if Product.objects.filter(sku__iexact=sku).exists():
            return Response({"detail": "El SKU ya pertenece a otro producto."}, status=status.HTTP_409_CONFLICT)
        payload = dict(request.data)
        payload.pop("internalCode", None)
        payload.pop("sku", None)
        payload.pop("name", None)
        try:
            with transaction.atomic():
                product = Product.objects.create(sku=sku, name=name, cost=0, price=0, stock=0, is_active=False)
                if payload:
                    product = apply_editor_product_patch(product=product, payload=payload, user=request.user)
        except ValidationError as exc:
            return _validation_error_response(exc)
        AdminAuditLog.objects.create(
            user=request.user,
            action="external_editor_create",
            target_type="product",
            target_id=str(product.pk),
        )
        return Response(serialize_editor_product(product), status=status.HTTP_201_CREATED)


class ApiEditorProductCloneView(ExternalEditorBaseView):
    def post(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        source = Product.objects.prefetch_related("categories").filter(pk=product_id).first()
        if not source:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        requested_sku = str(request.data.get("internalCode") or request.data.get("sku") or "").strip()
        base_sku = requested_sku or f"{source.sku}-COPIA"
        candidate = base_sku[:50]
        counter = 2
        while Product.objects.filter(sku__iexact=candidate).exists():
            suffix = f"-{counter}"
            candidate = f"{base_sku[:50 - len(suffix)]}{suffix}"
            counter += 1
        attributes = dict(source.attributes or {})
        attributes.pop("editor_deleted_at", None)
        attributes.pop("editor_deleted_by", None)
        clone = Product.objects.create(
            sku=candidate,
            name=str(request.data.get("name") or f"{source.name} (copia)").strip(),
            supplier=source.supplier,
            supplier_ref=source.supplier_ref,
            description=source.description,
            cost=source.cost,
            price=source.price,
            stock=source.stock,
            category=source.category,
            image=source.image,
            is_active=False,
            filter_1=source.filter_1,
            filter_2=source.filter_2,
            filter_3=source.filter_3,
            filter_4=source.filter_4,
            filter_5=source.filter_5,
            attributes=attributes,
        )
        clone.categories.set(source.categories.all())
        AdminAuditLog.objects.create(
            user=request.user,
            action="external_editor_clone",
            target_type="product",
            target_id=str(clone.pk),
            details={"source_product_id": source.pk},
        )
        return Response(serialize_editor_product(clone), status=status.HTTP_201_CREATED)


class ApiEditorProductImageView(ExternalEditorBaseView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        product = Product.objects.filter(pk=product_id).first()
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        upload = request.FILES.get("image")
        if not upload:
            return Response({"detail": "Selecciona una imagen."}, status=status.HTTP_400_BAD_REQUEST)
        if upload.size > 8 * 1024 * 1024:
            return Response({"detail": "La imagen no puede superar 8 MB."}, status=status.HTTP_400_BAD_REQUEST)
        if not str(upload.content_type or "").startswith("image/"):
            return Response({"detail": "El archivo debe ser una imagen."}, status=status.HTTP_400_BAD_REQUEST)
        product.image.save(upload.name, upload, save=True)
        product.refresh_from_db()
        return Response(serialize_editor_product(product))

    def delete(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        product = Product.objects.filter(pk=product_id).first()
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        product.image = None
        product.save(update_fields=["image", "updated_at"])
        return Response(serialize_editor_product(product))


class ApiEditorProductTrashView(ExternalEditorBaseView):
    def post(self, request, product_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        product = Product.objects.filter(pk=product_id).first()
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        action = str(request.data.get("action") or "trash").strip().lower()
        attributes = dict(product.attributes or {})
        if action == "restore":
            attributes.pop("editor_deleted_at", None)
            attributes.pop("editor_deleted_by", None)
        elif action == "trash":
            attributes["editor_deleted_at"] = timezone.now().isoformat()
            attributes["editor_deleted_by"] = request.user.username
            product.is_active = False
        else:
            return Response({"detail": "Accion de papelera invalida."}, status=status.HTTP_400_BAD_REQUEST)
        product.attributes = attributes
        product.save(update_fields=["attributes", "is_active", "updated_at"])
        AdminAuditLog.objects.create(
            user=request.user,
            action=f"external_editor_{action}",
            target_type="product",
            target_id=str(product.pk),
        )
        return Response(serialize_editor_product(product))


def _normalize_import_header(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(character for character in normalized if not unicodedata.combining(character)).strip().lower()


class ApiEditorSupplierImportPreviewView(ExternalEditorBaseView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Selecciona una lista CSV, TSV o XLSX."}, status=status.HTTP_400_BAD_REQUEST)
        extension = str(upload.name or "").lower().rsplit(".", 1)[-1]
        try:
            if extension in {"csv", "tsv", "txt"}:
                content = upload.read().decode("utf-8-sig", errors="replace")
                sample = content[:4096]
                if extension == "tsv":
                    delimiter = "\t"
                else:
                    try:
                        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
                    except csv.Error:
                        delimiter = ";"
                raw_rows = list(csv.DictReader(io.StringIO(content), delimiter=delimiter))
            elif extension == "xlsx":
                from openpyxl import load_workbook

                workbook = load_workbook(upload, read_only=True, data_only=True)
                sheet = workbook.active
                iterator = sheet.iter_rows(values_only=True)
                headers = [str(value or "") for value in next(iterator, [])]
                raw_rows = [dict(zip(headers, values)) for values in iterator]
            else:
                return Response({"detail": "Formato no soportado; usa CSV, TSV o XLSX."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({"detail": f"No se pudo leer la lista: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        aliases = {
            "sku": {"sku", "codigo", "codigo sku", "cod sku", "cod. sku", "internalcode"},
            "cost": {"costo", "cost", "precio costo"},
            "salePrice": {"precio", "venta", "saleprice", "precio venta"},
            "supplier": {"proveedor", "supplier"},
            "stock": {"stock", "existencia"},
        }
        normalized_rows = []
        for row in raw_rows[:MAX_SELECTION_IDS]:
            normalized = {_normalize_import_header(key): value for key, value in row.items()}
            item = {}
            for target, names in aliases.items():
                for name in names:
                    if name in normalized and normalized[name] not in (None, ""):
                        item[target] = normalized[name]
                        break
            if item.get("sku"):
                normalized_rows.append(item)

        skus = [str(row["sku"]).strip() for row in normalized_rows]
        product_map = {product.sku.casefold(): product for product in Product.objects.filter(sku__in=skus)}
        items = []
        unmatched = []
        for row in normalized_rows:
            sku = str(row.pop("sku")).strip()
            product = product_map.get(sku.casefold())
            if not product:
                unmatched.append(sku)
                continue
            changes = {key: value for key, value in row.items() if value not in (None, "")}
            if changes:
                items.append({"productId": product.pk, "sku": product.sku, "name": product.name, "changes": changes})
        return Response(
            {
                "fileName": upload.name,
                "totalRows": len(raw_rows),
                "matchedRows": len(items),
                "unmatchedRows": len(unmatched),
                "unmatchedSkus": unmatched[:100],
                "items": items,
                "preview": items[:100],
            }
        )


class ApiEditorSupplierPriceListCollectionView(ExternalEditorBaseView):
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        companies = list(get_user_companies(request.user))
        batches = (
            SupplierPriceListBatch.objects.filter(company__in=companies)
            .select_related("supplier", "company", "profile", "created_by", "applied_by")
            .order_by("-created_at")[:100]
        )
        return Response(
            {
                "companies": [
                    {"id": company.pk, "name": company.name, "slug": company.slug}
                    for company in companies
                ],
                "suppliers": [
                    {"id": supplier.pk, "name": supplier.name}
                    for supplier in Supplier.objects.filter(is_active=True).order_by("name")
                ],
                "profiles": [
                    {
                        "id": profile.pk,
                        "supplierId": profile.supplier_id,
                        "name": profile.name,
                        "sheetName": profile.sheet_name,
                        "headerRow": profile.header_row,
                        "columnMapping": profile.column_mapping,
                        "defaultCurrency": profile.default_currency,
                    }
                    for profile in SupplierImportProfile.objects.filter(is_active=True)
                    .select_related("supplier")
                    .order_by("supplier__name", "name")
                ],
                "items": [_serialize_supplier_price_list_batch(batch) for batch in batches],
                "mappingFields": [
                    {"field": field, "label": label} for field, label in MAPPING_FIELDS
                ],
            }
        )

    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Selecciona un archivo XLSX o CSV."}, status=400)
        extension = Path(str(upload.name or "")).suffix.lower()
        if extension not in {".xlsx", ".csv"}:
            return Response({"detail": "Formato no soportado. Usa XLSX o CSV."}, status=400)
        max_size = int(getattr(settings, "IMPORT_MAX_FILE_SIZE_BYTES", 10 * 1024 * 1024))
        if upload.size and upload.size > max_size:
            return Response({"detail": "El archivo supera el limite permitido."}, status=413)

        supplier = Supplier.objects.filter(
            pk=request.data.get("supplierId"), is_active=True
        ).first()
        if not supplier:
            return Response({"detail": "Selecciona un proveedor valido."}, status=400)
        company = _resolve_request_company(request)
        requested_company_id = request.data.get("companyId")
        if requested_company_id:
            company = get_user_companies(request.user).filter(pk=requested_company_id).first()
        if not company:
            return Response({"detail": "Selecciona una empresa habilitada."}, status=400)
        profile = None
        if request.data.get("profileId"):
            profile = SupplierImportProfile.objects.filter(
                pk=request.data.get("profileId"), supplier=supplier, is_active=True
            ).first()
            if not profile:
                return Response({"detail": "El perfil no pertenece al proveedor."}, status=400)

        digest = hash_uploaded_file(upload)
        if SupplierPriceListBatch.objects.filter(
            supplier=supplier,
            file_sha256=digest,
            status=SupplierPriceListBatch.STATUS_APPLIED,
        ).exists():
            return Response({"detail": "Este archivo ya fue aplicado para el proveedor."}, status=409)
        batch = SupplierPriceListBatch.objects.create(
            supplier=supplier,
            company=company,
            profile=profile,
            source_file=upload,
            original_filename=Path(upload.name).name[:255],
            file_sha256=digest,
            file_size=upload.size or 0,
            sheet_name=profile.sheet_name if profile else "",
            header_row=profile.header_row if profile else 1,
            column_mapping=profile.column_mapping if profile else {},
            default_currency=profile.default_currency if profile else "ARS",
            created_by=request.user,
        )
        try:
            inspection = inspect_source_file(
                batch.source_file.path,
                sheet_name=batch.sheet_name,
                header_row=batch.header_row,
                coordinate_mode=(
                    mapping_uses_coordinates(batch.column_mapping)
                    if batch.column_mapping
                    else True
                ),
            )
        except (ValidationError, OSError) as exc:
            source = batch.source_file
            batch.delete()
            source.delete(save=False)
            return _validation_error_response(exc)
        return Response(
            {"batch": _serialize_supplier_price_list_batch(batch), "inspection": inspection},
            status=status.HTTP_201_CREATED,
        )


class ApiEditorSupplierPriceListDetailView(ExternalEditorBaseView):
    def get(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        rows = batch.rows.select_related("matched_product", "product_supplier")
        change_filter = str(request.query_params.get("change") or "").strip()
        decision_filter = str(request.query_params.get("decision") or "").strip()
        if change_filter in dict(SupplierPriceListRow.CHANGE_CHOICES):
            rows = rows.filter(change_type=change_filter)
        if decision_filter in dict(SupplierPriceListRow.DECISION_CHOICES):
            rows = rows.filter(decision=decision_filter)
        try:
            page_size = min(200, max(1, int(request.query_params.get("pageSize", 100))))
            page_number = max(1, int(request.query_params.get("page", 1)))
        except (TypeError, ValueError):
            return Response({"detail": "Paginacion invalida."}, status=400)
        paginator = Paginator(rows.order_by("row_number"), page_size)
        page = paginator.get_page(page_number)
        return Response(
            {
                "batch": _serialize_supplier_price_list_batch(batch),
                "rows": [_serialize_supplier_price_list_row(row) for row in page.object_list],
                "page": page.number,
                "pageSize": page_size,
                "totalRows": paginator.count,
                "totalPages": paginator.num_pages,
            }
        )


class ApiEditorSupplierPriceListInspectView(ExternalEditorBaseView):
    def post(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        try:
            coordinate_mode = _parse_bool_param(request.data.get("coordinateMode"))
            if coordinate_mode is None:
                coordinate_mode = (
                    mapping_uses_coordinates(batch.column_mapping)
                    if batch.column_mapping
                    else True
                )
            inspection = inspect_source_file(
                batch.source_file.path,
                sheet_name=str(request.data.get("sheetName") or batch.sheet_name or ""),
                header_row=int(request.data.get("headerRow") or batch.header_row or 1),
                coordinate_mode=coordinate_mode,
            )
        except (ValidationError, OSError, TypeError, ValueError) as exc:
            return _validation_error_response(exc)
        return Response({"inspection": inspection})


class ApiEditorSupplierPriceListPreviewView(ExternalEditorBaseView):
    def post(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        mapping = request.data.get("mapping")
        if not isinstance(mapping, dict):
            return Response({"detail": "El mapeo de columnas no es valido."}, status=400)
        sheet_name = str(request.data.get("sheetName") or batch.sheet_name or "")
        try:
            header_row = int(request.data.get("headerRow") or batch.header_row or 1)
        except (TypeError, ValueError):
            return Response({"detail": "La fila inicial no es valida."}, status=400)
        currency = str(request.data.get("defaultCurrency") or batch.default_currency or "ARS").upper()
        if currency not in {"ARS", "USD", "EUR"}:
            return Response({"detail": "La moneda predeterminada no es valida."}, status=400)
        batch.default_currency = currency
        batch.save(update_fields=["default_currency", "updated_at"])
        try:
            batch = generate_supplier_price_list_preview(
                batch,
                mapping=mapping,
                sheet_name=sheet_name,
                header_row=header_row,
            )
        except ValidationError as exc:
            return _validation_error_response(exc)

        profile_name = str(request.data.get("profileName") or "").strip()
        if request.data.get("saveProfile") and profile_name:
            profile, _created = SupplierImportProfile.objects.update_or_create(
                supplier=batch.supplier,
                name=profile_name[:120],
                defaults={
                    "sheet_name": batch.sheet_name,
                    "header_row": batch.header_row,
                    "column_mapping": batch.column_mapping,
                    "default_currency": batch.default_currency,
                    "is_active": True,
                    "updated_by": request.user,
                    "created_by": request.user,
                },
            )
            batch.profile = profile
            batch.save(update_fields=["profile", "updated_at"])
        return Response({"batch": _serialize_supplier_price_list_batch(batch)})


class ApiEditorSupplierPriceListDecisionView(ExternalEditorBaseView):
    def post(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        decisions = request.data.get("decisions") or {}
        if request.data.get("resolveReviews") == "skip":
            decisions = {
                row_id: SupplierPriceListRow.DECISION_SKIP
                for row_id in batch.rows.filter(
                    decision=SupplierPriceListRow.DECISION_REVIEW
                ).values_list("id", flat=True)
            }
        if not isinstance(decisions, dict):
            return Response({"detail": "Las decisiones no son validas."}, status=400)
        try:
            normalized = {int(row_id): str(value) for row_id, value in decisions.items()}
            changed = update_row_decisions(batch, normalized)
        except (ValidationError, TypeError, ValueError) as exc:
            return _validation_error_response(exc)
        batch.refresh_from_db()
        return Response({"changed": changed, "batch": _serialize_supplier_price_list_batch(batch)})


class ApiEditorSupplierPriceListApplyView(ExternalEditorBaseView):
    def post(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        denied = _supplier_import_permission_response(request, changes_prices=True)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        expected = f"APLICAR LISTA {batch.pk}"
        if str(request.data.get("confirmation") or "").strip().upper() != expected:
            return Response({"detail": f'Escribe exactamente "{expected}".'}, status=400)
        try:
            batch = apply_supplier_price_list(
                batch,
                user=request.user,
                pricing_mode=request.data.get("pricingMode"),
            )
        except ValidationError as exc:
            return _validation_error_response(exc)
        return Response({"batch": _serialize_supplier_price_list_batch(batch)})


class ApiEditorSupplierPriceListRollbackView(ExternalEditorBaseView):
    def post(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        denied = _supplier_import_permission_response(request, changes_prices=True)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        expected = f"REVERTIR LISTA {batch.pk}"
        if str(request.data.get("confirmation") or "").strip().upper() != expected:
            return Response({"detail": f'Escribe exactamente "{expected}".'}, status=400)
        try:
            batch = rollback_supplier_price_list(batch, user=request.user)
        except ValidationError as exc:
            return _validation_error_response(exc)
        return Response({"batch": _serialize_supplier_price_list_batch(batch)})


class ApiEditorSupplierPriceListReportView(ExternalEditorBaseView):
    def get(self, request, batch_id):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        denied = _supplier_import_permission_response(request)
        if denied:
            return denied
        batch = _editor_supplier_batch(request, batch_id)
        if not batch:
            return Response({"detail": "Lista no encontrada."}, status=404)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response.write("\ufeff")
        response["Content-Disposition"] = (
            f'attachment; filename="lista_proveedor_{batch.pk}_reporte.csv"'
        )
        headers = [
            "fila", "tipo", "codigo_proveedor", "descripcion", "sku_identificado",
            "producto_identificado", "metodo", "confianza", "cambio", "costo_anterior",
            "costo_propuesto", "diferencia", "diferencia_porcentaje", "moneda", "decision",
            "aplicado", "advertencias",
        ]
        writer = csv.writer(response, delimiter=";")
        writer.writerow(headers)
        for item in report_rows(batch):
            values = []
            for header in headers:
                value = item.get(header, "")
                if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                    value = "'" + value
                values.append(value)
            writer.writerow(values)
        return response


class ApiEditorSavedViewListView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        queryset = ExternalEditorSavedView.objects.filter(created_by=request.user)
        return Response({"items": [_serialize_editor_saved_view(item) for item in queryset]})

    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        name = str(request.data.get("name") or "").strip()
        filters = request.data.get("filters")
        if not name or not isinstance(filters, dict):
            return Response({"detail": "Indica un nombre y filtros validos."}, status=status.HTTP_400_BAD_REQUEST)
        saved_view, _created = ExternalEditorSavedView.objects.update_or_create(
            created_by=request.user,
            name=name[:120],
            defaults={"filters": filters},
        )
        return Response(_serialize_editor_saved_view(saved_view), status=status.HTTP_201_CREATED)


class ApiEditorSavedViewDetailView(ExternalEditorBaseView):
    def delete(self, request, saved_view_id):
        deleted, _ = ExternalEditorSavedView.objects.filter(pk=saved_view_id, created_by=request.user).delete()
        if not deleted:
            return Response({"detail": "Vista no encontrada."}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class ApiEditorDraftListView(ExternalEditorBaseView):
    def get(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        queryset = ExternalEditorDraft.objects.filter(created_by=request.user)
        return Response({"items": [_serialize_editor_draft(item) for item in queryset[:100]]})

    def post(self, request):
        if not self.editor_is_enabled():
            return _external_editor_disabled_response()
        name = str(request.data.get("name") or "").strip()
        changes = request.data.get("changes")
        if not name:
            return Response({"detail": "Indica un nombre para el borrador."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            preview_external_editor_job({"items": changes})
        except ValidationError as exc:
            return _validation_error_response(exc)
        draft = ExternalEditorDraft.objects.create(
            created_by=request.user,
            name=name[:160],
            changes=changes,
        )
        return Response(_serialize_editor_draft(draft), status=status.HTTP_201_CREATED)


class ApiEditorDraftDetailView(ExternalEditorBaseView):
    def get_draft(self, request, draft_id):
        return ExternalEditorDraft.objects.filter(pk=draft_id, created_by=request.user).first()

    def get(self, request, draft_id):
        draft = self.get_draft(request, draft_id)
        if not draft:
            return Response({"detail": "Borrador no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        return Response(_serialize_editor_draft(draft))

    def patch(self, request, draft_id):
        draft = self.get_draft(request, draft_id)
        if not draft:
            return Response({"detail": "Borrador no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        if draft.status != ExternalEditorDraft.STATUS_DRAFT:
            return Response({"detail": "Solo se pueden modificar borradores pendientes."}, status=status.HTTP_409_CONFLICT)
        if "name" in request.data:
            draft.name = str(request.data.get("name") or draft.name).strip()[:160]
        if "changes" in request.data:
            try:
                preview_external_editor_job({"items": request.data.get("changes")})
            except ValidationError as exc:
                return _validation_error_response(exc)
            draft.changes = request.data.get("changes")
        draft.save()
        return Response(_serialize_editor_draft(draft))

    def delete(self, request, draft_id):
        draft = self.get_draft(request, draft_id)
        if not draft:
            return Response({"detail": "Borrador no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        draft.status = ExternalEditorDraft.STATUS_CANCELLED
        draft.save(update_fields=["status", "updated_at"])
        return Response(status=status.HTTP_204_NO_CONTENT)


class ApiEditorDraftPublishView(ApiEditorDraftDetailView):
    def post(self, request, draft_id):
        if not self.editor_writes_are_enabled():
            return _external_editor_writes_disabled_response()
        draft = self.get_draft(request, draft_id)
        if not draft:
            return Response({"detail": "Borrador no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        if draft.status != ExternalEditorDraft.STATUS_DRAFT:
            return Response({"detail": "El borrador ya no esta pendiente."}, status=status.HTTP_409_CONFLICT)
        payload = {"items": draft.changes}
        if _external_editor_payload_changes_prices(payload) and not has_capability(request.user, CAP_CHANGE_PRICES):
            return Response({"detail": "No tienes permiso para publicar cambios de precios."}, status=status.HTTP_403_FORBIDDEN)
        try:
            job, created = create_external_editor_job(
                payload=payload,
                user=request.user,
                idempotency_key=request.headers.get("Idempotency-Key") or f"draft-{draft.pk}-{uuid.uuid4()}",
            )
            job = _dispatch_external_editor_job(job, created)
        except (ValidationError, ExternalEditorJobPayloadConflict) as exc:
            return _validation_error_response(exc)
        draft.status = ExternalEditorDraft.STATUS_PUBLISHED
        draft.published_job = job
        draft.published_at = timezone.now()
        draft.save(update_fields=["status", "published_job", "published_at", "updated_at"])
        return Response({"draft": _serialize_editor_draft(draft), "job": serialize_external_editor_job(job)})


class ApiCategoryListView(ApiV1BaseListView):
    serializer_class = CategorySerializer
    throttle_scope = "api_v1_catalog"

    def get_queryset(self):
        queryset = Category.objects.select_related("parent").order_by("order", "name")

        user = self.request.user
        if not user.is_staff or not has_capability(user, CAP_GLOBAL_SEARCH):
            queryset = queryset.filter(is_active=True)

        q = (self.request.query_params.get("q") or "").strip()
        if q:
            queryset = queryset.filter(
                Q(name__icontains=q)
                | Q(slug__icontains=q)
                | Q(parent__name__icontains=q)
            )

        parent = self.request.query_params.get("parent")
        if parent:
            if parent == "root":
                queryset = queryset.filter(parent__isnull=True)
            elif str(parent).isdigit():
                queryset = queryset.filter(parent_id=int(parent))

        active_param = _parse_bool_param(self.request.query_params.get("active"))
        if active_param is not None:
            queryset = queryset.filter(is_active=active_param)

        return queryset


class ApiProductListView(ApiV1BaseListView):
    throttle_scope = "api_v1_catalog"

    def get_queryset(self):
        queryset = Product.objects.select_related("category", "supplier_ref").prefetch_related("categories").order_by("name")

        user = self.request.user
        if user.is_staff and has_capability(user, CAP_GLOBAL_SEARCH):
            active_param = _parse_bool_param(self.request.query_params.get("active"))
            if active_param is not None:
                queryset = queryset.filter(is_active=active_param)
        else:
            queryset = Product.catalog_visible(queryset=queryset, include_uncategorized=False)

        q = (self.request.query_params.get("q") or "").strip()
        if q:
            queryset = queryset.filter(
                Q(sku__icontains=q)
                | Q(name__icontains=q)
                | Q(supplier__icontains=q)
                | Q(supplier_ref__name__icontains=q)
                | Q(supplier_offers__supplier__name__icontains=q)
                | Q(supplier_offers__supplier_code__icontains=q)
                | Q(supplier_offers__supplier_description__icontains=q)
                | Q(filter_1__icontains=q)
                | Q(filter_2__icontains=q)
                | Q(filter_3__icontains=q)
                | Q(filter_4__icontains=q)
                | Q(filter_5__icontains=q)
            ).distinct()

        supplier = (self.request.query_params.get("supplier") or "").strip()
        if supplier:
            queryset = queryset.filter(
                Q(supplier__icontains=supplier)
                | Q(supplier_ref__name__icontains=supplier)
                | Q(supplier_offers__supplier__name__icontains=supplier)
            ).distinct()

        category = (self.request.query_params.get("category") or "").strip()
        if category:
            category_filters = Q(category__slug=category) | Q(categories__slug=category)
            if category.isdigit():
                category_id = int(category)
                category_filters |= Q(category_id=category_id) | Q(categories__id=category_id)
            queryset = queryset.filter(category_filters).distinct()

        in_stock_param = _parse_bool_param(self.request.query_params.get("in_stock"))
        if in_stock_param is True:
            queryset = queryset.filter(stock__gt=0)
        elif in_stock_param is False:
            queryset = queryset.filter(stock__lte=0)

        return queryset

    def get_serializer_class(self):
        if self.request.user.is_staff and has_capability(
            self.request.user,
            CAP_CHANGE_PRICES,
        ):
            return ProductStaffSerializer
        return ProductClientSerializer


class ApiClientListView(ApiV1BaseListView):
    permission_classes = [IsStaffUser, HasAdminCapability]
    required_capability = CAP_GLOBAL_SEARCH
    serializer_class = ClientProfileSerializer
    throttle_scope = "api_v1_admin"

    def get_queryset(self):
        queryset = ClientProfile.objects.select_related("user").order_by("company_name")
        company = _resolve_request_company(self.request)
        if company is None:
            return queryset.none()
        queryset = queryset.filter(
            company_links__company=company,
            company_links__is_active=True,
        ).distinct()

        q = (self.request.query_params.get("q") or "").strip()
        if q:
            queryset = queryset.filter(
                Q(company_name__icontains=q)
                | Q(cuit_dni__icontains=q)
                | Q(user__username__icontains=q)
                | Q(user__email__icontains=q)
            )

        approved_param = _parse_bool_param(self.request.query_params.get("approved"))
        if approved_param is not None:
            queryset = queryset.filter(is_approved=approved_param)

        return queryset


class ApiMyClientProfileView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_default"

    def get(self, request):
        profile = (
            ClientProfile.objects.select_related("user")
            .filter(user=request.user)
            .first()
        )
        if not profile:
            return Response({"detail": "Perfil de cliente no encontrado."}, status=404)
        serializer = ClientProfileSerializer(profile, context={"request": request})
        return Response(serializer.data)


class ApiOrderListView(ApiV1BaseListView):
    permission_classes = [permissions.IsAuthenticated, HasRequiredCapabilityWhenStaff]
    required_staff_capability = CAP_MANAGE_ORDERS
    serializer_class = OrderListSerializer
    throttle_scope = "api_v1_default"

    def get_queryset(self):
        user = self.request.user
        queryset = Order.objects.select_related("user", "company").prefetch_related("items").order_by("-created_at")
        company = _resolve_request_company(self.request)
        if company is None:
            return queryset.none()
        queryset = queryset.filter(company=company)

        if not user.is_staff:
            queryset = queryset.filter(user=user)

        status = (self.request.query_params.get("status") or "").strip().lower()
        if status:
            queryset = queryset.filter(status=status)

        if user.is_staff:
            user_id = (self.request.query_params.get("user_id") or "").strip()
            if user_id and user_id.isdigit():
                queryset = queryset.filter(user_id=int(user_id))

        return queryset


class ApiOrderQueueView(ApiV1BaseListView):
    permission_classes = [IsStaffUser, HasAdminCapability]
    required_capability = CAP_MANAGE_ORDERS
    serializer_class = OrderListSerializer
    throttle_scope = "api_v1_admin"

    def get_queryset(self):
        queryset = Order.objects.select_related("user", "company").prefetch_related("items").order_by("-updated_at")
        company = _resolve_request_company(self.request)
        if company is None:
            return queryset.none()
        queryset = queryset.filter(company=company)
        filtered_qs, role = get_order_queue_queryset_for_user(queryset, self.request.user)
        self._resolved_role = role

        status = (self.request.query_params.get("status") or "").strip().lower()
        if status:
            filtered_qs = filtered_qs.filter(status=status)
        return filtered_qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page if page is not None else queryset, many=True)

        role = getattr(self, "_resolved_role", None) or resolve_user_order_role(request.user)
        roles = get_user_order_roles(request.user)
        queue_statuses = get_role_queue_statuses(roles or role)
        counts = {
            status: queryset.filter(status=status).count()
            for status in queue_statuses
        }

        payload = {
            "role": role,
            "roles": roles,
            "queue_statuses": queue_statuses,
            "counts": counts,
            "results": serializer.data,
        }

        if page is not None:
            paginated = self.get_paginated_response(serializer.data)
            paginated.data["role"] = role
            paginated.data["roles"] = roles
            paginated.data["queue_statuses"] = queue_statuses
            paginated.data["counts"] = counts
            return paginated
        return Response(payload)


class ApiOrderWorkflowView(APIView):
    permission_classes = [permissions.IsAuthenticated, HasRequiredCapabilityWhenStaff]
    required_staff_capability = CAP_MANAGE_ORDERS
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_default"

    def get(self, request, order_id):
        company = _resolve_request_company(request)
        order_qs = Order.objects.filter(pk=order_id)
        if company is None:
            order_qs = order_qs.none()
        else:
            order_qs = order_qs.filter(company=company)
        if not request.user.is_staff:
            order_qs = order_qs.filter(user=request.user)
        order = order_qs.first()
        if not order:
            return Response({"detail": "Pedido no encontrado."}, status=404)

        if not request.user.is_staff and order.user_id != request.user.id:
            return Response({"detail": "No autorizado."}, status=403)

        allowed_statuses = get_allowed_next_statuses_for_user(request.user, order)
        return Response(
            {
                "order_id": order.id,
                "current_status": order.status,
                "allowed_next_statuses": allowed_statuses,
                "resolved_role": resolve_user_order_role(request.user),
                "resolved_roles": get_user_order_roles(request.user),
            },
        )


class RateLimitedObtainAuthToken(ObtainAuthToken):
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"


class ApiWebhookEndpointListCreateView(APIView):
    permission_classes = [IsStaffUser, HasAdminCapability]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"
    required_capability = CAP_MANAGE_INTEGRATIONS

    def get(self, request):
        company = _resolve_request_company(request)
        if not company:
            return Response({"detail": "Empresa activa o company_id requerido."}, status=400)
        rows = WebhookEndpoint.objects.filter(company=company).order_by("name", "id")
        return Response({
            "results": [
                {
                    "id": row.pk,
                    "name": row.name,
                    "target_url": row.target_url,
                    "events": row.events,
                    "is_active": row.is_active,
                    "created_at": row.created_at,
                }
                for row in rows
            ]
        })

    def post(self, request):
        company = _resolve_request_company(request)
        if not company:
            return Response({"detail": "Empresa activa o company_id requerido."}, status=400)
        endpoint = WebhookEndpoint(
            company=company,
            name=str(request.data.get("name", "")).strip(),
            target_url=str(request.data.get("target_url", "")).strip(),
            events=request.data.get("events") or [],
            is_active=(
                _parse_bool_param(request.data.get("is_active"))
                if _parse_bool_param(request.data.get("is_active")) is not None
                else True
            ),
            created_by=request.user,
        )
        supplied_secret = str(request.data.get("secret", "") or "").strip()
        if supplied_secret:
            endpoint.secret = supplied_secret
        try:
            endpoint.save()
        except ValidationError as exc:
            details = exc.message_dict if hasattr(exc, "message_dict") else exc.messages
            return Response({"detail": details}, status=400)
        return Response(
            {
                "id": endpoint.pk,
                "name": endpoint.name,
                "target_url": endpoint.target_url,
                "events": endpoint.events,
                "is_active": endpoint.is_active,
                "secret": endpoint.secret,
                "secret_notice": "Guarda este secreto ahora; no vuelve a mostrarse en listados.",
            },
            status=201,
        )


class ApiWebhookEndpointDetailView(APIView):
    permission_classes = [IsStaffUser, HasAdminCapability]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"
    required_capability = CAP_MANAGE_INTEGRATIONS

    def _get_endpoint(self, request, endpoint_id):
        company = _resolve_request_company(request)
        if not company:
            return None
        return WebhookEndpoint.objects.filter(company=company, pk=endpoint_id).first()

    def patch(self, request, endpoint_id):
        endpoint = self._get_endpoint(request, endpoint_id)
        if not endpoint:
            return Response({"detail": "Webhook no encontrado."}, status=404)
        for field in ("name", "target_url", "events", "is_active"):
            if field in request.data:
                value = request.data[field]
                if field == "is_active":
                    parsed_value = _parse_bool_param(value)
                    if parsed_value is None:
                        return Response({"detail": {"is_active": "Valor booleano invalido."}}, status=400)
                    value = parsed_value
                setattr(endpoint, field, value)
        if request.data.get("rotate_secret"):
            from core.models import generate_webhook_secret

            endpoint.secret = generate_webhook_secret()
        try:
            endpoint.save()
        except ValidationError as exc:
            details = exc.message_dict if hasattr(exc, "message_dict") else exc.messages
            return Response({"detail": details}, status=400)
        payload = {
            "id": endpoint.pk,
            "name": endpoint.name,
            "target_url": endpoint.target_url,
            "events": endpoint.events,
            "is_active": endpoint.is_active,
        }
        if request.data.get("rotate_secret"):
            payload["secret"] = endpoint.secret
        return Response(payload)

    def delete(self, request, endpoint_id):
        endpoint = self._get_endpoint(request, endpoint_id)
        if not endpoint:
            return Response({"detail": "Webhook no encontrado."}, status=404)
        endpoint.delete()
        return Response(status=204)


class ApiWebhookDeliveryListView(APIView):
    permission_classes = [IsStaffUser, HasAdminCapability]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "api_v1_admin"
    required_capability = CAP_MANAGE_INTEGRATIONS

    def get(self, request, endpoint_id):
        company = _resolve_request_company(request)
        endpoint = WebhookEndpoint.objects.filter(company=company, pk=endpoint_id).first() if company else None
        if not endpoint:
            return Response({"detail": "Webhook no encontrado."}, status=404)
        rows = WebhookDelivery.objects.filter(endpoint=endpoint).order_by("-created_at")[:100]
        return Response({
            "results": [
                {
                    "id": row.pk,
                    "event_id": row.event_id,
                    "event_type": row.event_type,
                    "status": row.status,
                    "attempts_count": row.attempts_count,
                    "response_status": row.response_status,
                    "last_error": row.last_error,
                    "next_retry_at": row.next_retry_at,
                    "created_at": row.created_at,
                }
                for row in rows
            ]
        })


class ApiSchemaView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        base = request.build_absolute_uri("/api/v1/")
        return Response({
            "openapi": "3.0.3",
            "info": {"title": "FLEXS API", "version": "1.0.0"},
            "servers": [{"url": base}],
            "authentication": {"header": "Authorization: Token <token>"},
            "tenant_scope": (
                "Enviar company_id (o company) cuando el usuario tenga varias empresas. "
                "Sin una empresa autorizada, los recursos empresariales no devuelven datos."
            ),
            "staff_capabilities": {
                "clients": CAP_GLOBAL_SEARCH,
                "orders": CAP_MANAGE_ORDERS,
                "orders_queue": CAP_MANAGE_ORDERS,
                "order_workflow": CAP_MANAGE_ORDERS,
                "webhooks": CAP_MANAGE_INTEGRATIONS,
                "product_cost": CAP_CHANGE_PRICES,
            },
            "paths": {
                "/health/": {"get": {"summary": "Estado de API y base"}},
                "/catalog/products/": {"get": {"summary": "Listado y busqueda de productos"}},
                "/catalog/categories/": {"get": {"summary": "Categorias"}},
                "/editor/products/": {"get": {"summary": "Grilla oficial del editor externo"}},
                "/editor/products/selection-ids/": {"get": {"summary": "Seleccion filtrada global"}},
                "/editor/products/{id}/": {
                    "get": {"summary": "Detalle editable"},
                    "patch": {"summary": "Edicion individual con control de version"},
                },
                "/editor/bulk/preview/": {"post": {"summary": "Vista previa de edicion masiva"}},
                "/editor/bulk/": {"post": {"summary": "Trabajo masivo idempotente"}},
                "/editor/jobs/{id}/": {"get": {"summary": "Progreso y resultado del trabajo"}},
                "/editor/jobs/{id}/rollback/": {"post": {"summary": "Reversion protegida"}},
                "/clients/": {"get": {"summary": "Clientes de la empresa"}},
                "/orders/": {"get": {"summary": "Pedidos de la empresa"}},
                "/orders/queue/": {"get": {"summary": "Cola por rol"}},
                "/webhooks/": {"get": {"summary": "Listar webhooks"}, "post": {"summary": "Crear webhook"}},
                "/webhooks/{id}/": {"patch": {"summary": "Editar o rotar secreto"}, "delete": {"summary": "Eliminar webhook"}},
                "/webhooks/{id}/deliveries/": {"get": {"summary": "Ultimas entregas"}},
            },
            "webhook_signature": "HMAC-SHA256 de '<timestamp>.<raw_body>' con el secreto; header X-FLEXS-Signature.",
            "webhook_events": [value for value, _label in WebhookEndpoint.EVENT_CHOICES],
        })


@login_required
@capability_required(CAP_MANAGE_INTEGRATIONS)
def api_docs(request):
    return render(request, "core/api_docs.html")
