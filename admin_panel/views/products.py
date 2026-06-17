"""
Admin Panel views - Custom admin interface.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.forms import PasswordResetForm, SetPasswordForm
from django.contrib.auth.password_validation import validate_password
from django.contrib import messages
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db import transaction, connection, IntegrityError
from django.db import DatabaseError
from django.db.models import (
    Q,
    Case,
    Count,
    Sum,
    Max,
    Avg,
    F,
    When,
    IntegerField,
    CharField,
    DecimalField,
    ExpressionWrapper,
    Value,
    Prefetch,
    OuterRef,
    Subquery,
)
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.text import slugify
from django.utils.http import url_has_allowed_host_and_scheme
import json
import os
import re
import unicodedata
from io import BytesIO, StringIO
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import urlencode, parse_qs
import csv
from openpyxl import Workbook, load_workbook

from catalog.models import (
    Product,
    Category,
    CategoryAttribute,
    CategoryProductOrder,
    ClampMeasureRequest,
    Supplier,
    PriceList,
)
from accounts.models import (
    AccountRequest,
    ClientCategory,
    ClientCompany,
    ClientPayment,
    ClientProfile,
    ClientTransaction,
)
from accounts.services.ledger import (
    create_adjustment_transaction,
    sync_order_charge_transaction,
)
from accounts.services.movement_lifecycle import (
    apply_transaction_state_transition,
    can_transition_transaction_state,
    is_transaction_reopen_locked as service_is_transaction_reopen_locked,
    movement_allows_print as service_movement_allows_print,
)
from orders.models import (
    ClampQuotation,
    Order,
    OrderItem,
    OrderProposal,
    OrderProposalItem,
    OrderRequest,
    OrderRequestEvent,
    OrderRequestItem,
    OrderStatusHistory,
)
from orders.services.workflow import (
    ROLE_ADMIN,
    ROLE_DEPOSITO,
    ROLE_FACTURACION,
    ROLE_VENTAS,
    can_user_transition_order,
    resolve_user_order_role,
)
from orders.services.request_workflow import (
    confirm_order_request,
    convert_request_to_order,
    create_order_proposal,
    record_order_request_event,
    reject_order_request,
)
from core.models import (
    AdminCompanyAccess,
    Company,
    DocumentSeries,
    FISCAL_BILLABLE_DOC_TYPES,
    FISCAL_DOC_TYPE_FA,
    FISCAL_DOC_TYPE_FB,
    FISCAL_DOC_TYPE_FC,
    FISCAL_INVOICE_DOC_TYPES,
    FISCAL_ISSUE_MODE_ARCA_WSFE,
    FISCAL_ISSUE_MODE_EXTERNAL_SAAS,
    FISCAL_ISSUE_MODE_MANUAL,
    FISCAL_STATUS_AUTHORIZED,
    FISCAL_STATUS_EXTERNAL_RECORDED,
    FISCAL_STATUS_PENDING_RETRY,
    FISCAL_STATUS_READY_TO_ISSUE,
    FISCAL_STATUS_REJECTED,
    FISCAL_STATUS_SUBMITTING,
    FISCAL_STATUS_VOIDED,
    FiscalDocument,
    FiscalPointOfSale,
    InternalDocument,
    SALES_BEHAVIOR_COTIZACION,
    SALES_DOCUMENT_BEHAVIOR_CHOICES,
    SALES_BEHAVIOR_FACTURA,
    SALES_BEHAVIOR_NOTA_CREDITO,
    SALES_BEHAVIOR_NOTA_DEBITO,
    SALES_BEHAVIOR_PEDIDO,
    SALES_BEHAVIOR_PRESUPUESTO,
    SALES_BEHAVIOR_RECIBO,
    SALES_BEHAVIOR_REMITO,
    SALES_BILLING_MODE_INTERNAL_DOCUMENT,
    SALES_BILLING_MODE_MANUAL_FISCAL,
    SalesDocumentType,
    SiteSettings,
    CatalogAnalyticsEvent,
    AdminAuditLog,
    ImportExecution,
    CatalogExcelTemplate,
    CatalogExcelTemplateSheet,
    CatalogExcelTemplateColumn,
    StockMovement,
    Warehouse,
)
from core.services.company_context import (
    admin_company_access_table_available,
    get_active_company,
    get_default_company,
    get_default_client_origin_company,
    get_preferred_client_company,
    get_user_companies,
    set_active_company,
    user_has_company_access,
)
from django.contrib.auth.models import Group, User
from admin_panel.forms.import_forms import ProductImportForm, ClientImportForm, CategoryImportForm
from admin_panel.forms.category_forms import CategoryForm
from admin_panel.fiscal_views import fiscal_health_view, fiscal_report_view
from admin_panel.forms.export_forms import (
    CatalogExcelTemplateForm,
    CatalogExcelTemplateSheetForm,
    CatalogExcelTemplateColumnForm,
)
from admin_panel.forms.sales_document_type_forms import SalesDocumentTypeForm, WarehouseForm
from catalog.services.product_importer import ProductImporter
from accounts.services.client_importer import ClientImporter
from catalog.services.category_importer import CategoryImporter
from catalog.services.abrazadera_importer import AbrazaderaImporter
from catalog.services.supplier_sync import ensure_supplier, clean_supplier_name
from catalog.services.clamp_code import (
    DIAMETER_HUMAN_TO_COMPACT_DEFAULT,
    generarCodigo,
    parsearCodigo,
)
from catalog.services.clamp_quoter import (
    CLAMP_LAMINATED_ALLOWED_DIAMETERS,
    CLAMP_PRICE_LISTS,
    CLAMP_WEIGHT_MAP,
    calculate_clamp_quote,
    get_allowed_diameter_options,
    parse_decimal_value,
    parse_int_value,
)
from catalog.services.clamp_request_products import (
    publish_clamp_request_product,
)
from catalog.services.category_assignment import (
    normalize_category_ids,
    assign_categories_to_product,
    add_category_to_products,
    replace_categories_for_products,
    remove_category_from_products,
)
from catalog.services.category_block_conversion import (
    convert_category_blocks_to_subcategories,
    rollback_category_block_conversion,
)
from core.services.import_manager import ImportTaskManager
from core.services.background_jobs import dispatch_import_job
from core.services.fiscal import (
    is_company_fiscal_ready,
    is_invoice_ready,
    resolve_payment_due_date,
)
from core.services.fiscal_notifications import send_fiscal_document_email
from core.services.arca_client import ArcaConfigurationError, ArcaTemporaryError, ArcaWsfeClient
from core.services.fiscal_documents import (
    close_fiscal_document,
    create_local_fiscal_document_from_order,
    reopen_fiscal_document,
    register_external_fiscal_document_for_order,
    void_fiscal_document,
)

from core.services.documents import (
    ensure_document_for_adjustment,
    ensure_document_for_order,
    ensure_document_for_payment,
)
from core.services.sales_documents import (
    create_fiscal_document_from_sales_type,
    create_internal_document_from_sales_type,
    resolve_sales_document_type,
)
from core.services.advanced_search import (
    apply_compact_text_search,
    apply_text_search,
    apply_parsed_text_search,
    compact_search_token,
    parse_text_search_query,
    sanitize_search_token,
)
from core.services.catalog_excel_exporter import build_catalog_workbook, build_export_filename
from core.services.audit import log_admin_action, log_admin_change, model_snapshot
from core.services.pricing import resolve_effective_price_list
import traceback
import logging
from core.decorators import superuser_required_for_modifications

logger = logging.getLogger(__name__)
PRIMARY_SUPERADMIN_USERNAME = getattr(settings, "ADMIN_PRIMARY_SUPERADMIN_USERNAME", "josueflexs")
ADMIN_ROLE_CHOICES = [
    (ROLE_ADMIN, "Administracion"),
    (ROLE_VENTAS, "Ventas"),
    (ROLE_DEPOSITO, "Deposito"),
    (ROLE_FACTURACION, "Facturacion"),
]
ADMIN_ROLE_LABELS = dict(ADMIN_ROLE_CHOICES)
FISCAL_PRINT_DOC_META = {
    "FA": {"letter": "A", "code": "001"},
    "FB": {"letter": "B", "code": "006"},
    "FC": {"letter": "C", "code": "011"},
    "NCA": {"letter": "A", "code": "003"},
    "NCB": {"letter": "B", "code": "008"},
    "NCC": {"letter": "C", "code": "013"},
    "NDA": {"letter": "A", "code": "002"},
    "NDB": {"letter": "B", "code": "007"},
    "NDC": {"letter": "C", "code": "012"},
}
FISCAL_PRINT_COPY_LABELS = {
    "original": "ORIGINAL",
    "duplicado": "DUPLICADO",
    "triplicado": "TRIPLICADO",
}
INVOICE_FISCAL_DOC_TYPES = tuple(sorted(FISCAL_INVOICE_DOC_TYPES))
BILLABLE_FISCAL_DOC_TYPES = tuple(sorted(FISCAL_BILLABLE_DOC_TYPES))
EMITTABLE_FISCAL_DOC_TYPES = tuple(choice[0] for choice in FiscalDocument.DOC_TYPE_CHOICES)
CLIENT_FACTURABLE_STATUSES = {
    Order.STATUS_CONFIRMED,
    Order.STATUS_PREPARING,
    Order.STATUS_SHIPPED,
    Order.STATUS_DELIVERED,
}
CLIENT_REMITO_READY_STATUSES = {
    Order.STATUS_SHIPPED,
    Order.STATUS_DELIVERED,
}
ORDER_INTERNAL_DOC_STATUS_RULES = {
    DocumentSeries.DOC_COT: {
        Order.STATUS_DRAFT,
        Order.STATUS_CONFIRMED,
        Order.STATUS_PREPARING,
        Order.STATUS_SHIPPED,
        Order.STATUS_DELIVERED,
    },
    DocumentSeries.DOC_PED: {
        Order.STATUS_CONFIRMED,
        Order.STATUS_PREPARING,
        Order.STATUS_SHIPPED,
        Order.STATUS_DELIVERED,
    },
    DocumentSeries.DOC_REM: CLIENT_REMITO_READY_STATUSES,
}



from .helpers import *



# ===================== PRODUCTS =====================

PRODUCT_LIST_ORDER_FIELDS = {
    "-updated_at",
    "updated_at",
    "name",
    "-name",
    "price",
    "-price",
    "stock",
    "-stock",
    "sku",
    "-sku",
}


def _clean_product_list_order(value):
    value = str(value or "").strip()
    if value in PRODUCT_LIST_ORDER_FIELDS:
        return value
    return "-updated_at"


def _apply_product_list_order(products, order):
    if order == "sku":
        return _sort_products_for_view_by_sku(products, "sku_asc")
    if order == "-sku":
        return _sort_products_for_view_by_sku(products, "sku_desc")
    return products.order_by(order)


def _attach_import_duplicate_alert(product):
    attrs = product.attributes or {}
    duplicate_keys = {
        ProductImporter.DUPLICATE_FLAG_KEY,
        ProductImporter.DUPLICATE_COUNT_KEY,
        ProductImporter.DUPLICATE_DETAIL_KEY,
        ProductImporter.DUPLICATE_ORIGINAL_ROW_KEY,
    }
    product.import_duplicate_alert = attrs.get(ProductImporter.DUPLICATE_FLAG_KEY) == "Si"
    product.import_duplicate_count = attrs.get(ProductImporter.DUPLICATE_COUNT_KEY) or ""
    product.import_duplicate_detail = attrs.get(ProductImporter.DUPLICATE_DETAIL_KEY) or ""
    product.import_duplicate_original_row = attrs.get(ProductImporter.DUPLICATE_ORIGINAL_ROW_KEY) or ""
    product.visible_import_attributes = {
        key: value for key, value in attrs.items() if key not in duplicate_keys
    }
    return product


@staff_member_required
def product_list(request):
    """Product list with search, filters, and pagination."""
    products, search, current_category_id, active_filter = get_product_queryset(request.GET)
    
    # Ordering
    order = _clean_product_list_order(request.GET.get('order', '-updated_at'))
    products = _apply_product_list_order(products, order)

    search_result_limit = 300
    filtered_total_count = len(products) if isinstance(products, list) else products.count()
    search_total_matches = filtered_total_count if search else 0
    search_results_truncated = False
    if search and search_total_matches > search_result_limit:
        products = products[:search_result_limit]
        search_results_truncated = True

    # Pagination
    paginator = Paginator(products, search_result_limit if search else 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    enrich_products_with_category_state(page_obj.object_list)
    for product in page_obj.object_list:
        _attach_import_duplicate_alert(product)
    
    category_options = get_cached_category_options(only_active=True, include_inactive_suffix=False)
    filter_chips = build_product_filter_chips(request.GET, category_options)
    
    context = {
        'page_obj': page_obj,
        'category_options': category_options,
        'search': search,
        'current_category_id': current_category_id,
        'active_filter': active_filter,
        'order_by': order,
        'active_filter_chips': filter_chips,
        'total_count': filtered_total_count,
        'pagination_count': len(page_obj.object_list),
        'search_result_limit': search_result_limit,
        'search_total_matches': search_total_matches,
        'search_results_truncated': search_results_truncated,
    }
    return render(request, 'admin_panel/products/list.html', context)


@staff_member_required
@superuser_required_for_modifications
def product_create(request):
    """Create new product."""
    category_options = get_cached_category_options(only_active=True, include_inactive_suffix=False)
    supplier_suggestions = list(Supplier.objects.order_by('name').values_list('name', flat=True)[:400])
    
    if request.method == 'POST':
        try:
            sku = request.POST.get('sku', '').strip()
            name = request.POST.get('name', '').strip()
            supplier_name = clean_supplier_name(request.POST.get('supplier', ''))
            supplier_obj = ensure_supplier(supplier_name) if supplier_name else None
            price = request.POST.get('price', '0')
            cost = request.POST.get('cost', '0')
            stock = request.POST.get('stock', '0')
            price_value = parse_admin_decimal_input(price, 'Precio', min_value='0')
            cost_value = parse_admin_decimal_input(cost, 'Costo', min_value='0')
            stock_value = parse_int_value(stock, 'Stock', min_value=0)
            primary_category_id = request.POST.get('category', '')
            selected_category_ids = normalize_category_ids(request.POST.getlist('categories'))
            description = request.POST.get('description', '').strip()
            attributes_payload = request.POST.get('attributes_json', '{}')
            uploaded_image = request.FILES.get('image')
            settings = SiteSettings.get_settings()

            if uploaded_image:
                _validate_admin_image_upload(uploaded_image)

            try:
                attributes_data = json.loads(attributes_payload or '{}')
            except json.JSONDecodeError:
                attributes_data = {}

            if (
                settings.require_primary_category_for_multicategory
                and len(selected_category_ids) > 1
                and not str(primary_category_id).isdigit()
            ):
                messages.error(
                    request,
                    "Debes definir una categoria principal cuando vinculas multiples categorias.",
                )
                return render(request, 'admin_panel/products/form.html', {
                    'category_options': category_options,
                    'selected_category_ids': selected_category_ids,
                    'supplier_suggestions': supplier_suggestions,
                        'action': 'Crear',
                    })

            missing_required, missing_recommended = validate_attributes_for_category(
                primary_category_id,
                attributes_data,
            )
            if missing_required:
                messages.error(
                    request,
                    f'Faltan atributos obligatorios para la categoria principal: {", ".join(missing_required)}.',
                )
                return render(request, 'admin_panel/products/form.html', {
                    'category_options': category_options,
                    'selected_category_ids': selected_category_ids,
                    'supplier_suggestions': supplier_suggestions,
                    'action': 'Crear',
                })
            if missing_recommended:
                messages.warning(
                    request,
                    f'Atributos recomendados sin completar: {", ".join(missing_recommended)}.',
                )
            
            if Product.objects.filter(sku=sku).exists():
                messages.error(request, f'Ya existe un producto con SKU "{sku}"')
            else:
                product = Product.objects.create(
                    sku=sku,
                    name=name,
                    supplier=supplier_name,
                    supplier_ref=supplier_obj,
                    cost=cost_value,
                    price=price_value,
                    stock=stock_value,
                    category_id=int(primary_category_id) if str(primary_category_id).isdigit() else None,
                    description=description,
                    attributes=attributes_data,
                    image=uploaded_image,
                )
                assign_categories_to_product(product, selected_category_ids, primary_category_id)
                blocks_payload = request.POST.get('product_blocks_json', '{}')
                try:
                    blocks_data = json.loads(blocks_payload or '{}')
                except Exception:
                    blocks_data = {}
                for cid_str, block_label in blocks_data.items():
                    if cid_str.isdigit():
                        cid = int(cid_str)
                        clean_block = " ".join(str(block_label or "").split())[:120]
                        CategoryProductOrder.objects.filter(
                            product=product,
                            category_id=cid
                        ).update(block_label=clean_block)
                log_admin_action(
                    request,
                    action="product_create",
                    target_type="product",
                    target_id=product.pk,
                    details={
                        "sku": product.sku,
                        "supplier": product.supplier,
                        "categories": selected_category_ids,
                    },
                )
                messages.success(request, f'Producto "{sku}" creado exitosamente.')
                return redirect('admin_product_list')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'admin_panel/products/form.html', {
        'category_options': category_options,
        'selected_category_ids': [],
        'supplier_suggestions': supplier_suggestions,
        'action': 'Crear',
        'existing_blocks_json': '{}',
    })


@staff_member_required
@superuser_required_for_modifications
def product_edit(request, pk):
    """Edit existing product."""
    product = get_object_or_404(Product, pk=pk)
    _attach_import_duplicate_alert(product)
    category_options = get_cached_category_options(only_active=True, include_inactive_suffix=False)
    existing_blocks = {
        row.category_id: row.block_label or ''
        for row in CategoryProductOrder.objects.filter(product=product)
    }
    existing_blocks_json = json.dumps(existing_blocks)
    supplier_suggestions = list(Supplier.objects.order_by('name').values_list('name', flat=True)[:400])
    
    if request.method == 'POST':
        try:
            product.sku = request.POST.get('sku', '').strip()
            product.name = request.POST.get('name', '').strip()
            product.supplier = clean_supplier_name(request.POST.get('supplier', ''))
            product.supplier_ref = ensure_supplier(product.supplier) if product.supplier else None
            product.cost = parse_admin_decimal_input(request.POST.get('cost', '0'), 'Costo', min_value='0')
            product.price = parse_admin_decimal_input(request.POST.get('price', '0'), 'Precio', min_value='0')
            product.stock = parse_int_value(request.POST.get('stock', '0'), 'Stock', min_value=0)
            product.description = request.POST.get('description', '').strip()
            product.is_active = request.POST.get('is_active') == 'on'
            uploaded_image = request.FILES.get('image')
            remove_image = request.POST.get('remove_image') == 'on'
            old_image_name = str(product.image.name or '').strip() if product.image else ''
            new_image_applied = False

            if uploaded_image:
                _validate_admin_image_upload(uploaded_image)

            primary_category_id = request.POST.get('category', '')
            selected_category_ids = normalize_category_ids(request.POST.getlist('categories'))
            settings = SiteSettings.get_settings()
            if (
                settings.require_primary_category_for_multicategory
                and len(selected_category_ids) > 1
                and not str(primary_category_id).isdigit()
            ):
                messages.error(
                    request,
                    "Debes definir una categoria principal cuando vinculas multiples categorias.",
                )
                return render(request, 'admin_panel/products/form.html', {
                    'product': product,
                    'category_options': category_options,
                    'selected_category_ids': selected_category_ids,
                    'supplier_suggestions': supplier_suggestions,
                    'action': 'Editar',
                    'existing_blocks_json': existing_blocks_json,
                })

            product.category_id = int(primary_category_id) if str(primary_category_id).isdigit() else None
            
            # Update attributes
            attributes_data = product.attributes or {}
            attributes_json = request.POST.get('attributes_json', '{}')
            if attributes_json:
                try:
                    attributes_data = json.loads(attributes_json)
                except json.JSONDecodeError:
                    attributes_data = product.attributes or {}

            missing_required, missing_recommended = validate_attributes_for_category(
                primary_category_id,
                attributes_data,
            )
            if missing_required:
                messages.error(
                    request,
                    f'Faltan atributos obligatorios para la categoria principal: {", ".join(missing_required)}.',
                )
                return render(request, 'admin_panel/products/form.html', {
                    'product': product,
                    'category_options': category_options,
                    'selected_category_ids': selected_category_ids,
                    'supplier_suggestions': supplier_suggestions,
                    'action': 'Editar',
                    'existing_blocks_json': existing_blocks_json,
                })
            if missing_recommended:
                messages.warning(
                    request,
                    f'Atributos recomendados sin completar: {", ".join(missing_recommended)}.',
                )

            product.attributes = attributes_data

            if remove_image and not uploaded_image:
                product.image = None
                new_image_applied = True
            if uploaded_image:
                product.image = uploaded_image
                new_image_applied = True
            
            product.save()
            assign_categories_to_product(product, selected_category_ids, primary_category_id)
            blocks_payload = request.POST.get('product_blocks_json', '{}')
            try:
                blocks_data = json.loads(blocks_payload or '{}')
            except Exception:
                blocks_data = {}
            for cid_str, block_label in blocks_data.items():
                if cid_str.isdigit():
                    cid = int(cid_str)
                    clean_block = " ".join(str(block_label or "").split())[:120]
                    CategoryProductOrder.objects.filter(
                        product=product,
                        category_id=cid
                    ).update(block_label=clean_block)
            if new_image_applied and old_image_name:
                _delete_orphan_product_image(old_image_name)
            log_admin_action(
                request,
                action="product_edit",
                target_type="product",
                target_id=product.pk,
                details={
                    "sku": product.sku,
                    "supplier": product.supplier,
                    "categories": selected_category_ids,
                },
            )
            messages.success(request, f'Producto "{product.sku}" actualizado.')
            return redirect('admin_product_list')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            
    selected_category_ids = list(product.categories.values_list('id', flat=True))
    if not selected_category_ids and product.category_id:
        selected_category_ids = [product.category_id]

    return render(request, 'admin_panel/products/form.html', {
        'product': product,
        'category_options': category_options,
        'selected_category_ids': selected_category_ids,
        'supplier_suggestions': supplier_suggestions,
        'action': 'Editar',
        'existing_blocks_json': existing_blocks_json,
    })


@staff_member_required
@superuser_required_for_modifications
def product_delete(request, pk):
    """Delete single product."""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        sku = product.sku
        product_id = product.pk
        product.delete()
        log_admin_action(
            request,
            action="product_delete",
            target_type="product",
            target_id=product_id,
            details={"sku": sku},
        )
        messages.success(request, f'Producto "{sku}" eliminado.')
        return redirect('admin_product_list')
        
    return render(request, 'admin_panel/delete_confirm.html', {
        'object': f"{product.name} ({product.sku})",
        'cancel_url': reverse('admin_product_list')
    })


@staff_member_required
@require_POST
@superuser_required_for_modifications
def product_toggle_active(request):
    """Toggle product active status (AJAX)."""
    try:
        data = json.loads(request.body)
        product_ids = data.get('ids', [])
        active = data.get('active', True)
        
        Product.objects.filter(id__in=product_ids).update(is_active=active)
        log_admin_action(
            request,
            action="product_toggle_active",
            target_type="product_bulk",
            details={"ids": product_ids, "active": bool(active)},
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{len(product_ids)} productos actualizados'
        })
    except Exception as e:
        logger.exception("Error toggling product active status")
        return JsonResponse({'success': False, 'error': 'No se pudieron actualizar los productos.'}, status=400)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def product_bulk_category_update(request):
    """Bulk categorize selected products."""
    try:
        raw_post_body = request.body
        category_id = request.POST.get('category_id')
        mode = request.POST.get('mode', 'append')
        select_all_pages = request.POST.get('select_all_pages') == 'true'

        if not category_id:
            messages.warning(request, 'No se selecciono una categoria.')
            return _redirect_admin_product_list_with_filters(request)

        if select_all_pages:
            products_to_update, _, _, _ = get_product_queryset(request.POST)
        else:
            product_ids = extract_target_product_ids_from_post(request.POST, raw_post_body)
            if not product_ids:
                logger.warning(
                    "product_bulk_category_update without selected products | user=%s | keys=%s | product_ids=%s | product_ids_csv=%s",
                    getattr(request.user, "username", "unknown"),
                    list(request.POST.keys()),
                    request.POST.getlist("product_ids"),
                    request.POST.get("product_ids_csv", ""),
                )
                messages.warning(request, 'No se seleccionaron productos.')
                return _redirect_admin_product_list_with_filters(request)
            products_to_update = Product.objects.filter(id__in=product_ids)

        target_ids = list(products_to_update.values_list('id', flat=True))
        if mode == 'replace':
            count = replace_categories_for_products(target_ids, category_id)
        else:
            count = add_category_to_products(target_ids, category_id)

        category = Category.objects.get(pk=category_id)
        if mode == 'replace':
            messages.success(request, f'{count} productos recategorizados a "{category.name}".')
        else:
            messages.success(request, f'{count} productos vinculados a "{category.name}".')

        log_admin_action(
            request,
            action="product_bulk_category_update",
            target_type="category",
            target_id=category.pk,
            details={"mode": mode, "count": count},
        )

    except Exception as e:
        messages.error(request, f'Error al actualizar categorias: {str(e)}')

    return _redirect_admin_product_list_with_filters(request)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def product_bulk_status_update(request):
    """Bulk activate/deactivate selected products."""
    raw_post_body = request.body
    set_active_raw = str(request.POST.get('set_active', '')).strip()
    if set_active_raw not in {'0', '1'}:
        messages.warning(request, 'Accion de estado invalida.')
        return _redirect_admin_product_list_with_filters(request)

    set_active = set_active_raw == '1'
    select_all_pages = request.POST.get('select_all_pages') == 'true'

    if select_all_pages:
        products_to_update, _, _, _ = get_product_queryset(request.POST)
    else:
        product_ids = extract_target_product_ids_from_post(request.POST, raw_post_body)
        if not product_ids:
            logger.warning(
                "product_bulk_status_update without selected products | user=%s | keys=%s | product_ids=%s | product_ids_csv=%s",
                getattr(request.user, "username", "unknown"),
                list(request.POST.keys()),
                request.POST.getlist("product_ids"),
                request.POST.get("product_ids_csv", ""),
            )
            messages.warning(request, 'No se seleccionaron productos.')
            return _redirect_admin_product_list_with_filters(request)
        products_to_update = Product.objects.filter(id__in=product_ids)

    target_ids = list(products_to_update.values_list('id', flat=True))
    if not target_ids:
        messages.info(request, 'No hubo productos para actualizar.')
        return _redirect_admin_product_list_with_filters(request)

    count = Product.objects.filter(id__in=target_ids).update(is_active=set_active)
    action_label = 'activados' if set_active else 'inactivados'
    messages.success(request, f'{count} productos {action_label}.')
    log_admin_action(
        request,
        action='product_bulk_status_update',
        target_type='product_bulk',
        details={'count': count, 'set_active': set_active, 'select_all_pages': select_all_pages},
    )
    return _redirect_admin_product_list_with_filters(request)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def product_bulk_image_update(request):
    """Bulk assign/clear product images for selected or filtered products."""
    image_mode = str(request.POST.get('image_mode', 'set')).strip().lower()
    select_all_pages = request.POST.get('select_all_pages') == 'true'
    only_missing = request.POST.get('only_missing') == '1'

    if image_mode not in {'set', 'clear'}:
        messages.warning(request, 'Modo de imagen invalido.')
        return _redirect_admin_product_list_with_filters(request)

    if select_all_pages:
        products_to_update, _, _, _ = get_product_queryset(request.POST)
    else:
        product_ids = extract_target_product_ids_from_post(request.POST, b"")
        if not product_ids:
            logger.warning(
                "product_bulk_image_update without selected products | user=%s | keys=%s | product_ids=%s | product_ids_csv=%s",
                getattr(request.user, "username", "unknown"),
                list(request.POST.keys()),
                request.POST.getlist("product_ids"),
                request.POST.get("product_ids_csv", ""),
            )
            messages.warning(request, 'No se seleccionaron productos.')
            return _redirect_admin_product_list_with_filters(request)
        products_to_update = Product.objects.filter(id__in=product_ids)

    # get_product_queryset() includes select_related/prefetch_related for list rendering;
    # reset those joins here to safely use deferred fields in bulk updates.
    products_to_update = products_to_update.select_related(None).prefetch_related(None)

    if only_missing and image_mode == 'set':
        products_to_update = products_to_update.filter(Q(image__isnull=True) | Q(image=''))

    if image_mode == 'set':
        uploaded_file = request.FILES.get('image_file')
        try:
            shared_image_name = _store_bulk_product_image(uploaded_file)
        except ValueError as exc:
            messages.error(request, str(exc))
            return _redirect_admin_product_list_with_filters(request)
        except Exception:
            messages.error(request, 'No se pudo almacenar la imagen. Intenta nuevamente.')
            return _redirect_admin_product_list_with_filters(request)

        # Bulk update to avoid worker timeouts when updating thousands of products.
        updated_count = (
            products_to_update
            .exclude(image=shared_image_name)
            .update(image=shared_image_name, updated_at=timezone.now())
        )

        if updated_count == 0:
            _delete_orphan_product_image(shared_image_name)
            messages.info(request, 'No hubo cambios: todos los productos ya tenian esa imagen.')
            return _redirect_admin_product_list_with_filters(request)

        messages.success(request, f'Imagen aplicada a {updated_count} productos.')
        log_admin_action(
            request,
            action='product_bulk_image_update',
            target_type='product_bulk',
            details={
                'count': updated_count,
                'mode': image_mode,
                'select_all_pages': select_all_pages,
                'only_missing': only_missing,
                'image': shared_image_name,
            },
        )
        return _redirect_admin_product_list_with_filters(request)

    # image_mode == "clear"
    updated_count = (
        products_to_update
        .exclude(Q(image__isnull=True) | Q(image=''))
        .update(image='', updated_at=timezone.now())
    )

    if updated_count:
        messages.success(request, f'Se quitaron imagenes en {updated_count} productos.')
    else:
        messages.info(request, 'No habia imagenes para quitar.')
    log_admin_action(
        request,
        action='product_bulk_image_update',
        target_type='product_bulk',
        details={
            'count': updated_count,
            'mode': image_mode,
            'select_all_pages': select_all_pages,
        },
    )
    return _redirect_admin_product_list_with_filters(request)


# ===================== SUPPLIERS =====================

@staff_member_required
def supplier_list(request):
    """Supplier directory with KPI summary."""
    search = sanitize_search_token(request.GET.get('q', ''))
    only_active = request.GET.get('only_active') == '1'

    suppliers_qs = Supplier.objects.all()

    if only_active:
        suppliers_qs = suppliers_qs.filter(is_active=True)
    suppliers_qs, search = apply_admin_text_search(
        suppliers_qs,
        search,
        ["name", "normalized_name", "slug"],
    )
    suppliers_qs = suppliers_qs.annotate(
        products_count=Count('products', distinct=True),
        active_products_count=Count('products', filter=Q(products__is_active=True), distinct=True),
        stock_total=Sum('products__stock'),
    ).order_by('name')

    uncategorized_products_count = Product.objects.filter(
        Q(supplier='') | Q(supplier__isnull=True) | Q(supplier_ref__isnull=True)
    ).count()

    paginator = Paginator(suppliers_qs, 40)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        'admin_panel/suppliers/list.html',
        {
            'page_obj': page_obj,
            'search': search,
            'only_active': only_active,
            'total_suppliers': suppliers_qs.count(),
            'uncategorized_products_count': uncategorized_products_count,
        },
    )


@staff_member_required
def supplier_detail(request, supplier_id):
    """List products belonging to one supplier."""
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    products, search, active_filter = build_supplier_products_queryset(supplier, request.GET)

    paginator = Paginator(products, 40)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    enrich_products_with_category_state(page_obj.object_list)

    metrics = products.aggregate(
        stock_total=Sum('stock'),
        avg_price=Avg('price'),
    )
    audit_logs = AdminAuditLog.objects.filter(
        target_type='supplier',
        target_id=str(supplier.pk),
    ).select_related('user').order_by('-created_at')[:12]

    return render(
        request,
        'admin_panel/suppliers/detail.html',
        {
            'supplier': supplier,
            'page_obj': page_obj,
            'search': search,
            'active_filter': active_filter,
            'total_products': products.count(),
            'stock_total': metrics.get('stock_total') or 0,
            'avg_price': metrics.get('avg_price') or Decimal('0.00'),
            'audit_logs': audit_logs,
        },
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def supplier_bulk_action(request, supplier_id):
    """
    Apply bulk operations to products of one supplier.
    """
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    products, _, _ = build_supplier_products_queryset(supplier, request.POST)

    action = request.POST.get('action', '').strip()
    percent_raw = request.POST.get('percent', '').strip()
    affected = 0

    if action in ('activate', 'deactivate'):
        is_active = action == 'activate'
        affected = products.update(is_active=is_active)
        messages.success(request, f'{affected} productos actualizados para {supplier.name}.')
    elif action in ('increase_pct', 'decrease_pct'):
        try:
            percent = Decimal(percent_raw)
            if percent <= 0:
                raise InvalidOperation()
        except Exception:
            messages.error(request, 'Porcentaje invalido. Debe ser mayor que cero.')
            return redirect(f"{reverse('admin_supplier_detail', args=[supplier.pk])}?q={request.POST.get('q', '')}&active={request.POST.get('active', '')}")

        factor = (Decimal('1') + (percent / Decimal('100'))) if action == 'increase_pct' else (
            Decimal('1') - (percent / Decimal('100'))
        )
        if factor <= 0:
            messages.error(request, 'El porcentaje deja los precios en cero o negativo.')
            return redirect(f"{reverse('admin_supplier_detail', args=[supplier.pk])}?q={request.POST.get('q', '')}&active={request.POST.get('active', '')}")

        to_update = []
        for product in products:
            product.price = (product.price * factor).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            to_update.append(product)

        if to_update:
            Product.objects.bulk_update(to_update, ['price'], batch_size=500)
        affected = len(to_update)
        messages.success(request, f'Precios actualizados en {affected} productos de {supplier.name}.')
    else:
        messages.warning(request, 'Accion invalida.')
        return redirect(f"{reverse('admin_supplier_detail', args=[supplier.pk])}?q={request.POST.get('q', '')}&active={request.POST.get('active', '')}")

    log_admin_action(
        request,
        action='supplier_bulk_action',
        target_type='supplier',
        target_id=supplier.pk,
        details={
            'supplier': supplier.name,
            'action': action,
            'affected': affected,
            'percent': percent_raw or None,
            'filters': {
                'q': request.POST.get('q', ''),
                'active': request.POST.get('active', ''),
            },
        },
    )

    return redirect(f"{reverse('admin_supplier_detail', args=[supplier.pk])}?q={request.POST.get('q', '')}&active={request.POST.get('active', '')}")


@staff_member_required
def supplier_export(request, supplier_id):
    """
    Export supplier products to CSV/XLSX.
    """
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    export_format = request.GET.get('format', 'xlsx').strip().lower()
    products, _, _ = build_supplier_products_queryset(supplier, request.GET)

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="proveedor_{supplier.slug}.csv"'
        writer = csv.writer(response)
        writer.writerow(['SKU', 'Nombre', 'Proveedor', 'Precio', 'Stock', 'Descripcion'])
        for product in products:
            writer.writerow([
                product.sku,
                product.name,
                supplier.name,
                str(product.price),
                product.stock,
                product.description or '',
            ])
        log_admin_action(
            request,
            action='supplier_export',
            target_type='supplier',
            target_id=supplier.pk,
            details={'format': 'csv', 'rows': products.count()},
        )
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'
    sheet.append(['SKU', 'Nombre', 'Proveedor', 'Precio', 'Stock', 'Descripcion'])
    for product in products:
        sheet.append([
            product.sku,
            product.name,
            supplier.name,
            float(product.price),
            product.stock,
            product.description or '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="proveedor_{supplier.slug}.xlsx"'
    workbook.save(response)

    log_admin_action(
        request,
        action='supplier_export',
        target_type='supplier',
        target_id=supplier.pk,
        details={'format': 'xlsx', 'rows': products.count()},
    )
    return response


@staff_member_required
def supplier_print(request, supplier_id):
    """
    Print-friendly supplier report (can be saved as PDF from browser).
    """
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    products, _, _ = build_supplier_products_queryset(supplier, request.GET)
    metrics = products.aggregate(stock_total=Sum('stock'))
    log_admin_action(
        request,
        action='supplier_export_print',
        target_type='supplier',
        target_id=supplier.pk,
        details={'rows': products.count()},
    )
    return render(
        request,
        'admin_panel/suppliers/print.html',
        {
            'supplier': supplier,
            'products': products,
            'stock_total': metrics.get('stock_total') or 0,
            'generated_at': timezone.now(),
        },
    )


@staff_member_required
def supplier_unassigned(request):
    """
    Products without supplier assigned.
    """
    products = Product.objects.select_related('category').prefetch_related('categories').filter(
        Q(supplier='') | Q(supplier__isnull=True) | Q(supplier_ref__isnull=True)
    ).order_by('name')

    products, search = apply_admin_text_search(
        products,
        request.GET.get('q', ''),
        ["sku", "name", "description", "supplier", "supplier_ref__name"],
    )

    paginator = Paginator(products, 40)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    enrich_products_with_category_state(page_obj.object_list)

    return render(
        request,
        'admin_panel/suppliers/unassigned.html',
        {
            'page_obj': page_obj,
            'search': search,
            'total_products': products.count(),
        },
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def supplier_toggle_active(request, supplier_id):
    """
    Activate/deactivate supplier entity.
    """
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    supplier.is_active = request.POST.get('is_active') == '1'
    supplier.save(update_fields=['is_active', 'updated_at'])
    log_admin_action(
        request,
        action='supplier_toggle_active',
        target_type='supplier',
        target_id=supplier.pk,
        details={'supplier': supplier.name, 'is_active': supplier.is_active},
    )
    messages.success(request, f'Proveedor {supplier.name} actualizado.')
    return redirect('admin_supplier_detail', supplier_id=supplier.pk)


# ===================== CATEGORIES =====================

@staff_member_required
def category_list(request):
    """Category list."""
    search = sanitize_search_token(request.GET.get('q', ''))
    status = request.GET.get('status', 'all').strip().lower()
    focus_raw = request.GET.get('focus', '').strip()

    all_categories = list(
        Category.objects.select_related('parent').order_by('order', 'name')
    )
    all_category_map = {c.id: c for c in all_categories}

    if status == 'active':
        filtered_categories = [c for c in all_categories if c.is_active]
    elif status == 'inactive':
        filtered_categories = [c for c in all_categories if not c.is_active]
    elif status == 'public':
        filtered_categories = [
            c for c in all_categories if c.is_active and getattr(c, 'visible_in_catalog', True)
        ]
    elif status == 'hidden':
        filtered_categories = [c for c in all_categories if not getattr(c, 'visible_in_catalog', True)]
    else:
        filtered_categories = all_categories

    category_map = {c.id: c for c in filtered_categories}
    child_index = {}
    for cat in filtered_categories:
        child_index.setdefault(cat.parent_id, []).append(cat.id)

    if search:
        parsed_search = normalize_admin_search_query(search)
        include_terms = [*parsed_search.get('phrases', []), *parsed_search.get('include_terms', [])]
        exclude_terms = parsed_search.get('exclude_terms', [])

        def _matches(cat):
            haystack = cat.name.lower()
            include_ok = all(term.lower() in haystack for term in include_terms) if include_terms else True
            exclude_ok = all(term.lower() not in haystack for term in exclude_terms)
            return include_ok and exclude_ok

        matched_ids = {c.id for c in filtered_categories if _matches(c)}
        visible_ids = set(matched_ids)

        # Keep the path to root visible so matches remain understandable in the tree.
        for cid in list(matched_ids):
            node = category_map.get(cid)
            parent = node.parent if node else None
            while parent:
                if parent.id not in category_map:
                    break
                visible_ids.add(parent.id)
                parent = parent.parent

        # Keep full branches when a parent node matches.
        pending = list(matched_ids)
        while pending:
            current = pending.pop()
            for child_id in child_index.get(current, []):
                if child_id not in visible_ids:
                    visible_ids.add(child_id)
                    pending.append(child_id)

        filtered_categories = [c for c in filtered_categories if c.id in visible_ids]
        category_map = {c.id: c for c in filtered_categories}

    children_map = {}
    for cat in filtered_categories:
        children_map.setdefault(cat.parent_id, []).append(cat)

    for child_list in children_map.values():
        child_list.sort(key=lambda c: (c.order, c.name.lower()))

    tree_rows = []

    def walk(node, depth):
        children = children_map.get(node.id, [])
        tree_rows.append({
            'category': node,
            'depth': depth,
            'has_children': bool(children),
            'children_count': len(children),
        })
        for child in children:
            walk(child, depth + 1)

    roots = [cat for cat in filtered_categories if cat.parent_id not in category_map]
    roots.sort(key=lambda c: (c.order, c.name.lower()))
    for root in roots:
        walk(root, 0)

    for row in tree_rows:
        category = row['category']
        tree_ids = category.get_descendant_ids(include_self=True)
        category.tree_products_count = Product.objects.filter(
            Q(categories__id__in=tree_ids) | Q(category_id__in=tree_ids)
        ).distinct().count()
        category.direct_products_count = category.products_m2m.count()

    integrity_issues = detect_category_integrity_issues(all_categories)
    catalog_diagnostics = build_category_catalog_diagnostics(all_categories)
    public_product_counts = catalog_diagnostics.get('public_category_product_counts', {})
    for row in tree_rows:
        category = row['category']
        category.public_products_count = public_product_counts.get(category.id, 0)

    focus_category_id = None
    auto_expand_ids = []
    if focus_raw.isdigit():
        candidate_id = int(focus_raw)
        if candidate_id in all_category_map:
            focus_category_id = candidate_id
            node = all_category_map[candidate_id]
            while node and node.parent_id:
                auto_expand_ids.append(node.parent_id)
                node = all_category_map.get(node.parent_id)

    return render(request, 'admin_panel/categories/list.html', {
        'tree_rows': tree_rows,
        'visible_total': len(tree_rows),
        'search': search,
        'status': status,
        'integrity_issues': integrity_issues,
        'catalog_diagnostics': catalog_diagnostics,
        'move_parent_options': build_category_options(
            all_categories,
            include_inactive_suffix=True,
        ),
        'focus_category_id': focus_category_id,
        'auto_expand_ids_json': json.dumps(auto_expand_ids),
    })


@staff_member_required
@superuser_required_for_modifications
def products_uncategorized(request):
    """View to show and quickly categorize products without category."""
    if request.method == 'POST':
        try:
            body_str = ""
            decode_error = None
            payload = {}
            if request.body:
                try:
                    body_str = request.body.decode('utf-8')
                    payload = json.loads(body_str)
                except Exception as e:
                    decode_error = str(e)

            if payload:
                action = payload.get('action', 'categorize')
                product_ids = payload.get('product_ids', [])
                category_id = payload.get('category_id')
                categories_ids = payload.get('categories', [])
                product_id = payload.get('product_id')
                name_val = payload.get('name')
                sku_val = payload.get('sku')
            else:
                action = request.POST.get('action', 'categorize')
                product_ids = request.POST.getlist('product_ids')
                category_id = request.POST.get('category_id')
                categories_ids = request.POST.getlist('categories')
                product_id = request.POST.get('product_id')
                name_val = request.POST.get('name')
                sku_val = request.POST.get('sku')

            # Process actions
            if action == 'quick_edit':
                if not product_id:
                    return JsonResponse({'success': False, 'error': 'Falta el ID del producto.'})
                prod = Product.objects.filter(pk=product_id).first()
                if not prod:
                    return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
                
                if name_val is not None:
                    name_clean = name_val.strip()
                    if not name_clean:
                        return JsonResponse({'success': False, 'error': 'El nombre no puede estar vacío.'})
                    prod.name = name_clean
                
                if sku_val is not None:
                    sku_clean = sku_val.strip()
                    if not sku_clean:
                        return JsonResponse({'success': False, 'error': 'El SKU no puede estar vacío.'})
                    if Product.objects.filter(sku=sku_clean).exclude(pk=product_id).exists():
                        return JsonResponse({'success': False, 'error': f'El SKU "{sku_clean}" ya está en uso.'})
                    prod.sku = sku_clean

                prod.save()
                log_admin_action(
                    request,
                    action="products_uncategorized_quick_edit",
                    target_type="product",
                    target_id=prod.pk,
                    details={'name': prod.name, 'sku': prod.sku},
                )
                return JsonResponse({'success': True})

            elif action == 'bulk_deactivate':
                product_ids = [int(x) for x in product_ids if str(x).isdigit()]
                if not product_ids:
                    return JsonResponse({'success': False, 'error': 'No se seleccionaron productos.'})
                products_qs = Product.objects.filter(pk__in=product_ids)
                count = products_qs.count()
                for prod in products_qs:
                    prod.is_active = False
                    prod.save()
                    log_admin_action(
                        request,
                        action="products_uncategorized_deactivate",
                        target_type="product",
                        target_id=prod.pk,
                    )
                return JsonResponse({'success': True, 'count': count})

            elif action == 'bulk_delete':
                product_ids = [int(x) for x in product_ids if str(x).isdigit()]
                if not product_ids:
                    return JsonResponse({'success': False, 'error': 'No se seleccionaron productos.'})
                products_qs = Product.objects.filter(pk__in=product_ids)
                count = products_qs.count()
                for prod in products_qs:
                    pk = prod.pk
                    prod.delete()
                    log_admin_action(
                        request,
                        action="products_uncategorized_delete",
                        target_type="product",
                        target_id=pk,
                    )
                return JsonResponse({'success': True, 'count': count})

            elif action == 'categorize':
                product_ids = [int(x) for x in product_ids if str(x).isdigit()]
                categories_ids = [int(x) for x in categories_ids if str(x).isdigit()]
                category_id = int(category_id) if category_id and str(category_id).isdigit() else None

                if not product_ids:
                    return JsonResponse({'success': False, 'error': 'No se seleccionaron productos.'})

                primary_cat = None
                if category_id:
                    primary_cat = Category.objects.filter(pk=category_id).first()

                products_qs = Product.objects.filter(pk__in=product_ids)
                
                for prod in products_qs:
                    if primary_cat:
                        prod.category = primary_cat
                    if categories_ids:
                        cats = Category.objects.filter(pk__in=categories_ids)
                        prod.categories.set(cats)
                        if not prod.category and cats.exists():
                            prod.category = cats.first()
                    elif category_id and not categories_ids:
                        prod.categories.set([primary_cat])
                    
                    prod.save()
                    
                    log_admin_action(
                        request,
                        action="products_uncategorized_assign",
                        target_type="product",
                        target_id=prod.pk,
                        details={'msg': f'Categorización rápida. Primaria: {prod.category_id}, M2M: {list(categories_ids)}'},
                    )
                    
                return JsonResponse({'success': True, 'count': products_qs.count()})
            else:
                return JsonResponse({'success': False, 'error': f'Acción no válida: {action}'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    # GET Request
    q = request.GET.get('q', '').strip()
    supplier_id = request.GET.get('supplier', '').strip()
    
    products = Product.objects.filter(category__isnull=True, categories__isnull=True)
    
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q) | Q(supplier__icontains=q))
    if supplier_id and supplier_id.isdigit():
        products = products.filter(supplier_ref_id=int(supplier_id))
        
    products = products.select_related('supplier_ref').order_by('name')
    
    # Smart suggestions matching
    all_cats = list(Category.objects.filter(is_active=True))
    sorted_cats = sorted(all_cats, key=lambda c: len(c.name), reverse=True)
    
    for prod in products:
        prod_name_lower = prod.name.lower()
        suggested = None
        for cat in sorted_cats:
            if len(cat.name) >= 3 and cat.name.lower() in prod_name_lower:
                suggested = cat
                break
        prod.suggested_category = suggested
        
    category_options = get_cached_category_options(only_active=True, include_inactive_suffix=False)
    suppliers = Supplier.objects.all().order_by('name')
    
    return render(request, 'admin_panel/categories/uncategorized.html', {
        'products': products,
        'category_options': category_options,
        'suppliers': suppliers,
        'q': q,
        'selected_supplier_id': int(supplier_id) if supplier_id.isdigit() else None,
    })


@staff_member_required
@require_POST
@superuser_required_for_modifications
def category_reorder(request):
    """
    Reorder categories by a user-provided list of IDs.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        ordered_ids = payload.get("ordered_ids", [])
        if not isinstance(ordered_ids, list) or not ordered_ids:
            return JsonResponse({"success": False, "error": "Datos invalidos."}, status=400)

        normalized_ids = normalize_category_ids(ordered_ids)
        categories = {cat.id: cat for cat in Category.objects.filter(id__in=normalized_ids)}
        updates = []
        for index, category_id in enumerate(normalized_ids):
            category = categories.get(category_id)
            if category is None:
                continue
            category.order = (index + 1) * 10
            category.updated_at = timezone.now()
            updates.append(category)

        if updates:
            Category.objects.bulk_update(updates, ["order", "updated_at"], batch_size=500)

        log_admin_action(
            request,
            action="category_reorder",
            target_type="category_bulk",
            details={"ordered_ids": normalized_ids[:100], "count": len(updates)},
        )
        return JsonResponse({"success": True, "updated": len(updates)})
    except Exception as exc:
        logger.exception("Error reordering categories")
        return JsonResponse({"success": False, "error": str(exc)}, status=400)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def category_sort_roots_alpha(request):
    """
    Sort only root categories alphabetically, keeping ABRAZADERAS as the first root.
    Subcategories keep their parent and current order.
    """
    root_categories = list(Category.objects.filter(parent__isnull=True))

    def root_sort_key(category):
        public_or_internal_name = (category.public_name or category.name or "").strip()
        normalized_names = {
            slugify(category.name or ""),
            slugify(category.public_name or ""),
        }
        return (
            0 if "abrazaderas" in normalized_names else 1,
            public_or_internal_name.casefold(),
            category.id,
        )

    ordered_roots = sorted(root_categories, key=root_sort_key)
    now = timezone.now()
    updates = []
    for index, category in enumerate(ordered_roots):
        order_value = (index + 1) * 10
        if category.order == order_value and category.public_order == order_value:
            continue
        category.order = order_value
        category.public_order = order_value
        category.updated_at = now
        updates.append(category)

    if updates:
        Category.objects.bulk_update(updates, ["order", "public_order", "updated_at"], batch_size=500)

    log_admin_action(
        request,
        action="category_sort_roots_alpha",
        target_type="category_bulk",
        details={
            "scope": "root_categories",
            "updated": len(updates),
            "total_roots": len(root_categories),
            "first": ordered_roots[0].name if ordered_roots else None,
        },
    )
    messages.success(
        request,
        f"Se ordenaron {len(root_categories)} categorias principales. ABRAZADERAS queda primera y las subcategorias no se tocaron.",
    )
    return redirect("admin_category_list")


@staff_member_required
@require_POST
@superuser_required_for_modifications
def category_bulk_status(request):
    """
    Bulk maintenance actions for selected categories.
    """
    action = request.POST.get("bulk_action", "").strip().lower()
    selected_ids = normalize_category_ids(request.POST.getlist("category_ids"))
    target_raw = request.POST.get("target_category_id", "").strip()
    q = request.POST.get("q", "").strip()
    status = request.POST.get("status", "all").strip().lower()

    redirect_url = reverse("admin_category_list")
    if q or status != "all":
        params = {}
        if q:
            params["q"] = q
        if status:
            params["status"] = status
        redirect_url = f"{redirect_url}?{urlencode(params)}"

    if not selected_ids:
        messages.warning(request, "No se seleccionaron categorias.")
        return redirect(redirect_url)

    valid_actions = {
        "activate",
        "deactivate",
        "show_catalog",
        "hide_catalog",
        "feature",
        "unfeature",
        "move",
        "clear_parent",
        "delete_empty",
        "delete_keep_products",
        "merge_delete",
    }
    if action not in valid_actions:
        messages.error(request, "Accion masiva invalida.")
        return redirect(redirect_url)

    categories_map = {
        category.id: category
        for category in Category.objects.select_related("parent").filter(id__in=selected_ids)
    }
    if not categories_map:
        messages.warning(request, "No se encontraron categorias validas para actualizar.")
        return redirect(redirect_url)

    selected_set = set(categories_map.keys())
    ordered_categories = sorted(
        categories_map.values(),
        key=lambda cat: len(cat.get_ancestor_ids(include_self=True)),
    )

    def _root_selected_categories():
        roots = []
        for category in ordered_categories:
            ancestor_ids = category.get_ancestor_ids(include_self=False)
            if not any(ancestor_id in selected_set for ancestor_id in ancestor_ids):
                roots.append(category)
        return roots

    def _deep_selected_categories():
        return sorted(
            ordered_categories,
            key=lambda cat: len(cat.get_ancestor_ids(include_self=True)),
            reverse=True,
        )

    def _linked_product_count(category):
        return Product.objects.filter(
            Q(category_id=category.pk) | Q(categories__id=category.pk)
        ).distinct().count()

    def _resolve_target_category():
        if not target_raw:
            return None, "Selecciona una categoria destino."
        if not target_raw.isdigit():
            return None, "Categoria destino invalida."
        target = Category.objects.filter(pk=int(target_raw)).first()
        if not target:
            return None, "No se encontro la categoria destino."
        if target.pk in selected_set:
            return None, "La categoria destino no puede estar dentro de la seleccion."
        return target, ""

    if action == "deactivate":
        impacted_ids = set()
        for category in ordered_categories:
            impacted_ids.update(category.get_descendant_ids(include_self=True))

        active_before_ids = set(
            Category.objects.filter(id__in=impacted_ids, is_active=True).values_list("id", flat=True)
        )

        for category in ordered_categories:
            if category.is_active:
                category.is_active = False
                category.save()

        active_after_ids = set(
            Category.objects.filter(id__in=impacted_ids, is_active=True).values_list("id", flat=True)
        )
        deactivated_total = len(active_before_ids - active_after_ids)
        direct_selected = len(
            [cid for cid in categories_map.keys() if cid in active_before_ids]
        )
        cascaded = max(deactivated_total - direct_selected, 0)

        messages.success(
            request,
            f"Categorias desactivadas: {deactivated_total} en total "
            f"({direct_selected} seleccionadas y {cascaded} por cascada).",
        )
        log_admin_action(
            request,
            action="category_bulk_deactivate",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "deactivated_total": deactivated_total,
                "cascaded": cascaded,
            },
        )
        return redirect(redirect_url)

    if action == "hide_catalog":
        impacted_ids = set()
        for category in ordered_categories:
            impacted_ids.update(category.get_descendant_ids(include_self=True))

        visible_before_ids = set(
            Category.objects.filter(
                id__in=impacted_ids,
                visible_in_catalog=True,
            ).values_list("id", flat=True)
        )

        for category in ordered_categories:
            if category.visible_in_catalog:
                category.visible_in_catalog = False
                category.save()

        visible_after_ids = set(
            Category.objects.filter(
                id__in=impacted_ids,
                visible_in_catalog=True,
            ).values_list("id", flat=True)
        )
        hidden_total = len(visible_before_ids - visible_after_ids)
        messages.success(request, f"Categorias ocultadas del catalogo cliente: {hidden_total}.")
        log_admin_action(
            request,
            action="category_bulk_hide_catalog",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "hidden_total": hidden_total,
            },
        )
        return redirect(redirect_url)

    if action == "show_catalog":
        visible_before = {
            category.id: category.visible_in_catalog for category in ordered_categories
        }
        blocked = 0
        with transaction.atomic():
            for category in ordered_categories:
                category.visible_in_catalog = True
                category.save()

        refreshed = Category.objects.filter(id__in=categories_map.keys())
        shown = 0
        for category in refreshed:
            if not visible_before.get(category.id, False) and category.visible_in_catalog:
                shown += 1
            if not category.visible_in_catalog:
                blocked += 1

        if blocked:
            messages.warning(
                request,
                f"Se hicieron visibles {shown} categorias. "
                f"{blocked} siguen ocultas porque dependen de un padre oculto.",
            )
        else:
            messages.success(request, f"Categorias visibles para cliente: {shown}.")
        log_admin_action(
            request,
            action="category_bulk_show_catalog",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "shown": shown,
                "blocked": blocked,
            },
        )
        return redirect(redirect_url)

    if action in {"feature", "unfeature"}:
        featured_value = action == "feature"
        updated = Category.objects.filter(id__in=selected_ids).update(
            is_featured=featured_value,
            updated_at=timezone.now(),
        )
        label = "destacadas" if featured_value else "quitadas de destacados"
        messages.success(request, f"Categorias {label}: {updated}.")
        log_admin_action(
            request,
            action=f"category_bulk_{action}",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "updated": updated,
            },
        )
        return redirect(redirect_url)

    if action in {"move", "clear_parent"}:
        new_parent = None
        if action == "move":
            new_parent, error = _resolve_target_category()
            if error:
                messages.error(request, error)
                return redirect(redirect_url)

        moved = 0
        skipped = []
        with transaction.atomic():
            for category in _root_selected_categories():
                if not category.can_move_to(new_parent):
                    skipped.append(category.name)
                    continue
                category.move_to(new_parent)
                moved += 1

        destination = new_parent.name if new_parent else "raiz"
        if skipped:
            messages.warning(
                request,
                f"Se movieron {moved} categorias a {destination}. "
                f"Se omitieron {len(skipped)} por riesgo de ciclo.",
            )
        else:
            messages.success(request, f"Se movieron {moved} categorias a {destination}.")
        log_admin_action(
            request,
            action="category_bulk_move",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "moved": moved,
                "target_id": new_parent.pk if new_parent else None,
                "skipped": skipped[:50],
            },
        )
        return redirect(redirect_url)

    if action == "delete_empty":
        deleted = 0
        skipped_with_products = 0
        skipped_with_children = 0
        deleted_names = []
        with transaction.atomic():
            for category in _deep_selected_categories():
                if _linked_product_count(category):
                    skipped_with_products += 1
                    continue
                if Category.objects.filter(parent_id=category.pk).exists():
                    skipped_with_children += 1
                    continue
                deleted_names.append(category.name)
                category.delete()
                deleted += 1

        if deleted:
            messages.success(request, f"Categorias vacias eliminadas: {deleted}.")
        if skipped_with_products or skipped_with_children:
            messages.warning(
                request,
                "Se protegieron categorias con productos o subcategorias: "
                f"{skipped_with_products} con productos, {skipped_with_children} con subcategorias.",
            )
        log_admin_action(
            request,
            action="category_bulk_delete_empty",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "deleted": deleted,
                "deleted_names": deleted_names[:100],
                "skipped_with_products": skipped_with_products,
                "skipped_with_children": skipped_with_children,
            },
        )
        return redirect(redirect_url)

    if action == "delete_keep_products":
        linked_products_count = Product.objects.filter(
            Q(category_id__in=selected_ids) | Q(categories__id__in=selected_ids)
        ).distinct().count()
        unselected_children_count = Category.objects.filter(
            parent_id__in=selected_ids,
        ).exclude(id__in=selected_ids).count()
        deleted = 0
        deleted_names = []
        with transaction.atomic():
            for category in _deep_selected_categories():
                deleted_names.append(category.name)
                category.delete()
                deleted += 1

        messages.success(
            request,
            f"Categorias eliminadas: {deleted}. Productos conservados: {linked_products_count}.",
        )
        if unselected_children_count:
            messages.warning(
                request,
                f"{unselected_children_count} subcategorias quedaron en la raiz para no borrarlas.",
            )
        log_admin_action(
            request,
            action="category_bulk_delete_keep_products",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "deleted": deleted,
                "deleted_names": deleted_names[:100],
                "linked_products_count": linked_products_count,
                "unselected_children_count": unselected_children_count,
            },
        )
        return redirect(redirect_url)

    if action == "merge_delete":
        target, error = _resolve_target_category()
        if error:
            messages.error(request, error)
            return redirect(redirect_url)

        merged = 0
        moved_primary_links = 0
        moved_m2m_links = 0
        moved_children = 0
        skipped = []
        with transaction.atomic():
            for category in _deep_selected_categories():
                if not target or target.pk in category.get_descendant_ids(include_self=True):
                    skipped.append(category.name)
                    continue

                primary_count = Product.objects.filter(category_id=category.pk).update(category_id=target.pk)
                moved_primary_links += primary_count

                linked_products = Product.objects.filter(categories=category).distinct()
                linked_count = linked_products.count()
                for product in linked_products.iterator():
                    product.categories.add(target)
                    product.categories.remove(category)
                moved_m2m_links += linked_count

                children = list(Category.objects.filter(parent_id=category.pk))
                for child in children:
                    if child.pk == target.pk:
                        continue
                    child.move_to(target)
                    moved_children += 1

                category.delete()
                merged += 1

        if skipped:
            messages.warning(
                request,
                f"Se fusionaron {merged} categorias en {target.name}. "
                f"Se omitieron {len(skipped)} por relacion jerarquica invalida.",
            )
        else:
            messages.success(
                request,
                f"Se fusionaron {merged} categorias en {target.name}. "
                f"Productos movidos: {moved_primary_links + moved_m2m_links}.",
            )
        log_admin_action(
            request,
            action="category_bulk_merge_delete",
            target_type="category_bulk",
            details={
                "selected_count": len(categories_map),
                "selected_ids": list(categories_map.keys())[:200],
                "target_id": target.pk,
                "merged": merged,
                "moved_primary_links": moved_primary_links,
                "moved_m2m_links": moved_m2m_links,
                "moved_children": moved_children,
                "skipped": skipped[:50],
            },
        )
        return redirect(redirect_url)

    # action == "activate"
    selected_before_active = {
        category.id: category.is_active for category in ordered_categories
    }

    for category in ordered_categories:
        if not category.is_active:
            category.is_active = True
            category.save()

    refreshed = {
        category.id: category
        for category in Category.objects.select_related("parent").filter(id__in=categories_map.keys())
    }
    activated = sum(
        1
        for cid, was_active in selected_before_active.items()
        if not was_active and refreshed.get(cid) and refreshed[cid].is_active
    )
    blocked = sum(
        1
        for cid, category in refreshed.items()
        if not category.is_active
    )

    if blocked:
        messages.warning(
            request,
            f"Se activaron {activated} categorias. "
            f"{blocked} no pudieron activarse porque tienen un padre inactivo.",
        )
    else:
        messages.success(request, f"Se activaron {activated} categorias seleccionadas.")

    log_admin_action(
        request,
        action="category_bulk_activate",
        target_type="category_bulk",
        details={
            "selected_count": len(categories_map),
            "selected_ids": list(categories_map.keys())[:200],
            "activated": activated,
            "blocked": blocked,
        },
    )
    return redirect(redirect_url)


@staff_member_required
@superuser_required_for_modifications
def category_create(request):
    """Create category."""
    parent_from_query = None
    parent_query_id = request.GET.get('parent', '').strip()
    if parent_query_id.isdigit():
        parent_from_query = Category.objects.filter(pk=int(parent_query_id)).first()

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        form.fields['parent'].queryset = Category.objects.order_by('order', 'name')
        if form.is_valid():
            category = form.save()
            log_admin_action(
                request,
                action="category_create",
                target_type="category",
                target_id=category.pk,
                details={"name": category.name, "parent_id": category.parent_id},
            )
            messages.success(request, f'Categoria "{category.name}" creada.')
            return redirect('admin_category_list')
    else:
        initial = {}
        if parent_from_query:
            initial['parent'] = parent_from_query.pk
        form = CategoryForm(initial=initial)
        form.fields['parent'].queryset = Category.objects.order_by('order', 'name')

    selected_parent_id = form['parent'].value()
    selected_parent = None
    if str(selected_parent_id).isdigit():
        selected_parent = Category.objects.filter(pk=int(selected_parent_id)).first()

    parent_options = build_category_options(
        form.fields['parent'].queryset,
        include_inactive_suffix=True,
    )

    return render(request, 'admin_panel/categories/form.html', {
        'form': form,
        'parent_options': parent_options,
        'selected_parent_id': str(selected_parent_id or ''),
        'selected_parent': selected_parent,
        'action': 'Crear',
    })


@staff_member_required
@superuser_required_for_modifications
@require_POST
def category_create_ajax(request):
    """Create a new category via AJAX and return the updated tree options."""
    name = request.POST.get('name', '').strip()
    parent_id = request.POST.get('parent_id', '').strip()

    if not name:
        return JsonResponse({'success': False, 'error': 'El nombre de la categoría es obligatorio.'})

    parent = None
    if parent_id.isdigit():
        parent = Category.objects.filter(pk=int(parent_id)).first()
        if not parent:
            return JsonResponse({'success': False, 'error': 'La categoría padre seleccionada no existe.'})

    try:
        # Create category with safe default slug and active status
        category = Category.objects.create(
            name=name,
            parent=parent,
            is_active=True,
            visible_in_catalog=True,
        )

        log_admin_action(
            request,
            action="category_create_ajax",
            target_type="category",
            target_id=category.pk,
            details={"name": category.name, "parent_id": category.parent_id},
        )

        # Clear and reload cache of category choices
        category_options = get_cached_category_options(only_active=True, include_inactive_suffix=False)

        return JsonResponse({
            'success': True,
            'new_category_id': category.pk,
            'categories': category_options
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al guardar la categoría: {str(e)}'})


@staff_member_required
@superuser_required_for_modifications
def category_edit(request, pk):
    """Edit category."""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        form.fields['parent'].queryset = Category.objects.exclude(pk=pk).order_by('order', 'name')
        if form.is_valid():
            updated = form.save()
            log_admin_action(
                request,
                action="category_edit",
                target_type="category",
                target_id=updated.pk,
                details={
                    "name": updated.name,
                    "parent_id": updated.parent_id,
                    "is_active": updated.is_active,
                },
            )
            messages.success(request, f'Categoría "{category.name}" actualizada.')
            return redirect('admin_category_list')
    else:
        form = CategoryForm(instance=category)
        # Exclude self from parents to avoid recursion
        form.fields['parent'].queryset = Category.objects.exclude(pk=pk).order_by('order', 'name')

    selected_parent_id = form['parent'].value()
    selected_parent = None
    if str(selected_parent_id).isdigit():
        selected_parent = Category.objects.filter(pk=int(selected_parent_id)).first()

    parent_options = build_category_options(
        form.fields['parent'].queryset,
        include_inactive_suffix=True,
    )
    deactivation_impact = calculate_category_deactivation_impact(category)

    return render(request, 'admin_panel/categories/form.html', {
        'form': form,
        'parent_options': parent_options,
        'selected_parent_id': str(selected_parent_id or ''),
        'selected_parent': selected_parent,
        'deactivation_impact': deactivation_impact,
        'category': category, # Keep category in context for attributes links
        'action': 'Editar',
    })


@staff_member_required
@superuser_required_for_modifications
@require_POST
def category_move(request, pk):
    """
    Move one category node to a new parent category.
    The full subtree moves with it.
    """
    category = get_object_or_404(Category, pk=pk)
    parent_raw = request.POST.get('parent_id', '').strip()

    new_parent = None
    if parent_raw:
        if not parent_raw.isdigit():
            messages.error(request, 'Categoria padre invalida.')
            return redirect('admin_category_list')
        new_parent = get_object_or_404(Category, pk=int(parent_raw))

    if category.parent_id == (new_parent.pk if new_parent else None):
        messages.info(request, f'La categoria "{category.name}" ya estaba en esa ubicacion.')
        return redirect(f"{reverse('admin_category_list')}?focus={category.pk}")

    if not category.can_move_to(new_parent):
        messages.error(
            request,
            'Movimiento invalido: no puedes mover una categoria dentro de si misma o de una subcategoria suya.',
        )
        return redirect(f"{reverse('admin_category_list')}?focus={category.pk}")

    previous_parent = category.parent
    subtree_size = len(category.get_descendant_ids(include_self=True))
    forced_deactivation = bool(new_parent and not new_parent.is_active and category.is_active)

    category.move_to(new_parent)

    log_admin_action(
        request,
        action='category_move',
        target_type='category',
        target_id=category.pk,
        details={
            'name': category.name,
            'from_parent_id': previous_parent.pk if previous_parent else None,
            'to_parent_id': new_parent.pk if new_parent else None,
            'to_parent_name': new_parent.name if new_parent else '',
            'subtree_size': subtree_size,
            'forced_deactivation': forced_deactivation,
        },
    )

    destination = new_parent.name if new_parent else 'raiz'
    if forced_deactivation:
        messages.success(
            request,
            f'Categoria "{category.name}" movida a "{destination}". '
            'Se desactivo automaticamente junto con su arbol porque el nuevo padre esta inactivo.',
        )
    else:
        messages.success(
            request,
            f'Categoria "{category.name}" movida a "{destination}" con {subtree_size} nodo(s) en su arbol.',
        )
    return redirect(f"{reverse('admin_category_list')}?focus={category.pk}")


@staff_member_required
@superuser_required_for_modifications
def category_delete(request, pk):
    """Delete single category."""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        name = category.name
        category_id = category.pk
        linked_products_count = Product.objects.filter(
            Q(category_id=category.pk) | Q(categories__id=category.pk)
        ).distinct().count()
        child_count = Category.objects.filter(parent_id=category.pk).count()
        category.delete()
        log_admin_action(
            request,
            action="category_delete",
            target_type="category",
            target_id=category_id,
            details={
                "name": name,
                "linked_products_count": linked_products_count,
                "child_count": child_count,
            },
        )
        messages.success(
            request,
            f'Categoria "{name}" eliminada. Productos conservados: {linked_products_count}.',
        )
        if child_count:
            messages.warning(request, f"{child_count} subcategorias quedaron en la raiz.")
        return redirect('admin_category_list')
        
    return render(request, 'admin_panel/delete_confirm.html', {
        'object': f"Categoría: {category.name}",
        'cancel_url': reverse('admin_category_list')
    })


@staff_member_required
@superuser_required_for_modifications
def category_attribute_create(request, category_id):
    """Create new category attribute."""
    category = get_object_or_404(Category, pk=category_id)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            slug = request.POST.get('slug', '').strip()
            attr_type = request.POST.get('type', 'text')
            options = request.POST.get('options', '')
            required = request.POST.get('required') == 'on'
            is_recommended = request.POST.get('is_recommended') == 'on'
            regex_pattern = request.POST.get('regex_pattern', '').strip()
            
            # Simple validation for slug
            if CategoryAttribute.objects.filter(category=category, slug=slug).exists():
                messages.error(request, f'El slug "{slug}" ya existe en esta categoría.')
            else:
                CategoryAttribute.objects.create(
                    category=category,
                    name=name,
                    slug=slug,
                    type=attr_type,
                    options=options,
                    required=required,
                    is_recommended=is_recommended,
                    regex_pattern=regex_pattern
                )
                messages.success(request, f'Atributo "{name}" agregado.')
                return redirect('admin_category_edit', pk=category.pk)
        except Exception as e:
            messages.error(request, f'Error al crear atributo: {str(e)}')
    
    return render(request, 'admin_panel/categories/attribute_form.html', {
        'category': category,
        'action': 'Crear',
    })


@staff_member_required
@superuser_required_for_modifications
def category_attribute_edit(request, category_id, attribute_id):
    """Edit existing category attribute."""
    category = get_object_or_404(Category, pk=category_id)
    attribute = get_object_or_404(CategoryAttribute, pk=attribute_id, category=category)
    
    if request.method == 'POST':
        try:
            attribute.name = request.POST.get('name', '').strip()
            # Slug shouldn't change generally, but legal here if unique
            new_slug = request.POST.get('slug', '').strip()
            if new_slug != attribute.slug and CategoryAttribute.objects.filter(category=category, slug=new_slug).exists():
                messages.error(request, f'El slug "{new_slug}" ya existe.')
                return redirect(request.path)
            
            attribute.slug = new_slug
            attribute.type = request.POST.get('type', 'text')
            attribute.options = request.POST.get('options', '')
            attribute.required = request.POST.get('required') == 'on'
            attribute.is_recommended = request.POST.get('is_recommended') == 'on'
            attribute.regex_pattern = request.POST.get('regex_pattern', '').strip()
            attribute.save()
            
            messages.success(request, f'Atributo "{attribute.name}" actualizado.')
            return redirect('admin_category_edit', pk=category.pk)
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
            
    return render(request, 'admin_panel/categories/attribute_form.html', {
        'category': category,
        'attribute': attribute,
        'action': 'Editar',
    })


@staff_member_required
@superuser_required_for_modifications
def category_attribute_delete(request, category_id, attribute_id):
    """Delete a category attribute."""
    category = get_object_or_404(Category, pk=category_id)
    attribute = get_object_or_404(CategoryAttribute, pk=attribute_id, category=category)
    
    name = attribute.name
    attribute.delete()
    messages.success(request, f'Atributo "{name}" eliminado.')
    
    return redirect('admin_category_edit', pk=category.pk)


def _clean_order_block_label(value):
    return str(value or "").strip()[:120]


def _normalize_order_import_header(value):
    value = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    value = "".join(char for char in value if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", value).strip("_")


def _parse_order_position(value, fallback):
    if value in (None, ""):
        return fallback
    try:
        parsed = int(Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return fallback
    return max(parsed, 1)


def _parse_category_order_import_file(uploaded_file):
    if not uploaded_file:
        raise ValueError("Subi un archivo XLSX, XLS o CSV con columnas SKU y orden.")

    filename = (uploaded_file.name or "").lower()
    rows = []
    if filename.endswith(".csv"):
        raw = uploaded_file.read()
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(StringIO(text))
        rows = list(reader)
    else:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]

    if not rows:
        raise ValueError("El archivo esta vacio.")

    headers = [_normalize_order_import_header(value) for value in rows[0]]
    aliases = {
        "sku": {"sku", "codigo", "cod", "codigo_producto", "codigo_original"},
        "order": {"orden", "posicion", "position", "orden_catalogo", "orden_producto"},
        "block": {"bloque", "block", "grupo", "seccion"},
        "block_order": {"orden_bloque", "posicion_bloque", "block_order", "grupo_orden"},
    }

    indexes = {}
    for field, options in aliases.items():
        for index, header in enumerate(headers):
            if header in options:
                indexes[field] = index
                break

    if "sku" not in indexes:
        raise ValueError("No encontre columna SKU/Codigo en el archivo.")

    parsed_rows = []
    block_orders = {}
    for row_number, row in enumerate(rows[1:], start=2):
        def cell(field):
            index = indexes.get(field)
            if index is None or index >= len(row):
                return ""
            return row[index]

        sku = str(cell("sku") or "").strip()
        if not sku:
            continue

        block_label = _clean_order_block_label(cell("block"))
        if block_label and block_label not in block_orders:
            block_orders[block_label] = len(block_orders) + 1

        block_order_value = cell("block_order")
        block_order = _parse_order_position(block_order_value, block_orders.get(block_label, 0))
        position = _parse_order_position(cell("order"), len(parsed_rows) + 1)
        parsed_rows.append(
            {
                "row": row_number,
                "sku": sku,
                "position": position,
                "block_label": block_label,
                "block_order": block_order,
            }
        )

    if not parsed_rows:
        raise ValueError("No encontre filas con SKU para importar orden.")
    return parsed_rows


def _save_category_product_order_entries(category, ordered_entries):
    product_ids = [entry["product_id"] for entry in ordered_entries]
    linked_ids = set(
        Product.categories.through.objects.filter(
            category_id=category.id,
            product_id__in=product_ids,
        ).values_list("product_id", flat=True)
    )
    normalized_entries = [
        entry for entry in ordered_entries
        if entry["product_id"] in linked_ids
    ]
    if not normalized_entries:
        return 0

    existing_rows = {
        row.product_id: row
        for row in CategoryProductOrder.objects.filter(
            category=category,
            product_id__in=[entry["product_id"] for entry in normalized_entries],
        )
    }
    now = timezone.now()
    updates = []
    creates = []
    for index, entry in enumerate(normalized_entries, start=1):
        product_id = entry["product_id"]
        block_label = _clean_order_block_label(entry.get("block_label"))
        try:
            block_order = int(entry.get("block_order") or 0)
        except (TypeError, ValueError):
            block_order = 0
        sort_order = index * 10
        row = existing_rows.get(product_id)
        if row:
            row.block_label = block_label
            row.block_order = max(block_order, 0)
            row.sort_order = sort_order
            row.updated_at = now
            updates.append(row)
        else:
            creates.append(
                CategoryProductOrder(
                    category=category,
                    product_id=product_id,
                    block_label=block_label,
                    block_order=max(block_order, 0),
                    sort_order=sort_order,
                )
            )

    if creates:
        CategoryProductOrder.objects.bulk_create(creates, ignore_conflicts=True, batch_size=500)
    if updates:
        CategoryProductOrder.objects.bulk_update(
            updates,
            ["block_label", "block_order", "sort_order", "updated_at"],
            batch_size=500,
        )
    return len(normalized_entries)


def _save_category_product_block_order(category, raw_block_labels):
    cleaned_labels = []
    for raw_label in raw_block_labels:
        if raw_label == "__no_block__":
            label = ""
        else:
            label = _clean_order_block_label(raw_label)
        if label not in cleaned_labels:
            cleaned_labels.append(label)

    if not cleaned_labels:
        return 0

    block_order_by_label = {
        label: (index + 1) * 10
        for index, label in enumerate(cleaned_labels)
    }
    linked_ids = list(
        Product.categories.through.objects.filter(category_id=category.id)
        .values_list("product_id", flat=True)
    )
    if not linked_ids:
        return 0

    rows = list(
        CategoryProductOrder.objects.filter(
            category=category,
            product_id__in=linked_ids,
        )
    )
    rows_by_product = {row.product_id: row for row in rows}
    now = timezone.now()
    updates = []
    for row in rows:
        label = row.block_label or ""
        if label in block_order_by_label:
            row.block_order = block_order_by_label[label]
            row.updated_at = now
            updates.append(row)

    creates = []
    if "" in block_order_by_label:
        products_with_named_block = {
            row.product_id
            for row in rows
            if row.block_label
        }
        max_sort_order = (
            CategoryProductOrder.objects.filter(category=category)
            .aggregate(max_order=Max("sort_order"))
            .get("max_order")
            or 0
        )
        missing_blank_product_ids = [
            product_id
            for product_id in linked_ids
            if product_id not in products_with_named_block and product_id not in rows_by_product
        ]
        for index, product_id in enumerate(missing_blank_product_ids, start=1):
            creates.append(
                CategoryProductOrder(
                    category=category,
                    product_id=product_id,
                    block_label="",
                    block_order=block_order_by_label[""],
                    sort_order=max_sort_order + (index * 10),
                )
            )

    if creates:
        CategoryProductOrder.objects.bulk_create(creates, ignore_conflicts=True, batch_size=500)
    if updates:
        CategoryProductOrder.objects.bulk_update(updates, ["block_order", "updated_at"], batch_size=500)
    return len(updates) + len(creates)


def _assign_category_product_block(category, product_ids, block_label):
    cleaned_label = _clean_order_block_label(block_label)
    unique_product_ids = []
    seen_product_ids = set()
    for raw_product_id in product_ids:
        try:
            product_id = int(raw_product_id)
        except (TypeError, ValueError):
            continue
        if product_id and product_id not in seen_product_ids:
            seen_product_ids.add(product_id)
            unique_product_ids.append(product_id)

    if not unique_product_ids:
        return 0

    linked_ids = set(
        Product.categories.through.objects.filter(
            category_id=category.id,
            product_id__in=unique_product_ids,
        ).values_list("product_id", flat=True)
    )
    target_product_ids = [
        product_id for product_id in unique_product_ids
        if product_id in linked_ids
    ]
    if not target_product_ids:
        return 0

    block_order = 0
    if cleaned_label:
        block_order = (
            CategoryProductOrder.objects.filter(category=category, block_label=cleaned_label)
            .exclude(block_order=0)
            .order_by("block_order")
            .values_list("block_order", flat=True)
            .first()
        )
        if not block_order:
            max_block_order = (
                CategoryProductOrder.objects.filter(category=category, product__categories=category)
                .exclude(block_label="")
                .aggregate(max_order=Max("block_order"))
                .get("max_order")
                or 0
            )
            block_order = max_block_order + 10

    existing_rows = {
        row.product_id: row
        for row in CategoryProductOrder.objects.filter(
            category=category,
            product_id__in=target_product_ids,
        )
    }
    max_sort_order = (
        CategoryProductOrder.objects.filter(category=category)
        .aggregate(max_order=Max("sort_order"))
        .get("max_order")
        or 0
    )
    now = timezone.now()
    updates = []
    creates = []
    for product_id in target_product_ids:
        row = existing_rows.get(product_id)
        if row:
            row.block_label = cleaned_label
            row.block_order = block_order
            row.updated_at = now
            updates.append(row)
        else:
            max_sort_order += 10
            creates.append(
                CategoryProductOrder(
                    category=category,
                    product_id=product_id,
                    block_label=cleaned_label,
                    block_order=block_order,
                    sort_order=max_sort_order,
                )
            )

    if creates:
        CategoryProductOrder.objects.bulk_create(creates, ignore_conflicts=True, batch_size=500)
    if updates:
        CategoryProductOrder.objects.bulk_update(
            updates,
            ["block_label", "block_order", "updated_at"],
            batch_size=500,
        )
    return len(target_product_ids)


def _natural_sku_sort_key(value):
    text = unicodedata.normalize("NFKD", str(value or "").strip().upper())
    text = "".join(char for char in text if not unicodedata.combining(char))
    parts = []
    for part in re.split(r"(\d+)", text):
        if not part:
            continue
        if part.isdigit():
            parts.append((0, int(part), part))
        else:
            parts.append((1, part))
    return tuple(parts), text


def _clean_product_view_sort(value):
    value = str(value or "").strip().lower()
    if value in {"sku_asc", "sku_desc"}:
        return value
    return ""


def _sort_products_for_view_by_sku(products, view_sort):
    if view_sort not in {"sku_asc", "sku_desc"}:
        return products
    descending = view_sort == "sku_desc"
    return sorted(
        list(products),
        key=lambda product: (
            _natural_sku_sort_key(product.sku),
            _natural_sku_sort_key(product.name),
            product.id,
        ),
        reverse=descending,
    )


def _sort_category_products_by_sku(category, descending=False):
    linked_products = list(
        Product.objects.filter(categories=category)
        .only("id", "sku", "name")
        .distinct()
    )
    if not linked_products:
        return 0

    product_ids = [product.id for product in linked_products]
    existing_rows = {
        row.product_id: row
        for row in CategoryProductOrder.objects.filter(
            category=category,
            product_id__in=product_ids,
        )
    }

    current_items = []
    for product in linked_products:
        row = existing_rows.get(product.id)
        current_items.append({
            "product": product,
            "block_label": row.block_label if row else "",
            "block_order": max(row.block_order if row else 0, 0),
            "sort_order": max(row.sort_order if row else 999999999, 0),
        })

    blocks = []
    blocks_by_label = {}
    for item in sorted(
        current_items,
        key=lambda entry: (
            entry["block_order"],
            entry["sort_order"],
            str(entry["product"].name or "").upper(),
            str(entry["product"].sku or "").upper(),
            entry["product"].id,
        ),
    ):
        label = item["block_label"] or ""
        block = blocks_by_label.get(label)
        if block is None:
            block = {
                "label": label,
                "block_order": item["block_order"],
                "items": [],
            }
            blocks_by_label[label] = block
            blocks.append(block)
        block["items"].append(item)

    ordered_entries = []
    for block_index, block in enumerate(blocks, start=1):
        block_order = block["block_order"]
        if block["label"] and block_order <= 0:
            block_order = block_index * 10
        sorted_items = sorted(
            block["items"],
            key=lambda item: (
                _natural_sku_sort_key(item["product"].sku),
                _natural_sku_sort_key(item["product"].name),
                item["product"].id,
            ),
            reverse=descending,
        )
        ordered_entries.extend(
            {
                "product_id": item["product"].id,
                "block_label": block["label"],
                "block_order": block_order,
            }
            for item in sorted_items
        )

    return _save_category_product_order_entries(category, ordered_entries)


def _build_order_entries_from_payload(payload_entries):
    block_order_by_label = {}
    normalized_entries = []

    for index, item in enumerate(payload_entries, start=1):
        try:
            product_id = int(item.get("product_id") or item.get("id") or 0)
        except (AttributeError, TypeError, ValueError):
            continue
        if not product_id:
            continue

        block_label = _clean_order_block_label(item.get("block_label"))
        block_key = block_label or "__sin_bloque__"
        block_order = block_order_by_label.setdefault(block_key, (len(block_order_by_label) + 1) * 10)

        normalized_entries.append(
            {
                "product_id": product_id,
                "block_label": block_label,
                "block_order": max(block_order, 0),
                "position": index,
            }
        )

    return normalized_entries


def _get_latest_block_conversion_log(category):
    logs = AdminAuditLog.objects.filter(
        action="category_blocks_convert_to_subcategories",
        target_type="category",
        target_id=str(category.pk),
    ).order_by("-created_at")[:10]
    for log in logs:
        rollback_payload = (log.details or {}).get("rollback")
        if rollback_payload and not rollback_payload.get("rolled_back_at"):
            return log
    return None


@staff_member_required
@superuser_required_for_modifications
def category_manage_products(request, pk):
    """
    Manage direct category links for products (many-to-many).
    """
    category = get_object_or_404(Category, pk=pk)
    import_order_preview = None

    def get_filtered_queryset(req_data):
        qs = Product.objects.select_related('category').prefetch_related('categories').all()
        block_filter = _clean_order_block_label(
            req_data.get('block_filter') or req_data.get('block') or ''
        )

        qs, search = apply_admin_text_search(
            qs,
            req_data.get('q', ''),
            ["sku", "name", "description", "supplier", "supplier_ref__name"],
        )

        status = req_data.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)

        cat_filter = req_data.get('category_filter', 'current') or 'current'
        if block_filter:
            cat_filter = 'current'

        if cat_filter == 'current':
            qs = qs.filter(categories=category)
        elif cat_filter == 'none':
            qs = qs.filter(category__isnull=True, categories__isnull=True)
        elif cat_filter == 'all':
            pass
        elif cat_filter.isdigit():
            qs = qs.filter(categories__id=int(cat_filter))

        if block_filter:
            order_rows = CategoryProductOrder.objects.filter(category=category)
            if block_filter == "__no_block__":
                named_block_product_ids = order_rows.exclude(block_label="").values("product_id")
                blank_block_product_ids = order_rows.filter(block_label="").values("product_id")
                qs = qs.filter(categories=category).filter(
                    Q(id__in=blank_block_product_ids) | ~Q(id__in=named_block_product_ids)
                )
            else:
                qs = qs.filter(
                    categories=category,
                    id__in=order_rows.filter(block_label=block_filter).values("product_id"),
                )

        return qs.distinct(), search, status, cat_filter, block_filter

    def redirect_with_filters():
        query = {}
        for key in ("q", "category_filter", "status", "block_filter", "view_sort"):
            value = str(request.POST.get(key) or "").strip()
            if value:
                query[key] = value
        url = reverse("admin_category_products", args=[pk])
        if query:
            url = f"{url}?{urlencode(query)}"
        return redirect(url)

    if request.method == 'POST':
        raw_post_body = request.body
        action = request.POST.get('action', 'assign').strip()
        select_all_pages = request.POST.get('select_all_pages') == 'true'

        if action == "rollback_block_conversion":
            log_id = str(request.POST.get("conversion_log_id") or "").strip()
            conversion_log = None
            if log_id.isdigit():
                conversion_log = AdminAuditLog.objects.filter(
                    pk=int(log_id),
                    action="category_blocks_convert_to_subcategories",
                    target_type="category",
                    target_id=str(category.pk),
                ).first()
            if conversion_log is None:
                conversion_log = _get_latest_block_conversion_log(category)

            rollback_payload = (conversion_log.details or {}).get("rollback") if conversion_log else None
            if not conversion_log or not rollback_payload:
                messages.warning(request, "No hay una conversion reciente para deshacer.")
                return redirect('admin_category_products', pk=pk)
            if rollback_payload.get("rolled_back_at"):
                messages.info(request, "Esta conversion ya fue deshecha.")
                return redirect('admin_category_products', pk=pk)

            rollback_result = rollback_category_block_conversion(category, rollback_payload)
            details = dict(conversion_log.details or {})
            updated_rollback = dict(details.get("rollback") or {})
            updated_rollback["rolled_back_at"] = timezone.now().isoformat()
            updated_rollback["rollback_result"] = rollback_result.as_dict()
            details["rollback"] = updated_rollback
            conversion_log.details = details
            conversion_log.save(update_fields=["details"])

            log_admin_action(
                request,
                action="category_blocks_rollback_subcategories",
                target_type="category",
                target_id=category.pk,
                details={
                    "conversion_log_id": conversion_log.pk,
                    **rollback_result.as_dict(),
                },
            )
            messages.success(
                request,
                f"Conversion deshecha. Vinculos quitados: {rollback_result.product_links_removed}; "
                f"subcategorias eliminadas: {rollback_result.categories_deleted}; "
                f"ordenes eliminadas: {rollback_result.order_rows_deleted}.",
            )
            if rollback_result.categories_kept or rollback_result.skipped_modified_order_rows:
                messages.warning(
                    request,
                    "Algunas subcategorias o filas de orden se conservaron porque tuvieron cambios posteriores.",
                )
            return redirect('admin_category_products', pk=pk)

        if action == "convert_blocks":
            result = convert_category_blocks_to_subcategories(category)
            if result.blocks_found:
                messages.success(
                    request,
                    f"Bloques convertidos en subcategorias: {result.blocks_found}. "
                    f"Subcategorias creadas: {result.categories_created}; "
                    f"reutilizadas: {result.categories_reused}; "
                    f"productos vinculados: {result.products_processed}.",
                )
                if result.skipped_without_block:
                    messages.info(
                        request,
                        f"{result.skipped_without_block} productos quedaron en la categoria padre por no tener bloque.",
                    )
                if result.skipped_unlinked:
                    messages.warning(
                        request,
                        f"{result.skipped_unlinked} filas de orden no se convirtieron porque el producto ya no esta vinculado a esta categoria.",
                    )
            else:
                messages.info(
                    request,
                    "No hay bloques con productos vinculados para convertir en subcategorias.",
                )
            log_admin_action(
                request,
                action="category_blocks_convert_to_subcategories",
                target_type="category",
                target_id=category.pk,
                details={
                    "blocks_found": result.blocks_found,
                    "categories_created": result.categories_created,
                    "categories_reused": result.categories_reused,
                    "products_processed": result.products_processed,
                    "product_links_created": result.product_links_created,
                    "order_rows_created": result.order_rows_created,
                    "order_rows_updated": result.order_rows_updated,
                    "skipped_without_block": result.skipped_without_block,
                    "skipped_unlinked": result.skipped_unlinked,
                    "rollback": result.rollback_payload,
                },
            )
            return redirect('admin_category_products', pk=pk)

        if action == "sort_sku":
            direction = str(request.POST.get("direction") or "asc").strip().lower()
            descending = direction == "desc"
            updated = _sort_category_products_by_sku(category, descending=descending)
            direction_label = "Z-A" if descending else "A-Z"
            messages.success(
                request,
                f"Orden por SKU {direction_label} aplicado dentro de cada bloque: {updated} productos actualizados.",
            )
            log_admin_action(
                request,
                action="category_products_sort_sku",
                target_type="category",
                target_id=category.pk,
                details={
                    "direction": direction_label,
                    "updated": updated,
                },
            )
            return redirect('admin_category_products', pk=pk)

        if action == "import_order":
            try:
                parsed_rows = _parse_category_order_import_file(request.FILES.get("order_file"))
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect('admin_category_products', pk=pk)

            products_by_sku = {
                product.sku.strip().lower(): product
                for product in Product.objects.filter(categories=category).only("id", "sku", "name")
                if product.sku
            }

            entries = []
            preview_rows = []
            skipped = []
            for row in sorted(parsed_rows, key=lambda item: (item["block_order"], item["position"], item["row"])):
                product = products_by_sku.get(row["sku"].lower())
                if not product:
                    skipped.append(f"Fila {row['row']}: SKU {row['sku']} no existe en esta categoria")
                    continue
                entry = {
                    "product_id": product.pk,
                    "block_label": row["block_label"],
                    "block_order": row["block_order"],
                }
                entries.append(entry)
                if len(preview_rows) < 40:
                    preview_rows.append({
                        "sku": product.sku,
                        "name": product.name,
                        "position": row["position"],
                        "block_label": row["block_label"] or "Sin bloque",
                    })

            if request.POST.get("dry_run") == "true":
                import_order_preview = {
                    "rows": preview_rows,
                    "valid_count": len(entries),
                    "skipped": skipped[:20],
                    "skipped_count": len(skipped),
                }
                messages.info(
                    request,
                    f"Previsualizacion lista: {len(entries)} productos se ordenarian; {len(skipped)} filas se omitirian.",
                )
            else:
                updated = _save_category_product_order_entries(category, entries)
                if skipped:
                    messages.warning(request, f"{len(skipped)} filas fueron omitidas. Revisa SKU o categoria.")
                messages.success(request, f"Orden importado: {updated} productos actualizados.")
                return redirect('admin_category_products', pk=pk)

        else:
            if select_all_pages:
                products_to_update, _, _, _, _ = get_filtered_queryset(request.POST)
                target_ids = list(products_to_update.values_list('id', flat=True))
            else:
                target_ids = extract_target_product_ids_from_post(request.POST, raw_post_body)
                if not target_ids:
                    logger.warning(
                        "category_manage_products without selected products | user=%s | category=%s | keys=%s | action=%s | product_ids=%s | product_ids_csv=%s",
                        getattr(request.user, "username", "unknown"),
                        pk,
                        list(request.POST.keys()),
                        action,
                        request.POST.getlist("product_ids"),
                        request.POST.get("product_ids_csv", ""),
                    )
                    messages.warning(request, 'No se seleccionaron productos.')
                    return redirect_with_filters()

            count = 0
            if action == 'assign':
                count = add_category_to_products(target_ids, category.id)
                messages.success(request, f'{count} productos vinculados a "{category.name}".')
            elif action == 'remove':
                count = remove_category_from_products(target_ids, category.id)
                messages.success(request, f'{count} vinculos removidos de "{category.name}".')
            elif action == 'assign_block':
                block_label = _clean_order_block_label(
                    request.POST.get("new_block_label") or request.POST.get("existing_block_label") or ""
                )
                if not block_label:
                    messages.warning(request, "Elegi o escribi un bloque para asignar.")
                    return redirect_with_filters()
                count = _assign_category_product_block(category, target_ids, block_label)
                messages.success(request, f'{count} productos asignados al bloque "{block_label}".')
            elif action == 'clear_block':
                count = _assign_category_product_block(category, target_ids, "")
                messages.success(request, f'{count} productos quedaron sin bloque.')
            else:
                messages.warning(request, "Accion masiva no reconocida.")

            return redirect_with_filters()

    products, search, status, cat_filter, block_filter = get_filtered_queryset(request.GET)
    view_sort = _clean_product_view_sort(request.GET.get("view_sort"))

    can_reorder_products = cat_filter == 'current' and not search and not status and not view_sort

    if cat_filter == 'current':
        order_subquery = (
            CategoryProductOrder.objects.filter(category=category, product_id=OuterRef("pk"))
            .values("sort_order")[:1]
        )
        block_label_subquery = (
            CategoryProductOrder.objects.filter(category=category, product_id=OuterRef("pk"))
            .values("block_label")[:1]
        )
        block_order_subquery = (
            CategoryProductOrder.objects.filter(category=category, product_id=OuterRef("pk"))
            .values("block_order")[:1]
        )
        products = products.annotate(
            category_block_label=Coalesce(
                Subquery(block_label_subquery, output_field=CharField()),
                Value(""),
            ),
            category_block_order=Coalesce(
                Subquery(block_order_subquery, output_field=IntegerField()),
                Value(0),
            ),
            category_sort_order=Coalesce(
                Subquery(order_subquery, output_field=IntegerField()),
                Value(999999999),
            )
        ).order_by("category_block_order", "category_sort_order", "name", "sku", "id")
    else:
        products = products.order_by('name', 'sku', 'id')

    products = _sort_products_for_view_by_sku(products, view_sort)

    paginator = Paginator(products, 300 if can_reorder_products else 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_start_index = page_obj.start_index()
    total_count = paginator.count

    enrich_products_with_category_state(page_obj.object_list)
    for product in page_obj.object_list:
        product.is_linked_to_current = any(cat.id == category.id for cat in product.linked_categories)

    all_categories = Category.objects.exclude(pk=pk).select_related('parent').order_by('order', 'name')
    all_category_options = build_category_options(all_categories, include_inactive_suffix=True)
    block_summaries = []
    named_block_summaries = []
    block_filter_options = []
    selected_block_label = ""
    if cat_filter == 'current':
        block_summaries = list(
            CategoryProductOrder.objects.filter(category=category, product__categories=category)
            .values("block_label", "block_order")
            .annotate(count=Count("id"))
            .order_by("block_order", "block_label")
        )
        named_block_index = 0
        for block in block_summaries:
            block["display_label"] = block["block_label"] or "Sin bloque"
            block["filter_value"] = block["block_label"] or "__no_block__"
            if block["block_label"]:
                named_block_index += 1
                block["named_index"] = named_block_index
                named_block_summaries.append(block)

        total_linked_count = Product.objects.filter(categories=category).distinct().count()
        named_block_count = (
            CategoryProductOrder.objects.filter(category=category, product__categories=category)
            .exclude(block_label="")
            .values("product_id")
            .distinct()
            .count()
        )
        no_block_count = max(total_linked_count - named_block_count, 0)
        block_filter_options = [
            {
                "value": "",
                "display_label": "Todos los bloques",
                "count": total_linked_count,
            }
        ]
        if no_block_count:
            block_filter_options.append({
                "value": "__no_block__",
                "display_label": "Sin bloque",
                "count": no_block_count,
            })
        block_filter_options.extend(
            {
                "value": block["block_label"],
                "display_label": block["display_label"],
                "count": block["count"],
            }
            for block in block_summaries
            if block["block_label"]
        )
        if block_filter == "__no_block__":
            selected_block_label = "Sin bloque"
        elif block_filter:
            selected_block_label = block_filter

    latest_block_conversion_log = _get_latest_block_conversion_log(category)

    return render(request, 'admin_panel/categories/manage_products.html', {
        'category': category,
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'category_filter': cat_filter,
        'view_sort': view_sort,
        'all_category_options': all_category_options,
        'total_count': total_count,
        'pagination_count': len(page_obj.object_list),
        'page_start_index': page_start_index,
        'can_reorder_products': can_reorder_products,
        'block_summaries': block_summaries,
        'named_block_summaries': named_block_summaries,
        'block_filter_options': block_filter_options,
        'selected_block_filter': block_filter,
        'selected_block_label': selected_block_label,
        'import_order_preview': import_order_preview,
        'latest_block_conversion_log': latest_block_conversion_log,
    })


@staff_member_required
@require_POST
@superuser_required_for_modifications
def category_products_reorder(request, pk):
    """
    Reorder products inside one category without changing category assignment.
    """
    category = get_object_or_404(Category, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalido."}, status=400)

    if "block_order_labels" in payload:
        block_order_labels = payload.get("block_order_labels")
        if not isinstance(block_order_labels, list):
            return JsonResponse({"success": False, "error": "Orden de bloques invalido."}, status=400)
        updated = _save_category_product_block_order(category, block_order_labels)
        log_admin_action(
            request,
            action="category_product_blocks_reorder",
            target_type="category",
            target_id=category.id,
            details={
                "block_order_labels": block_order_labels,
                "updated": updated,
            },
        )
        return JsonResponse({"success": True, "updated": updated})

    ordered_entries = _build_order_entries_from_payload(payload.get("ordered_entries") or [])
    if not ordered_entries:
        ordered_ids = normalize_category_ids(payload.get("ordered_ids", []))
        ordered_entries = [
            {
                "product_id": product_id,
                "block_label": "",
                "block_order": 0,
            }
            for product_id in ordered_ids
        ]

    if not ordered_entries:
        return JsonResponse({"success": False, "error": "No hay productos para ordenar."}, status=400)

    try:
        base_order = int(payload.get("base_order", 10))
    except (TypeError, ValueError):
        base_order = 10
    base_order = max(base_order, 10)

    updated = _save_category_product_order_entries(category, ordered_entries)
    if not updated:
        return JsonResponse({"success": False, "error": "Los productos no pertenecen a esta categoria."}, status=400)

    log_admin_action(
        request,
        action="category_products_reorder",
        target_type="category",
        target_id=category.id,
        details={
            "ordered_ids": [entry["product_id"] for entry in ordered_entries[:100]],
            "count": updated,
            "base_order": base_order,
            "blocks": sorted({entry.get("block_label") or "" for entry in ordered_entries}),
        },
    )
    return JsonResponse({"success": True, "updated": updated})

# ===================== API =====================

@staff_member_required
def get_category_attributes(request, category_id):
    """API: Get attributes for a category."""
    attributes = CategoryAttribute.objects.filter(category_id=category_id).values(
        'name', 'slug', 'type', 'options', 'required', 'is_recommended', 'regex_pattern'
    )
    blocks = list(
        CategoryProductOrder.objects.filter(category_id=category_id)
        .exclude(block_label="")
        .values_list("block_label", flat=True)
        .distinct()
        .order_by("block_label")
    )
    return JsonResponse({
        'attributes': list(attributes),
        'blocks': blocks
    })


@staff_member_required
@require_POST
def parse_product_description(request):
    """API: Parse description against category attributes."""
    try:
        data = json.loads(request.body)
        description = data.get('description', '')
        category_id = data.get('category_id')
        
        if not category_id:
            return JsonResponse({'success': False, 'error': 'Category ID required'})
            
        category = Category.objects.get(pk=category_id)
        # Instantiate dummy product to use extraction logic
        product = Product(description=description, category=category)
        extracted = product.extract_attributes_from_description()
        
        return JsonResponse({'success': True, 'attributes': extracted})
    except Exception as e:
        logger.exception("Error parsing product description")
        return JsonResponse({'success': False, 'error': 'No se pudo procesar la descripción.'}, status=400)


@staff_member_required
@require_POST
def parse_clamp_code_api(request):
    """API: Parse ABL/ABT code into attributes."""
    try:
        data = json.loads(request.body)
        code = str(data.get("code", "")).strip()
        known_widths = data.get("known_widths") or None
        known_lengths = data.get("known_lengths") or None

        if not code:
            return JsonResponse({"success": False, "error": "code es obligatorio."}, status=400)

        parsed = parsearCodigo(
            code,
            known_widths=known_widths,
            known_lengths=known_lengths,
        )
        return JsonResponse({"success": True, "result": parsed})
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    except Exception:
        logger.exception("Error parsing clamp code")
        return JsonResponse({"success": False, "error": "No se pudo parsear el codigo."}, status=500)

__all__ = ['product_list', 'product_create', 'product_edit', 'product_delete', 'product_toggle_active', 'product_bulk_category_update', 'product_bulk_status_update', 'product_bulk_image_update', 'supplier_list', 'supplier_detail', 'supplier_bulk_action', 'supplier_export', 'supplier_print', 'supplier_unassigned', 'supplier_toggle_active', 'category_list', 'category_reorder', 'category_sort_roots_alpha', 'category_bulk_status', 'category_create', 'category_create_ajax', 'category_edit', 'category_move', 'category_delete', 'category_attribute_create', 'category_attribute_edit', 'category_attribute_delete', 'category_manage_products', 'category_products_reorder', 'get_category_attributes', 'parse_product_description', 'parse_clamp_code_api', 'products_uncategorized']
