from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Q, Max
from django.core.paginator import Paginator
from django.urls import reverse
from django.utils import timezone
import json

from catalog.models import (
    Category,
    Brand,
    BrandAlias,
    BrandCatalogBatch,
    BrandCatalogRule,
    BrandRubro,
    BrandSubrubro,
    BrandSubrubroProductOrder,
    BrandRubroProductOrder,
    Product,
)
from catalog.services.brand_cataloging import (
    BrandSuggestionEngine,
    assign_products_to_brand_catalog,
    brand_quality_metrics,
    remove_products_from_brand_catalog,
    uncataloged_products,
    undo_brand_catalog_batch,
)
from admin_panel.forms.brand_forms import (
    BrandAliasForm,
    BrandCatalogRuleForm,
    BrandForm,
    BrandRubroForm,
    BrandSubrubroForm,
)
from admin_panel.views.helpers import get_cached_category_options
from core.services.audit import log_admin_action
from core.decorators import superuser_required_for_modifications
from core.services.advanced_search import sanitize_search_token


@staff_member_required
def brand_list(request):
    """View to list brands, rubros, and subrubros in a tree hierarchy."""
    search = sanitize_search_token(request.GET.get('q', ''))
    status = request.GET.get('status', 'all').strip().lower()

    # Pre-fetch the hierarchy
    brands_qs = Brand.objects.all()
    if status == 'active':
        brands_qs = brands_qs.filter(is_active=True)
    elif status == 'inactive':
        brands_qs = brands_qs.filter(is_active=False)

    if search:
        brands_qs = brands_qs.filter(name__icontains=search)

    brands = brands_qs.prefetch_related("rubros__subrubros").order_by("order", "name")

    return render(request, 'admin_panel/brands/brand_list.html', {
        'brands': brands,
        'search': search,
        'status': status,
        'quality_metrics': brand_quality_metrics(),
        'recent_batches': BrandCatalogBatch.objects.select_related(
            "brand", "brand_rubro", "brand_subrubro", "created_by"
        )[:5],
    })


@staff_member_required
@superuser_required_for_modifications
def brand_create(request):
    """Create brand view."""
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES)
        if form.is_valid():
            brand = form.save()
            log_admin_action(
                request,
                action="brand_create",
                target_type="brand",
                target_id=brand.pk,
                details={"name": brand.name},
            )
            messages.success(request, f'Marca "{brand.name}" creada con éxito.')
            return redirect('admin_brand_list')
    else:
        form = BrandForm()

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Crear',
        'title': 'Crear Nueva Marca',
    })


@staff_member_required
@superuser_required_for_modifications
def brand_edit(request, pk):
    """Edit brand view."""
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES, instance=brand)
        if form.is_valid():
            brand = form.save()
            log_admin_action(
                request,
                action="brand_edit",
                target_type="brand",
                target_id=brand.pk,
                details={"name": brand.name},
            )
            messages.success(request, f'Marca "{brand.name}" actualizada con éxito.')
            return redirect('admin_brand_list')
    else:
        form = BrandForm(instance=brand)

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Editar',
        'title': f'Editar Marca: {brand.name}',
    })


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_delete(request, pk):
    """Delete brand view."""
    brand = get_object_or_404(Brand, pk=pk)
    name = brand.name
    brand.delete()
    log_admin_action(
        request,
        action="brand_delete",
        target_type="brand",
        target_id=pk,
        details={"name": name},
    )
    messages.success(request, f'Marca "{name}" eliminada.')
    return redirect('admin_brand_list')


@staff_member_required
@superuser_required_for_modifications
def brand_rubro_create(request):
    """Create BrandRubro view."""
    brand_id = request.GET.get('brand', '').strip()
    initial = {}
    if brand_id.isdigit():
        initial['brand'] = int(brand_id)

    if request.method == 'POST':
        form = BrandRubroForm(request.POST, request.FILES)
        if form.is_valid():
            rubro = form.save()
            log_admin_action(
                request,
                action="brand_rubro_create",
                target_type="brand_rubro",
                target_id=rubro.pk,
                details={"name": rubro.name, "brand_id": rubro.brand_id},
            )
            messages.success(request, f'Rubro de marca "{rubro.name}" creado.')
            return redirect('admin_brand_list')
    else:
        form = BrandRubroForm(initial=initial)

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Crear Rubro',
        'title': 'Crear Rubro de Marca',
    })


@staff_member_required
@superuser_required_for_modifications
def brand_rubro_edit(request, pk):
    """Edit BrandRubro view."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    if request.method == 'POST':
        form = BrandRubroForm(request.POST, request.FILES, instance=rubro)
        if form.is_valid():
            rubro = form.save()
            log_admin_action(
                request,
                action="brand_rubro_edit",
                target_type="brand_rubro",
                target_id=rubro.pk,
                details={"name": rubro.name, "brand_id": rubro.brand_id},
            )
            messages.success(request, f'Rubro de marca "{rubro.name}" actualizado.')
            return redirect('admin_brand_list')
    else:
        form = BrandRubroForm(instance=rubro)

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Editar Rubro',
        'title': f'Editar Rubro: {rubro.name}',
    })


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_rubro_delete(request, pk):
    """Delete BrandRubro view."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    name = rubro.name
    rubro.delete()
    log_admin_action(
        request,
        action="brand_rubro_delete",
        target_type="brand_rubro",
        target_id=pk,
        details={"name": name},
    )
    messages.success(request, f'Rubro "{name}" eliminado.')
    return redirect('admin_brand_list')


@staff_member_required
@superuser_required_for_modifications
def brand_subrubro_create(request):
    """Create BrandSubrubro view."""
    rubro_id = request.GET.get('rubro', '').strip()
    initial = {}
    if rubro_id.isdigit():
        initial['brand_rubro'] = int(rubro_id)

    if request.method == 'POST':
        form = BrandSubrubroForm(request.POST, request.FILES)
        if form.is_valid():
            subrubro = form.save()
            log_admin_action(
                request,
                action="brand_subrubro_create",
                target_type="brand_subrubro",
                target_id=subrubro.pk,
                details={"name": subrubro.name, "brand_rubro_id": subrubro.brand_rubro_id},
            )
            messages.success(request, f'Subrubro de marca "{subrubro.name}" creado.')
            return redirect('admin_brand_list')
    else:
        form = BrandSubrubroForm(initial=initial)

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Crear Subrubro',
        'title': 'Crear Subrubro de Marca',
    })


@staff_member_required
@superuser_required_for_modifications
def brand_subrubro_edit(request, pk):
    """Edit BrandSubrubro view."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    if request.method == 'POST':
        form = BrandSubrubroForm(request.POST, request.FILES, instance=subrubro)
        if form.is_valid():
            subrubro = form.save()
            log_admin_action(
                request,
                action="brand_subrubro_edit",
                target_type="brand_subrubro",
                target_id=subrubro.pk,
                details={"name": subrubro.name, "brand_rubro_id": subrubro.brand_rubro_id},
            )
            messages.success(request, f'Subrubro de marca "{subrubro.name}" actualizado.')
            return redirect('admin_brand_list')
    else:
        form = BrandSubrubroForm(instance=subrubro)

    return render(request, 'admin_panel/brands/brand_form.html', {
        'form': form,
        'action': 'Editar Subrubro',
        'title': f'Editar Subrubro: {subrubro.name}',
    })


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_subrubro_delete(request, pk):
    """Delete BrandSubrubro view."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    name = subrubro.name
    subrubro.delete()
    log_admin_action(
        request,
        action="brand_subrubro_delete",
        target_type="brand_subrubro",
        target_id=pk,
        details={"name": name},
    )
    messages.success(request, f'Subrubro "{name}" eliminado.')
    return redirect('admin_brand_list')


def _brand_workspace_is_ajax(request):
    return (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.GET.get("ajax") == "1"
    )


def _brand_workspace_target(target):
    if isinstance(target, BrandSubrubro):
        return {
            "kind": "subrubro",
            "brand": target.brand_rubro.brand,
            "rubro": target.brand_rubro,
            "subrubro": target,
            "order_model": BrandSubrubroProductOrder,
            "target_field": "brand_subrubro",
            "target": target,
        }
    return {
        "kind": "rubro",
        "brand": target.brand,
        "rubro": target,
        "subrubro": None,
        "order_model": BrandRubroProductOrder,
        "target_field": "brand_rubro",
        "target": target,
    }


def _brand_workspace_associated_ids(target_info):
    return set(
        target_info["order_model"].objects.filter(
            **{target_info["target_field"]: target_info["target"]}
        ).values_list("product_id", flat=True)
    )


def _brand_workspace_filtered_products(filters, target_info):
    products = (
        Product.objects.select_related("category", "supplier_ref")
        .prefetch_related(
            "categories",
            "brand_rubro_orders__brand_rubro__brand",
            "brand_subrubro_orders__brand_subrubro__brand_rubro__brand",
        )
        .order_by("name", "sku")
    )
    q = sanitize_search_token(filters.get("q", ""))
    category_id = str(filters.get("category_id", "") or "").strip()
    supplier = sanitize_search_token(filters.get("supplier", ""))
    product_status = str(filters.get("status", "all") or "all").strip().lower()
    stock_status = str(filters.get("stock", "all") or "all").strip().lower()
    assignment = str(filters.get("assignment", "all") or "all").strip().lower()

    if q:
        products = products.filter(
            Q(sku__icontains=q)
            | Q(name__icontains=q)
            | Q(supplier__icontains=q)
            | Q(supplier_ref__name__icontains=q)
        )
    if category_id.isdigit():
        category = Category.objects.filter(pk=int(category_id)).first()
        if category:
            category_ids = category.get_descendant_ids(include_self=True)
            products = products.filter(
                Q(category_id__in=category_ids) | Q(categories__id__in=category_ids)
            )
    if supplier:
        products = products.filter(
            Q(supplier__iexact=supplier) | Q(supplier_ref__name__iexact=supplier)
        )
    if product_status == "active":
        products = products.filter(is_active=True)
    elif product_status == "inactive":
        products = products.filter(is_active=False)
    if stock_status == "available":
        products = products.filter(stock__gt=0)
    elif stock_status == "empty":
        products = products.filter(stock__lte=0)
    elif stock_status == "tracked":
        products = products.filter(tracks_stock=True)

    associated_ids = _brand_workspace_associated_ids(target_info)
    if assignment == "associated":
        products = products.filter(pk__in=associated_ids)
    elif assignment == "available":
        products = products.exclude(pk__in=associated_ids)
    elif assignment == "unassigned":
        products = products.filter(
            brand_rubro_orders__isnull=True,
            brand_subrubro_orders__isnull=True,
        )
    elif assignment == "other":
        products = products.filter(
            Q(brand_rubro_orders__isnull=False)
            | Q(brand_subrubro_orders__isnull=False)
        ).exclude(pk__in=associated_ids)
    return products.distinct(), associated_ids


def _brand_workspace_product_payload(product, target_info, associated_ids, engine):
    subrubro_rows = list(product.brand_subrubro_orders.all())
    subrubro_parent_ids = {
        row.brand_subrubro.brand_rubro_id for row in subrubro_rows
    }
    assignments = []
    for row in subrubro_rows:
        target = row.brand_subrubro
        assignments.append(
            {
                "kind": "subrubro",
                "brand_id": target.brand_rubro.brand_id,
                "label": f"{target.brand_rubro.brand.name} > {target.brand_rubro.name} > {target.name}",
            }
        )
    for row in product.brand_rubro_orders.all():
        target = row.brand_rubro
        if target.pk in subrubro_parent_ids:
            continue
        assignments.append(
            {
                "kind": "rubro",
                "brand_id": target.brand_id,
                "label": f"{target.brand.name} > {target.name}",
            }
        )

    suggestions = engine.suggest(product, limit=3)
    target_suggestion = None
    for suggestion in suggestions:
        matches_target = (
            suggestion["brand"].pk == target_info["brand"].pk
            and (
                target_info["kind"] == "rubro"
                and (
                    suggestion["rubro"] is None
                    or suggestion["rubro"].pk == target_info["rubro"].pk
                )
                or target_info["kind"] == "subrubro"
                and suggestion["subrubro"] is not None
                and suggestion["subrubro"].pk == target_info["subrubro"].pk
            )
        )
        if matches_target:
            target_suggestion = suggestion
            break
    best_suggestion = target_suggestion or (suggestions[0] if suggestions else None)
    primary_category = product.category
    if primary_category is None:
        linked_categories = list(product.categories.all())
        primary_category = linked_categories[0] if linked_categories else None
    try:
        image_url = product.image.url if product.image else ""
    except ValueError:
        image_url = ""

    return {
        "id": product.pk,
        "sku": product.sku,
        "name": product.name,
        "supplier": (
            product.supplier_ref.name
            if product.supplier_ref_id
            else product.supplier or "Sin proveedor"
        ),
        "category": (
            primary_category.get_full_path() if primary_category else "Sin categoria"
        ),
        "image_url": image_url,
        "is_active": product.is_active,
        "stock": product.stock,
        "tracks_stock": product.tracks_stock,
        "is_associated": product.pk in associated_ids,
        "assignments": assignments,
        "has_conflict": any(
            item["brand_id"] != target_info["brand"].pk for item in assignments
        ),
        "suggestion": (
            {
                "destination": best_suggestion["destination"],
                "confidence": best_suggestion["confidence"],
                "reason": best_suggestion["reason"],
                "matches_target": best_suggestion is target_suggestion,
            }
            if best_suggestion
            else None
        ),
    }


def _brand_workspace_search_response(request, target_info):
    products, associated_ids = _brand_workspace_filtered_products(
        request.GET,
        target_info,
    )
    total_count = products.count()
    page_raw = str(request.GET.get("page", "1") or "1").strip()
    page_number = max(int(page_raw), 1) if page_raw.isdigit() else 1
    page_size = 30
    start = (page_number - 1) * page_size
    subset = list(products[start : start + page_size + 1])
    has_more = len(subset) > page_size
    engine = BrandSuggestionEngine()
    results = [
        _brand_workspace_product_payload(
            product,
            target_info,
            associated_ids,
            engine,
        )
        for product in subset[:page_size]
    ]
    return JsonResponse(
        {
            "success": True,
            "results": results,
            "has_more": has_more,
            "total_count": total_count,
            "page": page_number,
            "associated_count": len(associated_ids),
        }
    )


def _brand_workspace_context(target_info):
    row_filter = {target_info["target_field"]: target_info["target"]}
    order_rows = (
        target_info["order_model"].objects.filter(**row_filter)
        .select_related("product", "product__category", "product__supplier_ref")
        .order_by("sort_order", "product__name")
    )
    batches = BrandCatalogBatch.objects.filter(
        brand=target_info["brand"],
        brand_rubro=target_info["rubro"],
    )
    if target_info["subrubro"] is not None:
        batches = batches.filter(brand_subrubro=target_info["subrubro"])
    else:
        batches = batches.filter(brand_subrubro__isnull=True)

    kind = target_info["kind"]
    pk = target_info["target"].pk
    suppliers = list(
        Product.objects.exclude(supplier="")
        .values_list("supplier", flat=True)
        .distinct()
        .order_by("supplier")[:250]
    )
    return {
        "brand": target_info["brand"],
        "rubro": target_info["rubro"],
        "subrubro": target_info["subrubro"],
        "target_kind": kind,
        "target_name": target_info["target"].name,
        "order_rows": order_rows,
        "category_options": get_cached_category_options(
            only_active=True,
            include_inactive_suffix=False,
        ),
        "suppliers": suppliers,
        "recent_batches": batches.select_related("created_by", "undone_by")[:6],
        "workspace_search_url": reverse(
            f"admin_brand_{kind}_products",
            args=[pk],
        ),
        "workspace_bulk_add_url": reverse(
            f"admin_brand_{kind}_bulk_assign",
            args=[pk],
        ),
        "workspace_bulk_remove_url": reverse(
            f"admin_brand_{kind}_bulk_remove",
            args=[pk],
        ),
        "workspace_reorder_url": reverse(
            f"admin_brand_{kind}_products_reorder",
            args=[pk],
        ),
        "workspace_remove_url": reverse(
            f"admin_brand_{kind}_remove_product",
            args=[pk],
        ),
        "workspace_sync_url": reverse(
            f"admin_brand_{kind}_sync",
            args=[pk],
        ),
        "workspace_undo_url_template": reverse(
            "admin_brand_catalog_batch_undo",
            args=[0],
        ).replace("/0/", "/__batch__/", 1),
    }


@staff_member_required
def brand_subrubro_products(request, pk):
    """Product catalog workspace for one brand subrubro."""
    subrubro = get_object_or_404(
        BrandSubrubro.objects.select_related("brand_rubro__brand"),
        pk=pk,
    )
    target_info = _brand_workspace_target(subrubro)
    if _brand_workspace_is_ajax(request):
        return _brand_workspace_search_response(request, target_info)
    return render(
        request,
        "admin_panel/brands/brand_subrubro_products.html",
        _brand_workspace_context(target_info),
    )


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_add_product(request, pk):
    """Manually add a product to a brand subrubro."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    product_id = request.POST.get('product_id', '').strip()
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'
    
    if product_id.isdigit():
        product = get_object_or_404(Product, pk=int(product_id))
        
        max_order = BrandSubrubroProductOrder.objects.filter(
            brand_subrubro=subrubro
        ).aggregate(Max("sort_order"))["sort_order__max"] or 0
        
        row, created = BrandSubrubroProductOrder.objects.get_or_create(
            brand_subrubro=subrubro,
            product=product,
            defaults={"sort_order": max_order + 10}
        )
        if is_ajax:
            return JsonResponse({
                "success": True,
                "created": created,
                "product": {
                    "id": product.id,
                    "sku": product.sku,
                    "name": product.name
                }
            })
            
        if created:
            messages.success(request, f'Producto "{product.name}" agregado.')
        else:
            messages.info(request, f'El producto "{product.name}" ya existe en este subrubro.')
            
    elif is_ajax:
        return JsonResponse({"success": False, "error": "ID de producto inválido."}, status=400)
        
    return redirect('admin_brand_subrubro_products', pk=subrubro.pk)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_remove_product(request, pk):
    """Remove a product from a brand subrubro."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    product_id = request.POST.get('product_id', '').strip()
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'
    
    if product_id.isdigit():
        product = get_object_or_404(Product, pk=int(product_id))
        BrandSubrubroProductOrder.objects.filter(
            brand_subrubro=subrubro,
            product=product
        ).delete()
        
        if is_ajax:
            return JsonResponse({
                "success": True,
                "product_id": product.id
            })
            
        messages.success(request, f'Producto "{product.name}" removido.')
        
    elif is_ajax:
        return JsonResponse({"success": False, "error": "ID de producto inválido."}, status=400)
        
    return redirect('admin_brand_subrubro_products', pk=subrubro.pk)


def _brand_workspace_payload(request):
    if "application/json" in (request.content_type or ""):
        try:
            return json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return {}
    payload = request.POST.dict()
    payload["product_ids"] = request.POST.getlist("product_ids")
    return payload


def _brand_workspace_resolve_product_ids(payload, target_info):
    raw_ids = payload.get("product_ids", [])
    if isinstance(raw_ids, str):
        raw_ids = raw_ids.split(",")
    product_ids = [
        int(product_id)
        for product_id in raw_ids
        if str(product_id).strip().isdigit()
    ]
    if payload.get("select_all") in {True, "1", "true"}:
        filters = payload.get("filters") or {}
        products, _associated_ids = _brand_workspace_filtered_products(
            filters,
            target_info,
        )
        total_count = products.count()
        if total_count > 20000:
            raise ValueError("La seleccion supera 20.000 productos. Aplica mas filtros.")
        product_ids = list(products.values_list("pk", flat=True))
    return sorted(set(product_ids))


def _brand_workspace_batch_payload(batch, target_info):
    created_count = (
        len(batch.created_subrubro_row_ids or [])
        if target_info["kind"] == "subrubro"
        else len(batch.created_rubro_row_ids or [])
    )
    moved_product_ids = {
        row["product_id"]
        for row in [
            *(batch.removed_rubro_rows or []),
            *(batch.removed_subrubro_rows or []),
        ]
    }
    return {
        "success": True,
        "batch_id": batch.pk,
        "selected_count": len(batch.product_ids or []),
        "created_count": created_count,
        "existing_count": max(len(batch.product_ids or []) - created_count, 0),
        "moved_count": len(moved_product_ids),
        "can_undo": batch.can_undo,
    }


def _brand_workspace_assign(request, target):
    target_info = _brand_workspace_target(target)
    payload = _brand_workspace_payload(request)
    try:
        product_ids = _brand_workspace_resolve_product_ids(payload, target_info)
        mode = str(payload.get("mode", "add") or "add").strip().lower()
        operation = (
            BrandCatalogBatch.OPERATION_RULE
            if payload.get("used_suggestion") in {True, "1", 1}
            else BrandCatalogBatch.OPERATION_ASSIGN
        )
        batch = assign_products_to_brand_catalog(
            product_ids=product_ids,
            brand=target_info["brand"],
            rubro=target_info["rubro"],
            subrubro=target_info["subrubro"],
            user=request.user,
            observation=payload.get("observation", ""),
            operation=operation,
            mode=mode,
        )
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    log_admin_action(
        request,
        action=f"brand_{target_info['kind']}_bulk_assign",
        target_type=f"brand_{target_info['kind']}",
        target_id=target.pk,
        details={
            "batch_id": batch.pk,
            "mode": mode,
            "product_ids": batch.product_ids[:100],
            "count": len(batch.product_ids),
            "observation": batch.observation,
        },
    )
    return JsonResponse(_brand_workspace_batch_payload(batch, target_info))


def _brand_workspace_remove(request, target):
    target_info = _brand_workspace_target(target)
    payload = _brand_workspace_payload(request)
    try:
        product_ids = _brand_workspace_resolve_product_ids(payload, target_info)
        batch = remove_products_from_brand_catalog(
            product_ids=product_ids,
            brand=target_info["brand"],
            rubro=target_info["rubro"],
            subrubro=target_info["subrubro"],
            user=request.user,
            observation=payload.get("observation", ""),
        )
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    log_admin_action(
        request,
        action=f"brand_{target_info['kind']}_bulk_remove",
        target_type=f"brand_{target_info['kind']}",
        target_id=target.pk,
        details={
            "batch_id": batch.pk,
            "product_ids": batch.product_ids[:100],
            "count": len(batch.product_ids),
            "observation": batch.observation,
        },
    )
    response = _brand_workspace_batch_payload(batch, target_info)
    response["removed_count"] = len(batch.product_ids)
    return JsonResponse(response)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_bulk_assign(request, pk):
    subrubro = get_object_or_404(
        BrandSubrubro.objects.select_related("brand_rubro__brand"),
        pk=pk,
    )
    return _brand_workspace_assign(request, subrubro)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_bulk_remove(request, pk):
    subrubro = get_object_or_404(
        BrandSubrubro.objects.select_related("brand_rubro__brand"),
        pk=pk,
    )
    return _brand_workspace_remove(request, subrubro)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_products_reorder(request, pk):
    """AJAX endpoint to save manual product order in a brand subrubro."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalido."}, status=400)

    ordered_ids = payload.get("ordered_ids", [])
    if not ordered_ids:
        return JsonResponse({"success": False, "error": "No hay productos para ordenar."}, status=400)

    ordered_ids = [int(x) for x in ordered_ids if str(x).isdigit()]

    existing_rows = {
        row.product_id: row
        for row in BrandSubrubroProductOrder.objects.filter(
            brand_subrubro=subrubro,
            product_id__in=ordered_ids
        )
    }

    updates = []
    creates = []

    for index, product_id in enumerate(ordered_ids, start=1):
        sort_order = index * 10
        row = existing_rows.get(product_id)
        if row:
            row.sort_order = sort_order
            updates.append(row)
        else:
            creates.append(
                BrandSubrubroProductOrder(
                    brand_subrubro=subrubro,
                    product_id=product_id,
                    sort_order=sort_order
                )
            )

    if creates:
        BrandSubrubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)
    if updates:
        BrandSubrubroProductOrder.objects.bulk_update(updates, ["sort_order"])

    log_admin_action(
        request,
        action="brand_subrubro_products_reorder",
        target_type="brand_subrubro",
        target_id=subrubro.id,
        details={"ordered_ids": ordered_ids[:100], "count": len(ordered_ids)}
    )

    return JsonResponse({"success": True, "count": len(ordered_ids)})


def _brand_workspace_sync_candidates(target_info):
    category_ids = set()
    if target_info["subrubro"] is not None:
        helper_categories = target_info["subrubro"].helper_categories.all()
        for category in helper_categories:
            category_ids.update(
                category.get_descendant_ids(include_self=True, only_active=True)
            )
    else:
        for subrubro in target_info["rubro"].subrubros.filter(is_active=True).prefetch_related(
            "helper_categories"
        ):
            for category in subrubro.helper_categories.all():
                category_ids.update(
                    category.get_descendant_ids(include_self=True, only_active=True)
                )
    if not category_ids:
        raise ValueError(
            "No hay categorias ayudantes configuradas para preparar la sincronizacion."
        )

    match_query = Q()
    terms = [target_info["brand"].name]
    terms.extend(
        target_info["brand"].aliases.filter(is_active=True).values_list(
            "value",
            flat=True,
        )
    )
    for term in {str(item).strip() for item in terms if str(item).strip()}:
        match_query |= (
            Q(name__icontains=term)
            | Q(sku__icontains=term)
            | Q(supplier__icontains=term)
            | Q(supplier_ref__name__icontains=term)
        )

    rules = BrandCatalogRule.objects.filter(
        is_active=True,
        brand=target_info["brand"],
    )
    if target_info["subrubro"] is not None:
        rules = rules.filter(
            Q(brand_subrubro=target_info["subrubro"])
            | Q(brand_subrubro__isnull=True)
        )
    elif target_info["rubro"] is not None:
        rules = rules.filter(
            Q(brand_rubro=target_info["rubro"])
            | Q(brand_rubro__isnull=True)
        )
    for rule in rules:
        lookup = "icontains"
        if rule.match_mode == BrandCatalogRule.MATCH_PREFIX:
            lookup = "istartswith"
        fields = {
            BrandCatalogRule.FIELD_NAME: ["name"],
            BrandCatalogRule.FIELD_SKU: ["sku"],
            BrandCatalogRule.FIELD_SUPPLIER: ["supplier", "supplier_ref__name"],
            BrandCatalogRule.FIELD_ANY: [
                "name",
                "sku",
                "supplier",
                "supplier_ref__name",
            ],
        }.get(rule.source_field, [])
        for field in fields:
            match_query |= Q(**{f"{field}__{lookup}": rule.pattern})

    return (
        Product.objects.filter(is_active=True)
        .filter(
            Q(category_id__in=category_ids) | Q(categories__id__in=category_ids)
        )
        .filter(match_query)
        .select_related("category", "supplier_ref")
        .prefetch_related(
            "categories",
            "brand_rubro_orders__brand_rubro__brand",
            "brand_subrubro_orders__brand_subrubro__brand_rubro__brand",
        )
        .order_by("name", "sku")
        .distinct()
    )


def _brand_workspace_sync_action(request, target):
    target_info = _brand_workspace_target(target)
    payload = _brand_workspace_payload(request)
    action = str(payload.get("action", "") or "").strip().lower()
    try:
        candidates = _brand_workspace_sync_candidates(target_info)
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    associated_ids = _brand_workspace_associated_ids(target_info)
    candidate_ids = list(candidates.values_list("pk", flat=True))
    new_ids = [product_id for product_id in candidate_ids if product_id not in associated_ids]

    if action == "preview":
        engine = BrandSuggestionEngine()
        preview_products = list(candidates[:25])
        preview = [
            _brand_workspace_product_payload(
                product,
                target_info,
                associated_ids,
                engine,
            )
            for product in preview_products
        ]
        conflict_product_ids = set(
            BrandRubroProductOrder.objects.filter(product_id__in=candidate_ids)
            .exclude(brand_rubro__brand=target_info["brand"])
            .values_list("product_id", flat=True)
        )
        conflict_product_ids.update(
            BrandSubrubroProductOrder.objects.filter(product_id__in=candidate_ids)
            .exclude(brand_subrubro__brand_rubro__brand=target_info["brand"])
            .values_list("product_id", flat=True)
        )
        return JsonResponse(
            {
                "success": True,
                "candidate_count": len(candidate_ids),
                "new_count": len(new_ids),
                "associated_count": len(candidate_ids) - len(new_ids),
                "conflict_count": len(conflict_product_ids),
                "preview": preview,
                "preview_truncated": len(candidate_ids) > len(preview),
            }
        )

    if action != "confirm":
        return None
    try:
        batch = assign_products_to_brand_catalog(
            product_ids=candidate_ids,
            brand=target_info["brand"],
            rubro=target_info["rubro"],
            subrubro=target_info["subrubro"],
            user=request.user,
            observation=payload.get("observation", ""),
            operation=BrandCatalogBatch.OPERATION_RULE,
            mode=str(payload.get("mode", "add") or "add").strip().lower(),
        )
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    log_admin_action(
        request,
        action=f"brand_{target_info['kind']}_sync_confirm",
        target_type=f"brand_{target_info['kind']}",
        target_id=target.pk,
        details={
            "batch_id": batch.pk,
            "candidate_count": len(candidate_ids),
            "observation": batch.observation,
        },
    )
    response = _brand_workspace_batch_payload(batch, target_info)
    response["candidate_count"] = len(candidate_ids)
    return JsonResponse(response)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_sync(request, pk):
    """AJAX endpoint to auto-populate brand subrubro using helper categories and brand keyword."""
    subrubro = get_object_or_404(
        BrandSubrubro.objects.select_related("brand_rubro__brand"),
        pk=pk,
    )
    if str(_brand_workspace_payload(request).get("action", "")).strip().lower() in {
        "preview",
        "confirm",
    }:
        return _brand_workspace_sync_action(request, subrubro)
    brand_name = subrubro.brand_rubro.brand.name.upper()
    categories = subrubro.helper_categories.all()
    if not categories:
        return JsonResponse({"success": False, "error": "No hay categorías ayudantes configuradas para este subrubro."})

    # Collect helper categories and all their active descendant category IDs
    all_cat_ids = []
    for cat in categories:
        all_cat_ids.extend(cat.get_descendant_ids(include_self=True, only_active=True))
    all_cat_ids = list(set(all_cat_ids))

    # Query active products in these categories matching the brand name
    products = Product.objects.filter(
        is_active=True
    ).filter(
        Q(category_id__in=all_cat_ids) | Q(categories__id__in=all_cat_ids)
    ).filter(
        Q(name__icontains=brand_name) | Q(sku__icontains=brand_name)
    ).distinct()

    existing_product_ids = set(
        subrubro.product_order_rows.values_list("product_id", flat=True)
    )

    max_order = subrubro.product_order_rows.aggregate(
        Max("sort_order")
    )["sort_order__max"] or 0

    created_count = 0
    creates = []
    added_list = []
    for prod in products:
        if prod.id not in existing_product_ids:
            max_order += 10
            creates.append(
                BrandSubrubroProductOrder(
                    brand_subrubro=subrubro,
                    product=prod,
                    sort_order=max_order
                )
            )
            added_list.append({
                "id": prod.id,
                "sku": prod.sku,
                "name": prod.name
            })
            created_count += 1

    if creates:
        BrandSubrubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)

    log_admin_action(
        request,
        action="brand_subrubro_sync",
        target_type="brand_subrubro",
        target_id=subrubro.id,
        details={"brand": brand_name, "added_count": created_count}
    )

    return JsonResponse({
        "success": True,
        "added_count": created_count,
        "added_products": added_list
    })


@staff_member_required
def brand_rubro_products(request, pk):
    """Product catalog workspace for one brand rubro."""
    rubro = get_object_or_404(
        BrandRubro.objects.select_related("brand"),
        pk=pk,
    )
    target_info = _brand_workspace_target(rubro)
    if _brand_workspace_is_ajax(request):
        return _brand_workspace_search_response(request, target_info)
    return render(
        request,
        "admin_panel/brands/brand_rubro_products.html",
        _brand_workspace_context(target_info),
    )


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_add_product(request, pk):
    """Manually add a product to a brand rubro."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    product_id = request.POST.get('product_id', '').strip()
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'
    
    if product_id.isdigit():
        product = get_object_or_404(Product, pk=int(product_id))
        
        max_order = BrandRubroProductOrder.objects.filter(
            brand_rubro=rubro
        ).aggregate(Max("sort_order"))["sort_order__max"] or 0
        
        row, created = BrandRubroProductOrder.objects.get_or_create(
            brand_rubro=rubro,
            product=product,
            defaults={"sort_order": max_order + 10}
        )
        if is_ajax:
            return JsonResponse({
                "success": True,
                "created": created,
                "product": {
                    "id": product.id,
                    "sku": product.sku,
                    "name": product.name
                }
            })
            
        if created:
            messages.success(request, f'Producto "{product.name}" agregado.')
        else:
            messages.info(request, f'El producto "{product.name}" ya existe en este rubro.')
            
    elif is_ajax:
        return JsonResponse({"success": False, "error": "ID de producto inválido."}, status=400)
        
    return redirect('admin_brand_rubro_products', pk=rubro.pk)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_remove_product(request, pk):
    """Remove a product from a brand rubro."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    product_id = request.POST.get('product_id', '').strip()
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'
    
    if product_id.isdigit():
        product = get_object_or_404(Product, pk=int(product_id))
        BrandRubroProductOrder.objects.filter(
            brand_rubro=rubro,
            product=product
        ).delete()
        
        if is_ajax:
            return JsonResponse({
                "success": True,
                "product_id": product.id
            })
            
        messages.success(request, f'Producto "{product.name}" removido.')
        
    elif is_ajax:
        return JsonResponse({"success": False, "error": "ID de producto inválido."}, status=400)
        
    return redirect('admin_brand_rubro_products', pk=rubro.pk)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_bulk_assign(request, pk):
    rubro = get_object_or_404(
        BrandRubro.objects.select_related("brand"),
        pk=pk,
    )
    return _brand_workspace_assign(request, rubro)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_bulk_remove(request, pk):
    rubro = get_object_or_404(
        BrandRubro.objects.select_related("brand"),
        pk=pk,
    )
    return _brand_workspace_remove(request, rubro)


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_products_reorder(request, pk):
    """AJAX endpoint to save manual product order in a brand rubro."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalido."}, status=400)

    ordered_ids = payload.get("ordered_ids", [])
    if not ordered_ids:
        return JsonResponse({"success": False, "error": "No hay productos para ordenar."}, status=400)

    ordered_ids = [int(x) for x in ordered_ids if str(x).isdigit()]

    existing_rows = {
        row.product_id: row
        for row in BrandRubroProductOrder.objects.filter(
            brand_rubro=rubro,
            product_id__in=ordered_ids
        )
    }

    updates = []
    creates = []

    for index, product_id in enumerate(ordered_ids, start=1):
        sort_order = index * 10
        row = existing_rows.get(product_id)
        if row:
            row.sort_order = sort_order
            updates.append(row)
        else:
            creates.append(
                BrandRubroProductOrder(
                    brand_rubro=rubro,
                    product_id=product_id,
                    sort_order=sort_order
                )
            )

    if creates:
        BrandRubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)
    if updates:
        BrandRubroProductOrder.objects.bulk_update(updates, ["sort_order"])

    log_admin_action(
        request,
        action="brand_rubro_products_reorder",
        target_type="brand_rubro",
        target_id=rubro.id,
        details={"ordered_ids": ordered_ids[:100], "count": len(ordered_ids)}
    )

    return JsonResponse({"success": True, "count": len(ordered_ids)})


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_sync(request, pk):
    """AJAX endpoint to auto-populate brand rubro using helper categories of its subrubros and brand keyword."""
    rubro = get_object_or_404(
        BrandRubro.objects.select_related("brand"),
        pk=pk,
    )
    if str(_brand_workspace_payload(request).get("action", "")).strip().lower() in {
        "preview",
        "confirm",
    }:
        return _brand_workspace_sync_action(request, rubro)
    brand_name = rubro.brand.name.upper()
    
    # Collect all helper categories from all active subrubros
    subrubros = rubro.subrubros.filter(is_active=True)
    all_cat_ids = []
    for sub in subrubros:
        for cat in sub.helper_categories.all():
            all_cat_ids.extend(cat.get_descendant_ids(include_self=True, only_active=True))
            
    all_cat_ids = list(set(all_cat_ids))
    if not all_cat_ids:
        return JsonResponse({"success": False, "error": "No hay categorías de subrubros configuradas para auto-poblar este rubro."})

    # Query active products in these categories matching the brand name
    products = Product.objects.filter(
        is_active=True
    ).filter(
        Q(category_id__in=all_cat_ids) | Q(categories__id__in=all_cat_ids)
    ).filter(
        Q(name__icontains=brand_name) | Q(sku__icontains=brand_name)
    ).distinct()

    existing_product_ids = set(
        rubro.product_order_rows.values_list("product_id", flat=True)
    )

    max_order = rubro.product_order_rows.aggregate(
        Max("sort_order")
    )["sort_order__max"] or 0

    created_count = 0
    creates = []
    added_list = []
    for prod in products:
        if prod.id not in existing_product_ids:
            max_order += 10
            creates.append(
                BrandRubroProductOrder(
                    brand_rubro=rubro,
                    product=prod,
                    sort_order=max_order
                )
            )
            added_list.append({
                "id": prod.id,
                "sku": prod.sku,
                "name": prod.name
            })
            created_count += 1

    if creates:
        BrandRubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)

    log_admin_action(
        request,
        action="brand_rubro_sync",
        target_type="brand_rubro",
        target_id=rubro.id,
        details={"brand": brand_name, "added_count": created_count}
    )

    return JsonResponse({
        "success": True,
        "added_count": created_count,
        "added_products": added_list
    })


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_rubro_bulk_add_category(request, pk):
    """AJAX endpoint to bulk associate all products of a main catalog category to a BrandRubro."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    category_id_str = request.POST.get('category_id', '').strip()
    
    if not category_id_str.isdigit():
        return JsonResponse({"success": False, "error": "ID de categoría inválido."}, status=400)
        
    category_id = int(category_id_str)
    category = get_object_or_404(Category, pk=category_id)
    
    # Get all active descendant category IDs
    descendant_ids = category.get_descendant_ids(include_self=True, only_active=True)
    
    # Query all active products in these categories that are NOT already associated with the rubro
    existing_product_ids = set(
        rubro.product_order_rows.values_list("product_id", flat=True)
    )
    
    products = Product.objects.filter(
        is_active=True,
        category_id__in=descendant_ids
    ).exclude(
        id__in=existing_product_ids
    ).distinct().order_by("name", "sku")
    
    max_order = rubro.product_order_rows.aggregate(
        Max("sort_order")
    )["sort_order__max"] or 0
    
    creates = []
    created_count = 0
    added_list = []
    for prod in products:
        max_order += 10
        creates.append(
            BrandRubroProductOrder(
                brand_rubro=rubro,
                product=prod,
                sort_order=max_order
            )
        )
        added_list.append({
            "id": prod.id,
            "sku": prod.sku,
            "name": prod.name
        })
        created_count += 1
        
    if creates:
        BrandRubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)
        
    log_admin_action(
        request,
        action="brand_rubro_bulk_add_category",
        target_type="brand_rubro",
        target_id=rubro.id,
        details={"category_id": category_id, "category_name": category.name, "added_count": created_count}
    )
    
    return JsonResponse({
        "success": True, 
        "added_count": created_count,
        "added_products": added_list
    })


@staff_member_required
@require_POST
@superuser_required_for_modifications
def brand_subrubro_bulk_add_category(request, pk):
    """AJAX endpoint to bulk associate all products of a main catalog category to a BrandSubrubro."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    category_id_str = request.POST.get('category_id', '').strip()
    
    if not category_id_str.isdigit():
        return JsonResponse({"success": False, "error": "ID de categoría inválido."}, status=400)
        
    category_id = int(category_id_str)
    category = get_object_or_404(Category, pk=category_id)
    
    # Get all active descendant category IDs
    descendant_ids = category.get_descendant_ids(include_self=True, only_active=True)
    
    # Query all active products in these categories that are NOT already associated with the subrubro
    existing_product_ids = set(
        subrubro.product_order_rows.values_list("product_id", flat=True)
    )
    
    products = Product.objects.filter(
        is_active=True,
        category_id__in=descendant_ids
    ).exclude(
        id__in=existing_product_ids
    ).distinct().order_by("name", "sku")
    
    max_order = subrubro.product_order_rows.aggregate(
        Max("sort_order")
    )["sort_order__max"] or 0
    
    creates = []
    created_count = 0
    added_list = []
    for prod in products:
        max_order += 10
        creates.append(
            BrandSubrubroProductOrder(
                brand_subrubro=subrubro,
                product=prod,
                sort_order=max_order
            )
        )
        added_list.append({
            "id": prod.id,
            "sku": prod.sku,
            "name": prod.name
        })
        created_count += 1
        
    if creates:
        BrandSubrubroProductOrder.objects.bulk_create(creates, ignore_conflicts=True)
        
    # Also associate to parent rubro if not already present
    rubro = subrubro.brand_rubro
    existing_rub_product_ids = set(
        rubro.product_order_rows.values_list("product_id", flat=True)
    )
    rub_max_order = rubro.product_order_rows.aggregate(Max("sort_order"))["sort_order__max"] or 0
    rub_creates = []
    for prod in products:
        if prod.id not in existing_rub_product_ids:
            rub_max_order += 10
            rub_creates.append(
                BrandRubroProductOrder(
                    brand_rubro=rubro,
                    product=prod,
                    sort_order=rub_max_order
                )
            )
            
    if rub_creates:
        BrandRubroProductOrder.objects.bulk_create(rub_creates, ignore_conflicts=True)
        
    log_admin_action(
        request,
        action="brand_subrubro_bulk_add_category",
        target_type="brand_subrubro",
        target_id=subrubro.id,
        details={"category_id": category_id, "category_name": category.name, "added_count": created_count}
    )
    
    return JsonResponse({
        "success": True, 
        "added_count": created_count,
        "added_products": added_list
    })


@staff_member_required
def brand_rubro_preview_category_bulk(request, pk):
    """AJAX endpoint to preview stats before bulk associating products of a category to a BrandRubro."""
    rubro = get_object_or_404(BrandRubro, pk=pk)
    category_id_str = request.GET.get('category_id', '').strip()
    
    if not category_id_str.isdigit():
        return JsonResponse({"success": False, "error": "ID de categoría inválido."}, status=400)
        
    category_id = int(category_id_str)
    category = get_object_or_404(Category, pk=category_id)
    
    descendant_ids = category.get_descendant_ids(include_self=True, only_active=True)
    
    existing_product_ids = set(
        rubro.product_order_rows.values_list("product_id", flat=True)
    )
    
    total_products = Product.objects.filter(
        is_active=True,
        category_id__in=descendant_ids
    ).distinct()
    
    total_count = total_products.count()
    associated_count = total_products.filter(id__in=existing_product_ids).count()
    new_count = total_count - associated_count
    
    return JsonResponse({
        "success": True,
        "total_count": total_count,
        "associated_count": associated_count,
        "new_count": new_count
    })


@staff_member_required
def brand_subrubro_preview_category_bulk(request, pk):
    """AJAX endpoint to preview stats before bulk associating products of a category to a BrandSubrubro."""
    subrubro = get_object_or_404(BrandSubrubro, pk=pk)
    category_id_str = request.GET.get('category_id', '').strip()
    
    if not category_id_str.isdigit():
        return JsonResponse({"success": False, "error": "ID de categoría inválido."}, status=400)
        
    category_id = int(category_id_str)
    category = get_object_or_404(Category, pk=category_id)
    
    descendant_ids = category.get_descendant_ids(include_self=True, only_active=True)
    
    existing_product_ids = set(
        subrubro.product_order_rows.values_list("product_id", flat=True)
    )
    
    total_products = Product.objects.filter(
        is_active=True,
        category_id__in=descendant_ids
    ).distinct()
    
    total_count = total_products.count()
    associated_count = total_products.filter(id__in=existing_product_ids).count()
    new_count = total_count - associated_count
    
    return JsonResponse({
        "success": True,
        "total_count": total_count,
        "associated_count": associated_count,
        "new_count": new_count
    })


@staff_member_required
def brand_catalog_inbox(request):
    """Review uncataloged products and apply deterministic suggestions in bulk."""
    search = sanitize_search_token(request.GET.get("q", ""))
    category_id = str(request.GET.get("category", "")).strip()
    supplier = sanitize_search_token(request.GET.get("supplier", ""))
    product_status = str(request.GET.get("product_status", "active")).strip().lower()

    products = uncataloged_products(
        Product.objects.select_related("category", "supplier_ref").prefetch_related("categories")
    )
    if product_status == "active":
        products = products.filter(is_active=True)
    elif product_status == "inactive":
        products = products.filter(is_active=False)
    if search:
        products = products.filter(
            Q(sku__icontains=search)
            | Q(name__icontains=search)
            | Q(supplier__icontains=search)
            | Q(supplier_ref__name__icontains=search)
        )
    if category_id.isdigit():
        category = Category.objects.filter(pk=int(category_id)).first()
        if category:
            category_ids = category.get_descendant_ids(include_self=True)
            products = products.filter(
                Q(category_id__in=category_ids) | Q(categories__id__in=category_ids)
            )
    if supplier:
        products = products.filter(
            Q(supplier__icontains=supplier) | Q(supplier_ref__name__icontains=supplier)
        )

    products = products.order_by("name", "sku").distinct()
    page_obj = Paginator(products, 30).get_page(request.GET.get("page"))
    engine = BrandSuggestionEngine()
    for product in page_obj.object_list:
        product.catalog_suggestions = engine.suggest(product)

    brands = Brand.objects.filter(is_active=True).prefetch_related(
        "rubros__subrubros"
    ).order_by("order", "name")
    suppliers = (
        Product.objects.exclude(supplier="")
        .values_list("supplier", flat=True)
        .distinct()
        .order_by("supplier")[:250]
    )

    return render(
        request,
        "admin_panel/brands/catalog_inbox.html",
        {
            "page_obj": page_obj,
            "quality_metrics": brand_quality_metrics(),
            "brands": brands,
            "category_options": get_cached_category_options(
                only_active=False,
                include_inactive_suffix=True,
            ),
            "suppliers": suppliers,
            "search": search,
            "category_id": category_id,
            "supplier": supplier,
            "product_status": product_status,
            "recent_batches": BrandCatalogBatch.objects.select_related(
                "brand", "brand_rubro", "brand_subrubro", "created_by"
            )[:8],
        },
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_catalog_assign(request):
    brand_id = str(request.POST.get("brand_id", "")).strip()
    rubro_id = str(request.POST.get("rubro_id", "")).strip()
    subrubro_id = str(request.POST.get("subrubro_id", "")).strip()
    if not brand_id.isdigit() or not rubro_id.isdigit():
        messages.error(request, "Selecciona una marca y un rubro validos.")
        return redirect(request.META.get("HTTP_REFERER") or "admin_brand_catalog_inbox")

    brand = get_object_or_404(Brand, pk=int(brand_id))
    rubro = get_object_or_404(BrandRubro, pk=int(rubro_id))
    subrubro = (
        get_object_or_404(BrandSubrubro, pk=int(subrubro_id))
        if subrubro_id.isdigit()
        else None
    )
    product_ids = request.POST.getlist("product_ids")
    observation = request.POST.get("observation", "")
    operation = (
        BrandCatalogBatch.OPERATION_RULE
        if request.POST.get("used_suggestion") == "1"
        else BrandCatalogBatch.OPERATION_ASSIGN
    )

    try:
        batch = assign_products_to_brand_catalog(
            product_ids=product_ids,
            brand=brand,
            rubro=rubro,
            subrubro=subrubro,
            user=request.user,
            observation=observation,
            operation=operation,
            mode=request.POST.get("mode", "add"),
        )
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(request.META.get("HTTP_REFERER") or "admin_brand_catalog_inbox")

    log_admin_action(
        request,
        action="brand_catalog_assign",
        target_type="brand_catalog_batch",
        target_id=batch.pk,
        details={
            "brand_id": brand.pk,
            "rubro_id": rubro.pk,
            "subrubro_id": subrubro.pk if subrubro else None,
            "product_ids": batch.product_ids,
            "observation": batch.observation,
            "mode": request.POST.get("mode", "add"),
        },
    )
    destination = str(subrubro or rubro)
    messages.success(
        request,
        f"{len(batch.product_ids)} producto(s) catalogado(s) en {destination}. Lote #{batch.pk}.",
    )
    return redirect("admin_brand_catalog_inbox")


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_catalog_batch_undo(request, pk):
    batch = get_object_or_404(BrandCatalogBatch, pk=pk)
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    try:
        batch = undo_brand_catalog_batch(batch, user=request.user)
    except ValueError as exc:
        if is_ajax:
            return JsonResponse({"success": False, "error": str(exc)}, status=400)
        messages.error(request, str(exc))
        return redirect(request.META.get("HTTP_REFERER") or "admin_brand_catalog_inbox")

    log_admin_action(
        request,
        action="brand_catalog_undo",
        target_type="brand_catalog_batch",
        target_id=batch.pk,
        details={"product_ids": batch.product_ids},
    )
    if is_ajax:
        return JsonResponse(
            {
                "success": True,
                "batch_id": batch.pk,
                "status": batch.status,
                "restored_count": len(batch.product_ids or []),
            }
        )
    messages.success(request, f"Lote #{batch.pk} deshecho sin modificar asociaciones anteriores.")
    return redirect(request.META.get("HTTP_REFERER") or "admin_brand_catalog_inbox")


def _brand_catalog_settings_context(*, alias_form=None, rule_form=None):
    catalog_tree = []
    for brand in Brand.objects.prefetch_related("rubros__subrubros").order_by("order", "name"):
        catalog_tree.append(
            {
                "id": brand.pk,
                "rubros": [
                    {
                        "id": rubro.pk,
                        "subrubros": [subrubro.pk for subrubro in rubro.subrubros.all()],
                    }
                    for rubro in brand.rubros.all()
                ],
            }
        )
    return {
        "alias_form": alias_form or BrandAliasForm(),
        "rule_form": rule_form or BrandCatalogRuleForm(),
        "aliases": BrandAlias.objects.select_related("brand").order_by("brand__name", "value"),
        "rules": BrandCatalogRule.objects.select_related(
            "brand", "brand_rubro", "brand_subrubro"
        ).order_by("-priority", "-confidence", "brand__name"),
        "quality_metrics": brand_quality_metrics(),
        "catalog_tree": catalog_tree,
    }


@staff_member_required
def brand_catalog_settings(request):
    return render(
        request,
        "admin_panel/brands/catalog_settings.html",
        _brand_catalog_settings_context(),
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_alias_create(request):
    form = BrandAliasForm(request.POST)
    if form.is_valid():
        alias = form.save()
        log_admin_action(
            request,
            action="brand_alias_create",
            target_type="brand_alias",
            target_id=alias.pk,
            details={"brand_id": alias.brand_id, "value": alias.value},
        )
        messages.success(request, f'Alias "{alias.value}" agregado a {alias.brand.name}.')
        return redirect("admin_brand_catalog_settings")
    return render(
        request,
        "admin_panel/brands/catalog_settings.html",
        _brand_catalog_settings_context(alias_form=form),
        status=400,
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_alias_delete(request, pk):
    alias = get_object_or_404(BrandAlias, pk=pk)
    details = {"brand_id": alias.brand_id, "value": alias.value}
    alias.delete()
    log_admin_action(
        request,
        action="brand_alias_delete",
        target_type="brand_alias",
        target_id=pk,
        details=details,
    )
    messages.success(request, "Alias eliminado.")
    return redirect("admin_brand_catalog_settings")


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_catalog_rule_create(request):
    form = BrandCatalogRuleForm(request.POST)
    if form.is_valid():
        rule = form.save()
        log_admin_action(
            request,
            action="brand_catalog_rule_create",
            target_type="brand_catalog_rule",
            target_id=rule.pk,
            details={
                "brand_id": rule.brand_id,
                "rubro_id": rule.brand_rubro_id,
                "subrubro_id": rule.brand_subrubro_id,
                "pattern": rule.pattern,
            },
        )
        messages.success(request, f'Regla para "{rule.pattern}" creada.')
        return redirect("admin_brand_catalog_settings")
    return render(
        request,
        "admin_panel/brands/catalog_settings.html",
        _brand_catalog_settings_context(rule_form=form),
        status=400,
    )


@staff_member_required
@superuser_required_for_modifications
@require_POST
def brand_catalog_rule_delete(request, pk):
    rule = get_object_or_404(BrandCatalogRule, pk=pk)
    details = {"brand_id": rule.brand_id, "pattern": rule.pattern}
    rule.delete()
    log_admin_action(
        request,
        action="brand_catalog_rule_delete",
        target_type="brand_catalog_rule",
        target_id=pk,
        details=details,
    )
    messages.success(request, "Regla eliminada.")
    return redirect("admin_brand_catalog_settings")


@staff_member_required
def brand_catalog_export(request):
    """Export uncataloged products and their best suggestion for offline review."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    products = list(
        uncataloged_products(
            Product.objects.filter(is_active=True)
            .select_related("category", "supplier_ref")
            .prefetch_related("categories")
        ).order_by("name", "sku")
    )
    engine = BrandSuggestionEngine()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pendientes por marca"
    headers = [
        "ID",
        "SKU",
        "Producto",
        "Proveedor",
        "Categoria principal",
        "Marca sugerida",
        "Rubro sugerido",
        "Subrubro sugerido",
        "Confianza",
        "Motivo",
        "Marca a asignar",
        "Rubro a asignar",
        "Subrubro a asignar",
        "Observacion",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="FF6B35")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for product in products:
        suggestion = (engine.suggest(product, limit=1) or [None])[0]
        primary_category = product.get_primary_category()
        sheet.append(
            [
                product.pk,
                product.sku,
                product.name,
                product.supplier_ref.name if product.supplier_ref_id else product.supplier,
                primary_category.get_full_path() if primary_category else "",
                suggestion["brand"].name if suggestion else "",
                suggestion["rubro"].name if suggestion and suggestion["rubro"] else "",
                suggestion["subrubro"].name if suggestion and suggestion["subrubro"] else "",
                suggestion["confidence"] if suggestion else "",
                suggestion["reason"] if suggestion else "",
                "",
                "",
                "",
                "",
            ]
        )

    widths = [10, 18, 52, 26, 42, 22, 22, 24, 12, 46, 22, 22, 24, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="catalogacion_marcas_pendiente_{timezone.localdate():%Y%m%d}.xlsx"'
    )
    return response
