"""
Admin Panel views - Custom admin interface.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.forms import PasswordResetForm, SetPasswordForm
from django.contrib.auth.password_validation import validate_password
from django.contrib import messages
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db import transaction, connection, IntegrityError
from django.db import DatabaseError
from django.db.models import (
    Q,
    Case,
    Count,
    Sum,
    Max,
    Avg,
    F,
    When,
    IntegerField,
    DecimalField,
    ExpressionWrapper,
    Value,
    Prefetch,
)
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.text import slugify
from django.utils.http import url_has_allowed_host_and_scheme
import json
import os
import re
from io import BytesIO, StringIO
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import urlencode, parse_qs
import csv
from openpyxl import Workbook

from catalog.models import Product, Category, CategoryAttribute, ClampMeasureRequest, Supplier, PriceList
from accounts.models import (
    AccountRequest,
    ClientCategory,
    ClientCompany,
    ClientPayment,
    ClientProfile,
    ClientTransaction,
)
from accounts.services.ledger import (
    create_adjustment_transaction,
    sync_order_charge_transaction,
)
from accounts.services.movement_lifecycle import (
    apply_transaction_state_transition,
    can_transition_transaction_state,
    is_transaction_reopen_locked as service_is_transaction_reopen_locked,
    movement_allows_print as service_movement_allows_print,
)
from orders.models import (
    ClampQuotation,
    Order,
    OrderItem,
    OrderProposal,
    OrderProposalItem,
    OrderRequest,
    OrderRequestEvent,
    OrderRequestItem,
    OrderStatusHistory,
)
from orders.services.workflow import (
    ROLE_ADMIN,
    ROLE_DEPOSITO,
    ROLE_FACTURACION,
    ROLE_VENTAS,
    can_user_transition_order,
    resolve_user_order_role,
)
from orders.services.request_workflow import (
    confirm_order_request,
    convert_request_to_order,
    create_order_proposal,
    record_order_request_event,
    reject_order_request,
)
from core.models import (
    AdminCompanyAccess,
    Company,
    DocumentSeries,
    FISCAL_BILLABLE_DOC_TYPES,
    FISCAL_DOC_TYPE_FA,
    FISCAL_DOC_TYPE_FB,
    FISCAL_DOC_TYPE_FC,
    FISCAL_INVOICE_DOC_TYPES,
    FISCAL_ISSUE_MODE_ARCA_WSFE,
    FISCAL_ISSUE_MODE_EXTERNAL_SAAS,
    FISCAL_ISSUE_MODE_MANUAL,
    FISCAL_STATUS_AUTHORIZED,
    FISCAL_STATUS_EXTERNAL_RECORDED,
    FISCAL_STATUS_PENDING_RETRY,
    FISCAL_STATUS_READY_TO_ISSUE,
    FISCAL_STATUS_REJECTED,
    FISCAL_STATUS_SUBMITTING,
    FISCAL_STATUS_VOIDED,
    FiscalDocument,
    FiscalPointOfSale,
    InternalDocument,
    SALES_BEHAVIOR_COTIZACION,
    SALES_DOCUMENT_BEHAVIOR_CHOICES,
    SALES_BEHAVIOR_FACTURA,
    SALES_BEHAVIOR_NOTA_CREDITO,
    SALES_BEHAVIOR_NOTA_DEBITO,
    SALES_BEHAVIOR_PEDIDO,
    SALES_BEHAVIOR_PRESUPUESTO,
    SALES_BEHAVIOR_RECIBO,
    SALES_BEHAVIOR_REMITO,
    SALES_BILLING_MODE_INTERNAL_DOCUMENT,
    SALES_BILLING_MODE_MANUAL_FISCAL,
    SalesDocumentType,
    SiteSettings,
    CatalogAnalyticsEvent,
    AdminAuditLog,
    ImportExecution,
    CatalogExcelTemplate,
    CatalogExcelTemplateSheet,
    CatalogExcelTemplateColumn,
    StockMovement,
    Warehouse,
)
from core.services.company_context import (
    admin_company_access_table_available,
    get_active_company,
    get_default_company,
    get_default_client_origin_company,
    get_preferred_client_company,
    get_user_companies,
    set_active_company,
    user_has_company_access,
)
from django.contrib.auth.models import Group, User
from admin_panel.forms.import_forms import ProductImportForm, ClientImportForm, CategoryImportForm
from admin_panel.forms.category_forms import CategoryForm
from admin_panel.fiscal_views import fiscal_health_view, fiscal_report_view
from admin_panel.forms.export_forms import (
    CatalogExcelTemplateForm,
    CatalogExcelTemplateSheetForm,
    CatalogExcelTemplateColumnForm,
)
from admin_panel.forms.sales_document_type_forms import SalesDocumentTypeForm, WarehouseForm
from catalog.services.product_importer import ProductImporter
from catalog.services.import_templates import (
    build_import_template_filename,
    build_product_import_template_workbook,
)
from accounts.services.client_importer import ClientImporter
from catalog.services.category_importer import CategoryImporter
from catalog.services.abrazadera_importer import AbrazaderaImporter
from catalog.services.supplier_sync import ensure_supplier, clean_supplier_name
from catalog.services.clamp_code import (
    DIAMETER_HUMAN_TO_COMPACT_DEFAULT,
    generarCodigo,
    parsearCodigo,
)
from catalog.services.clamp_quoter import (
    CLAMP_LAMINATED_ALLOWED_DIAMETERS,
    CLAMP_PRICE_LISTS,
    CLAMP_WEIGHT_MAP,
    calculate_clamp_quote,
    get_allowed_diameter_options,
    parse_decimal_value,
    parse_int_value,
)
from catalog.services.clamp_request_products import (
    publish_clamp_request_product,
)
from catalog.services.category_assignment import (
    normalize_category_ids,
    assign_categories_to_product,
    add_category_to_products,
    replace_categories_for_products,
    remove_category_from_products,
)
from core.services.import_manager import ImportTaskManager
from core.services.importer import sanitize_import_data
from core.services.background_jobs import dispatch_import_job
from core.services.fiscal import (
    is_company_fiscal_ready,
    is_invoice_ready,
    resolve_payment_due_date,
)
from core.services.fiscal_notifications import send_fiscal_document_email
from core.services.arca_client import ArcaConfigurationError, ArcaTemporaryError, ArcaWsfeClient
from core.services.fiscal_documents import (
    close_fiscal_document,
    create_local_fiscal_document_from_order,
    reopen_fiscal_document,
    register_external_fiscal_document_for_order,
    void_fiscal_document,
)

from core.services.documents import (
    ensure_document_for_adjustment,
    ensure_document_for_order,
    ensure_document_for_payment,
)
from core.services.sales_documents import (
    create_fiscal_document_from_sales_type,
    create_internal_document_from_sales_type,
    resolve_sales_document_type,
)
from core.services.advanced_search import (
    apply_compact_text_search,
    apply_text_search,
    apply_parsed_text_search,
    compact_search_token,
    parse_text_search_query,
    sanitize_search_token,
)
from core.services.catalog_excel_exporter import build_catalog_workbook, build_export_filename
from core.services.audit import log_admin_action, log_admin_change, model_snapshot
from core.services.pricing import resolve_effective_price_list
import traceback
import logging
from core.decorators import superuser_required_for_modifications

logger = logging.getLogger(__name__)
PRIMARY_SUPERADMIN_USERNAME = getattr(settings, "ADMIN_PRIMARY_SUPERADMIN_USERNAME", "josueflexs")
ADMIN_ROLE_CHOICES = [
    (ROLE_ADMIN, "Administracion"),
    (ROLE_VENTAS, "Ventas"),
    (ROLE_DEPOSITO, "Deposito"),
    (ROLE_FACTURACION, "Facturacion"),
]
ADMIN_ROLE_LABELS = dict(ADMIN_ROLE_CHOICES)
FISCAL_PRINT_DOC_META = {
    "FA": {"letter": "A", "code": "001"},
    "FB": {"letter": "B", "code": "006"},
    "FC": {"letter": "C", "code": "011"},
    "NCA": {"letter": "A", "code": "003"},
    "NCB": {"letter": "B", "code": "008"},
    "NCC": {"letter": "C", "code": "013"},
    "NDA": {"letter": "A", "code": "002"},
    "NDB": {"letter": "B", "code": "007"},
    "NDC": {"letter": "C", "code": "012"},
}
FISCAL_PRINT_COPY_LABELS = {
    "original": "ORIGINAL",
    "duplicado": "DUPLICADO",
    "triplicado": "TRIPLICADO",
}
INVOICE_FISCAL_DOC_TYPES = tuple(sorted(FISCAL_INVOICE_DOC_TYPES))
BILLABLE_FISCAL_DOC_TYPES = tuple(sorted(FISCAL_BILLABLE_DOC_TYPES))
EMITTABLE_FISCAL_DOC_TYPES = tuple(choice[0] for choice in FiscalDocument.DOC_TYPE_CHOICES)
CLIENT_FACTURABLE_STATUSES = {
    Order.STATUS_CONFIRMED,
    Order.STATUS_PREPARING,
    Order.STATUS_SHIPPED,
    Order.STATUS_DELIVERED,
}
CLIENT_REMITO_READY_STATUSES = {
    Order.STATUS_SHIPPED,
    Order.STATUS_DELIVERED,
}
ORDER_INTERNAL_DOC_STATUS_RULES = {
    DocumentSeries.DOC_COT: {
        Order.STATUS_DRAFT,
        Order.STATUS_CONFIRMED,
        Order.STATUS_PREPARING,
        Order.STATUS_SHIPPED,
        Order.STATUS_DELIVERED,
    },
    DocumentSeries.DOC_PED: {
        Order.STATUS_CONFIRMED,
        Order.STATUS_PREPARING,
        Order.STATUS_SHIPPED,
        Order.STATUS_DELIVERED,
    },
    DocumentSeries.DOC_REM: CLIENT_REMITO_READY_STATUSES,
}



from .helpers import *



# ===================== IMPORTERS =====================

def _import_row_errors(result, limit=50):
    return [
        {
            'row': row.row_number,
            'message': '; '.join(str(error) for error in (row.errors or [])),
            'data': sanitize_import_data(getattr(row, 'data', {}) or {}),
        }
        for row in result.row_results
        if not row.success
    ][:limit]


def _resolve_import_classes(import_type):
    if import_type == 'products':
        return ProductImportForm, ProductImporter
    if import_type == 'clients':
        return ClientImportForm, ClientImporter
    if import_type == 'categories':
        return CategoryImportForm, CategoryImporter
    if import_type == 'abrazaderas':
        return ProductImportForm, AbrazaderaImporter
    return None, None


def _cell_value(value):
    value = sanitize_import_data(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _extract_duplicate_rows(result):
    rows = []
    for row in result.row_results:
        message = '; '.join(str(error) for error in (row.errors or []))
        data = sanitize_import_data(getattr(row, 'data', {}) or {})
        if not data.get('_duplicate_sku') and 'duplicado dentro del archivo' not in message.lower():
            continue

        first_data = sanitize_import_data(data.get('_duplicate_first_data') or {})
        rows.append({
            'sku': data.get('_duplicate_sku') or data.get('sku') or data.get('codigo') or '',
            'first_row': data.get('_duplicate_first_row') or '',
            'duplicate_row': row.row_number,
            'first_name': first_data.get('nombre') or first_data.get('descripcion') or '',
            'duplicate_name': data.get('nombre') or data.get('descripcion') or '',
            'first_supplier': first_data.get('proveedor') or '',
            'duplicate_supplier': data.get('proveedor') or '',
            'first_category': first_data.get('rubro') or first_data.get('categoria') or '',
            'duplicate_category': data.get('rubro') or data.get('categoria') or '',
            'first_price': first_data.get('precio_final') or first_data.get('precio') or '',
            'duplicate_price': data.get('precio_final') or data.get('precio') or '',
            'first_cost': first_data.get('costo') or '',
            'duplicate_cost': data.get('costo') or '',
            'message': message,
            'suggestion': 'Si son el mismo producto, deja una sola fila. Si son productos distintos, asigna un Codigo unico al duplicado.',
        })
    return rows


def _import_duplicate_warnings(result, limit=50):
    return [
        {
            'row': row.row_number,
            'message': '; '.join(str(error) for error in (row.errors or []))
            or 'SKU duplicado dentro del archivo.',
            'data': sanitize_import_data(getattr(row, 'data', {}) or {}),
        }
        for row in result.row_results
        if (getattr(row, 'data', {}) or {}).get('_duplicate_sku')
    ][:limit]


def _build_import_diagnostic_workbook(result, import_type):
    workbook = Workbook()
    duplicate_sheet = workbook.active
    duplicate_sheet.title = 'Duplicados'

    duplicate_headers = [
        'Codigo',
        'Fila original',
        'Nombre original',
        'Proveedor original',
        'Rubro original',
        'Precio final original',
        'Costo original',
        'Fila duplicada',
        'Nombre duplicado',
        'Proveedor duplicado',
        'Rubro duplicado',
        'Precio final duplicado',
        'Costo duplicado',
        'Mensaje',
        'Sugerencia',
    ]
    duplicate_sheet.append(duplicate_headers)

    duplicate_rows = _extract_duplicate_rows(result)
    for item in duplicate_rows:
        duplicate_sheet.append([
            _cell_value(item['sku']),
            _cell_value(item['first_row']),
            _cell_value(item['first_name']),
            _cell_value(item['first_supplier']),
            _cell_value(item['first_category']),
            _cell_value(item['first_price']),
            _cell_value(item['first_cost']),
            _cell_value(item['duplicate_row']),
            _cell_value(item['duplicate_name']),
            _cell_value(item['duplicate_supplier']),
            _cell_value(item['duplicate_category']),
            _cell_value(item['duplicate_price']),
            _cell_value(item['duplicate_cost']),
            _cell_value(item['message']),
            _cell_value(item['suggestion']),
        ])

    if not duplicate_rows:
        duplicate_sheet.append(['Sin SKUs duplicados detectados.'])

    error_sheet = workbook.create_sheet('Todos los errores')
    error_sheet.append(['Fila', 'Codigo', 'Nombre', 'Proveedor', 'Rubro', 'Precio final', 'Costo', 'Mensaje'])
    for row in result.row_results:
        if getattr(row, 'success', False):
            continue
        data = sanitize_import_data(getattr(row, 'data', {}) or {})
        error_sheet.append([
            _cell_value(row.row_number),
            _cell_value(data.get('_duplicate_sku') or data.get('sku') or data.get('codigo') or ''),
            _cell_value(data.get('nombre') or data.get('descripcion') or ''),
            _cell_value(data.get('proveedor') or ''),
            _cell_value(data.get('rubro') or data.get('categoria') or ''),
            _cell_value(data.get('precio_final') or data.get('precio') or ''),
            _cell_value(data.get('costo') or ''),
            _cell_value('; '.join(str(error) for error in (row.errors or []))),
        ])

    summary = workbook.create_sheet('Resumen', 0)
    summary.append(['Tipo de importacion', import_type])
    summary.append(['Filas', result.total_rows])
    summary.append(['Creados simulados', result.created])
    summary.append(['Actualizados simulados', result.updated])
    summary.append(['Errores', result.errors])
    summary.append(['Duplicados', len(duplicate_rows)])
    summary.append([])
    summary.append(['Lectura recomendada'])
    summary.append(['1', 'Revisar la hoja Duplicados.'])
    summary.append(['2', 'Si dos filas son el mismo producto, dejar una sola.'])
    summary.append(['3', 'Si son productos distintos, cambiar el Codigo del duplicado por uno unico.'])
    summary.append(['4', 'Volver a correr Vista previa antes de importar en real.'])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or '')))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
        sheet.freeze_panes = 'A2'

    return workbook


@user_passes_test(is_primary_superadmin)
@require_POST
def import_diagnostic_download(request, import_type):
    """Build an XLSX diagnostic report for dry-run validation errors."""
    FormClass, ImporterClass = _resolve_import_classes(import_type)
    if not ImporterClass:
        return JsonResponse({'success': False, 'error': 'Tipo de importacion no valido.'}, status=400)
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Selecciona un archivo para diagnosticar.'}, status=400)

    uploaded_file = request.FILES['file']
    temp_dir = os.path.join(settings.BASE_DIR, 'media', 'temp_imports')
    os.makedirs(temp_dir, exist_ok=True)
    file_basename = os.path.basename(uploaded_file.name)
    stamp = timezone.now().strftime('%Y%m%d%H%M%S%f')
    file_path = os.path.join(temp_dir, f'diagnostic_{stamp}_{file_basename}')

    try:
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        importer = ImporterClass(file_path)
        preview = importer.run(dry_run=True)
        workbook = _build_import_diagnostic_workbook(preview, import_type)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename_stamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = (
            f'attachment; filename="diagnostico_importacion_{import_type}_{filename_stamp}.xlsx"'
        )
        return response
    except Exception:
        logger.exception('Error building import diagnostic report')
        return JsonResponse({'success': False, 'error': 'No se pudo generar el diagnostico.'}, status=500)
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except OSError:
            pass


def run_background_import(task_id, execution_id, import_type, ImporterClass, file_path, dry_run):
    """Function to run in a separate thread."""
    execution = ImportExecution.objects.filter(pk=execution_id).first()
    try:
        preflight_errors = []
        if not dry_run:
            preflight_importer = ImporterClass(file_path)
            preflight_result = preflight_importer.run(dry_run=True)
            preflight_errors = _import_row_errors(preflight_result)
            if preflight_result.has_errors:
                result_data = {
                    'created': 0,
                    'updated': 0,
                    'errors': preflight_result.errors,
                    'has_errors': True,
                    'row_errors': preflight_errors,
                    'execution_id': execution_id,
                    'import_type': import_type,
                    'message': 'Validacion previa fallida. No se aplicaron cambios.',
                }
                ImportTaskManager.fail_task(task_id, 'La validacion previa detecto errores.')
                if execution:
                    execution.status = ImportExecution.STATUS_FAILED
                    execution.result_summary = result_data
                    execution.error_count = preflight_result.errors
                    execution.finished_at = timezone.now()
                    execution.save(update_fields=['status', 'result_summary', 'error_count', 'finished_at'])
                return

        def progress_callback(current, total):
            ImportTaskManager.update_progress(task_id, current, total, f"Procesando fila {current} de {total}")

        importer = ImporterClass(file_path)
        result = importer.run(dry_run=dry_run, progress_callback=progress_callback)

        created_refs = collect_created_refs(import_type, result.row_results) if not dry_run else []
        result_data = {
            'created': result.created,
            'updated': result.updated,
            'errors': result.errors,
            'has_errors': result.has_errors,
            'row_errors': _import_row_errors(result),
            'duplicate_count': len(_extract_duplicate_rows(result)),
            'duplicate_warnings': _import_duplicate_warnings(result),
            'preflight_errors': preflight_errors,
            'execution_id': execution_id,
            'import_type': import_type,
        }

        ImportTaskManager.complete_task(task_id, result_data)

        if execution:
            execution.status = ImportExecution.STATUS_COMPLETED
            execution.created_count = result.created
            execution.updated_count = result.updated
            execution.error_count = result.errors
            execution.result_summary = result_data
            execution.created_refs = created_refs
            execution.finished_at = timezone.now()
            execution.save(
                update_fields=[
                    'status',
                    'created_count',
                    'updated_count',
                    'error_count',
                    'result_summary',
                    'created_refs',
                    'finished_at',
                ]
            )
    except Exception as e:
        traceback.print_exc()
        ImportTaskManager.fail_task(task_id, str(e))
        if execution:
            execution.status = ImportExecution.STATUS_FAILED
            execution.result_summary = {'error': str(e)}
            execution.finished_at = timezone.now()
            execution.save(update_fields=['status', 'result_summary', 'finished_at'])
    finally:
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
        except OSError:
            pass


def import_status(request, task_id):
    """API to poll status."""
    if not request.user.is_authenticated:
        return JsonResponse(
            {'status': 'failed', 'message': 'Sesion expirada. Inicia sesion nuevamente.'},
            status=401,
        )
    if not request.user.is_staff:
        return JsonResponse(
            {'status': 'failed', 'message': 'No tienes permisos para consultar esta importacion.'},
            status=403,
        )

    status = ImportTaskManager.get_status(task_id)
    if status:
        return JsonResponse(status)

    execution_id = request.GET.get('execution_id', '').strip()
    if execution_id.isdigit():
        try:
            execution = ImportExecution.objects.filter(pk=int(execution_id)).first()
        except DatabaseError:
            return JsonResponse({
                'status': 'processing',
                'current': 0,
                'total': 1,
                'message': 'Esperando base de datos...',
            })

        if execution:
            if execution.status == ImportExecution.STATUS_COMPLETED:
                result = execution.result_summary or {
                    'created': execution.created_count,
                    'updated': execution.updated_count,
                    'errors': execution.error_count,
                    'execution_id': execution.pk,
                    'import_type': execution.import_type,
                }
                return JsonResponse({'status': 'completed', 'result': result})

            if execution.status == ImportExecution.STATUS_FAILED:
                error_msg = str((execution.result_summary or {}).get('error') or 'La importacion fallo.')
                return JsonResponse({'status': 'failed', 'message': error_msg})

            if execution.status == ImportExecution.STATUS_ROLLED_BACK:
                result = execution.result_summary or {
                    'created': execution.created_count,
                    'updated': execution.updated_count,
                    'errors': execution.error_count,
                    'execution_id': execution.pk,
                    'import_type': execution.import_type,
                }
                return JsonResponse({'status': 'completed', 'result': result})

            return JsonResponse({
                'status': 'processing',
                'current': 0,
                'total': 1,
                'message': 'Procesando en segundo plano...',
            })

    return JsonResponse({'status': 'unknown', 'message': 'No se encontro el estado de la importacion.'}, status=404)


@user_passes_test(is_primary_superadmin)
def import_dashboard(request):
    """Import dashboard / hub."""
    active_company = get_active_company(request)
    executions = ImportExecution.objects.select_related('user').order_by('-created_at')
    if active_company:
        executions = executions.filter(company=active_company)
    executions = executions[:40]
    return render(
        request,
        'admin_panel/importers/dashboard.html',
        {'executions': executions, 'active_company': active_company},
    )


@user_passes_test(is_primary_superadmin)
def import_template_download(request, import_type):
    if import_type not in ("products", "abrazaderas"):
        messages.error(request, "Este tipo de importacion no tiene plantilla descargable.")
        return redirect("admin_import_dashboard")

    workbook = build_product_import_template_workbook(import_type=import_type)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = build_import_template_filename(import_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@user_passes_test(is_primary_superadmin)
def import_process(request, import_type):
    """Handle file upload and processing for imports."""
    active_company = get_active_company(request)
    if not active_company:
        messages.error(request, "Selecciona una empresa activa antes de importar.")
        return redirect('select_company')
    if import_type == 'products':
        FormClass = ProductImportForm
        ImporterClass = ProductImporter
        template = 'admin_panel/importers/import_form.html'
    elif import_type == 'clients':
        FormClass = ClientImportForm
        ImporterClass = ClientImporter
        template = 'admin_panel/importers/import_form.html'
    elif import_type == 'categories':
        FormClass = CategoryImportForm
        ImporterClass = CategoryImporter
        template = 'admin_panel/importers/import_form.html'
    elif import_type == 'abrazaderas':
        FormClass = ProductImportForm
        ImporterClass = AbrazaderaImporter
        template = 'admin_panel/importers/import_form.html'
    else:
        messages.error(request, 'Tipo de importacion no valido.')
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES)
        if form.is_valid():
            try:
                uploaded_file = request.FILES['file']
                preview_only = request.POST.get('preview_only') == '1'
                temp_dir = os.path.join(settings.BASE_DIR, 'media', 'temp_imports')
                os.makedirs(temp_dir, exist_ok=True)

                file_basename = os.path.basename(uploaded_file.name)
                stamp = timezone.now().strftime('%Y%m%d%H%M%S%f')
                file_path = os.path.join(temp_dir, f'import_{stamp}_{file_basename}')
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)

                if preview_only:
                    importer = ImporterClass(file_path)
                    preview = importer.run(dry_run=True)
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    except OSError:
                        pass

                    return JsonResponse({
                        'success': True,
                        'preview': {
                            'total_rows': preview.total_rows,
                            'created': preview.created,
                            'updated': preview.updated,
                            'errors': preview.errors,
                            'has_errors': preview.has_errors,
                            'row_errors': _import_row_errors(preview),
                            'duplicate_count': len(_extract_duplicate_rows(preview)),
                            'duplicate_warnings': _import_duplicate_warnings(preview),
                        },
                    })

                dry_run = form.cleaned_data.get('dry_run', True)
                confirm_apply = form.cleaned_data.get('confirm_apply', False)
                if not dry_run and not confirm_apply:
                    return JsonResponse(
                        {
                            'success': False,
                            'error': 'Debes confirmar explicitamente la aplicacion real antes de ejecutar.',
                        },
                        status=400,
                    )
                task_id = ImportTaskManager.start_task()
                execution = ImportExecution.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    company=active_company,
                    import_type=import_type,
                    file_name=file_basename,
                    dry_run=dry_run,
                    status=ImportExecution.STATUS_PROCESSING,
                    metrics={},
                    supplier_name="",
                )

                importer_class_path = f"{ImporterClass.__module__}.{ImporterClass.__name__}"
                dispatch_result = dispatch_import_job(
                    task_id=task_id,
                    execution_id=execution.pk,
                    import_type=import_type,
                    importer_class_path=importer_class_path,
                    file_path=file_path,
                    dry_run=dry_run,
                )

                log_admin_action(
                    request,
                    action='import_start',
                    target_type='import_execution',
                    target_id=execution.pk,
                    details={
                        'import_type': import_type,
                        'dry_run': dry_run,
                        'confirm_apply': bool(confirm_apply),
                        'file_name': file_basename,
                        'backend': dispatch_result.get('backend', 'thread'),
                    },
                )

                return JsonResponse({
                    'success': True,
                    'task_id': task_id,
                    'execution_id': execution.pk,
                    'message': 'Iniciando importacion...',
                })
            except Exception:
                logger.exception('Error starting import process')
                return JsonResponse({'success': False, 'error': 'No se pudo iniciar la importacion.'}, status=500)
    else:
        form = FormClass()

    return render(
        request,
        template,
        {
            'form': form,
            'import_type': import_type,
        },
    )


@user_passes_test(is_primary_superadmin)
@require_POST
@superuser_required_for_modifications
def import_rollback(request, execution_id):
    """Rollback records created by one import execution."""
    execution = get_object_or_404(ImportExecution, pk=execution_id)

    if execution.dry_run:
        messages.warning(request, 'No se puede aplicar rollback sobre una simulacion (dry run).')
        return redirect('admin_import_dashboard')
    if execution.status == ImportExecution.STATUS_ROLLED_BACK:
        messages.info(request, 'Este lote ya fue revertido.')
        return redirect('admin_import_dashboard')
    if execution.status != ImportExecution.STATUS_COMPLETED:
        messages.warning(request, 'Solo se pueden revertir importaciones completadas.')
        return redirect('admin_import_dashboard')

    refs = list(execution.created_refs or [])
    if not refs:
        messages.warning(request, 'No hay registros creados para revertir en este lote.')
        return redirect('admin_import_dashboard')

    deleted_count = 0
    try:
        with transaction.atomic():
            if execution.import_type in ('products', 'abrazaderas'):
                target_qs = Product.objects.filter(sku__in=refs)
            elif execution.import_type == 'categories':
                target_qs = Category.objects.filter(slug__in=refs)
            elif execution.import_type == 'clients':
                target_qs = User.objects.filter(username__in=refs)
            else:
                messages.error(request, 'Este tipo de importacion no soporta rollback automatico.')
                return redirect('admin_import_dashboard')

            deleted_count = target_qs.count()
            target_qs.delete()

            execution.status = ImportExecution.STATUS_ROLLED_BACK
            execution.rollback_at = timezone.now()
            execution.rollback_summary = {
                'deleted_count': deleted_count,
                'refs_total': len(refs),
                'refs_remaining': max(len(refs) - deleted_count, 0),
            }
            execution.save(update_fields=['status', 'rollback_at', 'rollback_summary'])

        log_admin_action(
            request,
            action='import_rollback',
            target_type='import_execution',
            target_id=execution.pk,
            details={
                'import_type': execution.import_type,
                'deleted_count': deleted_count,
                'refs_total': len(refs),
            },
        )
        messages.success(request, f'Rollback aplicado. Registros eliminados: {deleted_count}.')
    except Exception as exc:
        logger.exception('Rollback failed')
        messages.error(request, f'No se pudo completar el rollback: {exc}')

    return redirect('admin_import_dashboard')
# ===================== BULK DELETE ACTIONS =====================

@staff_member_required
@require_POST
@superuser_required_for_modifications
def product_delete_all(request):
    """Deletes ALL products if confirmation is correct."""
    confirmation = request.POST.get('confirmation', '').strip().lower()
    expected = "delete productos"
    
    if confirmation != expected:
        messages.error(request, f'Frase de confirmación incorrecta. Debe escribir: "{expected}"')
        return redirect('admin_product_list')
    
    count, _ = Product.objects.all().delete()
    log_admin_action(
        request,
        action='product_delete_all',
        target_type='product_bulk',
        details={'deleted_count': count},
    )
    messages.success(request, f'Se eliminaron {count} productos correctamente.')
    return redirect('admin_product_list')

@staff_member_required
@require_POST
@superuser_required_for_modifications
def client_delete_all(request):
    """Deactivate ALL clients without hard delete."""
    confirmation = request.POST.get('confirmation', '').strip().lower()
    expected = "delete clientes"
    
    if confirmation != expected:
        messages.error(request, f'Frase de confirmación incorrecta. Debe escribir: "{expected}"')
        return redirect('admin_client_list')

    active_user_ids = list(
        ClientProfile.objects.filter(user_id__isnull=False).values_list('user_id', flat=True)
    )
    deactivated_users = User.objects.filter(id__in=active_user_ids, is_active=True).update(is_active=False)
    deactivated_profiles = ClientProfile.objects.filter(is_approved=True).update(is_approved=False)

    log_admin_change(
        request,
        action='client_deactivate_all',
        target_type='client_bulk',
        before={},
        after={
            'deactivated_users': deactivated_users,
            'deactivated_profiles': deactivated_profiles,
        },
    )
    messages.success(
        request,
        f'Se desactivaron {deactivated_profiles} perfiles y {deactivated_users} usuarios de clientes.',
    )
    return redirect('admin_client_list')

@staff_member_required
@require_POST
@superuser_required_for_modifications
def category_delete_all(request):
    """Deletes ALL categories if confirmation is correct."""
    confirmation = request.POST.get('confirmation', '').strip().lower()
    expected = "delete categorias"
    
    if confirmation != expected:
        messages.error(request, f'Frase de confirmación incorrecta. Debe escribir: "{expected}"')
        return redirect('admin_category_list')
    
    count, _ = Category.objects.all().delete()
    log_admin_action(
        request,
        action='category_delete_all',
        target_type='category_bulk',
        details={'deleted_count': count},
    )
    messages.success(request, f'Se eliminaron {count} categorías correctamente.')
    return redirect('admin_category_list')

__all__ = ['run_background_import', 'import_status', 'import_dashboard', 'import_template_download', 'import_diagnostic_download', 'import_process', 'import_rollback', 'product_delete_all', 'client_delete_all', 'category_delete_all']
