import json
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.db import connection
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from admin_panel.forms.import_forms import ClientImportForm
from accounts.models import AccountRequest, ClientCategory, ClientCompany, ClientPayment, ClientProfile, ClientTransaction
from catalog.models import Category, ClampMeasureRequest, Product, Supplier
from catalog.services.clamp_quoter import calculate_clamp_quote
from core.models import (
    AdminAuditLog,
    AdminCompanyAccess,
    CatalogExcelTemplate,
    CatalogExcelTemplateSheet,
    CatalogExcelTemplateColumn,
    Company,
    FiscalDocument,
    FiscalDocumentItem,
    FiscalPointOfSale,
    InternalDocument,
    SalesDocumentType,
)
from core.services.company_context import get_default_company, get_default_client_origin_company
from orders.models import ClampQuotation, Order, OrderItem, OrderRequest, OrderStatusHistory
from orders.services.workflow import ROLE_FACTURACION, ROLE_VENTAS


class ClientOrderHistoryViewTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_history',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_historial',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Historial',
        )
        self.company = get_default_company()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('100.00'),
            total=Decimal('100.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DRAFT,
            subtotal=Decimal('200.00'),
            total=Decimal('190.00'),
            discount_amount=Decimal('10.00'),
            discount_percentage=Decimal('5.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_staff_can_open_client_order_history(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_client_order_history', args=[self.client_profile.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['client'].pk, self.client_profile.pk)
        self.assertEqual(response.context['summary']['orders_count'], 2)

    def test_status_filter_applies_in_client_order_history(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'status': Order.STATUS_CONFIRMED},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['summary']['orders_count'], 1)

    def test_client_order_history_uses_tabbed_hub_context(self):
        self.client.force_login(self.staff)
        self._activate_company()
        documents_response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'client_tab': 'documents'},
        )
        account_response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'client_tab': 'account'},
        )

        self.assertEqual(documents_response.status_code, 200)
        self.assertEqual(documents_response.context['client_tab'], 'account')
        self.assertEqual(documents_response.context['ledger_tab'], 'invoices')
        self.assertEqual(len(documents_response.context['client_tabs']), 6)
        self.assertEqual(len(documents_response.context['ledger_tabs']), 6)
        self.assertContains(documents_response, 'Facturas')
        self.assertContains(documents_response, 'Cuenta Corriente')

        self.assertEqual(account_response.status_code, 200)
        self.assertEqual(account_response.context['ledger_tab'], 'account')
        self.assertContains(account_response, 'Nuevo movimiento')
        self.assertContains(account_response, 'Recibo')
        self.assertContains(account_response, 'Cotizacion')

    def test_client_order_history_account_tab_uses_commercial_receipt_labels(self):
        payment = ClientPayment.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            amount=Decimal('75.00'),
            method=ClientPayment.METHOD_CASH,
            reference='Pago prueba historial',
        )
        receipt_document = InternalDocument.objects.select_related('sales_document_type').get(
            payment=payment,
            doc_type='REC',
        )
        sales_type = receipt_document.sales_document_type
        self.assertIsNotNone(sales_type)
        sales_type.name = 'Recibo comercial historial'
        sales_type.letter = 'RCH'
        sales_type.save(update_fields=['name', 'letter', 'updated_at'])
        receipt_document.number = 915
        receipt_document.sales_document_type = sales_type
        receipt_document.save(update_fields=['number', 'sales_document_type', 'updated_at'])

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'client_tab': 'account', 'ledger_tab': 'payments'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Recibo comercial historial')
        self.assertContains(response, receipt_document.display_number)
        self.assertContains(response, 'Interno')

    def test_client_order_history_account_tab_uses_invoice_language_for_order_charge(self):
        order = Order.objects.filter(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
        ).first()
        self.assertIsNotNone(order)
        point_of_sale = FiscalPointOfSale.objects.create(
            company=self.company,
            number='15',
            is_active=True,
            is_default=True,
        )
        sales_type = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
        ).first()
        invoice_document = FiscalDocument.objects.create(
            source_key='test-client-history-ledger-invoice',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=order,
            point_of_sale=point_of_sale,
            sales_document_type=sales_type,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'client_tab': 'account', 'ledger_tab': 'invoices'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['ledger_tab'], 'invoices')
        self.assertContains(response, invoice_document.display_number)
        self.assertContains(response, 'Fiscal')

    def test_client_order_history_payments_tab_shows_commercial_receipt_and_linked_invoice(self):
        order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DELIVERED,
            subtotal=Decimal('180.00'),
            total=Decimal('180.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        point_of_sale = FiscalPointOfSale.objects.create(
            company=self.company,
            number='27',
            is_active=True,
            is_default=True,
        )
        invoice_sales_type = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
        ).first()
        invoice_document = FiscalDocument.objects.create(
            source_key='test-client-history-payments-invoice',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=order,
            point_of_sale=point_of_sale,
            sales_document_type=invoice_sales_type,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            subtotal_net=Decimal('180.00'),
            total=Decimal('180.00'),
        )
        payment = ClientPayment.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            order=order,
            amount=Decimal('90.00'),
            method=ClientPayment.METHOD_TRANSFER,
            reference='Pago aplicado a factura',
        )
        receipt_document = InternalDocument.objects.select_related('sales_document_type').get(
            payment=payment,
            doc_type='REC',
        )
        receipt_sales_type = receipt_document.sales_document_type
        self.assertIsNotNone(receipt_sales_type)
        receipt_sales_type.name = 'Recibo comercial aplicado'
        receipt_sales_type.letter = 'RCA'
        receipt_sales_type.save(update_fields=['name', 'letter', 'updated_at'])
        receipt_document.number = 321
        receipt_document.sales_document_type = receipt_sales_type
        receipt_document.save(update_fields=['number', 'sales_document_type', 'updated_at'])

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_order_history', args=[self.client_profile.pk]),
            {'client_tab': 'payments'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['ledger_tab'], 'payments')
        self.assertContains(response, 'Recibo comercial aplicado')
        self.assertContains(response, receipt_document.display_number)
        self.assertContains(response, invoice_document.display_number)

    def test_client_transaction_state_actions_and_open_movements_section(self):
        tx = ClientTransaction.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            transaction_type=ClientTransaction.TYPE_ADJUSTMENT,
            amount=Decimal('57.00'),
            description='Movimiento manual abierto',
            source_key='test:client-history:movement-state',
        )

        self.client.force_login(self.staff)
        self._activate_company()
        account_url = (
            reverse('admin_client_order_history', args=[self.client_profile.pk])
            + '?client_tab=account'
        )

        response = self.client.get(account_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Movimientos abiertos')
        self.assertIn(tx.pk, [row['tx'].pk for row in response.context['open_movement_rows']])

        close_response = self.client.post(
            reverse('admin_client_transaction_set_state', args=[self.client_profile.pk, tx.pk]),
            data={'state': 'closed', 'next': account_url},
        )
        self.assertEqual(close_response.status_code, 302)
        tx.refresh_from_db()
        self.assertEqual(tx.movement_state, ClientTransaction.STATE_CLOSED)
        self.assertIsNotNone(tx.closed_at)

        reopen_response = self.client.post(
            reverse('admin_client_transaction_set_state', args=[self.client_profile.pk, tx.pk]),
            data={'state': 'open', 'next': account_url},
        )
        self.assertEqual(reopen_response.status_code, 302)
        tx.refresh_from_db()
        self.assertEqual(tx.movement_state, ClientTransaction.STATE_OPEN)
        self.assertIsNone(tx.closed_at)

        void_response = self.client.post(
            reverse('admin_client_transaction_set_state', args=[self.client_profile.pk, tx.pk]),
            data={'state': 'voided', 'next': account_url},
        )
        self.assertEqual(void_response.status_code, 302)
        tx.refresh_from_db()
        self.assertEqual(tx.movement_state, ClientTransaction.STATE_VOIDED)
        self.assertIsNotNone(tx.voided_at)

        response_after_void = self.client.get(account_url)
        self.assertEqual(response_after_void.status_code, 200)
        self.assertNotIn(tx.pk, [row['tx'].pk for row in response_after_void.context['open_movement_rows']])

    def test_invoice_movement_cannot_be_reopened_after_closing(self):
        order = Order.objects.filter(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
        ).first()
        self.assertIsNotNone(order)

        point_of_sale = FiscalPointOfSale.objects.create(
            company=self.company,
            number="61",
            is_active=True,
            is_default=False,
        )
        sales_type = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
        ).first()
        FiscalDocument.objects.create(
            source_key='test:invoice:movement:locked',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=order,
            point_of_sale=point_of_sale,
            sales_document_type=sales_type,
            doc_type='FA',
            issue_mode='manual',
            status='authorized',
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        tx = ClientTransaction.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            order=order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
            amount=Decimal('100.00'),
            source_key='test:invoice:movement:tx',
            movement_state=ClientTransaction.STATE_CLOSED,
            closed_at=timezone.now(),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        account_url = reverse('admin_client_order_history', args=[self.client_profile.pk]) + '?client_tab=account'

        response = self.client.post(
            reverse('admin_client_transaction_set_state', args=[self.client_profile.pk, tx.pk]),
            data={'state': 'open', 'next': account_url},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        tx.refresh_from_db()
        self.assertEqual(tx.movement_state, ClientTransaction.STATE_CLOSED)
        self.assertContains(response, 'no pueden volver a estado abierto')

    def test_remito_movement_cannot_be_reopened_after_closing(self):
        order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_SHIPPED,
            subtotal=Decimal('120.00'),
            total=Decimal('120.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        InternalDocument.objects.create(
            source_key='test:remito:movement:doc',
            doc_type='REM',
            number=401,
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=order,
        )
        tx = ClientTransaction.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            order=order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
            amount=Decimal('120.00'),
            source_key='test:remito:movement:tx',
            movement_state=ClientTransaction.STATE_CLOSED,
            closed_at=timezone.now(),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        account_url = reverse('admin_client_order_history', args=[self.client_profile.pk]) + '?client_tab=account'

        response = self.client.post(
            reverse('admin_client_transaction_set_state', args=[self.client_profile.pk, tx.pk]),
            data={'state': 'open', 'next': account_url},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        tx.refresh_from_db()
        self.assertEqual(tx.movement_state, ClientTransaction.STATE_CLOSED)
        self.assertContains(response, 'factura o remito')

    def test_internal_document_print_requires_closed_movement(self):
        order = Order.objects.filter(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
        ).first()
        self.assertIsNotNone(order)
        document = InternalDocument.objects.filter(
            order=order,
            doc_type='PED',
        ).order_by('-id').first()
        if not document:
            document = InternalDocument.objects.create(
                doc_type='PED',
                number=999001,
                company=self.company,
                client_company_ref=self.client_company,
                client_profile=self.client_profile,
                order=order,
            )

        tx = ClientTransaction.objects.filter(
            order=order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
        ).first()
        if not tx:
            tx = ClientTransaction.objects.create(
                client_profile=self.client_profile,
                company=self.company,
                order=order,
                transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
                amount=Decimal('100.00'),
                source_key='test:order:print-guard',
            )
        tx.movement_state = ClientTransaction.STATE_OPEN
        tx.save(update_fields=['movement_state', 'updated_at'])

        self.client.force_login(self.staff)
        self._activate_company()
        blocked = self.client.get(reverse('admin_internal_document_print', args=[document.pk]))
        self.assertEqual(blocked.status_code, 302)
        self.assertIn(reverse('admin_order_detail', args=[order.pk]), blocked.url)

        tx.movement_state = ClientTransaction.STATE_CLOSED
        tx.closed_at = timezone.now()
        tx.save(update_fields=['movement_state', 'closed_at', 'updated_at'])
        allowed = self.client.get(reverse('admin_internal_document_print', args=[document.pk]))
        self.assertEqual(allowed.status_code, 200)

    def test_fiscal_document_print_requires_closed_movement(self):
        order = Order.objects.filter(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
        ).first()
        self.assertIsNotNone(order)
        point_of_sale = FiscalPointOfSale.objects.create(
            company=self.company,
            number='44',
            is_active=True,
            is_default=False,
        )
        sales_type = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
        ).first()
        fiscal_document = FiscalDocument.objects.create(
            source_key='test-client-history-print-guard-fiscal',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=order,
            point_of_sale=point_of_sale,
            sales_document_type=sales_type,
            doc_type='FA',
            issue_mode='manual',
            status='external_recorded',
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )
        tx = ClientTransaction.objects.filter(
            order=order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
        ).first()
        if not tx:
            tx = ClientTransaction.objects.create(
                client_profile=self.client_profile,
                company=self.company,
                order=order,
                transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
                amount=Decimal('100.00'),
                source_key='test:order:print-guard-fiscal',
            )
        tx.movement_state = ClientTransaction.STATE_OPEN
        tx.save(update_fields=['movement_state', 'updated_at'])

        self.client.force_login(self.staff)
        self._activate_company()
        blocked = self.client.get(reverse('admin_fiscal_document_print', args=[fiscal_document.pk]))
        self.assertEqual(blocked.status_code, 302)
        self.assertIn(reverse('admin_fiscal_document_detail', args=[fiscal_document.pk]), blocked.url)

        tx.movement_state = ClientTransaction.STATE_CLOSED
        tx.closed_at = timezone.now()
        tx.save(update_fields=['movement_state', 'closed_at', 'updated_at'])
        allowed = self.client.get(reverse('admin_fiscal_document_print', args=[fiscal_document.pk]))
        self.assertEqual(allowed.status_code, 200)

    def test_quick_remito_redirects_to_latest_remito_document(self):
        shipped_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_SHIPPED,
            subtotal=Decimal('150.00'),
            total=Decimal('150.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={'action': 'remito', 'company_id': self.company.pk},
        )

        remito = InternalDocument.objects.filter(order=shipped_order, doc_type='REM').first()
        self.assertIsNotNone(remito)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('admin_internal_document_print', args=[remito.pk])}?copy=original",
        )

    def test_quick_invoice_opens_latest_unbilled_facturable_order(self):
        billable_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_PREPARING,
            subtotal=Decimal('220.00'),
            total=Decimal('220.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={'action': 'invoice', 'company_id': self.company.pk},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('admin_order_detail', args=[billable_order.pk]))

    def test_quick_quote_respects_selected_origin_channel(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={
                'action': 'quote',
                'company_id': self.company.pk,
                'origin_channel': Order.ORIGIN_WHATSAPP,
            },
        )

        created_order = Order.objects.filter(
            user=self.client_user,
            company=self.company,
        ).order_by('-pk').first()

        self.assertEqual(response.status_code, 302)
        self.assertIsNotNone(created_order)
        self.assertEqual(created_order.origin_channel, Order.ORIGIN_WHATSAPP)

    def test_quick_quote_related_sale_clones_source_order_items(self):
        source_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('240.00'),
            total=Decimal('228.00'),
            discount_amount=Decimal('12.00'),
            discount_percentage=Decimal('5.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        product = Product.objects.create(
            sku='REL-001',
            name='Producto Relacionado',
            price=Decimal('120.00'),
            stock=15,
            is_active=True,
        )
        OrderItem.objects.create(
            order=source_order,
            product=product,
            product_sku=product.sku,
            product_name=product.name,
            quantity=2,
            unit_price_base=Decimal('120.00'),
            discount_percentage_used=Decimal('5.00'),
            price_at_purchase=Decimal('114.00'),
            subtotal=Decimal('228.00'),
        )
        source_tx = ClientTransaction.objects.filter(
            order=source_order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
        ).order_by('-id').first()
        self.assertIsNotNone(source_tx)

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={
                'action': 'quote',
                'company_id': self.company.pk,
                'source_tx_id': source_tx.pk,
            },
        )

        related_order = (
            Order.objects.filter(user=self.client_user, company=self.company)
            .exclude(pk=source_order.pk)
            .order_by('-pk')
            .first()
        )
        self.assertIsNotNone(related_order)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('admin_order_detail', args=[related_order.pk]))
        self.assertIn('relacionada', (related_order.admin_notes or '').lower())

        related_items = list(related_order.items.all())
        self.assertEqual(len(related_items), 1)
        self.assertEqual(related_items[0].product_id, product.pk)
        self.assertEqual(related_items[0].quantity, 2)
        self.assertEqual(related_items[0].price_at_purchase, Decimal('114.00'))

    def test_quick_invoice_related_sale_uses_selected_transaction_order(self):
        source_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_PREPARING,
            subtotal=Decimal('310.00'),
            total=Decimal('310.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_PREPARING,
            subtotal=Decimal('520.00'),
            total=Decimal('520.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        source_tx = ClientTransaction.objects.filter(
            order=source_order,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
        ).order_by('-id').first()
        self.assertIsNotNone(source_tx)

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={
                'action': 'invoice',
                'company_id': self.company.pk,
                'source_tx_id': source_tx.pk,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('admin_order_detail', args=[source_order.pk]))

    def test_quick_credit_note_opens_latest_invoice_document(self):
        invoice_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DELIVERED,
            subtotal=Decimal('180.00'),
            total=Decimal('180.00'),
            client_company='Cliente Historial',
            client_company_ref=self.client_company,
        )
        point_of_sale = FiscalPointOfSale.objects.create(
            company=self.company,
            number='99',
            is_active=True,
            is_default=True,
        )
        invoice_document = FiscalDocument.objects.create(
            source_key='test-client-history-fa',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=invoice_order,
            point_of_sale=point_of_sale,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            subtotal_net=Decimal('180.00'),
            total=Decimal('180.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_quick_order', args=[self.client_profile.pk]),
            data={'action': 'credit_note', 'company_id': self.company.pk},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('admin_fiscal_document_detail', args=[invoice_document.pk]))


class DashboardHubRankingTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_dashboard_rankings',
            password='secret123',
            is_staff=True,
        )
        self.company = get_default_company()
        self.client_user = User.objects.create_user(
            username='cliente_dashboard_rankings',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Ranking',
        )
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.product = Product.objects.create(
            sku='RANK-001',
            name='Producto Ranking',
            price=Decimal('150.00'),
            stock=10,
            is_active=True,
        )
        self.point = FiscalPointOfSale.objects.create(
            company=self.company,
            number='11',
            is_active=True,
            is_default=True,
        )
        self.fiscal_document = FiscalDocument.objects.create(
            source_key='dashboard-ranking-invoice',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='external_recorded',
            issued_at=timezone.now(),
            subtotal_net=Decimal('150.00'),
            total=Decimal('150.00'),
        )
        FiscalDocumentItem.objects.create(
            fiscal_document=self.fiscal_document,
            line_number=1,
            product=self.product,
            sku=self.product.sku,
            description=self.product.name,
            quantity=Decimal('3.000'),
            unit_price_net=Decimal('50.00'),
            net_amount=Decimal('150.00'),
            total_amount=Decimal('150.00'),
        )
        ClientTransaction.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            transaction_type=ClientTransaction.TYPE_ORDER_CHARGE,
            source_key='dashboard-ranking-debt',
            amount=Decimal('150.00'),
            occurred_at=timezone.now(),
            description='Saldo de prueba',
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_dashboard_rankings_link_to_client_history_and_product(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(reverse('admin_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('admin_client_order_history', args=[self.client_profile.pk]))
        self.assertContains(response, reverse('admin_product_edit', args=[self.product.pk]))


class ClientCorePermissionsTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_clients_ops',
            password='secret123',
            is_staff=True,
            is_superuser=False,
        )
        self.primary_superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.client_user = User.objects.create_user(
            username='cliente_permiso',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Permisos',
        )

    def test_non_superadmin_staff_can_open_client_edit(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse('admin_client_edit', args=[self.client_profile.pk]))
        self.assertEqual(response.status_code, 200)

    def test_non_superadmin_staff_can_update_client_profile(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_client_edit', args=[self.client_profile.pk]),
            data={
                'company_name': 'Cliente Editado Staff',
                'cuit_dni': '20-11111111-1',
                'province': 'Buenos Aires',
                'address': 'Calle 123',
                'phone': '1111-2222',
                'discount': '7.5',
                'client_type': 'taller',
                'iva_condition': 'responsable_inscripto',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.client_profile.refresh_from_db()
        self.assertEqual(self.client_profile.company_name, 'Cliente Editado Staff')
        self.assertEqual(self.client_profile.discount, Decimal('7.5'))

    def test_non_superadmin_staff_cannot_change_client_password(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_client_password', args=[self.client_profile.pk]),
            data={'new_password1': 'nuevaClave123', 'new_password2': 'nuevaClave123'},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.client_user.refresh_from_db()
        self.assertTrue(self.client_user.check_password('secret123'))

    def test_non_superadmin_staff_cannot_delete_client(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_client_delete', args=[self.client_profile.pk]),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(ClientProfile.objects.filter(pk=self.client_profile.pk).exists())

    def test_primary_superadmin_can_open_client_edit(self):
        self.client.force_login(self.primary_superadmin)
        response = self.client.get(reverse('admin_client_edit', args=[self.client_profile.pk]))
        self.assertEqual(response.status_code, 200)

    def test_primary_superadmin_client_delete_deactivates_only(self):
        self.client.force_login(self.primary_superadmin)
        response = self.client.post(
            reverse('admin_client_delete', args=[self.client_profile.pk]),
            data={'cancel_reason': 'Baja administrativa'},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.client_profile.refresh_from_db()
        self.client_user.refresh_from_db()
        self.assertFalse(self.client_profile.is_approved)
        self.assertFalse(self.client_user.is_active)
        self.assertTrue(ClientProfile.objects.filter(pk=self.client_profile.pk).exists())


class PaymentPanelTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_payments_panel',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_payments_panel',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Panel Pagos',
        )
        self.company = get_default_company()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DRAFT,
            subtotal=Decimal('120.00'),
            total=Decimal('120.00'),
            client_company='Cliente Panel Pagos',
            client_company_ref=self.client_company,
        )
        self.order_confirmed = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('80.00'),
            total=Decimal('80.00'),
            client_company='Cliente Panel Pagos',
            client_company_ref=self.client_company,
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_staff_can_register_payment_from_panel(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_payment_list'),
            data={
                'action': 'create',
                'client_profile_id': self.client_profile.pk,
                'order_id': self.order.pk,
                'amount': '120.00',
                'method': ClientPayment.METHOD_TRANSFER,
                'paid_at': '2026-02-20T10:30',
                'reference': 'TRX-001',
                'notes': 'Pago completo',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        payment = ClientPayment.objects.filter(order=self.order, is_cancelled=False).first()
        self.assertIsNotNone(payment)
        self.assertEqual(payment.amount, Decimal('120.00'))

    def test_client_balance_uses_billed_documents_minus_payments(self):
        from accounts.services.ledger import sync_order_charge_transaction

        point = FiscalPointOfSale.objects.create(
            company=self.company,
            number='51',
            is_active=True,
            is_default=True,
        )
        sales_type = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
        ).first()
        FiscalDocument.objects.create(
            source_key='payment-panel-balance-invoice',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order_confirmed,
            point_of_sale=point,
            sales_document_type=sales_type,
            doc_type='FA',
            issue_mode='manual',
            status='external_recorded',
            issued_at=timezone.now(),
            subtotal_net=Decimal('80.00'),
            total=Decimal('80.00'),
        )
        sync_order_charge_transaction(order=self.order_confirmed)

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_payment_list'),
            data={
                'action': 'create',
                'client_profile_id': self.client_profile.pk,
                'order_id': self.order_confirmed.pk,
                'amount': '30.00',
                'method': ClientPayment.METHOD_CASH,
                'paid_at': '2026-02-20T11:00',
                'reference': 'EFE-001',
                'notes': '',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        self.client_profile.refresh_from_db()
        self.assertEqual(self.client_profile.get_total_orders_for_balance(), Decimal('80.00'))
        self.assertEqual(self.client_profile.get_total_paid(), Decimal('30.00'))
        self.assertEqual(self.client_profile.get_current_balance(), Decimal('50.00'))

    def test_client_balance_ignores_confirmed_order_without_invoice(self):
        self.assertEqual(self.client_profile.get_total_orders_for_balance(), Decimal('0.00'))
        self.assertEqual(self.client_profile.get_current_balance(), Decimal('0.00'))

    def test_payment_panel_shows_configured_receipt_label(self):
        self.client.force_login(self.staff)
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

        receipt_type = SalesDocumentType.objects.filter(
            company=self.company,
            code='recibo',
        ).first()
        self.assertIsNotNone(receipt_type)
        receipt_type.name = 'Recibo comercial'
        receipt_type.letter = 'RCB'
        receipt_type.save()

        response = self.client.post(
            reverse('admin_payment_list'),
            data={
                'action': 'create',
                'client_profile_id': self.client_profile.pk,
                'order_id': self.order.pk,
                'company_id': self.company.pk,
                'amount': '50.00',
                'method': ClientPayment.METHOD_TRANSFER,
                'paid_at': '2026-02-20T10:30',
                'reference': 'TRX-RECIBO-01',
                'notes': 'Pago parcial',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Recibo comercial')
        self.assertContains(response, 'RCB-')

    def test_payment_panel_can_apply_selected_receipt_type(self):
        self.client.force_login(self.staff)
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

        explicit_receipt_type = SalesDocumentType.objects.create(
            company=self.company,
            code='recibo-explicito',
            name='Recibo expreso',
            letter='RXP',
            enabled=True,
            document_behavior='Recibo',
            billing_mode='INTERNAL_DOCUMENT',
            internal_doc_type='REC',
            generate_stock_movement=False,
            generate_account_movement=False,
            group_equal_products=True,
            is_default=False,
            display_order=998,
        )

        response = self.client.post(
            reverse('admin_payment_list'),
            data={
                'action': 'create',
                'client_profile_id': self.client_profile.pk,
                'order_id': self.order.pk,
                'company_id': self.company.pk,
                'sales_document_type_id': explicit_receipt_type.pk,
                'amount': '60.00',
                'method': ClientPayment.METHOD_TRANSFER,
                'paid_at': '2026-02-20T12:15',
                'reference': 'TRX-RECIBO-02',
                'notes': 'Pago con tipo explicito',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        payment = ClientPayment.objects.filter(reference='TRX-RECIBO-02').first()
        self.assertIsNotNone(payment)
        receipt = InternalDocument.objects.filter(payment=payment, doc_type='REC').first()
        self.assertIsNotNone(receipt)
        self.assertEqual(receipt.sales_document_type_id, explicit_receipt_type.pk)
        self.assertContains(response, 'Recibo expreso')


class OrderAdminLookupTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_order_lookup',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_order_lookup',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Lookup',
        )
        self.company = get_default_company()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DRAFT,
            subtotal=Decimal('0.00'),
            total=Decimal('0.00'),
            client_company='Cliente Lookup',
            client_company_ref=self.client_company,
        )
        self.product = Product.objects.create(
            sku='BA03041',
            name='1/2 B.ARM FORD SOP BISAGRA CAPOT F-14000',
            price=Decimal('150.00'),
            cost=Decimal('80.00'),
            stock=5,
            is_active=True,
        )
        self.alt_product = Product.objects.create(
            sku='ABF201',
            name='ABRAZADERA FORJADA 18x82x220 SC',
            price=Decimal('120.00'),
            cost=Decimal('70.00'),
            stock=8,
            is_active=True,
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_staff_can_add_order_item_from_partial_product_text(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_order_item_add', args=[self.order.pk]),
            data={
                'sku': 'bisagra capot f14000',
                'quantity': '2',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        item = OrderItem.objects.filter(order=self.order, product=self.product).first()
        self.assertIsNotNone(item)
        self.assertEqual(item.quantity, 2)

    def test_staff_can_add_order_item_via_ajax_without_full_reload(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_order_item_add', args=[self.order.pk]),
            data={
                'sku': 'bisagra capot f14000',
                'quantity': '1',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload.get('ok'))
        self.assertEqual(payload.get('items_count'), 1)
        self.assertIn('items_tbody_html', payload)
        self.assertEqual(OrderItem.objects.filter(order=self.order).count(), 1)

    def test_staff_can_edit_order_item_and_replace_product(self):
        order_item = OrderItem.objects.create(
            order=self.order,
            product=self.product,
            product_sku=self.product.sku,
            product_name=self.product.name,
            quantity=2,
            unit_price_base=Decimal('150.00'),
            price_at_purchase=Decimal('150.00'),
            subtotal=Decimal('300.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_order_item_edit', args=[self.order.pk, order_item.pk]),
            data={
                'sku': self.alt_product.sku,
                'product_id': str(self.alt_product.pk),
                'quantity': '3',
                'price': '118.50',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        order_item.refresh_from_db()
        self.order.refresh_from_db()
        self.assertEqual(order_item.product_id, self.alt_product.pk)
        self.assertEqual(order_item.product_sku, self.alt_product.sku)
        self.assertEqual(order_item.product_name, self.alt_product.name)
        self.assertEqual(order_item.quantity, 3)
        self.assertEqual(order_item.price_at_purchase, Decimal('118.50'))
        self.assertEqual(order_item.subtotal, Decimal('355.50'))
        self.assertEqual(self.order.total, Decimal('355.50'))


class ConfiguredSalesDocumentTypeFlowTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_sales_doc_type',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_sales_doc_type',
            email='cliente_sales_doc_type@example.com',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Doc Type',
            document_type='cuit',
            document_number='20123456789',
            iva_condition='responsable_inscripto',
            fiscal_address='Av Test 123',
            fiscal_city='San Martin',
            fiscal_province='Buenos Aires',
            postal_code='1650',
        )
        self.company = get_default_company()
        self.company.legal_name = 'Flexs Test SA'
        self.company.cuit = '30-12345678-9'
        self.company.tax_condition = 'responsable_inscripto'
        self.company.fiscal_address = 'Indalecio Gomez 4215'
        self.company.fiscal_city = 'San Martin'
        self.company.fiscal_province = 'Buenos Aires'
        self.company.postal_code = '1650'
        self.company.point_of_sale_default = '1'
        self.company.save()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('100.00'),
            total=Decimal('100.00'),
            client_company='Cliente Doc Type',
            client_company_ref=self.client_company,
        )
        OrderItem.objects.create(
            order=self.order,
            product_name='Producto fiscal test',
            product_sku='FISC-001',
            quantity=1,
            price_at_purchase=Decimal('100.00'),
            subtotal=Decimal('100.00'),
        )
        self.point = FiscalPointOfSale.objects.filter(company=self.company).order_by('-is_default', 'number').first()
        if not self.point:
            self.point = FiscalPointOfSale.objects.create(
                company=self.company,
                number='1',
                is_active=True,
                is_default=True,
            )
        self.sales_document_type = (
            SalesDocumentType.objects.filter(
                company=self.company,
                document_behavior='Factura',
            )
            .order_by('-is_default', 'display_order', 'id')
            .first()
        )
        self.assertIsNotNone(self.sales_document_type)
        self.sales_document_type.point_of_sale = self.point
        self.sales_document_type.enabled = True
        self.sales_document_type.billing_mode = 'MANUAL_FISCAL_RECEIPT'
        self.sales_document_type.fiscal_doc_type = 'FA'
        self.sales_document_type.save(update_fields=[
            'point_of_sale',
            'enabled',
            'billing_mode',
            'fiscal_doc_type',
            'updated_at',
        ])

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_staff_can_register_external_fiscal_document_using_configured_type(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_order_fiscal_register_external', args=[self.order.pk]),
            data={
                'sales_document_type_id': str(self.sales_document_type.pk),
                'external_system': 'saas',
                'external_id': 'ext-123',
                'external_number': '0001-00000077',
            },
        )

        self.assertEqual(response.status_code, 302)
        fiscal_document = FiscalDocument.objects.get(order=self.order, external_id='ext-123')
        self.assertEqual(fiscal_document.sales_document_type_id, self.sales_document_type.pk)
        self.assertEqual(fiscal_document.point_of_sale_id, self.point.pk)
        self.assertEqual(fiscal_document.doc_type, 'FA')

    def test_fiscal_detail_uses_sales_record_layout(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-layout-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FB',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_fiscal_document_detail', args=[fiscal_document.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Datos basicos')
        self.assertContains(response, 'Productos de la venta')
        self.assertContains(response, 'Totales')

    def test_order_detail_uses_sales_record_sale_sheet_layout(self):
        OrderItem.objects.create(
            order=self.order,
            product_name='Producto layout',
            product_sku='LAYOUT-01',
            quantity=2,
            price_at_purchase=Decimal('50.00'),
            subtotal=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_order_detail', args=[self.order.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ficha de venta')
        self.assertContains(response, 'Datos basicos')
        self.assertContains(response, 'Productos de la venta')
        self.assertContains(response, 'Totales')
        self.assertNotContains(response, 'Resumen comercial')

    def test_order_invoice_open_creates_invoice_from_order(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_order_invoice_open', args=[self.order.pk]))

        self.assertEqual(response.status_code, 302)
        fiscal_document = FiscalDocument.objects.get(order=self.order)
        self.assertEqual(fiscal_document.sales_document_type.document_behavior, 'Factura')
        self.assertEqual(response.url, reverse('admin_fiscal_document_detail', args=[fiscal_document.pk]))

    def test_order_invoice_open_allows_pending_fiscal_data(self):
        self.sales_document_type.billing_mode = 'ELECTRONIC_AFIP_WSFE'
        self.sales_document_type.save(update_fields=['billing_mode', 'updated_at'])
        self.client_profile.document_type = ''
        self.client_profile.document_number = ''
        self.client_profile.fiscal_address = ''
        self.client_profile.save(update_fields=[
            'document_type',
            'document_number',
            'fiscal_address',
            'updated_at',
        ])

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_order_invoice_open', args=[self.order.pk]), follow=True)

        self.assertEqual(response.status_code, 200)
        fiscal_document = FiscalDocument.objects.get(order=self.order)
        self.assertEqual(fiscal_document.issue_mode, 'arca_wsfe')
        self.assertContains(response, 'Completa los datos fiscales antes de cerrar o emitir')

    def test_manual_fiscal_document_can_close_and_reopen(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-close-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        close_response = self.client.post(reverse('admin_fiscal_document_close', args=[fiscal_document.pk]))
        self.assertEqual(close_response.status_code, 302)
        fiscal_document.refresh_from_db()
        self.assertEqual(fiscal_document.status, 'external_recorded')

        reopen_response = self.client.post(reverse('admin_fiscal_document_reopen', args=[fiscal_document.pk]))
        self.assertEqual(reopen_response.status_code, 302)
        fiscal_document.refresh_from_db()
        self.assertEqual(fiscal_document.status, 'ready_to_issue')

    def test_manual_fiscal_close_sets_due_date_snapshot(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-due-date-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_fiscal_document_close', args=[fiscal_document.pk]))
        self.assertEqual(response.status_code, 302)
        fiscal_document.refresh_from_db()
        self.assertIsNotNone(fiscal_document.payment_due_date)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_fiscal_document_send_email_updates_trace_fields(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-email-send-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='external_recorded',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_fiscal_document_send_email', args=[fiscal_document.pk]))
        self.assertEqual(response.status_code, 302)
        fiscal_document.refresh_from_db()
        self.assertIsNotNone(fiscal_document.email_last_sent_at)
        self.assertEqual(fiscal_document.email_last_recipient, self.client_user.email)
        self.assertEqual(len(mail.outbox), 1)

    def test_ventas_role_cannot_close_fiscal_document(self):
        sales_only_user = User.objects.create_user(
            username='staff_sales_only',
            password='secret123',
            is_staff=True,
        )
        ventas_group, _ = Group.objects.get_or_create(name=ROLE_VENTAS)
        sales_only_user.groups.add(ventas_group)

        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-close-permission-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(sales_only_user)
        self._activate_company()
        response = self.client.post(
            reverse('admin_fiscal_document_close', args=[fiscal_document.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        fiscal_document.refresh_from_db()
        self.assertEqual(fiscal_document.status, 'ready_to_issue')
        self.assertContains(response, 'Requiere rol Facturacion o Administracion')

    def test_facturacion_role_can_open_fiscal_report(self):
        billing_user = User.objects.create_user(
            username='staff_facturacion_only',
            password='secret123',
            is_staff=True,
        )
        facturacion_group, _ = Group.objects.get_or_create(name=ROLE_FACTURACION)
        billing_user.groups.add(facturacion_group)

        self.client.force_login(billing_user)
        self._activate_company()
        response = self.client.get(reverse('admin_fiscal_report'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Reporte fiscal')

    def test_manual_fiscal_document_cannot_close_if_order_is_not_ready(self):
        self.client_profile.document_type = ''
        self.client_profile.document_number = ''
        self.client_profile.fiscal_address = ''
        self.client_profile.save(update_fields=[
            'document_type',
            'document_number',
            'fiscal_address',
            'updated_at',
        ])
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-close-not-ready-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_fiscal_document_close', args=[fiscal_document.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        fiscal_document.refresh_from_db()
        self.assertEqual(fiscal_document.status, 'ready_to_issue')
        self.assertContains(response, 'No se puede cerrar el comprobante')

    def test_manual_fiscal_document_can_be_voided(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-void-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_fiscal_document_void', args=[fiscal_document.pk]))

        self.assertEqual(response.status_code, 302)
        fiscal_document.refresh_from_db()
        self.assertEqual(fiscal_document.status, 'voided')

    def test_safe_fiscal_document_can_be_deleted(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-delete-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='ready_to_issue',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_fiscal_document_delete', args=[fiscal_document.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(FiscalDocument.objects.filter(pk=fiscal_document.pk).exists())

    def test_closed_fiscal_document_cannot_be_deleted(self):
        fiscal_document = FiscalDocument.objects.create(
            source_key='detail-manual-delete-blocked-test',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='external_recorded',
            sales_document_type=self.sales_document_type,
            subtotal_net=Decimal('100.00'),
            total=Decimal('100.00'),
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(
            reverse('admin_fiscal_document_delete', args=[fiscal_document.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(FiscalDocument.objects.filter(pk=fiscal_document.pk).exists())
        self.assertContains(response, 'No se puede eliminar el comprobante fiscal')

    def test_internal_document_can_be_deleted_when_unlinked(self):
        internal_document = InternalDocument.objects.create(
            source_key='internal-delete-test',
            doc_type='COT',
            number=412,
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
        )

        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.post(reverse('admin_internal_document_delete', args=[internal_document.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(InternalDocument.objects.filter(pk=internal_document.pk).exists())

    def test_internal_document_print_shows_configured_type_label(self):
        internal_type = SalesDocumentType.objects.create(
            company=self.company,
            code='remito-test',
            name='Remito comercial test',
            letter='REM',
            enabled=True,
            document_behavior='Remito',
            billing_mode='INTERNAL_DOCUMENT',
            internal_doc_type='REM',
            generate_stock_movement=False,
            generate_account_movement=False,
            group_equal_products=True,
            is_default=False,
            display_order=998,
        )
        internal_document = InternalDocument.objects.create(
            source_key='test:internal:print',
            doc_type='REM',
            number=12,
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            sales_document_type=internal_type,
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_internal_document_print', args=[internal_document.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Remito comercial test')
        self.assertContains(response, 'REM-00000012')

    def test_staff_can_generate_internal_document_using_configured_type(self):
        internal_type = SalesDocumentType.objects.create(
            company=self.company,
            code='cotizacion-manual-test',
            name='Cotizacion manual test',
            letter='COT',
            enabled=True,
            document_behavior='Cotizacion',
            billing_mode='INTERNAL_DOCUMENT',
            internal_doc_type='COT',
            generate_stock_movement=False,
            generate_account_movement=False,
            group_equal_products=True,
            is_default=False,
            display_order=997,
        )

        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_order_internal_document_create', args=[self.order.pk]),
            data={'sales_document_type_id': str(internal_type.pk)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        internal_document = InternalDocument.objects.get(order=self.order, doc_type='COT')
        self.assertEqual(internal_document.sales_document_type_id, internal_type.pk)
        self.assertContains(response, 'Cotizacion manual test')


class SalesDocumentTypeSettingsViewTests(TestCase):
    def setUp(self):
        self.superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.company = get_default_company()

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_primary_superadmin_can_open_sales_document_type_settings(self):
        self.client.force_login(self.superadmin)
        self._activate_company()

        response = self.client.get(reverse('admin_sales_document_type_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Tipos de venta y movimiento')
        self.assertContains(response, 'Cotizacion')

    def test_sales_type_alias_url_is_available(self):
        self.client.force_login(self.superadmin)
        self._activate_company()

        response = self.client.get(reverse('admin_sales_type_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Tipos de venta y movimiento')

    def test_primary_superadmin_can_delete_sales_document_type(self):
        self.client.force_login(self.superadmin)
        self._activate_company()
        sales_type = SalesDocumentType.objects.create(
            company=self.company,
            code='test-mov-delete',
            name='Movimiento temporal',
            letter='X',
            document_behavior='Pedido',
            billing_mode='INTERNAL_DOCUMENT',
            internal_doc_type='PED',
            enabled=True,
            display_order=999,
        )

        response = self.client.post(
            reverse('admin_sales_document_type_delete', args=[sales_type.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SalesDocumentType.objects.filter(pk=sales_type.pk).exists())
        self.assertContains(response, "Tipo de venta")
        self.assertContains(response, "Movimiento temporal")

    def test_create_sales_type_persists_advanced_fields(self):
        self.client.force_login(self.superadmin)
        self._activate_company()
        self.company.fiscal_city = 'San Martin'
        self.company.save(update_fields=['fiscal_city', 'updated_at'])

        response = self.client.post(
            reverse('admin_sales_document_type_create'),
            data={
                'code': 'factura-usd-test',
                'name': 'Factura Exportacion Test',
                'letter': 'E',
                'last_number': '7',
                'enabled': 'on',
                'document_behavior': 'Factura',
                'generate_stock_movement': 'on',
                'generate_account_movement': 'on',
                'group_equal_products': 'on',
                'prioritize_default_warehouse': 'on',
                'billing_mode': 'INTERNAL_DOCUMENT',
                'internal_doc_type': 'PED',
                'fiscal_doc_type': '',
                'use_document_situation': 'on',
                'default_sales_user_selector': 'CURRENT_USER',
                'print_address': 'Av. Industrial 123',
                'print_email': 'facturacion@example.com',
                'print_phones': '+54 11 1234-5678',
                'print_locality': 'San Martin',
                'print_signature': 'Mensaje comercial de prueba',
                'base_design': 'default',
                'notes': 'Nota interna',
                'is_default': '',
                'display_order': '11',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        created = SalesDocumentType.objects.get(code='factura-usd-test', company=self.company)
        self.assertTrue(created.use_document_situation)
        self.assertEqual(created.print_address, 'Av. Industrial 123')
        self.assertEqual(created.print_email, 'facturacion@example.com')
        self.assertEqual(created.print_phones, '+54 11 1234-5678')
        self.assertEqual(created.print_locality, 'San Martin')
        self.assertEqual(created.print_signature, 'Mensaje comercial de prueba')
        self.assertEqual(created.default_sales_user_label, 'El usuario que agrega la venta')
        self.assertEqual(created.base_design, 'default')
        self.assertEqual(created.notes, 'Nota interna')

    def test_create_sales_type_rejects_fiscal_mode_for_quote_behavior(self):
        self.client.force_login(self.superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_sales_document_type_create'),
            data={
                'code': 'cot-fiscal-invalida',
                'name': 'Cotizacion Fiscal Invalida',
                'letter': 'COT',
                'last_number': '0',
                'document_behavior': 'Cotizacion',
                'billing_mode': 'MANUAL_FISCAL_RECEIPT',
                'fiscal_doc_type': 'FB',
                'internal_doc_type': 'COT',
                'default_sales_user_selector': 'CURRENT_USER',
                'display_order': '12',
                'base_design': 'default',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Solo Factura / Nota de credito / Nota de debito permiten modo fiscal')
        self.assertFalse(
            SalesDocumentType.objects.filter(
                company=self.company,
                code='cot-fiscal-invalida',
            ).exists()
        )

    def test_create_sales_type_rejects_account_movement_for_quote(self):
        self.client.force_login(self.superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_sales_document_type_create'),
            data={
                'code': 'cot-cuenta-invalida',
                'name': 'Cotizacion con cuenta invalida',
                'letter': 'COT',
                'last_number': '0',
                'document_behavior': 'Cotizacion',
                'billing_mode': 'INTERNAL_DOCUMENT',
                'internal_doc_type': 'COT',
                'generate_account_movement': 'on',
                'default_sales_user_selector': 'CURRENT_USER',
                'display_order': '13',
                'base_design': 'default',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cotizacion y Presupuesto no deben impactar cuenta corriente')
        self.assertFalse(
            SalesDocumentType.objects.filter(
                company=self.company,
                code='cot-cuenta-invalida',
            ).exists()
        )

    def test_create_sales_type_accepts_point_of_sale_from_active_company(self):
        self.client.force_login(self.superadmin)
        self._activate_company()
        point = FiscalPointOfSale.objects.create(
            company=self.company,
            number='51',
            is_active=True,
            is_default=False,
        )

        response = self.client.post(
            reverse('admin_sales_document_type_create'),
            data={
                'code': 'fac-pos-valida',
                'name': 'Factura con POS valido',
                'letter': 'A',
                'point_of_sale': str(point.pk),
                'last_number': '0',
                'document_behavior': 'Factura',
                'billing_mode': 'MANUAL_FISCAL_RECEIPT',
                'fiscal_doc_type': 'FA',
                'default_sales_user_selector': 'CURRENT_USER',
                'display_order': '14',
                'base_design': 'default',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        created = SalesDocumentType.objects.filter(
            company=self.company,
            code='fac-pos-valida',
        ).first()
        self.assertIsNotNone(created)
        self.assertEqual(created.point_of_sale_id, point.pk)

    def test_create_sales_type_duplicate_default_behavior_shows_form_error(self):
        self.client.force_login(self.superadmin)
        self._activate_company()
        existing_default = SalesDocumentType.objects.filter(
            company=self.company,
            document_behavior='Factura',
            is_default=True,
        ).first()
        if not existing_default:
            existing_default = SalesDocumentType.objects.create(
                company=self.company,
                code='fac-default-base',
                name='Factura Default Base',
                letter='A',
                document_behavior='Factura',
                billing_mode='INTERNAL_DOCUMENT',
                internal_doc_type='PED',
                is_default=True,
                enabled=True,
                display_order=15,
            )

        response = self.client.post(
            reverse('admin_sales_document_type_create'),
            data={
                'code': 'fac-default-duplicado',
                'name': 'Factura Default Duplicada',
                'letter': 'A',
                'last_number': '0',
                'document_behavior': 'Factura',
                'billing_mode': 'INTERNAL_DOCUMENT',
                'internal_doc_type': 'PED',
                'is_default': 'on',
                'default_sales_user_selector': 'CURRENT_USER',
                'display_order': '16',
                'base_design': 'default',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ya existe un tipo de venta predeterminado para este comportamiento y canal.')
        self.assertFalse(
            SalesDocumentType.objects.filter(
                company=self.company,
                code='fac-default-duplicado',
            ).exists()
        )
        self.assertTrue(
            SalesDocumentType.objects.filter(
                pk=existing_default.pk,
                is_default=True,
            ).exists()
        )


class FiscalPrintTemplateTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_fiscal_print',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_fiscal_print',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Fiscal Print',
            document_type='cuit',
            document_number='20123456789',
            iva_condition='responsable_inscripto',
            fiscal_address='Av Siempre Viva 123',
            fiscal_city='San Martin',
            fiscal_province='Buenos Aires',
            postal_code='1650',
        )
        self.company = get_default_company()
        self.company.legal_name = 'Flexs Print SA'
        self.company.cuit = '30-12345678-9'
        self.company.tax_condition = 'responsable_inscripto'
        self.company.fiscal_address = 'Indalecio Gomez 4215'
        self.company.fiscal_city = 'San Martin'
        self.company.fiscal_province = 'Buenos Aires'
        self.company.postal_code = '1650'
        self.company.save()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('1000.00'),
            discount_percentage=Decimal('10.00'),
            discount_amount=Decimal('100.00'),
            total=Decimal('900.00'),
            client_company='Cliente Fiscal Print',
            client_company_ref=self.client_company,
            notes='Observacion visible en factura',
        )
        self.point = FiscalPointOfSale.objects.create(
            company=self.company,
            number='99',
            is_active=True,
            is_default=True,
        )
        self.document = FiscalDocument.objects.create(
            source_key='test:fiscal:print',
            company=self.company,
            client_company_ref=self.client_company,
            client_profile=self.client_profile,
            order=self.order,
            point_of_sale=self.point,
            doc_type='FA',
            issue_mode='manual',
            status='authorized',
            number=4869,
            subtotal_net=Decimal('1000.00'),
            discount_total=Decimal('100.00'),
            tax_total=Decimal('0.00'),
            total=Decimal('900.00'),
            cae='12345678901234',
        )
        FiscalDocumentItem.objects.create(
            fiscal_document=self.document,
            line_number=1,
            sku='BA03041',
            description='Buje de prueba para impresion',
            quantity=Decimal('2.000'),
            unit_price_net=Decimal('450.00'),
            discount_percentage=Decimal('0.00'),
            discount_amount=Decimal('0.00'),
            net_amount=Decimal('900.00'),
            iva_rate=Decimal('0.00'),
            iva_amount=Decimal('0.00'),
            total_amount=Decimal('900.00'),
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_fiscal_print_renders_saas_like_layout_blocks(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(
            reverse('admin_fiscal_document_print', args=[self.document.pk]),
            {'copy': 'duplicado'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Factura A')
        self.assertContains(response, 'DUPLICADO')
        self.assertContains(response, 'N 00099-00004869')
        self.assertContains(response, 'Condicion de venta')
        self.assertContains(response, 'Observaciones')


class ClampQuoterTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_clamp_quoter',
            password='secret123',
            is_staff=True,
        )

    def test_clamp_calculation_matches_expected_formula(self):
        result = calculate_clamp_quote({
            'client_name': 'Cliente Demo',
            'dollar_rate': '1000',
            'steel_price_usd': '1',
            'supplier_discount_pct': '0',
            'general_increase_pct': '23',
            'clamp_type': 'trefilada',
            'is_zincated': '0',
            'diameter': '1/2',
            'width_mm': '100',
            'length_mm': '200',
            'profile_type': 'SEMICURVA',
        })

        self.assertEqual(result['base_cost'], Decimal('622.91'))
        row_map = {row['key']: row for row in result['price_rows']}
        self.assertEqual(row_map['lista_1']['final_price'], Decimal('872.07'))
        self.assertEqual(row_map['facturacion']['final_price'], Decimal('1245.82'))

    def test_laminada_allows_only_configured_diameters(self):
        with self.assertRaises(ValueError):
            calculate_clamp_quote({
                'client_name': 'Cliente Demo',
                'dollar_rate': '1000',
                'steel_price_usd': '1',
                'supplier_discount_pct': '0',
                'general_increase_pct': '23',
                'clamp_type': 'laminada',
                'is_zincated': '0',
                'diameter': '1/2',
                'width_mm': '100',
                'length_mm': '200',
                'profile_type': 'PLANA',
            })

    def test_staff_can_save_clamp_quote(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_clamp_quoter'),
            data={
                'action': 'save_quote',
                'price_list_key': 'lista_2',
                'client_name': 'Cliente Guardado',
                'dollar_rate': '1000',
                'steel_price_usd': '1.2',
                'supplier_discount_pct': '5',
                'general_increase_pct': '23',
                'clamp_type': 'laminada',
                'is_zincated': '1',
                'diameter': '3/4',
                'width_mm': '80',
                'length_mm': '150',
                'profile_type': 'PLANA',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        saved = ClampQuotation.objects.filter(client_name='Cliente Guardado').first()
        self.assertIsNotNone(saved)
        self.assertEqual(saved.price_list, ClampQuotation.PRICE_LIST_2)
        self.assertGreater(saved.final_price, Decimal('0.00'))

    def test_staff_can_create_product_from_clamp_quoter_result(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_clamp_quoter'),
            data={
                'action': 'create_product',
                'price_list_key': 'facturacion',
                'client_name': 'Cliente Producto',
                'dollar_rate': '1000',
                'steel_price_usd': '1.2',
                'supplier_discount_pct': '5',
                'general_increase_pct': '23',
                'clamp_type': 'trefilada',
                'is_zincated': '0',
                'diameter': '3/4',
                'width_mm': '80',
                'length_mm': '220',
                'profile_type': 'SEMICURVA',
                'product_stock': '3',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        product = Product.objects.filter(name__icontains='ABRAZADERA TREFILADA DE 3/4 X 80 X 220 SEMICURVA').first()
        self.assertIsNotNone(product)
        self.assertEqual(product.stock, 3)
        self.assertTrue(product.is_active)
        self.assertEqual(product.supplier, 'COTIZADOR')
        self.assertTrue(product.categories.filter(name__icontains='ABRAZADERA').exists())
        self.assertIsNotNone(getattr(product, 'clamp_specs', None))

        expected = calculate_clamp_quote({
            'client_name': 'Cliente Producto',
            'dollar_rate': '1000',
            'steel_price_usd': '1.2',
            'supplier_discount_pct': '5',
            'general_increase_pct': '23',
            'clamp_type': 'trefilada',
            'is_zincated': '0',
            'diameter': '3/4',
            'width_mm': '80',
            'length_mm': '220',
            'profile_type': 'SEMICURVA',
        })
        expected_price_map = {row['key']: row['final_price'] for row in expected['price_rows']}
        self.assertEqual(product.cost, expected['base_cost'])
        self.assertEqual(product.price, expected_price_map['facturacion'])

    def test_clamp_code_parse_api(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('api_clamp_code_parse'),
            data=json.dumps({'code': 'ABT91685270P'}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['result']['diametro'], '9/16')
        self.assertEqual(payload['result']['ancho'], 85)
        self.assertEqual(payload['result']['largo'], 270)

    def test_clamp_code_generate_api(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('api_clamp_code_generate'),
            data=json.dumps({
                'tipo': 'ABT',
                'diametro': '3/4',
                'ancho': 80,
                'largo': 220,
                'forma': 'SEMICURVA',
            }),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['result']['codigo'], 'ABT3480220S')


class ClampMeasureRequestAdminTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_clamp_requests',
            password='secret123',
            is_staff=True,
        )
        self.clamp_request = ClampMeasureRequest.objects.create(
            client_name='Cliente Clamp',
            client_email='cliente@clamp.com',
            clamp_type='trefilada',
            is_zincated=False,
            diameter='3/4',
            width_mm=80,
            length_mm=220,
            profile_type='SEMICURVA',
            quantity=2,
            description='ABRAZADERA TREFILADA DE 3/4 X 80 X 220 SEMICURVA',
            generated_code='ABT3480220S',
            dollar_rate=Decimal('1000.00'),
            steel_price_usd=Decimal('1450.00'),
            supplier_discount_pct=Decimal('0.00'),
            general_increase_pct=Decimal('40.00'),
            base_cost=Decimal('1000.00'),
            selected_price_list='lista_1',
            estimated_final_price=Decimal('1400.00'),
            exists_in_catalog=False,
        )

    def test_staff_can_open_clamp_request_queue(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse('admin_clamp_request_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Solicitud')

    def test_staff_can_update_clamp_request_status(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_clamp_request_detail', args=[self.clamp_request.pk]),
            data={
                'status': ClampMeasureRequest.STATUS_QUOTED,
                'admin_note': 'Cotizacion enviada por WhatsApp.',
                'confirmed_price_list': 'lista_2',
                'client_response_note': 'Precio confirmado para entrega inmediata.',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.clamp_request.refresh_from_db()
        self.assertEqual(self.clamp_request.status, ClampMeasureRequest.STATUS_QUOTED)
        self.assertEqual(self.clamp_request.processed_by, self.staff)
        self.assertEqual(self.clamp_request.confirmed_price_list, 'lista_2')
        self.assertIsNotNone(self.clamp_request.confirmed_price)
        self.assertEqual(self.clamp_request.client_response_note, 'Precio confirmado para entrega inmediata.')

    def test_staff_can_modify_quote_criteria_and_recalculate_price(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_clamp_request_detail', args=[self.clamp_request.pk]),
            data={
                'status': ClampMeasureRequest.STATUS_COMPLETED,
                'dollar_rate': '1200.00',
                'steel_price_usd': '1550.00',
                'supplier_discount_pct': '5.00',
                'general_increase_pct': '35.00',
                'selected_price_list': 'lista_3',
                'confirmed_price_list': 'lista_3',
                'client_response_note': 'Precio actualizado con nuevos criterios.',
                'admin_note': 'Ajuste de criterios economicos.',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.clamp_request.refresh_from_db()
        self.assertEqual(self.clamp_request.status, ClampMeasureRequest.STATUS_COMPLETED)
        self.assertEqual(self.clamp_request.dollar_rate, Decimal('1200.00'))
        self.assertEqual(self.clamp_request.steel_price_usd, Decimal('1550.00'))
        self.assertEqual(self.clamp_request.supplier_discount_pct, Decimal('5.00'))
        self.assertEqual(self.clamp_request.general_increase_pct, Decimal('35.00'))
        self.assertEqual(self.clamp_request.selected_price_list, 'lista_3')
        self.assertEqual(self.clamp_request.confirmed_price_list, 'lista_3')

        expected = calculate_clamp_quote({
            'client_name': self.clamp_request.client_name,
            'dollar_rate': '1200.00',
            'steel_price_usd': '1550.00',
            'supplier_discount_pct': '5.00',
            'general_increase_pct': '35.00',
            'clamp_type': self.clamp_request.clamp_type,
            'is_zincated': self.clamp_request.is_zincated,
            'diameter': self.clamp_request.diameter,
            'width_mm': str(self.clamp_request.width_mm),
            'length_mm': str(self.clamp_request.length_mm),
            'profile_type': self.clamp_request.profile_type,
        })
        price_map = {row['key']: row['final_price'] for row in expected['price_rows']}
        self.assertEqual(self.clamp_request.estimated_final_price, price_map['lista_3'])
        self.assertEqual(self.clamp_request.confirmed_price, price_map['lista_3'])

    def test_staff_can_modify_technical_data_and_refresh_all_outputs(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_clamp_request_detail', args=[self.clamp_request.pk]),
            data={
                'status': ClampMeasureRequest.STATUS_PENDING,
                'clamp_type': 'laminada',
                'is_zincated': '1',
                'diameter': '3/4',
                'width_mm': '95',
                'length_mm': '260',
                'profile_type': 'CURVA',
                'quantity': '5',
                'dollar_rate': '1300.00',
                'steel_price_usd': '1450.00',
                'supplier_discount_pct': '0.00',
                'general_increase_pct': '40.00',
                'selected_price_list': 'lista_2',
                'confirmed_price_list': '',
                'confirmed_price': '',
                'admin_note': 'Ajuste tecnico por nueva medida.',
                'client_response_note': '',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.clamp_request.refresh_from_db()
        self.assertEqual(self.clamp_request.clamp_type, 'laminada')
        self.assertTrue(self.clamp_request.is_zincated)
        self.assertEqual(self.clamp_request.diameter, '3/4')
        self.assertEqual(self.clamp_request.width_mm, 95)
        self.assertEqual(self.clamp_request.length_mm, 260)
        self.assertEqual(self.clamp_request.profile_type, 'CURVA')
        self.assertEqual(self.clamp_request.quantity, 5)
        self.assertIn('LAMINADA', self.clamp_request.description)
        self.assertIn('95 X 260 CURVA', self.clamp_request.description)
        self.assertTrue(self.clamp_request.generated_code.startswith('ABL'))

        expected = calculate_clamp_quote({
            'client_name': self.clamp_request.client_name,
            'dollar_rate': '1300.00',
            'steel_price_usd': '1450.00',
            'supplier_discount_pct': '0.00',
            'general_increase_pct': '40.00',
            'clamp_type': 'laminada',
            'is_zincated': '1',
            'diameter': '3/4',
            'width_mm': '95',
            'length_mm': '260',
            'profile_type': 'CURVA',
        })
        price_map = {row['key']: row['final_price'] for row in expected['price_rows']}
        self.assertEqual(self.clamp_request.base_cost, expected['base_cost'])
        self.assertEqual(self.clamp_request.estimated_final_price, price_map['lista_2'])


class OrderClampPublishTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_publish_clamp',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_publish_clamp',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Publish Clamp',
        )
        self.company = get_default_company()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('1500.00'),
            total=Decimal('1500.00'),
            client_company='Cliente Publish Clamp',
            client_company_ref=self.client_company,
        )
        self.product = Product.objects.create(
            sku='ABT3481220S-R1',
            name='ABRAZADERA TREFILADA DE 3/4 X 81 X 220 SEMICURVA',
            price=Decimal('1500.00'),
            stock=2,
            is_active=False,
            attributes={
                'source': 'clamp_request',
                'clamp_request_id': 0,
            },
        )
        self.clamp_request = ClampMeasureRequest.objects.create(
            client_user=self.client_user,
            client_name='Cliente Publish Clamp',
            clamp_type='trefilada',
            is_zincated=False,
            diameter='3/4',
            width_mm=81,
            length_mm=220,
            profile_type='SEMICURVA',
            quantity=2,
            description='ABRAZADERA TREFILADA DE 3/4 X 81 X 220 SEMICURVA',
            generated_code='ABT3481220S',
            linked_product=self.product,
            dollar_rate=Decimal('1300.00'),
            steel_price_usd=Decimal('1450.00'),
            supplier_discount_pct=Decimal('0.00'),
            general_increase_pct=Decimal('40.00'),
            base_cost=Decimal('1000.00'),
            selected_price_list='lista_1',
            estimated_final_price=Decimal('1400.00'),
            confirmed_price_list='lista_2',
            confirmed_price=Decimal('1500.00'),
            status=ClampMeasureRequest.STATUS_COMPLETED,
            exists_in_catalog=False,
        )
        self.order_item = OrderItem.objects.create(
            order=self.order,
            product=self.product,
            clamp_request=self.clamp_request,
            product_sku=self.product.sku,
            product_name=self.product.name,
            quantity=1,
            price_at_purchase=Decimal('1500.00'),
            subtotal=Decimal('1500.00'),
        )

    def test_staff_can_publish_clamp_order_item_to_catalog(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_order_item_publish_clamp', args=[self.order.pk, self.order_item.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.clamp_request.refresh_from_db()
        self.order_item.refresh_from_db()

        self.assertTrue(self.product.is_active)
        self.assertIsNotNone(self.clamp_request.published_to_catalog_at)
        self.assertTrue(
            self.product.categories.filter(name__icontains='ABRAZADERA').exists()
        )
        self.assertEqual(self.order_item.product_id, self.product.pk)
        self.assertTrue(
            Category.objects.filter(name__icontains='ABRAZADERA').exists()
        )
        expected = calculate_clamp_quote({
            'client_name': self.clamp_request.client_name,
            'dollar_rate': str(self.clamp_request.dollar_rate),
            'steel_price_usd': str(self.clamp_request.steel_price_usd),
            'supplier_discount_pct': str(self.clamp_request.supplier_discount_pct),
            'general_increase_pct': str(self.clamp_request.general_increase_pct),
            'clamp_type': self.clamp_request.clamp_type,
            'is_zincated': self.clamp_request.is_zincated,
            'diameter': self.clamp_request.diameter,
            'width_mm': str(self.clamp_request.width_mm),
            'length_mm': str(self.clamp_request.length_mm),
            'profile_type': self.clamp_request.profile_type,
        })
        fact_map = {row['key']: row['final_price'] for row in expected['price_rows']}
        self.assertEqual(self.product.price, fact_map['facturacion'])

    def test_publish_uses_confirmed_facturacion_price_when_present(self):
        self.clamp_request.confirmed_price_list = 'facturacion'
        self.clamp_request.confirmed_price = Decimal('2278.54')
        self.clamp_request.save(update_fields=['confirmed_price_list', 'confirmed_price'])

        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_order_item_publish_clamp', args=[self.order.pk, self.order_item.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.assertEqual(self.product.price, Decimal('2278.54'))


class OrderDeleteTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_order_delete',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_order_delete',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Delete',
        )
        self.company = get_default_company()
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.product = Product.objects.create(
            sku='DEL-001',
            name='Producto Delete',
            price=Decimal('100.00'),
            cost=Decimal('50.00'),
            stock=10,
            is_active=True,
        )
        self.order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            status=Order.STATUS_DRAFT,
            subtotal=Decimal('100.00'),
            total=Decimal('100.00'),
            client_company='Cliente Delete',
            client_company_ref=self.client_company,
        )
        OrderItem.objects.create(
            order=self.order,
            product=self.product,
            product_sku=self.product.sku,
            product_name=self.product.name,
            quantity=1,
            price_at_purchase=Decimal('100.00'),
            subtotal=Decimal('100.00'),
        )
        OrderStatusHistory.objects.create(
            order=self.order,
            from_status=Order.STATUS_DRAFT,
            to_status=Order.STATUS_DRAFT,
            note='Creado para test',
            changed_by=self.staff,
        )
        ClientPayment.objects.create(
            client_profile=self.client_profile,
            order=self.order,
            company=self.company,
            amount=Decimal('20.00'),
            method=ClientPayment.METHOD_TRANSFER,
            created_by=self.staff,
        )

        # Legacy row that used to block deletion through NO ACTION FK.
        if "accounts_clientaccountdocument" in connection.introspection.table_names():
            now = timezone.now()
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO accounts_clientaccountdocument (
                        document_type, document_number, issue_date, due_date, total_amount,
                        notes, is_cancelled, cancelled_at, cancel_reason, created_at, updated_at,
                        client_profile_id, created_by_id, order_id
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        'factura',
                        'TST-001',
                        now.date().isoformat(),
                        None,
                        '100.00',
                        'doc test',
                        0,
                        None,
                        '',
                        now,
                        now,
                        self.client_profile.pk,
                        self.staff.pk,
                        self.order.pk,
                    ],
                )

    def test_order_delete_endpoint_cancels_order_without_deleting(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_order_delete', args=[self.order.pk]),
            data={'cancel_reason': 'Cancelacion de prueba'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.status, Order.STATUS_CANCELLED)
        self.assertTrue(Order.objects.filter(pk=self.order.pk).exists())
        self.assertTrue(OrderItem.objects.filter(order_id=self.order.pk).exists())
        self.assertTrue(OrderStatusHistory.objects.filter(order_id=self.order.pk).exists())

        payment = ClientPayment.objects.first()
        self.assertIsNotNone(payment)
        self.assertEqual(payment.order_id, self.order.pk)

        if "accounts_clientaccountdocument" in connection.introspection.table_names():
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT COUNT(1) FROM accounts_clientaccountdocument WHERE order_id = %s",
                    [self.order.pk],
                )
                linked_count = cursor.fetchone()[0]
            self.assertEqual(linked_count, 1)

    def test_hard_delete_is_blocked_when_order_has_payments(self):
        self.client.force_login(self.staff)
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

        response = self.client.post(
            reverse('admin_order_hard_delete', args=[self.order.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Order.objects.filter(pk=self.order.pk).exists())
        self.assertContains(response, 'No se puede eliminar el pedido')

    def test_hard_delete_removes_safe_order_and_reopens_source_request(self):
        source_request = OrderRequest.objects.create(
            user=self.client_user,
            company=self.company,
            client_company_ref=self.client_company,
            status=OrderRequest.STATUS_CONVERTED,
            requested_total=Decimal('50.00'),
            requested_subtotal=Decimal('50.00'),
            submitted_at=timezone.now(),
            converted_at=timezone.now(),
        )
        deletable_order = Order.objects.create(
            user=self.client_user,
            company=self.company,
            source_request=source_request,
            status=Order.STATUS_DRAFT,
            subtotal=Decimal('50.00'),
            total=Decimal('50.00'),
            client_company='Cliente Delete',
            client_company_ref=self.client_company,
        )
        OrderItem.objects.create(
            order=deletable_order,
            product=self.product,
            product_sku=self.product.sku,
            product_name=self.product.name,
            quantity=1,
            price_at_purchase=Decimal('50.00'),
            subtotal=Decimal('50.00'),
        )

        self.client.force_login(self.staff)
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

        response = self.client.post(
            reverse('admin_order_hard_delete', args=[deletable_order.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Order.objects.filter(pk=deletable_order.pk).exists())
        source_request.refresh_from_db()
        self.assertEqual(source_request.status, OrderRequest.STATUS_CONFIRMED)
        self.assertIsNone(source_request.converted_at)


class CategoryManageProductsTests(TestCase):
    def setUp(self):
        self.superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.category = Category.objects.create(name='Categoria Test', slug='categoria-test')
        self.product = Product.objects.create(
            sku='CAT-TEST-001',
            name='Producto Categoria Test',
            price=Decimal('100.00'),
            cost=Decimal('50.00'),
            stock=5,
            is_active=True,
        )

    def test_assign_selected_products_to_category(self):
        self.client.force_login(self.superadmin)
        response = self.client.post(
            reverse('admin_category_products', args=[self.category.pk]),
            data={
                'action': 'assign',
                'select_all_pages': 'false',
                'product_ids': [str(self.product.pk)],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.assertTrue(self.product.categories.filter(pk=self.category.pk).exists())


class ProductBulkCategoryFallbackTests(TestCase):
    def setUp(self):
        self.superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.category = Category.objects.create(name='Categoria Bulk', slug='categoria-bulk')
        self.product = Product.objects.create(
            sku='BULK-TEST-001',
            name='Producto Bulk Test',
            price=Decimal('100.00'),
            cost=Decimal('50.00'),
            stock=2,
            is_active=True,
        )

    def test_bulk_category_assign_accepts_csv_fallback(self):
        self.client.force_login(self.superadmin)
        response = self.client.post(
            reverse('admin_product_bulk_category'),
            data={
                'category_id': str(self.category.pk),
                'mode': 'append',
                'select_all_pages': 'false',
                'product_ids_csv': str(self.product.pk),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.assertTrue(self.product.categories.filter(pk=self.category.pk).exists())

    def test_assign_selected_products_to_category_using_csv_fallback(self):
        self.client.force_login(self.superadmin)
        response = self.client.post(
            reverse('admin_category_products', args=[self.category.pk]),
            data={
                'action': 'assign',
                'select_all_pages': 'false',
                'product_ids_csv': str(self.product.pk),
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.assertTrue(self.product.categories.filter(pk=self.category.pk).exists())


class AdminInputValidationTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_validation',
            password='secret123',
            is_staff=True,
        )
        self.client_user = User.objects.create_user(
            username='cliente_validation',
            password='secret123',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Validacion',
            discount=Decimal('0.00'),
        )

    def test_client_edit_accepts_discount_with_comma(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse('admin_client_edit', args=[self.client_profile.pk]),
            data={
                'company_name': 'Cliente Validacion',
                'cuit_dni': '',
                'province': '',
                'address': '',
                'phone': '',
                'discount': '10,5',
                'client_type': '',
                'iva_condition': '',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.client_profile.refresh_from_db()
        self.assertEqual(self.client_profile.discount, Decimal('10.5'))

    def test_request_approve_rejects_blank_password(self):
        self.client.force_login(self.staff)
        request_row = AccountRequest.objects.create(
            company_name='Empresa Sin Password',
            contact_name='Contacto',
            email='sinpass@example.com',
            phone='1234',
            status='pending',
        )

        response = self.client.post(
            reverse('admin_request_approve', args=[request_row.pk]),
            data={
                'username': 'usuario_sin_password',
                'password': '',
                'discount': '0',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(User.objects.filter(username='usuario_sin_password').exists())
        request_row.refresh_from_db()
        self.assertEqual(request_row.status, 'pending')

    def test_request_approve_accepts_discount_with_comma(self):
        self.client.force_login(self.staff)
        request_row = AccountRequest.objects.create(
            company_name='Empresa Decimal',
            contact_name='Contacto Decimal',
            email='decimal@example.com',
            phone='1234',
            status='pending',
        )

        response = self.client.post(
            reverse('admin_request_approve', args=[request_row.pk]),
            data={
                'username': 'usuario_decimal_ok',
                'password': 'ClaveSegura123!',
                'discount': '10,5',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        user = User.objects.filter(username='usuario_decimal_ok').first()
        self.assertIsNotNone(user)
        profile = ClientProfile.objects.get(user=user)
        self.assertEqual(profile.discount, Decimal('10.5'))
        request_row.refresh_from_db()
        self.assertEqual(request_row.status, 'approved')


class ClientManagementViewTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_client_management',
            password='secret123',
            is_staff=True,
        )
        self.company = get_default_company()
        self.default_client_company = get_default_client_origin_company()
        self.client_user = User.objects.create_user(
            username='cliente_management',
            password='secret123',
            email='cliente_old@example.com',
        )
        self.client_profile = ClientProfile.objects.create(
            user=self.client_user,
            company_name='Cliente Gestion',
            is_approved=True,
        )
        self.client_company = ClientCompany.objects.create(
            client_profile=self.client_profile,
            company=self.company,
            is_active=True,
        )
        self.company_b = Company.objects.create(name='Flexs Secundaria')

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_staff_can_open_client_create_form(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_client_create'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nuevo cliente')
        self.assertContains(response, 'Contrasena inicial')

    def test_client_create_form_defaults_to_default_client_company(self):
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(reverse('admin_client_create'))

        self.assertEqual(response.status_code, 200)
        expected_company = self.default_client_company or self.company
        self.assertEqual(response.context['form_values']['company_id'], str(expected_company.pk))
        self.assertIn(str(expected_company.pk), response.context['form_values']['linked_company_ids'])

    def test_staff_can_create_client_with_user_profile_and_company_link(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_create'),
            data={
                'username': 'cliente_nuevo_panel',
                'email': 'cliente_nuevo@example.com',
                'first_name': 'Nuevo',
                'last_name': 'Cliente',
                'password': 'ClaveSegura123!',
                'password_confirm': 'ClaveSegura123!',
                'user_is_active': 'on',
                'client_is_approved': 'on',
                'company_is_active': 'on',
                'company_id': str(self.company.pk),
                'company_name': 'Cliente Nuevo Panel',
                'discount': '12,5',
                'phone': '11-4444-5555',
                'iva_condition': 'responsable_inscripto',
                'client_type': 'taller',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        new_user = User.objects.filter(username='cliente_nuevo_panel').first()
        self.assertIsNotNone(new_user)
        self.assertEqual(new_user.email, 'cliente_nuevo@example.com')
        self.assertEqual(new_user.first_name, 'Nuevo')
        new_profile = ClientProfile.objects.get(user=new_user)
        self.assertEqual(new_profile.company_name, 'Cliente Nuevo Panel')
        self.assertTrue(new_profile.is_approved)
        client_link = ClientCompany.objects.filter(
            client_profile=new_profile,
            company=self.company,
            is_active=True,
        ).first()
        self.assertIsNotNone(client_link)
        self.assertEqual(client_link.discount_percentage, Decimal('12.5'))
        self.assertContains(response, 'Editar cliente')

    def test_staff_can_create_client_with_multiple_company_links(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_create'),
            data={
                'username': 'cliente_multi_empresa',
                'email': 'cliente_multi@example.com',
                'first_name': 'Multi',
                'last_name': 'Empresa',
                'password': 'ClaveSegura123!',
                'password_confirm': 'ClaveSegura123!',
                'user_is_active': 'on',
                'client_is_approved': 'on',
                'company_is_active': 'on',
                'company_id': str(self.company.pk),
                'linked_company_ids': [str(self.company.pk), str(self.company_b.pk)],
                'company_name': 'Cliente Multiempresa',
                'discount': '8',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        new_user = User.objects.get(username='cliente_multi_empresa')
        new_profile = ClientProfile.objects.get(user=new_user)
        self.assertEqual(
            ClientCompany.objects.filter(client_profile=new_profile, is_active=True).count(),
            2,
        )
        self.assertTrue(new_profile.can_operate_in_company(self.company_b))

    def test_staff_can_update_client_user_fields_from_client_edit(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_edit', args=[self.client_profile.pk]),
            data={
                'username': 'cliente_management_editado',
                'email': 'cliente_editado@example.com',
                'first_name': 'Miguel',
                'last_name': 'Editado',
                'company_id': str(self.company.pk),
                'company_name': 'Cliente Gestion Editado',
                'cuit_dni': '20-11111111-1',
                'discount': '5',
                'user_is_active': 'on',
                'client_is_approved': 'on',
                'company_is_active': 'on',
                'client_type': 'taller',
                'iva_condition': 'responsable_inscripto',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.client_user.refresh_from_db()
        self.client_profile.refresh_from_db()
        self.client_company.refresh_from_db()
        self.assertEqual(self.client_user.username, 'cliente_management_editado')
        self.assertEqual(self.client_user.email, 'cliente_editado@example.com')
        self.assertEqual(self.client_user.first_name, 'Miguel')
        self.assertEqual(self.client_user.last_name, 'Editado')
        self.assertEqual(self.client_profile.company_name, 'Cliente Gestion Editado')
        self.assertEqual(self.client_profile.cuit_dni, '20-11111111-1')
        self.assertTrue(self.client_user.is_active)
        self.assertTrue(self.client_profile.is_approved)
        self.assertTrue(self.client_company.is_active)

    def test_staff_can_enable_client_in_multiple_companies_from_edit_form(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.post(
            reverse('admin_client_edit', args=[self.client_profile.pk]),
            data={
                'username': 'cliente_management',
                'email': 'cliente_old@example.com',
                'first_name': 'Cliente',
                'last_name': 'Gestion',
                'company_id': str(self.company_b.pk),
                'linked_company_ids': [str(self.company.pk), str(self.company_b.pk)],
                'company_name': 'Cliente Gestion',
                'cuit_dni': '',
                'discount': '5',
                'user_is_active': 'on',
                'client_is_approved': 'on',
                'company_is_active': 'on',
                'client_type': '',
                'iva_condition': '',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        second_link = ClientCompany.objects.filter(
            client_profile=self.client_profile,
            company=self.company_b,
        ).first()
        self.assertIsNotNone(second_link)
        self.assertTrue(second_link.is_active)
        self.assertTrue(self.client_profile.can_operate_in_company(self.company_b))

    def test_scoped_staff_only_sees_allowed_company_in_client_form(self):
        AdminCompanyAccess.objects.create(
            user=self.staff,
            company=self.company,
            is_active=True,
        )
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(reverse('admin_client_create'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.company.name)
        self.assertNotContains(response, self.company_b.name)

    def test_client_edit_shows_company_summary_and_recent_audit(self):
        AdminAuditLog.objects.create(
            user=self.staff,
            action='client_update',
            target_type='client_profile',
            target_id=str(self.client_profile.pk),
            details={'after': {'company_name': self.client_profile.company_name}},
        )
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(reverse('admin_client_edit', args=[self.client_profile.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Resumen por empresa')
        self.assertContains(response, 'Auditoria reciente')

    def test_client_list_highlights_profiles_without_email(self):
        self.client_user.email = ""
        self.client_user.save(update_fields=["email"])
        self.client.force_login(self.staff)
        self._activate_company()

        response = self.client.get(reverse("admin_client_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin email (1)")
        self.assertContains(response, "cliente(s) sin email")

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_primary_superadmin_can_send_client_password_reset_email(self):
        primary_superadmin = User.objects.create_superuser(
            username="josueflexs",
            email="josue@example.com",
            password="secret123",
        )
        self.client.force_login(primary_superadmin)
        self._activate_company()
        mail.outbox = []

        response = self.client.post(
            reverse("admin_client_password_reset_email", args=[self.client_profile.pk]),
            data={"next": reverse("admin_client_order_history", args=[self.client_profile.pk])},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, [self.client_user.email])
        self.assertContains(response, "Se envio mail de recuperacion")


class AdminUserPermissionsViewTests(TestCase):
    def setUp(self):
        self.primary_superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.target_admin = User.objects.create_user(
            username='operador_interno',
            password='secret123',
            is_staff=True,
        )
        self.company_a = get_default_company()
        self.company_b = Company.objects.create(name='Ubolt Interna', slug='ubolt-interna')

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company_a.pk
        session.save()

    def test_superadmin_can_assign_role_and_company_scope(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_user_permissions', args=[self.target_admin.pk]),
            data={
                'is_active': 'on',
                'is_staff': 'on',
                'admin_roles': ['ventas', 'deposito'],
                'company_scope_mode': 'limited',
                'allowed_company_ids': [str(self.company_b.pk)],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_admin.refresh_from_db()
        self.assertTrue(self.target_admin.groups.filter(name='ventas').exists())
        self.assertTrue(self.target_admin.groups.filter(name='deposito').exists())
        self.assertFalse(self.target_admin.groups.filter(name='admin').exists())
        self.assertEqual(
            list(
                AdminCompanyAccess.objects.filter(user=self.target_admin, is_active=True).values_list('company_id', flat=True)
            ),
            [self.company_b.pk],
        )
        self.assertContains(response, 'Permisos actualizados')

    def test_superadmin_can_edit_admin_identity_fields(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_user_edit', args=[self.target_admin.pk]),
            data={
                'username': 'operador_editado',
                'email': 'operador_editado@example.com',
                'first_name': 'Operador',
                'last_name': 'Editado',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_admin.refresh_from_db()
        self.assertEqual(self.target_admin.username, 'operador_editado')
        self.assertEqual(self.target_admin.email, 'operador_editado@example.com')
        self.assertEqual(self.target_admin.first_name, 'Operador')
        self.assertEqual(self.target_admin.last_name, 'Editado')
        self.assertContains(response, 'Informacion actualizada')

    def test_superadmin_can_reset_admin_password(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_user_password_change', args=[self.target_admin.pk]),
            data={
                'new_password1': 'ClaveNuevaSegura123!',
                'new_password2': 'ClaveNuevaSegura123!',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_admin.refresh_from_db()
        self.assertTrue(self.target_admin.check_password('ClaveNuevaSegura123!'))
        self.assertContains(response, 'Contrasena actualizada')

    def test_superadmin_can_deactivate_operator_account(self):
        ventas_group, _ = Group.objects.get_or_create(name='ventas')
        self.target_admin.groups.add(ventas_group)
        AdminCompanyAccess.objects.create(
            user=self.target_admin,
            company=self.company_b,
            is_active=True,
        )
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_user_delete', args=[self.target_admin.pk]),
            data={'cancel_reason': 'Baja de operador'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_admin.refresh_from_db()
        self.assertFalse(self.target_admin.is_active)
        self.assertFalse(self.target_admin.is_staff)
        self.assertFalse(self.target_admin.is_superuser)
        self.assertFalse(self.target_admin.groups.filter(name='ventas').exists())
        self.assertFalse(
            AdminCompanyAccess.objects.filter(user=self.target_admin, is_active=True).exists()
        )
        self.assertContains(response, 'fue dada de baja del panel')

    def test_superadmin_cannot_delete_primary_account(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse('admin_user_delete', args=[self.primary_superadmin.pk]),
            data={'cancel_reason': 'Intento invalido'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.primary_superadmin.refresh_from_db()
        self.assertTrue(self.primary_superadmin.is_active)
        self.assertTrue(self.primary_superadmin.is_staff)
        self.assertTrue(self.primary_superadmin.is_superuser)

    @patch("admin_panel.views.admin_company_access_table_available", return_value=False)
    def test_admin_list_still_loads_when_admin_scope_table_is_missing(self, _mock_scope_table):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.get(reverse("admin_user_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.primary_superadmin.username)

    def test_admin_list_highlights_accounts_without_email(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.get(reverse("admin_user_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "sin email")
        self.assertContains(response, self.target_admin.username)

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_primary_superadmin_can_send_admin_password_reset_email(self):
        self.target_admin.email = "operador_interno@example.com"
        self.target_admin.save(update_fields=["email"])
        self.client.force_login(self.primary_superadmin)
        self._activate_company()
        mail.outbox = []

        response = self.client.post(
            reverse("admin_user_password_reset_email", args=[self.target_admin.pk]),
            data={"next": reverse("admin_user_list")},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, [self.target_admin.email])
        self.assertContains(response, "Se envio mail de recuperacion")

class CatalogExcelTemplateExportTests(TestCase):
    def setUp(self):
        self.primary_superadmin = User.objects.create_superuser(
            username='josueflexs',
            email='josue@example.com',
            password='secret123',
        )
        self.staff = User.objects.create_user(
            username='staff_export_excel',
            password='secret123',
            is_staff=True,
        )

        self.category = Category.objects.create(name='Export Categoria', slug='export-categoria')
        self.supplier = Supplier.objects.create(name='Proveedor Export')
        self.product = Product.objects.create(
            sku='EXP-001',
            name='Producto Exportable',
            supplier='Proveedor Export',
            supplier_ref=self.supplier,
            price=Decimal('1234.50'),
            cost=Decimal('900.00'),
            stock=7,
            category=self.category,
            is_active=True,
        )
        self.product.categories.add(self.category)

        self.template = CatalogExcelTemplate.objects.create(
            name='Plantilla Test Excel',
            slug='plantilla-test-excel',
            is_active=True,
            created_by=self.primary_superadmin,
            updated_by=self.primary_superadmin,
        )
        self.sheet = CatalogExcelTemplateSheet.objects.create(
            template=self.template,
            name='Productos',
            include_header=True,
            only_active_products=True,
            sort_by='name_asc',
        )
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key='sku', order=1)
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key='name', order=2)
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key='price', order=3)

    def test_staff_can_download_catalog_template_excel(self):
        self.client.force_login(self.staff)
        response = self.client.get(
            reverse('admin_catalog_excel_template_download', args=[self.template.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml.sheet', response['Content-Type'])

        wb = load_workbook(BytesIO(response.content))
        ws = wb['Productos']
        self.assertEqual(ws['A1'].value, 'SKU')
        self.assertEqual(ws['B1'].value, 'Nombre')
        self.assertEqual(ws['A2'].value, 'EXP-001')
        self.assertEqual(ws['B2'].value, 'Producto Exportable')

    def test_non_primary_superadmin_cannot_create_template(self):
        other_superadmin = User.objects.create_superuser(
            username='otroadmin',
            email='otro@example.com',
            password='secret123',
        )
        self.client.force_login(other_superadmin)
        response = self.client.post(
            reverse('admin_catalog_excel_template_create'),
            data={
                'name': 'Plantilla Bloqueada',
                'description': 'No deberia crearse',
                'is_active': 'on',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            CatalogExcelTemplate.objects.filter(name='Plantilla Bloqueada').exists()
        )

    def test_staff_can_open_export_templates_pages(self):
        self.client.force_login(self.staff)
        list_response = self.client.get(reverse('admin_catalog_excel_template_list'))
        detail_response = self.client.get(
            reverse('admin_catalog_excel_template_detail', args=[self.template.pk])
        )
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(detail_response.status_code, 200)

    def test_primary_superadmin_can_autogenerate_sheets_by_root_categories(self):
        root_a = Category.objects.create(name='Abrazaderas Auto', slug='abrazaderas-auto')
        Category.objects.create(name='Sub A', slug='sub-a', parent=root_a)
        root_b = Category.objects.create(name='Bujes Auto', slug='bujes-auto')
        Category.objects.create(name='Inactiva Auto', slug='inactiva-auto', is_active=False)

        self.client.force_login(self.primary_superadmin)
        response = self.client.post(
            reverse(
                'admin_catalog_excel_template_autogenerate_main_category_sheets',
                args=[self.template.pk],
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        active_roots = list(
            Category.objects.filter(parent__isnull=True, is_active=True).order_by('id')
        )
        for root in active_roots:
            sheet = (
                CatalogExcelTemplateSheet.objects.filter(
                    template=self.template,
                    categories=root,
                )
                .order_by('id')
                .first()
            )
            self.assertIsNotNone(sheet)
            self.assertTrue(sheet.include_descendant_categories)
            self.assertTrue(sheet.only_active_products)
            self.assertTrue(sheet.only_catalog_visible)
            keys = list(sheet.columns.filter(is_active=True).order_by('order').values_list('key', flat=True))
            self.assertEqual(keys, ['sku', 'name', 'price'])

        inactive_root = Category.objects.filter(slug='inactiva-auto').first()
        self.assertFalse(
            CatalogExcelTemplateSheet.objects.filter(
                template=self.template,
                categories=inactive_root,
            ).exists()
        )

    def test_autogenerate_can_include_inactive_root_categories(self):
        inactive_root = Category.objects.create(
            name='Inactiva Incluida',
            slug='inactiva-incluida',
            is_active=False,
        )
        self.client.force_login(self.primary_superadmin)
        response = self.client.post(
            reverse(
                'admin_catalog_excel_template_autogenerate_main_category_sheets',
                args=[self.template.pk],
            ),
            data={'include_inactive_categories': '1'},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        generated_sheet = (
            CatalogExcelTemplateSheet.objects.filter(
                template=self.template,
                categories=inactive_root,
            )
            .order_by('id')
            .first()
        )
        self.assertIsNotNone(generated_sheet)
        self.assertFalse(generated_sheet.only_catalog_visible)

    def test_only_one_template_is_published_for_clients(self):
        second_template = CatalogExcelTemplate.objects.create(
            name='Plantilla Cliente 2',
            slug='plantilla-cliente-2',
            is_active=True,
            is_client_download_enabled=True,
            created_by=self.primary_superadmin,
            updated_by=self.primary_superadmin,
        )
        self.template.refresh_from_db()
        second_template.refresh_from_db()

        self.assertFalse(self.template.is_client_download_enabled)
        self.assertTrue(second_template.is_client_download_enabled)


class ClientReportsViewTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff_reports',
            password='secret123',
            is_staff=True,
        )
        self.company = get_default_company()
        self.category_a = ClientCategory.objects.create(name='Mayorista Reportes A')
        self.category_b = ClientCategory.objects.create(name='Mayorista Reportes B')

        self.client_user_a = User.objects.create_user(
            username='cliente_report_a',
            email='a@example.com',
            password='secret123',
            first_name='Ana',
        )
        self.client_profile_a = ClientProfile.objects.create(
            user=self.client_user_a,
            company_name='Cliente Reporte A',
            cuit_dni='20-11111111-1',
            fiscal_city='San Martin',
            fiscal_province='Buenos Aires',
            fiscal_address='Calle A 123',
            phone='1111-1111',
            iva_condition='responsable_inscripto',
            client_category=self.category_a,
            is_approved=True,
        )
        self.client_company_a = ClientCompany.objects.create(
            client_profile=self.client_profile_a,
            company=self.company,
            client_category=self.category_a,
            is_active=True,
        )

        self.client_user_b = User.objects.create_user(
            username='cliente_report_b',
            email='b@example.com',
            password='secret123',
            first_name='Beto',
            is_active=False,
        )
        self.client_profile_b = ClientProfile.objects.create(
            user=self.client_user_b,
            company_name='Cliente Reporte B',
            cuit_dni='20-22222222-2',
            province='Sin especificar',
            phone='2222-2222',
            iva_condition='monotributista',
            client_category=self.category_b,
            is_approved=False,
        )
        self.client_company_b = ClientCompany.objects.create(
            client_profile=self.client_profile_b,
            company=self.company,
            client_category=self.category_b,
            is_active=False,
        )

        Order.objects.create(
            user=self.client_user_a,
            company=self.company,
            status=Order.STATUS_CONFIRMED,
            subtotal=Decimal('200.00'),
            total=Decimal('200.00'),
            client_company='Cliente Reporte A',
            client_company_ref=self.client_company_a,
        )
        Order.objects.create(
            user=self.client_user_a,
            company=self.company,
            status=Order.STATUS_DELIVERED,
            subtotal=Decimal('80.00'),
            total=Decimal('80.00'),
            client_company='Cliente Reporte A',
            client_company_ref=self.client_company_a,
        )
        ClientPayment.objects.create(
            client_profile=self.client_profile_a,
            company=self.company,
            amount=Decimal('50.00'),
            method=ClientPayment.METHOD_TRANSFER,
            reference='Pago reporte',
        )
        ClientTransaction.objects.create(
            client_profile=self.client_profile_b,
            company=self.company,
            amount=Decimal('35.00'),
            transaction_type=ClientTransaction.TYPE_ADJUSTMENT,
            description='Saldo pendiente inactivo',
        )

    def _activate_company(self):
        session = self.client.session
        session['active_company_id'] = self.company.pk
        session.save()

    def test_reports_hub_loads(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_client_reports_hub'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Lista de clientes')
        self.assertContains(response, 'Ranking de clientes')
        self.assertContains(response, 'Clientes deudores')

    def test_client_tools_hub_loads(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(reverse('admin_client_tools_hub'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Exportar clientes')
        self.assertContains(response, 'Importar o actualizar')
        self.assertContains(response, 'Solicitudes')

    def test_client_list_report_filters_and_renders_rows(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_report_list'),
            {
                'action': 'generate',
                'locality': 'San Martin',
                'category': str(self.category_a.pk),
                'state': 'enabled',
                'iva_condition': 'responsable_inscripto',
                'text_field': 'company_name',
                'text': 'Reporte A',
                'columns': ['locality', 'price_list', 'balance'],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cliente Reporte A')
        self.assertNotContains(response, 'Cliente Reporte B')
        self.assertContains(response, 'San Martin')

    def test_client_ranking_report_uses_orders(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_report_ranking'),
            {
                'action': 'generate',
                'date_range': 'all',
                'ranking': 'top_10',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cliente Reporte A')
        self.assertEqual(response.context['rows'][0]['total_sales'], Decimal('280.00'))

    def test_client_debtors_report_uses_balance_logic(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_report_debtors'),
            {
                'action': 'generate',
                'report_type': 'disabled_non_zero',
                'tolerance': '1.00',
                'currency': 'all',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cliente Reporte B')
        self.assertNotContains(response, 'Cliente Reporte A')

    def test_client_reports_support_csv_download(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_report_ranking'),
            {
                'action': 'download',
                'date_range': 'all',
                'ranking': 'top_10',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')
        self.assertIn('Cliente Reporte A', response.content.decode('utf-8-sig'))

    def test_client_export_operational_download(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_export'),
            {
                'action': 'download',
                'preset': 'operational',
                'encoding': 'utf8',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')
        content = response.content.decode('utf-8-sig')
        self.assertIn('Nro de cliente;Categoria de cliente;Estado;Nombre', content)
        self.assertIn('Cliente Reporte A', content)

    def test_client_export_import_compatible_download_uses_selected_encoding(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_export'),
            {
                'action': 'download',
                'preset': 'import_compatible',
                'encoding': 'latin1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=iso-8859-1')
        content = response.content.decode('iso-8859-1')
        self.assertIn('Usuario;Contrasena;Nombre;Email', content)
        self.assertIn('cliente_report_a', content)

    def test_client_reports_generate_standalone_output(self):
        self.client.force_login(self.staff)
        self._activate_company()
        response = self.client.get(
            reverse('admin_client_report_ranking'),
            {
                'action': 'generate',
                'standalone': '1',
                'date_range': 'all',
                'ranking': 'top_10',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Informe generado')
        self.assertContains(response, 'Ranking de clientes')
        self.assertContains(response, 'Imprimir')
        self.assertContains(response, 'FLEXS Admin')
        self.assertNotContains(response, 'Panel de clientes')


@override_settings(
    IMPORTS_FORCE_SYNC=True,
    FEATURE_BACKGROUND_JOBS_ENABLED=False,
    DEFAULT_CLIENT_IMPORT_COMPANY_SLUGS=["flexs", "ubolt"],
)
class ClientImportProcessTests(TestCase):
    def setUp(self):
        self.primary_superadmin = User.objects.create_superuser(
            username="josueflexs",
            email="josueflexs@example.com",
            password="secret123",
        )
        self.company = get_default_company()
        ClientCategory.objects.get_or_create(name="N°1")

    def _activate_company(self):
        session = self.client.session
        session["active_company_id"] = self.company.pk
        session.save()

    def _build_import_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clientes"
        sheet.append(
            [
                "Usuario",
                "Nombre",
                "Email",
                "Tipo de cliente",
                "Cond. IVA",
                "Descuento",
                "Provincia",
                "Domicilio",
                "Telefonos",
                "Contacto",
            ]
        )
        sheet.append(
            [
                "import_test_user_admin",
                "Cliente Importado Admin",
                "import_test_user_admin@example.com",
                "N°1",
                "Consumidor final",
                0,
                "Buenos Aires",
                "Calle 123",
                "123456",
                "Contacto Test",
            ]
        )
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "clientes.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_real_client_import_can_start_successfully(self):
        self.client.force_login(self.primary_superadmin)
        self._activate_company()

        response = self.client.post(
            reverse("admin_import_process", args=["clients"]),
            {
                "file": self._build_import_file(),
                "dry_run": "",
                "confirm_apply": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertIn("execution_id", payload)


class ImportFormSecurityTests(TestCase):
    def _build_xlsx_payload(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Importacion"
        sheet.append(["Usuario", "Nombre"])
        sheet.append(["cliente_seguro", "Cliente Seguro"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    @override_settings(IMPORT_MAX_FILE_SIZE_BYTES=32)
    def test_import_form_rejects_oversized_file(self):
        upload = SimpleUploadedFile(
            "clientes.xlsx",
            b"x" * 64,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        form = ClientImportForm(data={}, files={"file": upload})

        self.assertFalse(form.is_valid())
        self.assertIn("file", form.errors)

    def test_import_form_rejects_invalid_content_type(self):
        upload = SimpleUploadedFile(
            "clientes.xlsx",
            self._build_xlsx_payload(),
            content_type="text/plain",
        )

        form = ClientImportForm(data={}, files={"file": upload})

        self.assertFalse(form.is_valid())
        self.assertIn("file", form.errors)

    def test_import_form_accepts_valid_xlsx_payload(self):
        upload = SimpleUploadedFile(
            "clientes.xlsx",
            self._build_xlsx_payload(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        form = ClientImportForm(data={}, files={"file": upload})

        self.assertTrue(form.is_valid(), form.errors)
