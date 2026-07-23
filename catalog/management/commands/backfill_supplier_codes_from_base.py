import json
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from catalog.models import Product, ProductSupplier


class Command(BaseCommand):
    help = (
        "Concilia referencias de proveedor desde una hoja BASE. Por defecto solo genera "
        "una vista previa; usa --apply para guardar exclusivamente coincidencias no ambiguas."
    )

    def add_arguments(self, parser):
        parser.add_argument("base_path", help="Ruta del archivo XLSX que contiene la hoja BASE.")
        parser.add_argument("--sheet", default="BASE", help="Nombre de la hoja. Predeterminado: BASE.")
        parser.add_argument("--apply", action="store_true", help="Guardar las coincidencias seguras.")
        parser.add_argument(
            "--replace-existing",
            action="store_true",
            help="Permitir reemplazar códigos existentes distintos. Requiere --apply.",
        )
        parser.add_argument("--report", help="Ruta opcional para guardar el informe JSON.")

    @staticmethod
    def _header(value):
        return str(value or "").strip().casefold()

    @staticmethod
    def _text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def handle(self, *args, **options):
        base_path = Path(options["base_path"]).expanduser().resolve()
        if not base_path.is_file():
            raise CommandError(f"No existe el archivo: {base_path}")
        if options["replace_existing"] and not options["apply"]:
            raise CommandError("--replace-existing solo puede utilizarse junto con --apply.")

        try:
            workbook = load_workbook(base_path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f"No se pudo abrir el XLSX: {exc}") from exc
        sheet_name = options["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f'No existe la hoja "{sheet_name}".')
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        headers = {self._header(value): index for index, value in enumerate(header_values or ())}
        sku_index = headers.get("código", headers.get("codigo"))
        reference_index = headers.get("referencia")
        if sku_index is None or reference_index is None:
            raise CommandError("La hoja debe contener las columnas Código y Referencia.")

        products_by_sku = {
            self._text(product.sku).upper(): product
            for product in Product.objects.only("id", "sku")
        }
        preferred_offers = list(
            ProductSupplier.objects.filter(is_preferred=True)
            .select_related("supplier", "product")
            .order_by("id")
        )
        offers_by_product = {
            offer.product_id: offer
            for offer in preferred_offers
        }
        offers_by_id = {offer.pk: offer for offer in preferred_offers}
        candidates = defaultdict(set)
        source_rows = defaultdict(list)
        references_by_offer = defaultdict(set)
        missing_sku = []
        without_offer = []
        missing_reference = []

        for row_number, values in enumerate(iterator, start=2):
            sku = self._text(values[sku_index] if sku_index < len(values) else "").upper()
            reference = self._text(
                values[reference_index] if reference_index < len(values) else ""
            )
            if not sku:
                continue
            product = products_by_sku.get(sku)
            if not product:
                missing_sku.append({"row": row_number, "sku": sku})
                continue
            offer = offers_by_product.get(product.pk)
            if not offer:
                without_offer.append({"row": row_number, "sku": sku, "product_id": product.pk})
                continue
            if not reference:
                missing_reference.append({"row": row_number, "sku": sku, "product_id": product.pk})
                continue
            normalized = ProductSupplier.normalize_supplier_code(reference)
            key = (offer.supplier_id, normalized)
            candidates[key].add(offer.pk)
            source_rows[(offer.pk, normalized)].append(row_number)
            references_by_offer[offer.pk].add(normalized)
        workbook.close()

        ambiguous_offer_ids = set()
        ambiguous = []
        for (supplier_id, normalized), offer_ids in candidates.items():
            if len(offer_ids) <= 1:
                continue
            ambiguous_offer_ids.update(offer_ids)
            ambiguous.append(
                {
                    "supplier_id": supplier_id,
                    "supplier_code": normalized,
                    "offer_ids": sorted(offer_ids),
                }
            )
        conflicting_offer_references = []
        for offer_id, references in references_by_offer.items():
            if len(references) <= 1:
                continue
            ambiguous_offer_ids.add(offer_id)
            conflicting_offer_references.append(
                {"offer_id": offer_id, "references": sorted(references)}
            )

        safe = []
        existing_conflicts = []
        unchanged = []
        for (offer_id, normalized), rows in source_rows.items():
            if offer_id in ambiguous_offer_ids:
                continue
            offer = offers_by_id.get(offer_id)
            if not offer:
                continue
            if offer.normalized_supplier_code == normalized:
                unchanged.append(offer_id)
                continue
            if offer.normalized_supplier_code and not options["replace_existing"]:
                existing_conflicts.append(
                    {
                        "offer_id": offer.pk,
                        "sku": offer.product.sku,
                        "current_code": offer.supplier_code,
                        "proposed_code": normalized,
                        "rows": rows,
                    }
                )
                continue
            safe.append(
                {
                    "offer": offer,
                    "supplier_code": normalized,
                    "rows": rows,
                }
            )

        applied = 0
        if options["apply"] and safe:
            with transaction.atomic():
                locked = {
                    offer.pk: offer
                    for offer in ProductSupplier.objects.select_for_update().filter(
                        pk__in=[item["offer"].pk for item in safe]
                    )
                }
                updates = []
                for item in safe:
                    offer = locked.get(item["offer"].pk)
                    if not offer:
                        continue
                    if (
                        offer.normalized_supplier_code
                        and offer.normalized_supplier_code != item["supplier_code"]
                        and not options["replace_existing"]
                    ):
                        continue
                    offer.supplier_code = item["supplier_code"]
                    offer.normalized_supplier_code = item["supplier_code"]
                    offer.updated_at = timezone.now()
                    updates.append(offer)
                ProductSupplier.objects.bulk_update(
                    updates,
                    ["supplier_code", "normalized_supplier_code", "updated_at"],
                    batch_size=1000,
                )
                applied = len(updates)

        report = {
            "file": str(base_path),
            "sheet": sheet_name,
            "mode": "apply" if options["apply"] else "preview",
            "safe_matches": len(safe),
            "applied": applied,
            "unchanged": len(set(unchanged)),
            "ambiguous_groups": len(ambiguous),
            "ambiguous_offers": len(ambiguous_offer_ids),
            "offers_with_multiple_references": len(conflicting_offer_references),
            "existing_code_conflicts": len(existing_conflicts),
            "missing_products": len(missing_sku),
            "products_without_preferred_offer": len(without_offer),
            "missing_references": len(missing_reference),
            "ambiguous": ambiguous[:200],
            "multiple_reference_examples": conflicting_offer_references[:100],
            "existing_conflict_examples": existing_conflicts[:100],
            "missing_product_examples": missing_sku[:100],
            "without_offer_examples": without_offer[:100],
            "missing_reference_examples": missing_reference[:100],
        }
        output = json.dumps(report, ensure_ascii=False, indent=2)
        self.stdout.write(output)
        if options.get("report"):
            report_path = Path(options["report"]).expanduser().resolve()
            report_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text(output + "\n", encoding="utf-8")
            self.stdout.write(self.style.SUCCESS(f"Informe guardado en {report_path}"))
        if options["apply"]:
            self.stdout.write(self.style.SUCCESS(f"Referencias aplicadas: {applied}"))
        else:
            self.stdout.write(self.style.WARNING("Vista previa: no se modificaron datos."))
