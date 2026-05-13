from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.core.cache import cache
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from accounts.models import ClientProfile
from catalog.services.abrazadera_importer import AbrazaderaImporter
from catalog.services.clamp_code import generarCodigo, parsearCodigo
from catalog.services.product_importer import ProductImporter
from catalog.models import Category, CategoryProductOrder, ClampMeasureRequest, ClampSpecs, Product
from core.models import CatalogExcelTemplate, CatalogExcelTemplateColumn, CatalogExcelTemplateSheet
from core.services.catalog_excel_exporter import build_catalog_workbook
from orders.models import CartItem


def build_import_workbook(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Importacion"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class ClampCodeTests(SimpleTestCase):
    def test_parse_real_examples(self):
        parsed_abl = parsearCodigo("ABL1135400C")
        self.assertEqual(parsed_abl["prefijo"], "ABL")
        self.assertEqual(parsed_abl["tipo"], "LAMINADA")
        self.assertEqual(parsed_abl["diametro_compactado"], "1")
        self.assertEqual(parsed_abl["diametro"], "1")
        self.assertEqual(parsed_abl["ancho"], 135)
        self.assertEqual(parsed_abl["largo"], 400)
        self.assertEqual(parsed_abl["forma"], "CURVA")

        parsed_abt = parsearCodigo("ABT91685270P")
        self.assertEqual(parsed_abt["prefijo"], "ABT")
        self.assertEqual(parsed_abt["tipo"], "TREFILADA")
        self.assertEqual(parsed_abt["diametro_compactado"], "916")
        self.assertEqual(parsed_abt["diametro"], "9/16")
        self.assertEqual(parsed_abt["ancho"], 85)
        self.assertEqual(parsed_abt["largo"], 270)
        self.assertEqual(parsed_abt["forma"], "PLANA")

        parsed_semicurva = parsearCodigo("ABT3480220S")
        self.assertEqual(parsed_semicurva["diametro_compactado"], "34")
        self.assertEqual(parsed_semicurva["diametro"], "3/4")
        self.assertEqual(parsed_semicurva["ancho"], 80)
        self.assertEqual(parsed_semicurva["largo"], 220)
        self.assertEqual(parsed_semicurva["forma"], "SEMICURVA")

    def test_parse_variable_length_diameters(self):
        parsed = parsearCodigo("ABT111680220S")
        self.assertEqual(parsed["diametro_compactado"], "1116")
        self.assertEqual(parsed["diametro"], "11/16")
        self.assertEqual(parsed["ancho"], 80)
        self.assertEqual(parsed["largo"], 220)
        self.assertFalse(parsed["diametro_requiere_mapeo"])

    def test_generate_known_codes(self):
        code_1 = generarCodigo(
            tipo="ABT",
            diametro="9/16",
            ancho=85,
            largo=270,
            forma="P",
        )
        self.assertEqual(code_1, "ABT91685270P")

        code_2 = generarCodigo(
            tipo="TREFILADA",
            diametro="3/4",
            ancho=80,
            largo=220,
            forma="SEMICURVA",
        )
        self.assertEqual(code_2, "ABT3480220S")

        code_3 = generarCodigo(
            tipo="LAMINADA",
            diametro="1",
            ancho=135,
            largo=400,
            forma="CURVA",
        )
        self.assertEqual(code_3, "ABL1135400C")

    def test_generate_unknown_fraction_marks_mapping_pending(self):
        metadata = generarCodigo(
            tipo="ABT",
            diametro="1 1/8",
            ancho=90,
            largo=250,
            forma="P",
            with_metadata=True,
        )
        self.assertEqual(metadata["codigo"], "ABT11890250P")
        self.assertTrue(metadata["diametro_requiere_mapeo"])
        self.assertTrue(metadata["warnings"])

    def test_generate_unknown_fraction_in_strict_mode_raises(self):
        with self.assertRaises(ValueError):
            generarCodigo(
                tipo="ABT",
                diametro="1 1/8",
                ancho=90,
                largo=250,
                forma="P",
                strict_diameter_mapping=True,
            )


class ProductImportTests(TestCase):
    def test_product_import_accepts_header_file_without_supplier(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Stock", "Categoria", "Atributos"],
            [["IMP-001", "Producto importado", 100.5, 4, "Prueba", "Color:Rojo;Material:Acero"]],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="IMP-001")
        self.assertEqual(product.name, "Producto importado")
        self.assertEqual(product.price, Decimal("100.50"))
        self.assertEqual(product.stock, 4)
        self.assertTrue(product.categories.filter(name="Prueba").exists())
        self.assertEqual(product.attributes, {"Color": "Rojo", "Material": "Acero"})

    def test_product_import_does_not_create_categories_by_default(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Categoria"],
            [["IMP-NO-CAT", "Producto sin categoria creada", 100, "Categoria accidental"]],
        )

        result = ProductImporter(file_obj).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="IMP-NO-CAT")
        self.assertFalse(Category.objects.filter(name="Categoria accidental").exists())
        self.assertFalse(product.categories.exists())
        self.assertEqual(
            result.row_results[0].data["categorias_no_encontradas"],
            ["Categoria accidental"],
        )

    def test_product_import_preserves_existing_categories_by_default(self):
        manual_category = Category.objects.create(name="Categoria Manual", slug="categoria-manual")
        excel_category = Category.objects.create(name="Categoria Excel", slug="categoria-excel")
        product = Product.objects.create(
            sku="IMP-PRESERVE-CAT",
            name="Producto ordenado manualmente",
            price=Decimal("100.00"),
            stock=3,
            category=manual_category,
        )
        product.categories.add(manual_category)
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Categoria"],
            [["IMP-PRESERVE-CAT", "Producto actualizado", 150, "Categoria Excel"]],
        )

        result = ProductImporter(file_obj).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product.refresh_from_db()
        self.assertEqual(product.name, "Producto actualizado")
        self.assertEqual(product.price, Decimal("150.00"))
        self.assertEqual(product.category_id, manual_category.pk)
        self.assertTrue(product.categories.filter(pk=manual_category.pk).exists())
        self.assertFalse(product.categories.filter(pk=excel_category.pk).exists())
        self.assertTrue(result.row_results[0].data["categorias_preservadas"])

    def test_product_import_links_existing_category_to_new_products_by_default(self):
        category = Category.objects.create(name="Categoria Existente", slug="categoria-existente")
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Categoria"],
            [["IMP-NEW-CAT", "Producto nuevo con categoria", 100, "Categoria Existente"]],
        )

        result = ProductImporter(file_obj).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="IMP-NEW-CAT")
        self.assertEqual(product.category_id, category.pk)
        self.assertTrue(product.categories.filter(pk=category.pk).exists())

    def test_product_import_requires_explicit_permission_to_create_categories(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Categoria"],
            [["IMP-CREATE-BLOCKED", "Producto sin alta de categoria", 100, "Categoria Nueva"]],
        )

        result = ProductImporter(file_obj, category_mode="create").run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="IMP-CREATE-BLOCKED")
        self.assertFalse(Category.objects.filter(name="Categoria Nueva").exists())
        self.assertFalse(product.categories.exists())
        self.assertEqual(result.row_results[0].data["modo_categorias"], "existing")

    def test_product_import_parses_argentine_money_and_preserves_existing_blanks(self):
        supplier_category = Category.objects.create(name="Existentes")
        product = Product.objects.create(
            sku="IMP-002",
            name="Producto viejo",
            supplier="Proveedor Actual",
            price=Decimal("10.00"),
            stock=7,
            category=supplier_category,
        )
        product.categories.add(supplier_category)
        file_obj = build_import_workbook(
            ["Codigo", "Articulo", "Venta", "Costo", "Stock", "Proveedor"],
            [["IMP-002", "Producto nuevo", "$ 12.500,50", "ARS 1.250", "", ""]],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product.refresh_from_db()
        self.assertEqual(product.name, "Producto nuevo")
        self.assertEqual(product.price, Decimal("12500.50"))
        self.assertEqual(product.cost, Decimal("1250.00"))
        self.assertEqual(product.stock, 7)
        self.assertEqual(product.supplier, "Proveedor Actual")

    def test_product_import_updates_filters_for_existing_products(self):
        product = Product.objects.create(
            sku="IMP-FILTERS",
            name="Producto con filtros",
            price=Decimal("10.00"),
            stock=1,
            filter_1="Viejo",
        )
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Filtro 1", "Filtro 2", "Filtro 5"],
            [["IMP-FILTERS", "Producto con filtros", "", "Camion", "Suspension", "Pesado"]],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product.refresh_from_db()
        self.assertEqual(product.filter_1, "Camion")
        self.assertEqual(product.filter_2, "Suspension")
        self.assertEqual(product.filter_5, "Pesado")

    def test_product_import_builds_attributes_from_technical_columns(self):
        file_obj = build_import_workbook(
            [
                "Codigo Flexs",
                "Nombre producto",
                "Precio lista",
                "Stock",
                "Familia",
                "Subfamilia",
                "Estado",
                "Diametro",
                "Ancho",
                "Largo",
                "Forma",
                "Atributos",
            ],
            [
                [
                    "IMP-TECH",
                    "ABRAZADERA TREFILADA DE 1/2 X 80 X 220 CURVA",
                    "1.250,75",
                    12,
                    "Abrazaderas",
                    "Trefiladas",
                    "X",
                    "1/2",
                    80,
                    220,
                    "Curva",
                    "Color:Negro",
                ]
            ],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="IMP-TECH")
        self.assertTrue(product.is_active)
        self.assertEqual(product.price, Decimal("1250.75"))
        self.assertTrue(product.categories.filter(name="Trefiladas").exists())
        self.assertEqual(product.attributes["Color"], "Negro")
        self.assertEqual(product.attributes["Diametro"], "1/2")
        self.assertEqual(product.attributes["Ancho"], "80")
        self.assertEqual(product.attributes["Largo"], "220")
        self.assertEqual(product.attributes["Forma"], "Curva")

    def test_product_import_adapts_saas_export_columns(self):
        file_obj = build_import_workbook(
            [
                "Nº de producto",
                "Estado",
                "Rubro",
                "Nombre",
                "Código",
                "Código universal de producto (UPC)",
                "Código de proveedor",
                "Unidad",
                "Alicuota de IVA",
                "Proveedor",
                "Costo ($)",
                "Utilidad (%)",
                "Precio ($)",
                "Precio Final ($)",
                "Mostrar en tienda",
            ],
            [
                [
                    918,
                    "Habilitado",
                    "BUJE ARMADO",
                    "BUJE DEMO SAAS",
                    "SAAS-001",
                    "7791234567890",
                    "PROV-55",
                    "unidad",
                    "21%",
                    "MOVIGOM S.R.L.",
                    "1.000,00",
                    "1",
                    "10.000,00",
                    "12.100,00",
                    "no",
                ]
            ],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="SAAS-001")
        self.assertEqual(product.price, Decimal("12100.00"))
        self.assertEqual(product.cost, Decimal("1000.00"))
        self.assertTrue(product.is_active)
        self.assertEqual(product.attributes["Numero SaaS"], "918")
        self.assertEqual(product.attributes["Codigo proveedor"], "PROV-55")
        self.assertEqual(product.attributes["UPC"], "7791234567890")
        self.assertEqual(product.attributes["IVA"], "21%")
        self.assertEqual(product.attributes["Precio neto SaaS"], "10.000,00")
        self.assertEqual(product.attributes["Precio final SaaS"], "12.100,00")
        self.assertEqual(product.attributes["Origen importacion"], "SaaS Argentina")

    def test_product_import_places_saas_clamps_under_parent_category(self):
        file_obj = build_import_workbook(
            ["Rubro", "Nombre", "Código", "Precio Final ($)", "Proveedor"],
            [
                [
                    "ABRAZADERA DE 5/8",
                    "ABRAZADERA TREFILADA DE 5/8 X 80 X 220 CURVA",
                    "SAAS-ABR-001",
                    2500,
                    "ROCES",
                ]
            ],
        )

        result = ProductImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="SAAS-ABR-001")
        self.assertTrue(product.categories.filter(name__iexact="ABRAZADERAS").exists())
        subcategory = product.categories.get(name="ABRAZADERA DE 5/8")
        self.assertEqual(subcategory.parent.name, "ABRAZADERAS")

    def test_product_import_rejects_invalid_active_value(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Activo"],
            [["IMP-ACTIVE", "Producto", 10, "tal vez"]],
        )

        result = ProductImporter(file_obj).run(dry_run=True)

        self.assertEqual(result.errors, 1)
        self.assertIn("Activo invalido", result.row_results[0].errors[0])

    def test_product_import_reports_duplicate_sku_inside_file(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio"],
            [
                ["IMP-DUP", "Uno", 100],
                ["IMP-DUP", "Dos", 200],
            ],
        )

        result = ProductImporter(file_obj).run(dry_run=True)

        self.assertEqual(result.errors, 0)
        self.assertTrue(result.row_results[1].success)
        self.assertEqual(result.row_results[1].action, "skipped")
        self.assertIn("SKU duplicado", result.row_results[1].errors[0])
        self.assertEqual(result.row_results[1].data["_duplicate_sku"], "IMP-DUP")
        self.assertEqual(result.row_results[1].data["_duplicate_first_row"], 2)
        self.assertEqual(result.row_results[1].data["_duplicate_first_data"]["nombre"], "Uno")

    def test_product_import_records_duplicate_warning_on_product(self):
        file_obj = build_import_workbook(
            ["SKU", "Nombre", "Precio", "Proveedor", "Rubro"],
            [
                ["IMP-DUP-REAL", "Producto principal", 100, "Proveedor A", "Rubro A"],
                ["IMP-DUP-REAL", "Producto alternativo", 200, "Proveedor B", "Rubro B"],
            ],
        )

        result = ProductImporter(file_obj).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        self.assertEqual(result.created, 1)
        product = Product.objects.get(sku="IMP-DUP-REAL")
        self.assertEqual(product.name, "Producto principal")
        self.assertEqual(product.attributes["Duplicado en importacion"], "Si")
        self.assertEqual(product.attributes["Duplicados importacion"], "1")
        self.assertIn("Fila 3 contra fila 2", product.attributes["Detalle duplicados importacion"])
        self.assertIn("Producto alternativo", product.attributes["Detalle duplicados importacion"])

    def test_abrazadera_import_accepts_product_style_headers(self):
        file_obj = build_import_workbook(
            ["sku", "nombre", "precio", "stock", "categoria"],
            [["ABR-IMP", "ABRAZADERA TREFILADA DE 1/2 X 85 X 260 CURVA", "1.250,75", 12, "Abrazaderas"]],
        )

        result = AbrazaderaImporter(
            file_obj,
            category_mode="create",
            allow_category_creation=True,
        ).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        product = Product.objects.get(sku="ABR-IMP")
        self.assertEqual(product.price, Decimal("1250.75"))
        self.assertEqual(product.stock, 12)
        self.assertTrue(product.categories.filter(name="Abrazaderas").exists())
        self.assertTrue(hasattr(product, "clamp_specs"))

    def test_abrazadera_import_skips_non_clamp_rows(self):
        file_obj = build_import_workbook(
            ["Codigo", "Nombre", "Rubro", "Precio Final ($)"],
            [["BUJE-IMP", "BUJE DE GOMA DEMO", "BUJE DE GOMA", "1000"]],
        )

        result = AbrazaderaImporter(file_obj).run(dry_run=False)

        self.assertEqual(result.errors, 0)
        self.assertEqual(result.created, 0)
        self.assertFalse(Product.objects.filter(sku="BUJE-IMP").exists())


class ClampMeasureRequestFlowTests(TestCase):
    def setUp(self):
        self.client_user = User.objects.create_user(username="cliente_medidas", password="secret123")
        ClientProfile.objects.create(
            user=self.client_user,
            company_name="Cliente Medidas",
            is_approved=True,
        )
        self.client.force_login(self.client_user)

        self.category = Category.objects.create(name="Abrazaderas", is_active=True)
        self.product = Product.objects.create(
            sku="ABT3480220S",
            name="ABRAZADERA TREFILADA DE 3/4 X 80 X 220 SEMICURVA",
            price=Decimal("12000.00"),
            stock=8,
            is_active=True,
            category=self.category,
        )
        self.product.categories.add(self.category)
        ClampSpecs.objects.create(
            product=self.product,
            fabrication="TREFILADA",
            diameter="3/4",
            width=80,
            length=220,
            shape="SEMICURVA",
        )

    def test_calculate_detects_existing_measure(self):
        response = self.client.post(
            reverse("catalog_clamp_request"),
            data={
                "action": "check_exists",
                "client_name": "Cliente Test",
                "clamp_type": "trefilada",
                "diameter": "3/4",
                "width_mm": "80",
                "length_mm": "220",
                "profile_type": "SEMICURVA",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["has_matches"])
        self.assertEqual(response.context["matching_products"][0].sku, "ABT3480220S")

    def test_submit_creates_request_when_measure_not_found(self):
        response = self.client.post(
            reverse("catalog_clamp_request"),
            data={
                "action": "submit_request",
                "price_list_key": "lista_2",
                "client_name": "Cliente Nuevo",
                "client_email": "cliente@demo.com",
                "quantity": "3",
                "clamp_type": "trefilada",
                "diameter": "3/4",
                "width_mm": "81",
                "length_mm": "220",
                "profile_type": "SEMICURVA",
            },
        )

        self.assertEqual(response.status_code, 302)
        request_obj = ClampMeasureRequest.objects.get()
        self.assertEqual(request_obj.client_name, "Cliente Nuevo")
        self.assertEqual(request_obj.selected_price_list, "lista_1")
        self.assertFalse(request_obj.exists_in_catalog)

    def test_non_client_non_admin_cannot_access_clamp_measure_page(self):
        user = User.objects.create_user(username="usuario_sin_perfil", password="secret123")
        self.client.force_login(user)

        response = self.client.get(reverse("catalog_clamp_request"), follow=True)

        self.assertRedirects(response, reverse("catalog"))
        self.assertContains(response, "solo para clientes aprobados o administradores")

    def test_authenticated_client_sees_confirmed_price_in_history(self):
        user = User.objects.create_user(username="cliente_medida", password="secret123")
        ClientProfile.objects.create(
            user=user,
            company_name="Cliente Medida",
            is_approved=True,
        )
        ClampMeasureRequest.objects.create(
            client_user=user,
            client_name="Cliente Medida",
            clamp_type="trefilada",
            is_zincated=False,
            diameter="3/4",
            width_mm=81,
            length_mm=220,
            profile_type="SEMICURVA",
            quantity=1,
            description="ABRAZADERA TREFILADA DE 3/4 X 81 X 220 SEMICURVA",
            generated_code="ABT3481220S",
            dollar_rate=Decimal("1300.00"),
            steel_price_usd=Decimal("1450.00"),
            supplier_discount_pct=Decimal("0.00"),
            general_increase_pct=Decimal("40.00"),
            base_cost=Decimal("1000.00"),
            selected_price_list="lista_1",
            estimated_final_price=Decimal("1400.00"),
            confirmed_price_list="lista_3",
            confirmed_price=Decimal("1600.00"),
            status=ClampMeasureRequest.STATUS_QUOTED,
            client_response_note="Precio confirmado por ventas.",
        )

        self.client.force_login(user)
        response = self.client.get(reverse("catalog_clamp_request"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Precio confirmado por ventas.")
        self.assertEqual(response.context["client_requests"][0].confirmed_price, Decimal("1600.00"))

    def test_client_can_add_completed_request_to_cart(self):
        user = User.objects.create_user(username="cliente_add_carrito", password="secret123")
        ClientProfile.objects.create(
            user=user,
            company_name="Cliente Carrito",
            is_approved=True,
        )
        clamp_request = ClampMeasureRequest.objects.create(
            client_user=user,
            client_name="Cliente Carrito",
            clamp_type="trefilada",
            is_zincated=False,
            diameter="3/4",
            width_mm=81,
            length_mm=220,
            profile_type="SEMICURVA",
            quantity=2,
            description="ABRAZADERA TREFILADA DE 3/4 X 81 X 220 SEMICURVA",
            generated_code="ABT3481220S",
            dollar_rate=Decimal("1300.00"),
            steel_price_usd=Decimal("1450.00"),
            supplier_discount_pct=Decimal("0.00"),
            general_increase_pct=Decimal("40.00"),
            base_cost=Decimal("1000.00"),
            selected_price_list="lista_1",
            estimated_final_price=Decimal("1400.00"),
            confirmed_price_list="lista_2",
            confirmed_price=Decimal("1500.00"),
            status=ClampMeasureRequest.STATUS_COMPLETED,
        )

        self.client.force_login(user)
        response = self.client.post(
            reverse("catalog_clamp_request_add_to_cart", args=[clamp_request.pk]),
            data={"quantity": "3"},
        )

        self.assertEqual(response.status_code, 302)
        cart_item = CartItem.objects.select_related("product").get(cart__user=user)
        self.assertEqual(cart_item.quantity, 3)
        self.assertEqual(cart_item.clamp_request_id, clamp_request.pk)
        self.assertFalse(cart_item.product.is_active)

        clamp_request.refresh_from_db()
        self.assertIsNotNone(clamp_request.linked_product_id)
        self.assertEqual(clamp_request.linked_product_id, cart_item.product_id)
        self.assertIsNotNone(clamp_request.added_to_cart_at)

    def test_client_cannot_add_non_completed_request_to_cart(self):
        user = User.objects.create_user(username="cliente_no_completada", password="secret123")
        ClientProfile.objects.create(
            user=user,
            company_name="Cliente No Completada",
            is_approved=True,
        )
        clamp_request = ClampMeasureRequest.objects.create(
            client_user=user,
            client_name="Cliente No Completada",
            clamp_type="trefilada",
            is_zincated=False,
            diameter="3/4",
            width_mm=81,
            length_mm=220,
            profile_type="SEMICURVA",
            quantity=1,
            description="ABRAZADERA TREFILADA DE 3/4 X 81 X 220 SEMICURVA",
            generated_code="ABT3481220S",
            dollar_rate=Decimal("1300.00"),
            steel_price_usd=Decimal("1450.00"),
            supplier_discount_pct=Decimal("0.00"),
            general_increase_pct=Decimal("40.00"),
            base_cost=Decimal("1000.00"),
            selected_price_list="lista_1",
            estimated_final_price=Decimal("1400.00"),
            status=ClampMeasureRequest.STATUS_QUOTED,
        )

        self.client.force_login(user)
        response = self.client.post(
            reverse("catalog_clamp_request_add_to_cart", args=[clamp_request.pk]),
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(CartItem.objects.filter(cart__user=user).exists())


class CatalogAdvancedSearchTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name="Abrazaderas", slug="abrazaderas", is_active=True)

        self.product_exact = Product.objects.create(
            sku="ABT3480220S",
            name="ABRAZADERA TREFILADA DE 3/4 X 80 X 220 SEMICURVA",
            description="Uso suspension trasera",
            supplier="MOVIGOM",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=self.category,
        )
        self.product_exact.categories.add(self.category)
        ClampSpecs.objects.create(
            product=self.product_exact,
            fabrication="TREFILADA",
            diameter="3/4",
            width=80,
            length=220,
            shape="SEMICURVA",
        )

        self.product_other = Product.objects.create(
            sku="ABT3485220P",
            name="ABRAZADERA TREFILADA DE 3/4 X 85 X 220 PLANA",
            description="Aplicacion delantera",
            supplier="ROCES",
            price=Decimal("110.00"),
            stock=4,
            is_active=True,
            category=self.category,
        )
        self.product_other.categories.add(self.category)
        ClampSpecs.objects.create(
            product=self.product_other,
            fabrication="TREFILADA",
            diameter="3/4",
            width=85,
            length=220,
            shape="PLANA",
        )

    def test_search_parses_dimensions_and_type(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": "tipo:trefilada 3/4x80x220 semicurva"},
        )

        self.assertEqual(response.status_code, 200)
        page_items = list(response.context["page_obj"].object_list)
        self.assertEqual(len(page_items), 1)
        self.assertEqual(page_items[0].sku, "ABT3480220S")

    def test_search_supports_exclusion_term(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": "abrazadera -delantera"},
        )

        self.assertEqual(response.status_code, 200)
        skus = [product.sku for product in response.context["page_obj"].object_list]
        self.assertIn("ABT3480220S", skus)
        self.assertNotIn("ABT3485220P", skus)

    def test_search_relevance_prioritizes_exact_sku(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": "ABT3480220S", "order": "relevance"},
        )

        self.assertEqual(response.status_code, 200)
        first_product = response.context["page_obj"].object_list[0]
        self.assertEqual(first_product.sku, "ABT3480220S")

    def test_search_accepts_compact_diameter_with_unicode_separator(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": "tipo:t 34×80×220"},
        )

        self.assertEqual(response.status_code, 200)
        page_items = list(response.context["page_obj"].object_list)
        self.assertEqual(len(page_items), 1)
        self.assertEqual(page_items[0].sku, "ABT3480220S")

    def test_search_ignores_trailing_token_punctuation(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": "abrazadera, semicurva;"},
        )

        self.assertEqual(response.status_code, 200)
        skus = [product.sku for product in response.context["page_obj"].object_list]
        self.assertIn("ABT3480220S", skus)

    def test_search_normalizes_query_item_label(self):
        response = self.client.get(
            reverse("catalog"),
            {"q": 'Buscar "ABT3480220S"'},
        )

        self.assertEqual(response.status_code, 200)
        skus = [product.sku for product in response.context["page_obj"].object_list]
        self.assertIn("ABT3480220S", skus)


class CatalogCategoryVisibilityTests(TestCase):
    def setUp(self):
        cache.clear()

    def tearDown(self):
        cache.clear()

    def test_hidden_category_hides_product_from_public_catalog(self):
        hidden_category = Category.objects.create(
            name="Uso interno",
            slug="uso-interno",
            is_active=True,
            visible_in_catalog=False,
        )
        product = Product.objects.create(
            sku="INT-001",
            name="Producto interno",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=hidden_category,
        )
        product.categories.add(hidden_category)

        response = self.client.get(reverse("catalog"))

        self.assertEqual(response.status_code, 200)
        skus = [item.sku for item in response.context["page_obj"].object_list]
        self.assertNotIn("INT-001", skus)

    def test_public_tree_hides_empty_categories_and_uses_public_name(self):
        public_category = Category.objects.create(
            name="ELASTICOS Y SUSPENSION",
            public_name="Suspension",
            public_description="Linea comercial visible.",
            slug="elasticos-suspension",
            is_active=True,
            visible_in_catalog=True,
        )
        empty_category = Category.objects.create(
            name="VACIA INTERNA",
            public_name="Vacia cliente",
            slug="vacia-interna",
            is_active=True,
            visible_in_catalog=True,
        )
        product = Product.objects.create(
            sku="PUB-001",
            name="Producto visible",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=public_category,
        )
        product.categories.add(public_category)

        response = self.client.get(reverse("catalog"))

        self.assertEqual(response.status_code, 200)
        row_categories = [row["category"] for row in response.context["category_tree_rows"]]
        row_ids = {category.id for category in row_categories}
        self.assertIn(public_category.id, row_ids)
        self.assertNotIn(empty_category.id, row_ids)
        self.assertIn("Suspension", [category.display_name for category in row_categories])

    def test_hiding_parent_category_hides_public_descendants(self):
        parent = Category.objects.create(
            name="Padre publico",
            slug="padre-publico",
            is_active=True,
            visible_in_catalog=True,
        )
        child = Category.objects.create(
            name="Hijo publico",
            slug="hijo-publico",
            parent=parent,
            is_active=True,
            visible_in_catalog=True,
        )
        product = Product.objects.create(
            sku="CHILD-001",
            name="Producto hijo",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=child,
        )
        product.categories.add(child)

        parent.visible_in_catalog = False
        parent.save()
        child.refresh_from_db()

        response = self.client.get(reverse("catalog"))

        self.assertFalse(child.visible_in_catalog)
        skus = [item.sku for item in response.context["page_obj"].object_list]
        self.assertNotIn("CHILD-001", skus)

    def test_category_page_uses_manual_product_order_by_default(self):
        category = Category.objects.create(
            name="Orden publico",
            slug="orden-publico",
            is_active=True,
            visible_in_catalog=True,
        )
        first_product = Product.objects.create(
            sku="ORD-001",
            name="Producto primero alfabetico",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=category,
        )
        second_product = Product.objects.create(
            sku="ORD-002",
            name="ZZ Producto manual primero",
            price=Decimal("100.00"),
            stock=5,
            is_active=True,
            category=category,
        )
        first_product.categories.add(category)
        second_product.categories.add(category)
        CategoryProductOrder.objects.create(category=category, product=first_product, sort_order=20)
        CategoryProductOrder.objects.create(category=category, product=second_product, sort_order=10)

        response = self.client.get(reverse("catalog"), {"category": category.slug})

        self.assertEqual(response.status_code, 200)
        skus = [item.sku for item in response.context["page_obj"].object_list]
        self.assertEqual(skus[:2], ["ORD-002", "ORD-001"])
        self.assertEqual(response.context["order_by"], "manual")


class ProductDetailTemplateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="cliente_template", password="secret123")
        self.category = Category.objects.create(name="Plantillas", is_active=True)
        self.product = Product.objects.create(
            id=24950,
            sku="TPL-24950",
            name="Producto Template",
            price=Decimal("1000.00"),
            stock=5,
            is_active=True,
            category=self.category,
        )
        self.product.categories.add(self.category)

    def test_product_id_is_unlocalized_in_inline_js(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("product_detail", args=[self.product.sku]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "addToCart(24950)")
        self.assertContains(response, "toggleFavorite(24950, this)")


class CatalogClientExcelDownloadTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name="Categoria XLSX", slug="categoria-xlsx", is_active=True)
        self.product = Product.objects.create(
            sku="XLSX-001",
            name="Producto XLSX",
            supplier="Proveedor XLSX",
            price=Decimal("1234.56"),
            cost=Decimal("900.00"),
            stock=10,
            is_active=True,
            category=self.category,
        )
        self.product.categories.add(self.category)

        self.approved_user = User.objects.create_user(username="cliente_xlsx_ok", password="secret123")
        ClientProfile.objects.create(
            user=self.approved_user,
            company_name="Cliente XLSX",
            is_approved=True,
        )
        self.unapproved_user = User.objects.create_user(username="cliente_xlsx_no", password="secret123")
        ClientProfile.objects.create(
            user=self.unapproved_user,
            company_name="Cliente XLSX No",
            is_approved=False,
        )

        self.template = CatalogExcelTemplate.objects.create(
            name="Plantilla Cliente XLSX",
            is_active=True,
            is_client_download_enabled=True,
            client_download_label="Descargar version clientes",
        )
        self.sheet = CatalogExcelTemplateSheet.objects.create(
            template=self.template,
            name="Catalogo",
            include_header=True,
            only_active_products=True,
            sort_by="name_asc",
        )
        self.sheet.categories.add(self.category)
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key="name", order=2, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=self.sheet, key="price", order=3, is_active=True)

    def test_approved_client_can_download_published_catalog_excel(self):
        self.client.force_login(self.approved_user)
        response = self.client.get(reverse("catalog_client_excel_download"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])
        self.assertIn("no-store", response["Cache-Control"])
        self.assertIn("no-cache", response["Cache-Control"])
        self.assertIn("max-age=0", response["Cache-Control"])
        self.assertEqual(response["Pragma"], "no-cache")
        self.assertEqual(response["Expires"], "0")
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["Categoria XLSX"]
        self.assertEqual(worksheet["A1"].value, "SKU")
        self.assertEqual(worksheet["B2"].value, "Producto XLSX")

    def test_unapproved_client_cannot_download_catalog_excel(self):
        self.client.force_login(self.unapproved_user)
        response = self.client.get(reverse("catalog_client_excel_download"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "solo para clientes aprobados")

    def test_download_warns_when_no_template_published(self):
        self.template.is_client_download_enabled = False
        self.template.save(update_fields=["is_client_download_enabled", "updated_at"])

        self.client.force_login(self.approved_user)
        response = self.client.get(reverse("catalog_client_excel_download"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No hay una plantilla de Excel publicada para clientes")


class CatalogExcelGroupedExportTests(TestCase):
    def test_sheet_can_group_products_by_subcategory_blocks(self):
        root = Category.objects.create(name="Elasticos", order=1, is_active=True)
        child = Category.objects.create(name="Bujes", parent=root, order=2, is_active=True)
        direct_product = Product.objects.create(
            sku="ROOT-001",
            name="Producto categoria raiz",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=root,
        )
        direct_product.categories.add(root)
        child_product = Product.objects.create(
            sku="CHILD-001",
            name="Producto subcategoria",
            price=Decimal("200.00"),
            stock=2,
            is_active=True,
            category=child,
        )
        child_product.categories.add(child)

        template = CatalogExcelTemplate.objects.create(name="Catalogo agrupado")
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Elasticos",
            include_header=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet.categories.add(root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="price", order=3, is_active=True)

        workbook, stats = build_catalog_workbook(template)
        worksheet = workbook["Elasticos"]

        self.assertEqual(stats["rows_by_sheet"]["Elasticos"], 2)
        self.assertEqual(worksheet["A1"].value, "Sin subcategoria (1 productos)")
        self.assertEqual(worksheet["A2"].value, "SKU")
        self.assertEqual(worksheet["A3"].value, "ROOT-001")
        self.assertEqual(worksheet["A4"].value, "Bujes (1 productos)")
        self.assertEqual(worksheet["A5"].value, "SKU")
        self.assertEqual(worksheet["A6"].value, "CHILD-001")
        self.assertIsNone(worksheet.freeze_panes)

    def test_client_catalog_export_skips_inactive_category_blocks(self):
        root = Category.objects.create(name="Catalogo", order=1, is_active=True, visible_in_catalog=True)
        active_child = Category.objects.create(
            name="Visible",
            parent=root,
            order=1,
            is_active=True,
            visible_in_catalog=True,
        )
        inactive_child = Category.objects.create(
            name="Inactiva",
            parent=root,
            order=2,
            is_active=False,
            visible_in_catalog=False,
        )
        visible_product = Product.objects.create(
            sku="VISIBLE-001",
            name="Producto visible",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=active_child,
        )
        visible_product.categories.add(active_child)
        hidden_product = Product.objects.create(
            sku="HIDDEN-001",
            name="Producto oculto por categoria",
            price=Decimal("200.00"),
            stock=1,
            is_active=True,
            category=inactive_child,
        )
        hidden_product.categories.add(inactive_child)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo publico agrupado",
            is_client_download_enabled=True,
        )
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Catalogo",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=False,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet.categories.add(root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)
        worksheet = workbook["Catalogo"]
        values = [cell.value for row in worksheet.iter_rows() for cell in row if cell.value]

        self.assertEqual(stats["rows_by_sheet"]["Catalogo"], 1)
        self.assertIn("Visible (1 productos)", values)
        self.assertIn("VISIBLE-001", values)
        self.assertNotIn("Inactiva (1 productos)", values)
        self.assertNotIn("HIDDEN-001", values)

    def test_client_catalog_export_with_inactive_selected_root_does_not_export_all_products(self):
        inactive_root = Category.objects.create(
            name="Raiz inactiva",
            is_active=False,
            visible_in_catalog=False,
        )
        unrelated_root = Category.objects.create(
            name="Raiz visible",
            is_active=True,
            visible_in_catalog=True,
        )
        unrelated_product = Product.objects.create(
            sku="UNRELATED-001",
            name="Producto de otra categoria",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=unrelated_root,
        )
        unrelated_product.categories.add(unrelated_root)

        template = CatalogExcelTemplate.objects.create(name="Catalogo raiz inactiva")
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Raiz inactiva",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet.categories.add(inactive_root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)
        worksheet = workbook["Catalogo"]
        values = [cell.value for row in worksheet.iter_rows() for cell in row if cell.value]

        self.assertEqual(stats["rows_by_sheet"]["Catalogo"], 0)
        self.assertEqual(stats["skipped_sheets"], ["Raiz inactiva"])
        self.assertNotIn("Raiz inactiva", workbook.sheetnames)
        self.assertIn("No hay categorias visibles para exportar.", values)
        self.assertNotIn("UNRELATED-001", values)

    def test_public_export_skips_inactive_category_sheet_but_keeps_visible_sheets(self):
        visible_root = Category.objects.create(
            name="Visible",
            is_active=True,
            visible_in_catalog=True,
        )
        hidden_root = Category.objects.create(
            name="Oculta",
            is_active=False,
            visible_in_catalog=False,
        )
        product = Product.objects.create(
            sku="VISIBLE-SHEET-001",
            name="Producto visible",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=visible_root,
        )
        product.categories.add(visible_root)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo con hojas heredadas",
            is_client_download_enabled=True,
        )
        visible_sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Visible",
            include_header=True,
            only_active_products=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        visible_sheet.categories.add(visible_root)
        hidden_sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Oculta",
            include_header=True,
            only_active_products=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        hidden_sheet.categories.add(hidden_root)
        for sheet in (visible_sheet, hidden_sheet):
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)

        self.assertIn("Visible", workbook.sheetnames)
        self.assertNotIn("Oculta", workbook.sheetnames)
        self.assertEqual(stats["skipped_sheets"], ["Oculta"])
        self.assertEqual(stats["rows_by_sheet"]["Visible"], 1)

    def test_client_export_adds_missing_public_root_sheets(self):
        configured_root = Category.objects.create(
            name="Configurada",
            is_active=True,
            visible_in_catalog=True,
        )
        missing_root = Category.objects.create(
            name="Nueva visible",
            is_active=True,
            visible_in_catalog=True,
        )
        configured_product = Product.objects.create(
            sku="CONFIG-001",
            name="Producto configurado",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=configured_root,
        )
        configured_product.categories.add(configured_root)
        missing_product = Product.objects.create(
            sku="MISSING-001",
            name="Producto raiz nueva",
            price=Decimal("200.00"),
            stock=1,
            is_active=True,
            category=missing_root,
        )
        missing_product.categories.add(missing_root)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo incompleto",
            is_client_download_enabled=True,
        )
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Configurada",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet.categories.add(configured_root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)

        self.assertIn("Configurada", workbook.sheetnames)
        self.assertIn("Nueva visible", workbook.sheetnames)
        self.assertEqual(stats["rows_by_sheet"]["Configurada"], 1)
        self.assertEqual(stats["rows_by_sheet"]["Nueva visible"], 1)

    def test_client_export_orders_sheets_by_public_catalog_order(self):
        abrazaderas = Category.objects.create(
            name="ABRAZADERAS",
            order=10,
            public_order=10,
            is_active=True,
            visible_in_catalog=True,
        )
        bujes = Category.objects.create(
            name="BUJES",
            order=30,
            public_order=30,
            is_active=True,
            visible_in_catalog=True,
        )
        acero = Category.objects.create(
            name="ACERO",
            order=20,
            public_order=20,
            is_active=True,
            visible_in_catalog=True,
        )

        for category, sku in (
            (bujes, "BUJE-001"),
            (acero, "ACERO-001"),
            (abrazaderas, "ABR-001"),
        ):
            product = Product.objects.create(
                sku=sku,
                name=f"Producto {sku}",
                price=Decimal("100.00"),
                stock=1,
                is_active=True,
                category=category,
            )
            product.categories.add(category)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo orden hojas",
            is_client_download_enabled=True,
        )
        sheet_bujes = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="BUJES viejo",
            order=1,
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet_bujes.categories.add(bujes)
        sheet_abrazaderas = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="ABRAZADERAS viejo",
            order=2,
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet_abrazaderas.categories.add(abrazaderas)
        sheet_acero = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="ACERO viejo",
            order=3,
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet_acero.categories.add(acero)
        for sheet in (sheet_bujes, sheet_abrazaderas, sheet_acero):
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)

        self.assertEqual(workbook.sheetnames[:3], ["ABRAZADERAS", "ACERO", "BUJES"])
        self.assertEqual(list(stats["rows_by_sheet"].keys())[:3], ["ABRAZADERAS", "ACERO", "BUJES"])

    def test_client_export_skips_public_sheet_without_scope(self):
        public_root = Category.objects.create(
            name="ABRAZADERAS",
            is_active=True,
            visible_in_catalog=True,
        )
        product = Product.objects.create(
            sku="AB-001",
            name="Abrazadera visible",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=public_root,
        )
        product.categories.add(public_root)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo con hoja huerfana",
            is_client_download_enabled=True,
        )
        orphan_sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="CAÑO TENSOR",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=False,
            sort_by="sku_asc",
        )
        CatalogExcelTemplateColumn.objects.create(
            sheet=orphan_sheet,
            key="sku",
            order=1,
            is_active=True,
        )
        CatalogExcelTemplateColumn.objects.create(
            sheet=orphan_sheet,
            key="name",
            order=2,
            is_active=True,
        )

        workbook, stats = build_catalog_workbook(template)

        self.assertNotIn("CAÑO TENSOR", workbook.sheetnames)
        self.assertIn("ABRAZADERAS", workbook.sheetnames)
        self.assertIn("CAÑO TENSOR", stats["skipped_sheets"])
        self.assertEqual(stats["rows_by_sheet"]["ABRAZADERAS"], 1)

    def test_client_export_uses_one_canonical_category_for_multi_category_products(self):
        primary_root = Category.objects.create(
            name="ABRAZADERAS",
            is_active=True,
            visible_in_catalog=True,
        )
        secondary_root = Category.objects.create(
            name="CAÑO TENSOR",
            is_active=True,
            visible_in_catalog=True,
        )
        shared_product = Product.objects.create(
            sku="SHARED-001",
            name="Producto con dos categorias",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=primary_root,
        )
        shared_product.categories.add(primary_root, secondary_root)
        secondary_product = Product.objects.create(
            sku="TENSOR-001",
            name="Producto tensor",
            price=Decimal("200.00"),
            stock=1,
            is_active=True,
            category=secondary_root,
        )
        secondary_product.categories.add(secondary_root)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo sin duplicar",
            is_client_download_enabled=True,
        )
        primary_sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="ABRAZADERAS",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=False,
            sort_by="sku_asc",
        )
        primary_sheet.categories.add(primary_root)
        secondary_sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="CAÑO TENSOR",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=False,
            sort_by="sku_asc",
        )
        secondary_sheet.categories.add(secondary_root)
        for sheet in (primary_sheet, secondary_sheet):
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
            CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)

        self.assertIn("ABRAZADERAS", workbook.sheetnames)
        self.assertIn("CAÑO TENSOR", workbook.sheetnames)
        self.assertEqual(stats["rows_by_sheet"]["ABRAZADERAS"], 1)
        self.assertEqual(stats["rows_by_sheet"]["CAÑO TENSOR"], 1)

    def test_client_export_respects_manual_product_order_per_category(self):
        root = Category.objects.create(
            name="Orden publico",
            is_active=True,
            visible_in_catalog=True,
        )
        first_product = Product.objects.create(
            sku="ORD-001",
            name="A producto alfabetico",
            price=Decimal("100.00"),
            stock=1,
            is_active=True,
            category=root,
        )
        second_product = Product.objects.create(
            sku="ORD-002",
            name="Z producto manual primero",
            price=Decimal("200.00"),
            stock=1,
            is_active=True,
            category=root,
        )
        first_product.categories.add(root)
        second_product.categories.add(root)
        CategoryProductOrder.objects.create(category=root, product=first_product, sort_order=20)
        CategoryProductOrder.objects.create(category=root, product=second_product, sort_order=10)

        template = CatalogExcelTemplate.objects.create(
            name="Catalogo ordenado",
            is_client_download_enabled=True,
        )
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Orden publico",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=False,
            sort_by="name_asc",
        )
        sheet.categories.add(root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)
        worksheet = workbook["Orden publico"]

        self.assertEqual(stats["rows_by_sheet"]["Orden publico"], 2)
        self.assertEqual(worksheet["A2"].value, "ORD-002")
        self.assertEqual(worksheet["A3"].value, "ORD-001")

    def test_public_export_skips_active_empty_category_sheet(self):
        empty_root = Category.objects.create(
            name="Vacia visible",
            is_active=True,
            visible_in_catalog=True,
        )
        template = CatalogExcelTemplate.objects.create(
            name="Catalogo vacio",
            is_client_download_enabled=True,
        )
        sheet = CatalogExcelTemplateSheet.objects.create(
            template=template,
            name="Vacia visible",
            include_header=True,
            only_active_products=True,
            only_catalog_visible=True,
            include_descendant_categories=True,
            group_by_subcategories=True,
            sort_by="sku_asc",
        )
        sheet.categories.add(empty_root)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="sku", order=1, is_active=True)
        CatalogExcelTemplateColumn.objects.create(sheet=sheet, key="name", order=2, is_active=True)

        workbook, stats = build_catalog_workbook(template)

        self.assertNotIn("Vacia visible", workbook.sheetnames)
        self.assertEqual(stats["skipped_sheets"], ["Vacia visible"])
        self.assertIn("Catalogo", workbook.sheetnames)
        self.assertEqual(stats["rows_by_sheet"]["Catalogo"], 0)
