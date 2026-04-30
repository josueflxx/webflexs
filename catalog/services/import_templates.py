from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PRODUCT_TEMPLATE_HEADERS = [
    "SKU",
    "Nombre",
    "Descripcion",
    "Proveedor",
    "Costo",
    "Precio",
    "Stock",
    "Rubro",
    "Subrubro",
    "Categorias",
    "Activo",
    "Filtro 1",
    "Filtro 2",
    "Filtro 3",
    "Filtro 4",
    "Filtro 5",
    "Diametro",
    "Ancho",
    "Largo",
    "Forma",
    "Fabricacion",
    "Material",
    "Terminacion",
    "Aplicacion",
    "Codigo OEM",
    "Codigo Referencia",
    "Ubicacion",
    "Unidad",
    "Atributos",
]


PRODUCT_TEMPLATE_EXAMPLE_ROWS = [
    [
        "EL-001",
        "ELASTICO DELANTERO 8 HOJAS MERCEDES",
        "Descripcion comercial visible para busqueda interna y catalogo.",
        "Proveedor Ejemplo",
        50000,
        75000,
        4,
        "Elasticos",
        "Mercedes Benz",
        "Elasticos; Camion",
        "SI",
        "Suspension",
        "Camion",
        "Mercedes",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "Acero",
        "",
        "Delantero",
        "",
        "REF-001",
        "Deposito A",
        "unidad",
        "Color:Negro;Notas:Ejemplo",
    ],
    [
        "ABT1280220C",
        "ABRAZADERA TREFILADA DE 1/2 X 80 X 220 CURVA",
        "Abrazadera de ejemplo para validar parseo tecnico.",
        "COTIZADOR",
        1800,
        3250.5,
        12,
        "Abrazaderas",
        "Trefiladas",
        "Abrazaderas; Suspension",
        "X",
        "Abrazadera",
        "Trefilada",
        "1/2",
        "80",
        "220",
        "1/2",
        80,
        220,
        "Curva",
        "Trefilada",
        "Acero",
        "Zincado",
        "Suspension",
        "",
        "ABT1280220C",
        "Deposito A",
        "unidad",
        "Lista:1;Origen:Plantilla",
    ],
]


ABRAZADERA_TEMPLATE_HEADERS = [
    "SKU",
    "Nombre",
    "Descripcion",
    "Proveedor",
    "Costo",
    "Precio",
    "Stock",
    "Rubro",
    "Subrubro",
    "Categorias",
    "Activo",
    "Diametro",
    "Ancho",
    "Largo",
    "Forma",
    "Fabricacion",
    "Material",
    "Terminacion",
    "Codigo Referencia",
    "Atributos",
]


ABRAZADERA_TEMPLATE_EXAMPLE_ROWS = [
    [
        "ABL1135400C",
        "ABRAZADERA LAMINADA DE 1 X 135 X 400 CURVA",
        "Ejemplo laminada. El sistema intenta completar ficha tecnica por nombre/codigo.",
        "COTIZADOR",
        0,
        12500,
        1,
        "Abrazaderas",
        "Laminadas",
        "Abrazaderas; Laminadas",
        "SI",
        "1",
        135,
        400,
        "Curva",
        "Laminada",
        "Acero",
        "Natural",
        "ABL1135400C",
        "Lista:1;Uso:Acoplado",
    ],
    [
        "ABT71665100P",
        "ABRAZADERA TREFILADA DE 7/16 X 65 X 100 PLANA",
        "Ejemplo trefilada.",
        "COTIZADOR",
        0,
        2278.54,
        1,
        "Abrazaderas",
        "Trefiladas",
        "Abrazaderas; Trefiladas",
        "SI",
        "7/16",
        65,
        100,
        "Plana",
        "Trefilada",
        "Acero",
        "Zincado",
        "ABT71665100P",
        "Lista:1",
    ],
]


SAAS_TEMPLATE_HEADERS = [
    "Nº de producto",
    "Estado",
    "Disponible para la venta",
    "Disponible para la compra",
    "Disponible para integrar otros productos",
    "Compuesto por otros productos",
    "Rubro",
    "Nombre",
    "Código",
    "Código universal de producto (UPC)",
    "Código de proveedor",
    "Stock actual",
    "Stock ideal",
    "Stock mínimo",
    "Unidad",
    "Alicuota de IVA",
    "Proveedor",
    "Costo ($)",
    "Utilidad (%)",
    "Precio ($)",
    "Precio Final ($)",
    "Controla stock",
    "Stock negativo",
    "Mostrar en tienda",
    "Nº de publicación en MercadoLibre",
    "Nº de publicación adicional en MercadoLibre",
    "Descripción",
    "Descripción para la tienda",
    "Observaciones Internas",
]


SAAS_TEMPLATE_EXAMPLE_ROWS = [
    [
        1,
        "Habilitado",
        "SI",
        "SI",
        "NO",
        "NO",
        "BUJE ARMADO",
        "BUJE DEMO SAAS",
        "SAAS-001",
        "",
        "PROV-001",
        "",
        "",
        "",
        "unidad",
        "21%",
        "MOVIGOM S.R.L.",
        1000,
        1,
        10000,
        12100,
        "no",
        "si",
        "no",
        0,
        0,
        "",
        "",
        "",
    ],
    [
        2,
        "Habilitado",
        "SI",
        "SI",
        "NO",
        "NO",
        "ABRAZADERA DE 5/8",
        "ABRAZADERA TREFILADA DE 5/8 X 80 X 220 CURVA",
        "SAAS-ABR-001",
        "",
        "ABR-001",
        "",
        "",
        "",
        "unidad",
        "21%",
        "ROCES, RICARDO ALBERTO",
        0,
        1,
        2066.12,
        2500,
        "no",
        "si",
        "no",
        0,
        0,
        "",
        "",
        "",
    ],
]


def build_import_template_filename(import_type="products"):
    if import_type == "abrazaderas":
        return "plantilla_importacion_abrazaderas_webflexs.xlsx"
    return "plantilla_importacion_productos_webflexs.xlsx"


def _style_sheet(sheet, headers):
    header_fill = PatternFill("solid", fgColor="1F2933")
    required_fill = PatternFill("solid", fgColor="FF6B3D")
    header_font = Font(color="FFFFFF", bold=True)
    help_font = Font(color="6B7280", italic=True)

    required = {"SKU", "Nombre", "Precio"}
    for cell in sheet[1]:
        cell.fill = required_fill if cell.value in required else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col_index, header in enumerate(headers, start=1):
        width = max(12, min(34, len(str(header)) + 6))
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    note_row = sheet.max_row + 2
    sheet.cell(note_row, 1, "Columnas naranja: recomendadas para alta segura.")
    sheet.cell(note_row + 1, 1, "Activo acepta SI/NO, ACTIVO/INACTIVO, X o BAJA.")
    sheet.cell(note_row + 2, 1, "Categorias multiples se separan con punto y coma: Abrazaderas; Suspension.")
    sheet.cell(note_row + 3, 1, "Atributos acepta JSON o formato Campo:Valor;Campo:Valor.")
    for row_idx in range(note_row, note_row + 4):
        sheet.cell(row_idx, 1).font = help_font


def _add_rows(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, headers)


def _add_instructions(workbook):
    sheet = workbook.create_sheet("Instrucciones")
    rows = [
        ("Uso recomendado", "Primero cargar una copia en Dry run y corregir errores por fila antes de aplicar real."),
        ("SKU", "Obligatorio. Es la clave para crear o actualizar sin duplicar."),
        ("Nombre", "Obligatorio para productos nuevos. Se usa en catalogo, busqueda y parser de abrazaderas."),
        ("Precio", "Obligatorio para productos nuevos. Acepta 12500.50, 12.500,50 o $ 12.500,50."),
        ("Costo", "Opcional. Si esta vacio no pisa el costo existente."),
        ("Stock", "Opcional en actualizacion. Para nuevo producto vacio se toma como 0."),
        ("Proveedor", "Opcional. Si viene informado, se sincroniza con proveedores."),
        ("Rubro/Subrubro", "Crean o asignan categorias. Subrubro queda como hija de Rubro cuando ambos existen."),
        ("Categorias", "Acepta varias categorias separadas por ; | , > o /."),
        ("Activo", "SI/X/ACTIVO muestra en catalogo. NO/INACTIVO/BAJA lo oculta."),
        ("Filtros", "Campos operativos de busqueda interna. No son obligatorios."),
        ("Atributos", 'Para datos extra: Color:Negro;Medida:1/2 o JSON {"Color":"Negro"}.'),
        ("Abrazaderas", "Completar Diametro, Ancho, Largo, Forma y Fabricacion mejora el resultado tecnico."),
    ]
    sheet.append(["Campo", "Regla"])
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, ["Campo", "Regla"])
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 92


def _add_values_sheet(workbook):
    sheet = workbook.create_sheet("Valores")
    rows = [
        ("Activo", "SI, NO, X, ACTIVO, INACTIVO, BAJA"),
        ("Forma abrazadera", "Curva, Plana, Semicurva"),
        ("Fabricacion", "Trefilada, Laminada, Forjada"),
        ("Terminacion", "Natural, Zincado"),
        ("Separador categorias", "Abrazaderas; Trefiladas; Suspension"),
        ("Separador atributos", "Campo:Valor;Campo:Valor"),
    ]
    sheet.append(["Tipo", "Valores recomendados"])
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, ["Tipo", "Valores recomendados"])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 70


def _add_validations(sheet, headers):
    if "Activo" in headers:
        active_col = get_column_letter(headers.index("Activo") + 1)
        validation = DataValidation(type="list", formula1='"SI,NO,X,ACTIVO,INACTIVO,BAJA"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{active_col}2:{active_col}5000")
    if "Forma" in headers:
        shape_col = get_column_letter(headers.index("Forma") + 1)
        validation = DataValidation(type="list", formula1='"Curva,Plana,Semicurva"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{shape_col}2:{shape_col}5000")
    if "Fabricacion" in headers:
        fabrication_col = get_column_letter(headers.index("Fabricacion") + 1)
        validation = DataValidation(type="list", formula1='"Trefilada,Laminada,Forjada"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{fabrication_col}2:{fabrication_col}5000")


def build_product_import_template_workbook(import_type="products"):
    workbook = Workbook()
    main = workbook.active
    main.title = "Productos"
    _add_rows(main, PRODUCT_TEMPLATE_HEADERS, PRODUCT_TEMPLATE_EXAMPLE_ROWS)
    _add_validations(main, PRODUCT_TEMPLATE_HEADERS)

    clamps = workbook.create_sheet("Abrazaderas")
    _add_rows(clamps, ABRAZADERA_TEMPLATE_HEADERS, ABRAZADERA_TEMPLATE_EXAMPLE_ROWS)
    _add_validations(clamps, ABRAZADERA_TEMPLATE_HEADERS)

    saas = workbook.create_sheet("Formato SaaS")
    _add_rows(saas, SAAS_TEMPLATE_HEADERS, SAAS_TEMPLATE_EXAMPLE_ROWS)
    _add_validations(saas, SAAS_TEMPLATE_HEADERS)

    _add_instructions(workbook)
    _add_values_sheet(workbook)

    if import_type == "abrazaderas":
        workbook.active = workbook.sheetnames.index("Abrazaderas")
    return workbook
