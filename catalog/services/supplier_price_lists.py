"""Safe, auditable supplier price-list previews and applications."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from catalog.models import (
    Product,
    ProductSupplier,
    SupplierPriceListBatch,
    SupplierPriceListRow,
)
from catalog.services.product_suppliers import (
    calculate_final_cost,
    upsert_product_supplier_offer,
)
from core.models import ImportExecution


MAPPING_FIELDS = (
    ("supplier_code", "Codigo del proveedor"),
    ("internal_sku", "SKU interno"),
    ("description", "Descripcion"),
    ("cost", "Costo de lista"),
    ("currency", "Moneda"),
    ("discount_percentage", "Descuento %"),
    ("bonus_percentage", "Bonificacion %"),
    ("tax_percentage", "Impuesto %"),
    ("is_available", "Disponible"),
    ("lead_time_days", "Demora en dias"),
    ("price_list_date", "Fecha de lista"),
)
IDENTITY_FIELDS = {"supplier_code", "internal_sku", "description"}
OPTIONAL_TERM_FIELDS = {
    "discount_percentage",
    "bonus_percentage",
    "tax_percentage",
    "is_available",
    "lead_time_days",
    "price_list_date",
}
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
MONEY_QUANTUM = Decimal("0.0001")
COORDINATE_COLUMN_PREFIX = "__excel_column__:"
COORDINATE_SCAN_ROWS = 25


def hash_uploaded_file(uploaded_file):
    digest = hashlib.sha256()
    try:
        position = uploaded_file.tell()
    except Exception:
        position = None
    uploaded_file.seek(0)
    for chunk in uploaded_file.chunks():
        digest.update(chunk)
    uploaded_file.seek(position or 0)
    return digest.hexdigest()


def _extension(file_path):
    suffix = Path(str(file_path)).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValidationError("Formato no soportado. Usa un archivo XLSX o CSV.")
    return suffix


def _json_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return None
        if value.is_integer():
            return int(value)
        return str(value)
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _header_labels(values):
    labels = []
    seen = Counter()
    for index, value in enumerate(values, start=1):
        base = re.sub(r"\s+", " ", str(value or "").strip()) or f"Columna {index}"
        seen[base] += 1
        labels.append(base if seen[base] == 1 else f"{base} [{seen[base]}]")
    return labels


def _coordinate_key(index):
    return f"{COORDINATE_COLUMN_PREFIX}{index}"


def _coordinate_columns(width):
    return [
        {
            "index": index,
            "letter": get_column_letter(index),
            "key": _coordinate_key(index),
        }
        for index in range(1, width + 1)
    ]


def _header_columns(headers):
    return [
        {
            "index": index,
            "letter": get_column_letter(index),
            "key": header,
            "header": header,
        }
        for index, header in enumerate(headers, start=1)
    ]


def _last_used_column(values):
    for index in range(len(values), 0, -1):
        if values[index - 1] not in (None, ""):
            return index
    return 0


def _coordinate_inspection(*, sheets, selected, rows, start_row, sample_limit):
    scanned_rows = []
    width = 0
    scan_limit = max(COORDINATE_SCAN_ROWS, int(sample_limit or 0))
    for row_number, values in enumerate(rows, start=start_row):
        if not any(value not in (None, "") for value in values):
            continue
        values = tuple(values)
        scanned_rows.append((row_number, values))
        width = max(width, _last_used_column(values))
        if len(scanned_rows) >= scan_limit:
            break
    if not width:
        raise ValidationError(f"No se encontraron datos desde la fila {start_row}.")
    columns = _coordinate_columns(width)
    headers = [column["key"] for column in columns]
    samples = [
        {
            column["key"]: _json_value(values[column["index"] - 1])
            if column["index"] <= len(values)
            else None
            for column in columns
        }
        for _row_number, values in scanned_rows[:sample_limit]
    ]
    return {
        "mode": "coordinates",
        "sheets": sheets,
        "sheet_name": selected,
        "data_start_row": start_row,
        "headers": headers,
        "columns": columns,
        "samples": samples,
        "sample_row_numbers": [
            row_number for row_number, _values in scanned_rows[:sample_limit]
        ],
    }


def _open_csv(file_path):
    last_error = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            handle = open(file_path, "r", encoding=encoding, newline="")
            sample = handle.read(8192)
            handle.seek(0)
            nonempty_lines = [line for line in sample.splitlines() if line.strip()][:25]
            preferred_delimiter = next(
                (
                    delimiter
                    for delimiter in (";", "\t", "|")
                    if nonempty_lines and all(delimiter in line for line in nonempty_lines)
                ),
                None,
            )
            try:
                dialect = (
                    None
                    if preferred_delimiter
                    else csv.Sniffer().sniff(sample, delimiters=",;\t|")
                )
            except csv.Error:
                dialect = None
            if preferred_delimiter or dialect is None:
                delimiter = preferred_delimiter or (
                    ";" if sample.count(";") > sample.count(",") else ","
                )

                class FallbackDialect(csv.excel):
                    pass

                FallbackDialect.delimiter = delimiter
                dialect = FallbackDialect
            return handle, dialect
        except UnicodeDecodeError as exc:
            try:
                handle.close()
            except Exception:
                pass
            last_error = exc
    raise ValidationError("No se pudo leer la codificacion del CSV.") from last_error


def inspect_source_file(
    file_path,
    *,
    sheet_name="",
    header_row=1,
    sample_limit=5,
    coordinate_mode=False,
):
    """Return sheets, unique column labels and a few JSON-safe sample rows."""
    extension = _extension(file_path)
    header_row = int(header_row or 1)
    if header_row < 1 or header_row > 500:
        raise ValidationError("La fila debe estar entre 1 y 500.")

    if extension == ".xlsx":
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError("El archivo XLSX no se pudo abrir o esta danado.") from exc
        try:
            sheets = list(workbook.sheetnames)
            selected = sheet_name if sheet_name in sheets else (sheets[0] if sheets else "")
            if not selected:
                raise ValidationError("El libro no contiene hojas.")
            worksheet = workbook[selected]
            rows = worksheet.iter_rows(min_row=header_row, values_only=True)
            if coordinate_mode:
                return _coordinate_inspection(
                    sheets=sheets,
                    selected=selected,
                    rows=rows,
                    start_row=header_row,
                    sample_limit=sample_limit,
                )
            header_values = next(rows, None)
            if not header_values:
                raise ValidationError("No se encontro una fila de encabezado.")
            headers = _header_labels(header_values)
            samples = []
            for values in rows:
                if not any(value not in (None, "") for value in values):
                    continue
                samples.append(
                    {header: _json_value(value) for header, value in zip(headers, values)}
                )
                if len(samples) >= sample_limit:
                    break
            return {
                "mode": "headers",
                "sheets": sheets,
                "sheet_name": selected,
                "headers": headers,
                "columns": _header_columns(headers),
                "samples": samples,
            }
        finally:
            workbook.close()

    handle, dialect = _open_csv(file_path)
    try:
        rows = csv.reader(handle, dialect)
        for _ in range(header_row - 1):
            next(rows, None)
        if coordinate_mode:
            return _coordinate_inspection(
                sheets=["CSV"],
                selected="CSV",
                rows=rows,
                start_row=header_row,
                sample_limit=sample_limit,
            )
        header_values = next(rows, None)
        if not header_values:
            raise ValidationError("No se encontro una fila de encabezado en el CSV.")
        headers = _header_labels(header_values)
        samples = []
        for values in rows:
            if not any(str(value or "").strip() for value in values):
                continue
            samples.append({header: _json_value(value) for header, value in zip(headers, values)})
            if len(samples) >= sample_limit:
                break
        return {
            "mode": "headers",
            "sheets": ["CSV"],
            "sheet_name": "CSV",
            "headers": headers,
            "columns": _header_columns(headers),
            "samples": samples,
        }
    finally:
        handle.close()


def iter_source_rows(file_path, *, sheet_name="", header_row=1, coordinate_mode=False):
    extension = _extension(file_path)
    inspection = inspect_source_file(
        file_path,
        sheet_name=sheet_name,
        header_row=header_row,
        sample_limit=0,
        coordinate_mode=coordinate_mode,
    )
    headers = inspection["headers"]
    max_rows = int(getattr(settings, "SUPPLIER_PRICE_LIST_MAX_ROWS", 100000))

    if extension == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[inspection["sheet_name"]]
            first_data_row = int(header_row) if coordinate_mode else int(header_row) + 1
            for offset, values in enumerate(
                worksheet.iter_rows(min_row=first_data_row, values_only=True),
                start=first_data_row,
            ):
                if offset - first_data_row + 1 > max_rows:
                    raise ValidationError(f"La lista supera el limite de {max_rows} filas.")
                if not any(value not in (None, "") for value in values):
                    continue
                yield offset, {
                    header: _json_value(value) for header, value in zip(headers, values)
                }
        finally:
            workbook.close()
        return

    handle, dialect = _open_csv(file_path)
    try:
        rows = csv.reader(handle, dialect)
        rows_to_skip = int(header_row) - 1 if coordinate_mode else int(header_row)
        for _ in range(rows_to_skip):
            next(rows, None)
        first_data_row = int(header_row) if coordinate_mode else int(header_row) + 1
        for offset, values in enumerate(rows, start=first_data_row):
            if offset - first_data_row + 1 > max_rows:
                raise ValidationError(f"La lista supera el limite de {max_rows} filas.")
            if not any(str(value or "").strip() for value in values):
                continue
            yield offset, {header: _json_value(value) for header, value in zip(headers, values)}
    finally:
        handle.close()


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _normalized_key(value):
    value = unicodedata.normalize("NFKD", _text(value))
    value = "".join(char for char in value if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", value.upper()).strip()


def _decimal(value, field_label, *, allow_blank=False):
    text = _text(value)
    if not text:
        if allow_blank:
            return None
        raise ValidationError(f"Falta {field_label}.")
    negative_parentheses = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9,.-]", "", text.strip("()"))
    if not cleaned:
        raise ValidationError(f"{field_label} no es un numero valido.")
    if "," in cleaned and "." in cleaned:
        decimal_separator = "," if cleaned.rfind(",") > cleaned.rfind(".") else "."
        thousands_separator = "." if decimal_separator == "," else ","
        cleaned = cleaned.replace(thousands_separator, "").replace(decimal_separator, ".")
    elif "," in cleaned:
        parts = cleaned.split(",")
        cleaned = "".join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else "".join(parts[:-1]) + "." + parts[-1]
    elif cleaned.count(".") > 1:
        parts = cleaned.split(".")
        cleaned = "".join(parts) if len(parts[-1]) == 3 else "".join(parts[:-1]) + "." + parts[-1]
    elif "." in cleaned:
        left, right = cleaned.rsplit(".", 1)
        if len(right) == 3 and left not in ("0", "-0"):
            cleaned = left + right
    try:
        result = Decimal(cleaned)
    except (InvalidOperation, ValueError) as exc:
        raise ValidationError(f"{field_label} no es un numero valido.") from exc
    if negative_parentheses:
        result = -result
    if result < 0:
        raise ValidationError(f"{field_label} no puede ser negativo.")
    return result.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)


def _percentage(value, label):
    result = _decimal(value, label, allow_blank=True)
    if result is not None and result > 100:
        raise ValidationError(f"{label} no puede superar 100%.")
    return result


def _boolean(value):
    if isinstance(value, bool):
        return value
    normalized = _normalized_key(value)
    if normalized in {"SI", "S", "YES", "Y", "TRUE", "1", "DISPONIBLE", "ACTIVO"}:
        return True
    if normalized in {"NO", "N", "FALSE", "0", "NO DISPONIBLE", "INACTIVO"}:
        return False
    raise ValidationError("El valor de disponibilidad no es reconocible.")


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            pass
    text = _text(value)
    parsed = parse_date(text)
    if parsed:
        return parsed
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValidationError("La fecha de lista no es valida.")


def validate_mapping(mapping):
    mapping = {str(key): str(value or "").strip() for key, value in (mapping or {}).items()}
    if not mapping.get("supplier_code"):
        raise ValidationError("Debes seleccionar la columna del codigo del proveedor.")
    if not mapping.get("cost"):
        raise ValidationError("Debes seleccionar la columna del precio o costo.")
    if mapping["supplier_code"] == mapping["cost"]:
        raise ValidationError("El codigo y el precio deben usar columnas diferentes.")
    return {key: value for key, value in mapping.items() if value}


def mapping_uses_coordinates(mapping):
    return any(
        str(value or "").startswith(COORDINATE_COLUMN_PREFIX)
        for value in (mapping or {}).values()
    )


def _mapped_value(raw, mapping, field):
    column = mapping.get(field)
    return raw.get(column) if column else None


def _provided(raw, mapping, field):
    return bool(mapping.get(field) and _text(_mapped_value(raw, mapping, field)))


def _product_indexes(supplier):
    products = list(Product.objects.only("id", "sku", "name", "cost", "supplier_ref_id"))
    skus = defaultdict(list)
    descriptions = defaultdict(list)
    for product in products:
        skus[_normalized_key(product.sku)].append(product)
        descriptions[_normalized_key(product.name)].append(product)
    offers = list(
        ProductSupplier.objects.filter(supplier=supplier)
        .select_related("product")
        .order_by("id")
    )
    codes = {offer.normalized_supplier_code: offer for offer in offers if offer.normalized_supplier_code}
    offers_by_product = {offer.product_id: offer for offer in offers}
    return skus, descriptions, codes, offers_by_product, offers


def _offer_snapshot(offer):
    if not offer:
        return None
    return {
        "id": offer.pk,
        "supplier_code": offer.supplier_code,
        "supplier_description": offer.supplier_description,
        "current_cost": str(offer.current_cost),
        "currency": offer.currency,
        "discount_percentage": str(offer.discount_percentage),
        "bonus_percentage": str(offer.bonus_percentage),
        "tax_percentage": str(offer.tax_percentage),
        "minimum_purchase_quantity": offer.minimum_purchase_quantity,
        "is_available": offer.is_available,
        "lead_time_days": offer.lead_time_days,
        "price_list_date": offer.price_list_date.isoformat() if offer.price_list_date else None,
        "status": offer.status,
        "is_preferred": offer.is_preferred,
    }


def _match_product(*, supplier_code, internal_sku, description, skus, descriptions, codes):
    warnings = []
    code_offer = codes.get(ProductSupplier.normalize_supplier_code(supplier_code)) if supplier_code else None
    sku_matches = skus.get(_normalized_key(internal_sku), []) if internal_sku else []
    sku_product = sku_matches[0] if len(sku_matches) == 1 else None
    if len(sku_matches) > 1:
        warnings.append("El SKU interno coincide con mas de un producto.")
    if code_offer and sku_product and code_offer.product_id != sku_product.id:
        return None, None, "conflict", 0, warnings + [
            "El codigo del proveedor y el SKU interno apuntan a productos distintos."
        ]
    if code_offer:
        return code_offer.product, code_offer, "supplier_code_exact", 100, warnings
    if sku_product:
        return sku_product, None, "internal_sku_exact", 100, warnings
    description_matches = descriptions.get(_normalized_key(description), []) if description else []
    if len(description_matches) == 1:
        return description_matches[0], None, "description_exact_review", 80, warnings + [
            "Coincidencia solo por descripcion: requiere decision manual."
        ]
    if len(description_matches) > 1:
        warnings.append("La descripcion coincide con mas de un producto.")
    return None, None, "unmatched", 0, warnings


def _row_signature_payload(row):
    return {
        "row_number": row.row_number,
        "row_type": row.row_type,
        "raw_data": row.raw_data,
        "normalized_data": row.normalized_data,
        "matched_product_id": row.matched_product_id,
        "product_supplier_id": row.product_supplier_id,
        "supplier_code": row.supplier_code,
        "supplier_description": row.supplier_description,
        "match_method": row.match_method,
        "match_confidence": row.match_confidence,
        "change_type": row.change_type,
        "previous_cost": str(row.previous_cost) if row.previous_cost is not None else None,
        "proposed_cost": str(row.proposed_cost) if row.proposed_cost is not None else None,
        "proposed_final_cost": (
            str(row.proposed_final_cost) if row.proposed_final_cost is not None else None
        ),
        "difference_amount": (
            str(row.difference_amount) if row.difference_amount is not None else None
        ),
        "difference_percentage": (
            str(row.difference_percentage) if row.difference_percentage is not None else None
        ),
        "currency": row.currency,
        "discount_percentage": str(row.discount_percentage),
        "bonus_percentage": str(row.bonus_percentage),
        "tax_percentage": str(row.tax_percentage),
        "is_available": row.is_available,
        "lead_time_days": row.lead_time_days,
        "price_list_date": row.price_list_date.isoformat() if row.price_list_date else None,
        "warnings": row.warnings,
    }


def calculate_preview_signature(batch):
    digest = hashlib.sha256()
    digest.update(batch.file_sha256.encode("ascii"))
    digest.update(json.dumps(batch.column_mapping, sort_keys=True).encode("utf-8"))
    for row in batch.rows.order_by("row_number").iterator(chunk_size=1000):
        digest.update(
            json.dumps(_row_signature_payload(row), ensure_ascii=False, sort_keys=True).encode("utf-8")
        )
    return digest.hexdigest()


def _summary(batch):
    changes = Counter(batch.rows.values_list("change_type", flat=True))
    decisions = Counter(batch.rows.values_list("decision", flat=True))
    return {
        "total_rows": batch.rows.count(),
        "source_rows": batch.rows.filter(row_type=SupplierPriceListRow.TYPE_SOURCE).count(),
        "absent_rows": batch.rows.filter(row_type=SupplierPriceListRow.TYPE_ABSENT).count(),
        "changes": dict(changes),
        "decisions": dict(decisions),
    }


def generate_supplier_price_list_preview(batch, *, mapping, sheet_name="", header_row=1):
    """Replace a non-applied batch preview with normalized, matched comparison rows."""
    if batch.status == SupplierPriceListBatch.STATUS_APPLIED:
        raise ValidationError("Una lista ya aplicada no puede regenerarse.")
    if SupplierPriceListBatch.objects.filter(
        supplier=batch.supplier,
        file_sha256=batch.file_sha256,
        status=SupplierPriceListBatch.STATUS_APPLIED,
    ).exclude(pk=batch.pk).exists():
        raise ValidationError("Este mismo archivo ya fue aplicado para el proveedor.")

    mapping = validate_mapping(mapping)
    coordinate_mode = mapping_uses_coordinates(mapping)
    inspection = inspect_source_file(
        batch.source_file.path,
        sheet_name=sheet_name,
        header_row=header_row,
        sample_limit=0,
        coordinate_mode=coordinate_mode,
    )
    missing_columns = sorted(set(mapping.values()) - set(inspection["headers"]))
    if missing_columns:
        raise ValidationError("Columnas no encontradas: " + ", ".join(missing_columns))

    skus, descriptions, codes, offers_by_product, supplier_offers = _product_indexes(batch.supplier)
    rows_to_create = []
    seen_codes = set()
    seen_products = set()
    source_codes = set()

    for row_number, raw in iter_source_rows(
        batch.source_file.path,
        sheet_name=inspection["sheet_name"],
        header_row=header_row,
        coordinate_mode=coordinate_mode,
    ):
        warnings = []
        supplier_code = _text(_mapped_value(raw, mapping, "supplier_code"))
        internal_sku = _text(_mapped_value(raw, mapping, "internal_sku"))
        description = _text(_mapped_value(raw, mapping, "description"))
        normalized_code = ProductSupplier.normalize_supplier_code(supplier_code)
        if normalized_code:
            source_codes.add(normalized_code)

        normalized = {
            "supplier_code": supplier_code,
            "internal_sku": internal_sku,
            "description": description,
            "provided_fields": [],
        }
        change_type = SupplierPriceListRow.CHANGE_INVALID
        decision = SupplierPriceListRow.DECISION_SKIP
        matched_product = None
        product_supplier = None
        match_method = ""
        confidence = 0
        cost = None
        previous_cost = None
        difference = None
        difference_percentage = None

        try:
            cost = _decimal(_mapped_value(raw, mapping, "cost"), "el costo")
            if cost == 0:
                warnings.append("El costo es cero y debe revisarse antes de aplicar.")
            matched_product, product_supplier, match_method, confidence, match_warnings = _match_product(
                supplier_code=supplier_code,
                internal_sku=internal_sku,
                description=description,
                skus=skus,
                descriptions=descriptions,
                codes=codes,
            )
            warnings.extend(match_warnings)
            if normalized_code and normalized_code in seen_codes:
                match_method = "duplicate_source_code"
                confidence = 0
                warnings.append("Codigo de proveedor repetido dentro del archivo.")
            if normalized_code:
                seen_codes.add(normalized_code)

            if matched_product and matched_product.pk in seen_products:
                match_method = "duplicate_source_product"
                confidence = 0
                warnings.append("Otra fila del archivo ya identifica al mismo producto.")
            if matched_product:
                seen_products.add(matched_product.pk)

            if matched_product:
                product_supplier = product_supplier or offers_by_product.get(matched_product.id)
            previous_cost = product_supplier.current_cost if product_supplier else None
            normalized["previous_offer"] = _offer_snapshot(product_supplier)

            term_values = {}
            for field in OPTIONAL_TERM_FIELDS:
                if _provided(raw, mapping, field):
                    normalized["provided_fields"].append(field)
                    if field in {"discount_percentage", "bonus_percentage", "tax_percentage"}:
                        term_values[field] = _percentage(_mapped_value(raw, mapping, field), field)
                    elif field == "is_available":
                        term_values[field] = _boolean(_mapped_value(raw, mapping, field))
                    elif field == "lead_time_days":
                        parsed_days = _decimal(_mapped_value(raw, mapping, field), field)
                        if parsed_days != parsed_days.to_integral_value():
                            raise ValidationError("La demora debe ser una cantidad entera de dias.")
                        term_values[field] = int(parsed_days)
                    elif field == "price_list_date":
                        term_values[field] = _date(_mapped_value(raw, mapping, field))

            currency = _text(_mapped_value(raw, mapping, "currency")).upper() or batch.default_currency
            if currency in {"$", "PESO", "PESOS"}:
                currency = ProductSupplier.CURRENCY_ARS
            elif currency in {"U$S", "US$", "DOLAR", "DOLARES"}:
                currency = ProductSupplier.CURRENCY_USD
            if currency not in dict(ProductSupplier.CURRENCY_CHOICES):
                raise ValidationError(f"Moneda no soportada: {currency}.")

            discount = term_values.get(
                "discount_percentage",
                product_supplier.discount_percentage if product_supplier else Decimal("0"),
            )
            bonus = term_values.get(
                "bonus_percentage",
                product_supplier.bonus_percentage if product_supplier else Decimal("0"),
            )
            tax = term_values.get(
                "tax_percentage",
                product_supplier.tax_percentage if product_supplier else Decimal("0"),
            )
            available = term_values.get(
                "is_available", product_supplier.is_available if product_supplier else True
            )
            lead_days = term_values.get(
                "lead_time_days", product_supplier.lead_time_days if product_supplier else 0
            )
            list_date = term_values.get(
                "price_list_date", product_supplier.price_list_date if product_supplier else None
            )
            final_cost = calculate_final_cost(cost, discount, bonus, tax)

            normalized.update(
                {
                    "cost": str(cost),
                    "currency": currency,
                    "discount_percentage": str(discount),
                    "bonus_percentage": str(bonus),
                    "tax_percentage": str(tax),
                    "is_available": available,
                    "lead_time_days": lead_days,
                    "price_list_date": list_date.isoformat() if list_date else None,
                }
            )

            if match_method in {"duplicate_source_code", "duplicate_source_product"}:
                change_type = SupplierPriceListRow.CHANGE_REVIEW
                decision = SupplierPriceListRow.DECISION_REVIEW
            elif not matched_product:
                change_type = (
                    SupplierPriceListRow.CHANGE_REVIEW
                    if match_method in {"conflict", "duplicate_source_code"}
                    else SupplierPriceListRow.CHANGE_UNMATCHED
                )
                decision = SupplierPriceListRow.DECISION_REVIEW
            else:
                if previous_cost is None:
                    change_type = SupplierPriceListRow.CHANGE_NEW_RELATION
                elif cost > previous_cost:
                    change_type = SupplierPriceListRow.CHANGE_INCREASE
                elif cost < previous_cost:
                    change_type = SupplierPriceListRow.CHANGE_DECREASE
                else:
                    change_type = SupplierPriceListRow.CHANGE_UNCHANGED
                decision = (
                    SupplierPriceListRow.DECISION_REVIEW
                    if match_method == "description_exact_review" or cost == 0
                    else SupplierPriceListRow.DECISION_APPLY
                )
                if previous_cost is not None:
                    difference = (cost - previous_cost).quantize(MONEY_QUANTUM)
                    if previous_cost != 0:
                        difference_percentage = (
                            difference / previous_cost * Decimal("100")
                        ).quantize(MONEY_QUANTUM)

            row_kwargs = {
                "currency": currency,
                "discount_percentage": discount,
                "bonus_percentage": bonus,
                "tax_percentage": tax,
                "is_available": available,
                "lead_time_days": lead_days,
                "price_list_date": list_date,
                "proposed_final_cost": final_cost,
            }
        except ValidationError as exc:
            warnings.extend(exc.messages)
            row_kwargs = {
                "currency": batch.default_currency,
                "discount_percentage": Decimal("0"),
                "bonus_percentage": Decimal("0"),
                "tax_percentage": Decimal("0"),
                "is_available": True,
                "lead_time_days": 0,
                "price_list_date": None,
                "proposed_final_cost": None,
            }

        rows_to_create.append(
            SupplierPriceListRow(
                batch=batch,
                row_number=row_number,
                raw_data=raw,
                normalized_data=normalized,
                supplier_code=supplier_code,
                supplier_description=description,
                matched_product=matched_product,
                product_supplier=product_supplier,
                match_method=match_method,
                match_confidence=confidence,
                change_type=change_type,
                previous_cost=previous_cost,
                proposed_cost=cost,
                difference_amount=difference,
                difference_percentage=difference_percentage,
                warnings=warnings,
                decision=decision,
                **row_kwargs,
            )
        )

    if mapping.get("supplier_code"):
        next_row = max((row.row_number for row in rows_to_create), default=int(header_row)) + 1
        for offer in supplier_offers:
            if not offer.normalized_supplier_code or offer.normalized_supplier_code in source_codes:
                continue
            rows_to_create.append(
                SupplierPriceListRow(
                    batch=batch,
                    row_number=next_row,
                    row_type=SupplierPriceListRow.TYPE_ABSENT,
                    normalized_data={"supplier_code": offer.supplier_code},
                    supplier_code=offer.supplier_code,
                    supplier_description=offer.supplier_description,
                    matched_product=offer.product,
                    product_supplier=offer,
                    match_method="absent_from_source",
                    match_confidence=100,
                    change_type=SupplierPriceListRow.CHANGE_ABSENT,
                    previous_cost=offer.current_cost,
                    currency=offer.currency,
                    discount_percentage=offer.discount_percentage,
                    bonus_percentage=offer.bonus_percentage,
                    tax_percentage=offer.tax_percentage,
                    is_available=offer.is_available,
                    lead_time_days=offer.lead_time_days,
                    price_list_date=offer.price_list_date,
                    warnings=["La oferta existente no aparece en esta lista; no se dara de baja automaticamente."],
                    decision=SupplierPriceListRow.DECISION_REVIEW,
                )
            )
            next_row += 1

    if not rows_to_create:
        raise ValidationError("La lista no contiene filas de datos.")

    with transaction.atomic():
        locked = SupplierPriceListBatch.objects.select_for_update().get(pk=batch.pk)
        locked.rows.all().delete()
        SupplierPriceListRow.objects.bulk_create(rows_to_create, batch_size=1000)
        locked.sheet_name = inspection["sheet_name"]
        locked.header_row = int(header_row)
        locked.column_mapping = mapping
        locked.status = SupplierPriceListBatch.STATUS_PREVIEWED
        locked.previewed_at = timezone.now()
        locked.error_message = ""
        locked.save(
            update_fields=[
                "sheet_name", "header_row", "column_mapping", "status", "previewed_at",
                "error_message", "updated_at",
            ]
        )
        locked.preview_signature = calculate_preview_signature(locked)
        locked.summary = _summary(locked)
        locked.save(update_fields=["preview_signature", "summary", "updated_at"])
    batch.refresh_from_db()
    return batch


def allowed_decisions_for_row(row):
    if str(row.match_method or "").startswith("duplicate_source"):
        return {SupplierPriceListRow.DECISION_SKIP, SupplierPriceListRow.DECISION_REVIEW}
    if row.row_type == SupplierPriceListRow.TYPE_ABSENT or row.change_type in {
        SupplierPriceListRow.CHANGE_INVALID,
        SupplierPriceListRow.CHANGE_UNMATCHED,
    }:
        return {SupplierPriceListRow.DECISION_SKIP, SupplierPriceListRow.DECISION_REVIEW}
    if not row.matched_product_id or row.proposed_cost is None:
        return {SupplierPriceListRow.DECISION_SKIP, SupplierPriceListRow.DECISION_REVIEW}
    return {
        SupplierPriceListRow.DECISION_APPLY,
        SupplierPriceListRow.DECISION_SKIP,
        SupplierPriceListRow.DECISION_REVIEW,
    }


def update_row_decisions(batch, decisions):
    if batch.status != SupplierPriceListBatch.STATUS_PREVIEWED:
        raise ValidationError("Solo se pueden decidir filas de una previsualizacion activa.")
    rows = {row.pk: row for row in batch.rows.filter(pk__in=decisions.keys())}
    changed = []
    for row_id, decision in decisions.items():
        row = rows.get(int(row_id))
        if not row:
            continue
        if decision not in allowed_decisions_for_row(row):
            raise ValidationError(f"La decision no es valida para la fila {row.row_number}.")
        if row.decision != decision:
            row.decision = decision
            changed.append(row)
    if changed:
        now = timezone.now()
        for row in changed:
            row.updated_at = now
        SupplierPriceListRow.objects.bulk_update(changed, ["decision", "updated_at"])
    batch.summary = _summary(batch)
    batch.save(update_fields=["summary", "updated_at"])
    return len(changed)


def apply_supplier_price_list(batch, *, user, pricing_mode=None):
    """Apply approved rows atomically and append the existing immutable cost history."""
    try:
        with transaction.atomic():
            locked = SupplierPriceListBatch.objects.select_for_update().select_related("supplier", "company").get(
                pk=batch.pk
            )
            if locked.status != SupplierPriceListBatch.STATUS_PREVIEWED:
                raise ValidationError("El lote no esta listo para aplicar.")
            if SupplierPriceListBatch.objects.filter(
                supplier=locked.supplier,
                file_sha256=locked.file_sha256,
                status=SupplierPriceListBatch.STATUS_APPLIED,
            ).exclude(pk=locked.pk).exists():
                raise ValidationError("Este mismo archivo ya fue aplicado para el proveedor.")
            if locked.preview_signature != calculate_preview_signature(locked):
                raise ValidationError("La previsualizacion cambio; debes regenerarla antes de aplicar.")
            review_count = locked.rows.filter(decision=SupplierPriceListRow.DECISION_REVIEW).count()
            if review_count:
                raise ValidationError(
                    f"Quedan {review_count} filas en revision. Decide aplicar u omitir cada una."
                )
            pricing_mode = str(
                pricing_mode or locked.pricing_mode or SupplierPriceListBatch.PRICING_COST_ONLY
            ).strip()
            if pricing_mode not in dict(SupplierPriceListBatch.PRICING_MODE_CHOICES):
                raise ValidationError("El modo de actualizacion de precios no es valido.")

            execution = ImportExecution.objects.create(
                user=user if getattr(user, "is_authenticated", False) else None,
                company=locked.company,
                import_type="supplier_price_list",
                file_name=locked.original_filename,
                dry_run=False,
                status=ImportExecution.STATUS_PROCESSING,
                supplier=locked.supplier,
                supplier_name=locked.supplier.name,
                metrics=locked.summary,
                result_summary={"batch_id": locked.pk},
            )
            created_count = 0
            updated_count = 0
            history_count = 0
            repriced_count = 0
            pricing_skipped_count = 0
            rows = list(
                locked.rows.filter(decision=SupplierPriceListRow.DECISION_APPLY)
                .select_related("matched_product", "product_supplier")
                .order_by("row_number")
            )
            for row in rows:
                if (
                    row.row_type != SupplierPriceListRow.TYPE_SOURCE
                    or not row.matched_product_id
                    or SupplierPriceListRow.DECISION_APPLY not in allowed_decisions_for_row(row)
                ):
                    raise ValidationError(f"La fila {row.row_number} no se puede aplicar.")
                product = row.matched_product
                previous_product_cost = Decimal(product.cost or 0)
                previous_sale_price = Decimal(product.price or 0)
                previous_supplier_ref_id = product.supplier_ref_id
                previous_supplier_name = product.supplier
                current_offer = (
                    ProductSupplier.objects.select_for_update()
                    .filter(product=product, supplier=locked.supplier)
                    .first()
                )
                current_cost = current_offer.current_cost if current_offer else None
                if current_cost != row.previous_cost:
                    raise ValidationError(
                        f"El costo vigente de la fila {row.row_number} cambio desde la previsualizacion."
                    )
                expected_offer = (row.normalized_data or {}).get("previous_offer")
                if _offer_snapshot(current_offer) != expected_offer:
                    raise ValidationError(
                        f"Las condiciones comerciales de la fila {row.row_number} cambiaron desde la previsualizacion."
                    )
                offer, history = upsert_product_supplier_offer(
                    product=product,
                    supplier=locked.supplier,
                    current_cost=row.proposed_cost,
                    currency=row.currency,
                    supplier_code=row.supplier_code or (current_offer.supplier_code if current_offer else ""),
                    supplier_description=row.supplier_description or (
                        current_offer.supplier_description if current_offer else ""
                    ),
                    discount_percentage=row.discount_percentage,
                    bonus_percentage=row.bonus_percentage,
                    tax_percentage=row.tax_percentage,
                    minimum_purchase_quantity=(
                        current_offer.minimum_purchase_quantity if current_offer else 1
                    ),
                    is_available=row.is_available,
                    lead_time_days=row.lead_time_days,
                    price_list_date=row.price_list_date,
                    source="supplier_price_list",
                    source_file=locked.original_filename,
                    source_row=row.row_number,
                    import_execution=execution,
                    changed_by=user,
                    reason=f"Lista de proveedor #{locked.pk}",
                    is_preferred=current_offer.is_preferred if current_offer else None,
                    status=current_offer.status if current_offer else ProductSupplier.STATUS_ACTIVE,
                    match_confidence=row.match_confidence,
                    match_method=row.match_method,
                    notes=current_offer.notes if current_offer else "",
                )
                if current_offer:
                    updated_count += 1
                else:
                    created_count += 1
                if history:
                    history_count += 1
                new_sale_price = previous_sale_price
                pricing_warning = ""
                if pricing_mode == SupplierPriceListBatch.PRICING_PRESERVE_MARGIN and offer.is_preferred:
                    if previous_product_cost > 0:
                        multiplier = previous_sale_price / previous_product_cost
                        new_sale_price = (Decimal(product.cost or 0) * multiplier).quantize(
                            Decimal("0.01"),
                            rounding=ROUND_HALF_UP,
                        )
                        if product.price != new_sale_price:
                            product.price = new_sale_price
                            product.save(update_fields=["price", "updated_at"])
                            repriced_count += 1
                    else:
                        pricing_warning = (
                            "No se recalculo la venta porque el costo anterior del producto era cero."
                        )
                        pricing_skipped_count += 1

                normalized_data = dict(row.normalized_data or {})
                normalized_data["applied_offer"] = _offer_snapshot(offer)
                normalized_data["pricing_snapshot"] = {
                    "mode": pricing_mode,
                    "previous_product_cost": str(previous_product_cost),
                    "new_product_cost": str(product.cost),
                    "previous_sale_price": str(previous_sale_price),
                    "new_sale_price": str(new_sale_price),
                    "previous_supplier_ref_id": previous_supplier_ref_id,
                    "previous_supplier_name": previous_supplier_name,
                }
                warnings = list(row.warnings or [])
                if pricing_warning and pricing_warning not in warnings:
                    warnings.append(pricing_warning)
                row.product_supplier = offer
                row.cost_history = history
                row.applied = True
                row.normalized_data = normalized_data
                row.warnings = warnings
                row.save(
                    update_fields=[
                        "product_supplier", "cost_history", "applied", "normalized_data",
                        "warnings", "updated_at",
                    ]
                )

            locked.status = SupplierPriceListBatch.STATUS_APPLIED
            locked.pricing_mode = pricing_mode
            locked.applied_by = user
            locked.applied_at = timezone.now()
            locked.summary = {
                **_summary(locked),
                "created_offers": created_count,
                "updated_offers": updated_count,
                "cost_history_rows": history_count,
                "pricing_mode": pricing_mode,
                "repriced_products": repriced_count,
                "pricing_skipped": pricing_skipped_count,
            }
            locked.import_execution = execution
            locked.error_message = ""
            locked.save(
                update_fields=[
                    "status", "pricing_mode", "applied_by", "applied_at", "summary",
                    "import_execution", "error_message", "updated_at",
                ]
            )
            execution.status = ImportExecution.STATUS_COMPLETED
            execution.created_count = created_count
            execution.updated_count = updated_count
            execution.error_count = 0
            execution.metrics = locked.summary
            execution.result_summary = {"batch_id": locked.pk, **locked.summary}
            execution.finished_at = timezone.now()
            execution.save(
                update_fields=[
                    "status", "created_count", "updated_count", "error_count", "metrics",
                    "result_summary", "finished_at",
                ]
            )
        batch.refresh_from_db()
        return batch
    except Exception as exc:
        SupplierPriceListBatch.objects.filter(pk=batch.pk).exclude(
            status=SupplierPriceListBatch.STATUS_APPLIED
        ).update(error_message=str(exc)[:2000])
        raise


def rollback_supplier_price_list(batch, *, user):
    """Safely reverse an applied supplier batch when its current values still match."""
    with transaction.atomic():
        locked = (
            SupplierPriceListBatch.objects.select_for_update()
            .select_related("supplier", "company")
            .get(pk=batch.pk)
        )
        if locked.status != SupplierPriceListBatch.STATUS_APPLIED:
            raise ValidationError("Solo se puede revertir una lista aplicada.")

        reverted = 0
        conflicts = 0
        rows = list(
            locked.rows.filter(applied=True)
            .select_related("matched_product", "product_supplier")
            .order_by("-row_number")
        )
        for row in rows:
            product = Product.objects.select_for_update().get(pk=row.matched_product_id)
            current_offer = (
                ProductSupplier.objects.select_for_update()
                .filter(product=product, supplier=locked.supplier)
                .first()
            )
            warnings = list(row.warnings or [])
            normalized_data = dict(row.normalized_data or {})
            snapshot = dict(normalized_data.get("pricing_snapshot") or {})
            expected_offer = normalized_data.get("applied_offer")
            expected_price = snapshot.get("new_sale_price")
            if not current_offer or current_offer.current_cost != row.proposed_cost:
                warnings.append(
                    "No se revirtio: el costo del proveedor cambio despues de aplicar el lote."
                )
                row.warnings = warnings
                row.save(update_fields=["warnings", "updated_at"])
                conflicts += 1
                continue
            if expected_offer and _offer_snapshot(current_offer) != expected_offer:
                warnings.append(
                    "No se revirtio: las condiciones del proveedor cambiaron despues de aplicar."
                )
                row.warnings = warnings
                row.save(update_fields=["warnings", "updated_at"])
                conflicts += 1
                continue
            if expected_price not in (None, "") and Decimal(product.price) != Decimal(expected_price):
                warnings.append(
                    "No se revirtio: el precio de venta cambio despues de aplicar el lote."
                )
                row.warnings = warnings
                row.save(update_fields=["warnings", "updated_at"])
                conflicts += 1
                continue

            previous_offer = dict(normalized_data.get("previous_offer") or {})
            if row.previous_cost is None:
                current_offer.status = ProductSupplier.STATUS_INACTIVE
                current_offer.is_preferred = False
                current_offer.notes = (
                    f"{current_offer.notes}\nRevertido por lista de proveedor #{locked.pk}"
                ).strip()
                current_offer.save(update_fields=["status", "is_preferred", "notes", "updated_at"])
                product.cost = Decimal(snapshot.get("previous_product_cost") or product.cost)
                product.price = Decimal(snapshot.get("previous_sale_price") or product.price)
                product.supplier_ref_id = snapshot.get("previous_supplier_ref_id")
                product.supplier = str(snapshot.get("previous_supplier_name") or "")
                product.save(
                    update_fields=["cost", "price", "supplier_ref", "supplier", "updated_at"]
                )
            else:
                upsert_product_supplier_offer(
                    product=product,
                    supplier=locked.supplier,
                    current_cost=row.previous_cost,
                    currency=previous_offer.get("currency") or current_offer.currency,
                    supplier_code=previous_offer.get("supplier_code") or current_offer.supplier_code,
                    supplier_description=(
                        previous_offer.get("supplier_description")
                        or current_offer.supplier_description
                    ),
                    discount_percentage=previous_offer.get("discount_percentage", 0),
                    bonus_percentage=previous_offer.get("bonus_percentage", 0),
                    tax_percentage=previous_offer.get("tax_percentage", 0),
                    minimum_purchase_quantity=previous_offer.get("minimum_purchase_quantity", 1),
                    is_available=previous_offer.get("is_available", True),
                    lead_time_days=previous_offer.get("lead_time_days", 0),
                    price_list_date=parse_date(previous_offer.get("price_list_date") or ""),
                    source="supplier_price_list_rollback",
                    source_file=locked.original_filename,
                    source_row=row.row_number,
                    import_execution=locked.import_execution,
                    changed_by=user,
                    reason=f"Reversion lista de proveedor #{locked.pk}",
                    is_preferred=previous_offer.get("is_preferred", current_offer.is_preferred),
                    status=previous_offer.get("status") or current_offer.status,
                    match_confidence=row.match_confidence,
                    match_method="supplier_price_list_rollback",
                    notes=current_offer.notes,
                )
                if snapshot:
                    product.refresh_from_db(fields=["price"])
                    product.price = Decimal(snapshot.get("previous_sale_price") or product.price)
                    product.save(update_fields=["price", "updated_at"])

            normalized_data["rollback"] = {
                "at": timezone.now().isoformat(),
                "by": getattr(user, "username", ""),
            }
            row.normalized_data = normalized_data
            row.applied = False
            row.save(update_fields=["normalized_data", "applied", "updated_at"])
            reverted += 1

        locked.status = (
            SupplierPriceListBatch.STATUS_ROLLED_BACK
            if not conflicts
            else SupplierPriceListBatch.STATUS_ROLLBACK_PARTIAL
        )
        locked.rolled_back_by = user
        locked.rolled_back_at = timezone.now()
        locked.summary = {
            **(locked.summary or {}),
            "rollback_reverted": reverted,
            "rollback_conflicts": conflicts,
        }
        locked.save(
            update_fields=[
                "status", "rolled_back_by", "rolled_back_at", "summary", "updated_at",
            ]
        )
    batch.refresh_from_db()
    return batch


def report_rows(batch):
    for row in batch.rows.select_related("matched_product").order_by("row_number").iterator():
        yield {
            "fila": row.row_number,
            "tipo": row.get_row_type_display(),
            "codigo_proveedor": row.supplier_code,
            "descripcion": row.supplier_description,
            "sku_identificado": row.matched_product.sku if row.matched_product_id else "",
            "producto_identificado": row.matched_product.name if row.matched_product_id else "",
            "metodo": row.match_method,
            "confianza": row.match_confidence,
            "cambio": row.get_change_type_display(),
            "costo_anterior": row.previous_cost,
            "costo_propuesto": row.proposed_cost,
            "diferencia": row.difference_amount,
            "diferencia_porcentaje": row.difference_percentage,
            "moneda": row.currency,
            "decision": row.get_decision_display(),
            "aplicado": "Si" if row.applied else "No",
            "advertencias": " | ".join(row.warnings or []),
        }
